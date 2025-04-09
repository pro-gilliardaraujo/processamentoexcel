"""
Script para processamento completo de dados de monitoramento de colhedoras.
Lê arquivos TXT na pasta raiz, processa-os e gera arquivos Excel com planilhas auxiliares prontas.
"""

import pandas as pd
import numpy as np
import os
import glob
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Constantes
COLUNAS_REMOVER = [
    'Justificativa Corte Base Desligado',
    'Latitude',
    'Longitude',
    'Regional',
    'Tipo de Equipamento',
    'Unidade',
    'Centro de Custo'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento', 'Apertura do Rolo', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo', 'Corte Base Automatico/Manual',
    'Descricao Equipamento', 'Estado', 'Estado Operacional', 'Esteira Ligada',
    'Field Cruiser', 'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Implemento Ligado', 'Motor Ligado', 'Operacao', 'Operador', 'Pressao de Corte',
    'RPM Extrator', 'RPM Motor', 'RTK (Piloto Automatico)', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parada com Motor Ligado',
    'Horas Produtivas'
]

# Valores a serem filtrados
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO"]

def processar_arquivo_base(caminho_txt):
    """
    Processa o arquivo TXT e retorna o DataFrame com as transformações necessárias.
    
    Args:
        caminho_txt (str): Caminho do arquivo TXT de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    try:
        # Leitura do arquivo
        df = pd.read_csv(caminho_txt, sep=';', encoding='utf-8')
        print(f"Arquivo lido com sucesso! Total de linhas: {len(df)}")
        
        # Processamento de data e hora
        df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
        df = df.drop(columns=['Data/Hora'])
        
        # Conversão e cálculo de diferenças de hora
        df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S')
        df['Diferença_Hora'] = df['Hora'].diff().dt.total_seconds() / 3600
        df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: max(x, 0))
        
        # Nova regra: se Diferença_Hora > 0.50, então 0
        df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else x)
        
        # Cálculos adicionais
        RPM_MINIMO = 300  # Definindo constante para RPM mínimo
        df['Parada com Motor Ligado'] = ((df['Velocidade'] == 0) & (df['RPM Motor'] >= RPM_MINIMO)).astype(int)
        df['Horas Produtivas'] = df.apply(
            lambda row: row['Diferença_Hora'] if row['Grupo Operacao'] == 'Produtiva' else 0,
            axis=1
        )
        
        # Conversão de colunas binárias para valores numéricos (garantindo que sejam números)
        for col in ['Esteira Ligada', 'Motor Ligado', 'Field Cruiser', 'RTK (Piloto Automatico)', 'Implemento Ligado']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        
        # Limpeza e organização das colunas
        df = df.drop(columns=COLUNAS_REMOVER, errors='ignore')
        
        # Garantir que todas as colunas desejadas existam
        for col in COLUNAS_DESEJADAS:
            if col not in df.columns:
                df[col] = np.nan
        
        # Reorganizar as colunas na ordem desejada
        colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
        df = df[colunas_existentes]
        
        return df
        
    except FileNotFoundError:
        print(f"Erro: Arquivo não encontrado em {caminho_txt}")
        return None
    except pd.errors.EmptyDataError:
        print("Erro: O arquivo está vazio")
        return None
    except Exception as e:
        print(f"Erro inesperado: {str(e)}")
        return None

def calcular_base_calculo(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Horas totais
        horas_totais = dados_filtrados['Diferença_Hora'].sum()
        
        # Horas elevador (Esteira Ligada = 1)
        horas_elevador = dados_filtrados[dados_filtrados['Esteira Ligada'] == 1]['Diferença_Hora'].sum()
        
        # Percentual horas elevador
        percent_elevador = (horas_elevador / horas_totais * 100) if horas_totais > 0 else 0
        
        # RTK (Piloto Automático = 1 e Field Cruiser = 1)
        rtk = dados_filtrados[(dados_filtrados['RTK (Piloto Automatico)'] == 1) & 
                             (dados_filtrados['Field Cruiser'] == 1)]['Diferença_Hora'].sum()
        
        # Horas Produtivas
        horas_produtivas = dados_filtrados['Horas Produtivas'].sum()
        
        # % Utilização RTK
        utilizacao_rtk = (rtk / horas_produtivas * 100) if horas_produtivas > 0 else 0
        
        # Motor Ligado
        motor_ligado = dados_filtrados[dados_filtrados['Motor Ligado'] == 1]['Diferença_Hora'].sum()
        
        # % Eficiência Elevador
        eficiencia_elevador = (horas_elevador / motor_ligado * 100) if motor_ligado > 0 else 0
        
        # Parado com Motor Ligado
        parado_motor_ligado = dados_filtrados[dados_filtrados['Parada com Motor Ligado'] == 1]['Diferença_Hora'].sum()
        
        # % Parado com motor ligado
        percent_parado_motor = (parado_motor_ligado / motor_ligado * 100) if motor_ligado > 0 else 0
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas elevador': horas_elevador,
            '%': percent_elevador,
            'RTK': rtk,
            'Horas Produtivas': horas_produtivas,
            '% Utilização RTK': utilizacao_rtk,
            'Motor Ligado': motor_ligado,
            '% Eficiência Elevador': eficiencia_elevador,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor
        })
    
    return pd.DataFrame(resultados)

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento
    """
    # Filtramos os dados excluindo os operadores da lista
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df_filtrado['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        total_horas = dados_equip['Diferença_Hora'].sum()
        
        # Calcular horas de manutenção
        manutencao = dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum()
        
        # A disponibilidade mecânica é o percentual de tempo fora de manutenção
        disp_mecanica = (total_horas - manutencao) / total_horas if total_horas > 0 else 0
        
        resultados.append({
            'Frota': equipamento,
            'Disponibilidade': disp_mecanica
        })
    
    return pd.DataFrame(resultados)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Esta função NÃO aplica qualquer filtro de operador.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        total_horas = dados_equip['Diferença_Hora'].sum()
        
        # Calcular a diferença para 24 horas
        diferenca_24h = 24 - total_horas
        
        # Se a diferença for negativa, considera-se 0 (mais de 24h registradas)
        diferenca_24h = max(diferenca_24h, 0)
        
        resultados.append({
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        })
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica(base_calculo):
    """
    Calcula a eficiência energética por operador.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador
    """
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Eficiência Energética = horas elevador / motor ligado
        horas_elevador_sum = dados_op['Horas elevador'].sum()
        motor_ligado_sum = dados_op['Motor Ligado'].sum()
        
        eficiencia = horas_elevador_sum / motor_ligado_sum if motor_ligado_sum > 0 else 0
        
        resultados.append({
            'Operador': operador,
            'Eficiência': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_hora_elevador(base_calculo):
    """
    Calcula as horas de elevador por operador.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Horas de elevador por operador
    """
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Somar horas de elevador
        horas_elevador_sum = dados_op['Horas elevador'].sum()
        
        resultados.append({
            'Operador': operador,
            'Horas': horas_elevador_sum
        })
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso(base_calculo):
    """
    Calcula o percentual de motor ocioso por operador.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador
    """
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Motor Ocioso = Parado Com Motor Ligado / Motor Ligado
        parado_motor_sum = dados_op['Parado Com Motor Ligado'].sum()
        motor_ligado_sum = dados_op['Motor Ligado'].sum()
        
        percentual = parado_motor_sum / motor_ligado_sum if motor_ligado_sum > 0 else 0
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def calcular_uso_gps(base_calculo):
    """
    Calcula o percentual de uso de GPS por operador.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador
    """
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Uso GPS = RTK / Implemento Ligado
        rtk_sum = dados_op['RTK'].sum()
        horas_produtivas_sum = dados_op['Horas Produtivas'].sum()
        
        percentual = rtk_sum / horas_produtivas_sum if horas_produtivas_sum > 0 else 0
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def criar_excel_com_planilhas(df_base, base_calculo, disp_mecanica, eficiencia_energetica, 
                             hora_elevador, motor_ocioso, uso_gps, horas_por_frota, caminho_saida):
    """
    Cria um arquivo Excel com todas as planilhas auxiliares.
    
    Args:
        df_base (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
        disp_mecanica (DataFrame): Disponibilidade mecânica
        eficiencia_energetica (DataFrame): Eficiência energética
        hora_elevador (DataFrame): Horas de elevador
        motor_ocioso (DataFrame): Motor ocioso
        uso_gps (DataFrame): Uso GPS
        horas_por_frota (DataFrame): Horas totais registradas por frota
        caminho_saida (str): Caminho do arquivo Excel de saída
    """
    writer = pd.ExcelWriter(caminho_saida, engine='openpyxl')
    
    # Salvar cada DataFrame em uma planilha separada
    df_base.to_excel(writer, sheet_name='BASE', index=False)
    base_calculo.to_excel(writer, sheet_name='Base Calculo', index=False)
    
    # Planilhas auxiliares (formatadas conforme necessário)
    disp_mecanica.to_excel(writer, sheet_name='1_Disponibilidade Mecânica', index=False)
    eficiencia_energetica.to_excel(writer, sheet_name='2_Eficiência Energética', index=False)
    hora_elevador.to_excel(writer, sheet_name='3_Hora Elevador', index=False)
    motor_ocioso.to_excel(writer, sheet_name='4_Motor Ocioso', index=False)
    uso_gps.to_excel(writer, sheet_name='5_Uso GPS', index=False)
    horas_por_frota.to_excel(writer, sheet_name='Horas por Frota', index=False)
    
    # Aplicar formatação nas planilhas
    workbook = writer.book
    
    # Formatar planilha de Disponibilidade Mecânica
    worksheet = workbook['1_Disponibilidade Mecânica']
    for row in range(2, worksheet.max_row + 1):  # Começando da linha 2 (ignorando cabeçalho)
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Disponibilidade)
        cell.number_format = '0.00%'  # Formato de porcentagem
    
    # Formatar planilha de Eficiência Energética
    worksheet = workbook['2_Eficiência Energética']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Eficiência)
        cell.number_format = '0.00%'  # Formato de porcentagem
    
    # Formatar planilha de Motor Ocioso
    worksheet = workbook['4_Motor Ocioso']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
        cell.number_format = '0.00%'  # Formato de porcentagem
    
    # Formatar planilha de Uso GPS
    worksheet = workbook['5_Uso GPS']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
        cell.number_format = '0.00%'  # Formato de porcentagem
    
    # Formatar planilha de Horas por Frota
    worksheet = workbook['Horas por Frota']
    for row in range(2, worksheet.max_row + 1):
        # Coluna B (Horas Registradas)
        cell_b = worksheet.cell(row=row, column=2)
        cell_b.number_format = '0.00'  # Formato decimal com 2 casas
        
        # Coluna C (Diferença para 24h)
        cell_c = worksheet.cell(row=row, column=3)
        cell_c.number_format = '0.00'  # Formato decimal com 2 casas
    
    writer.close()
    print(f"Arquivo Excel salvo com sucesso em {caminho_saida}")

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT na pasta raiz do script.
    """
    # Obter o diretório onde está o script
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    if diretorio_atual == '':
        diretorio_atual = '.'
    
    # Encontrar todos os arquivos TXT
    arquivos_txt = glob.glob(os.path.join(diretorio_atual, "*.txt"))
    
    if not arquivos_txt:
        print("Nenhum arquivo TXT encontrado na pasta!")
        return
    
    print(f"Encontrados {len(arquivos_txt)} arquivos TXT para processar.")
    
    # Processar cada arquivo
    for arquivo_txt in arquivos_txt:
        nome_base = os.path.splitext(os.path.basename(arquivo_txt))[0]
        arquivo_saida = os.path.join(diretorio_atual, nome_base + ".xlsx")
        
        print(f"\nProcessando arquivo: {nome_base}.txt")
        
        # Processar o arquivo base
        df_base = processar_arquivo_base(arquivo_txt)
        if df_base is None:
            print(f"Erro ao processar {nome_base}.txt. Pulando para o próximo arquivo.")
            continue
        
        # Calcular a Base Calculo
        base_calculo = calcular_base_calculo(df_base)
        
        # Calcular as métricas auxiliares
        disp_mecanica = calcular_disponibilidade_mecanica(df_base)
        eficiencia_energetica = calcular_eficiencia_energetica(base_calculo)
        hora_elevador = calcular_hora_elevador(base_calculo)
        motor_ocioso = calcular_motor_ocioso(base_calculo)
        uso_gps = calcular_uso_gps(base_calculo)
        horas_por_frota = calcular_horas_por_frota(df_base)
        
        # Criar o arquivo Excel com todas as planilhas
        criar_excel_com_planilhas(
            df_base, base_calculo, disp_mecanica, eficiencia_energetica,
            hora_elevador, motor_ocioso, uso_gps, horas_por_frota, arquivo_saida
        )
        
        print(f"Arquivo {nome_base}.xlsx gerado com sucesso!")

if __name__ == "__main__":
    print("Iniciando processamento de arquivos...")
    processar_todos_arquivos()
    print("\nProcessamento concluído!") 