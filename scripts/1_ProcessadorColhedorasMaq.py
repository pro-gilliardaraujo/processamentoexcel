import pandas as pd
import numpy as np
import os
import glob
import zipfile
import tempfile
import shutil
import json
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from functools import reduce
import requests

# Configurações
processCsv = False  # Altere para True quando quiser processar arquivos CSV

# Configurações de produção por frente - CONFIGURE ANTES DE EXECUTAR
TONELADAS_FRENTE_03 = 882.92   # Toneladas Frente03
TONELADAS_FRENTE_04 = 1418.86   # Toneladas Frente04  
TONELADAS_FRENTE_08 = 1486.56   # Toneladas Frente08
TONELADAS_FRENTE_ZIRLENO = 0000  # Toneladas FrenteZirleno

# Mapeamento de frentes para toneladas
TONELADAS_POR_FRENTE = {
    'Frente03': TONELADAS_FRENTE_03,
    'Frente04': TONELADAS_FRENTE_04,
    'Frente08': TONELADAS_FRENTE_08,
    'FreteZirleno': TONELADAS_FRENTE_ZIRLENO,
    'Zirleno': TONELADAS_FRENTE_ZIRLENO,  # Alias para compatibilidade
}


# Configurações Supabase
SUPABASE_URL = "https://kjlwqezxzqjfhacmjhbh.supabase.co"
SUPABASE_ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtqbHdxZXp4enFqZmhhY21qaGJoIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Mzc1NDY3OTgsImV4cCI6MjA1MzEyMjc5OH0.bB58zKtOtIyd4pZl-lycUQFVyHsQK_6Rxe2XtYta_cY"

# Configurações de manobras
tempoMinimoManobras = 15  # Tempo mínimo para considerar uma manobra válida (em segundos)
velocidadeMinimaManobras = 0  # Velocidade mínima para considerar uma manobra válida (em km/h)

# Configurações de validação de dados
TEMPO_MINIMO_VALIDO = 0.001  # 3.6 segundos em horas - tempo mínimo para considerar um registro válido

# Constantes
COLUNAS_REMOVER = [
    'Justificativa Corte Base Desligado',
    'Regional',
    'Tipo de Equipamento',
    'Unidade',
    'Centro de Custo',
    'Trabalhando em File',
    'Trabalhando Frente Dividida',
    'Trabalhando em Fila',
    'Apertura do Rolo',
    'Codigo da Operacao',
    'Codigo Frente (digitada)',
    'Corporativo',
    'Descricao Equipamento',
    'Fazenda',
    'Zona',
    'Talhao'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento',
    'Apertura do Rolo', 'Codigo da Operacao', 'Codigo Frente (digitada)', 'Corporativo',
    'Corte Base Automatico/Manual', 'Descricao Equipamento', 'Estado', 'Estado Operacional',
    # Grupo de colunas solicitado imediatamente após Estado Operacional
    'Operador', 'Grupo Operacao', 'Operacao', 'Diferença_Hora', 'Horimetro',
    # Demais colunas
    'Esteira Ligada', 'Field Cruiser', 'Grupo Equipamento/Frente',
    'Implemento Ligado', 'Motor Ligado', 'Pressao de Corte',
    'Corte Base Automatico/Manual', 'RPM Extrator', 'RPM Motor',
    'RTK (Piloto Automatico)', 'Fazenda', 'Zona', 'Talhao', 'Velocidade',
    'Parada com Motor Ligado', 'Latitude', 'Longitude'
]

# Remover colunas duplicadas entre remover e desejadas
COLUNAS_DESEJADAS = [col for col in COLUNAS_DESEJADAS if col not in COLUNAS_REMOVER]

# Valores a serem filtrados
OPERADORES_EXCLUIR = []

# Mapeamento de valores booleanos para 1/0
MAPEAMENTO_BOOLEANO = {
    'VERDADEIRO': 1, 'FALSO': 0,
    'TRUE': 1, 'FALSE': 0,
    'LIGADO': 1, 'DESLIGADO': 0,
    True: 1, False: 0,
    1: 1, 0: 0
}

def carregar_config_calculos():
    """
    Retorna as configurações de cálculos embutidas diretamente no código.
    As configurações são as mesmas definidas no arquivo calculos_config.json.
    """
    # Configurações embutidas diretamente no código
    config = {
        "CD": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "8490 - LAVAGEM",
                    "MANUTENCAO",
                    "LAVAGEM",
                    "INST CONFIG TECNOL EMBARCADAS",
                    "1055 - MANOBRA"
                ],
                "grupos_operacao_excluidos": [
                    "Manutenção", 
                    "Inaptidão"
                ]
            },
            "equipamentos_excluidos": []
        },
        "TT": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "9016 - ENCH SISTEMA FREIO",
                    "6340 - BASCULANDO  TRANSBORDAGEM",
                    "9024 - DESATOLAMENTO",
                    "MANUTENCAO",
                    "MANUTENÇÃO",
                    "INST CONFIG TECNOL EMBARCADAS",
                    "1055 - MANOBRA"
                ],
                "grupos_operacao_excluidos": [
                    "Manutenção", 
                    "Inaptidão"
                ]
            },
            "equipamentos_excluidos": []
        }
    }
    
    print("Usando configurações embutidas no código, ignorando o arquivo calculos_config.json")
    return config

def calcular_motor_ocioso_novo(df):
    """
    Calcula o tempo de motor ocioso de acordo com as novas regras:
    1. Intervalo é fechado quando encontra 'Parada com Motor Ligado = 0' com duração > 1 minuto
    2. Soma sequências de 'Parada com Motor Ligado = 1' com 'Parada com Motor Ligado = 0' < 1 minuto
    3. Se o total > 1 minuto, subtrai 1 minuto; se menor, descarta o intervalo
    
    Args:
        df (DataFrame): DataFrame com os dados de operação
        
    Returns:
        DataFrame: DataFrame com a coluna 'Motor Ocioso' atualizada
    """
    # Converter a coluna de diferença para minutos
    df['Diferença_Minutos'] = df['Diferença_Hora'] * 60
    
    # Inicializar colunas
    df['Motor Ocioso'] = 0
    df['Em_Intervalo'] = False
    df['Soma_Intervalo'] = 0
    
    # Variáveis para controle do intervalo atual
    em_intervalo = False
    soma_intervalo = 0
    inicio_intervalo = None
    
    # Iterar sobre as linhas do DataFrame
    for i in range(len(df)):
        parada_motor = df.iloc[i]['Parada com Motor Ligado']
        diferenca = df.iloc[i]['Diferença_Minutos']
        
        # Se não estamos em um intervalo
        if not em_intervalo:
            # Se encontrar Parada com Motor Ligado = 1, inicia novo intervalo
            if parada_motor == 1:
                em_intervalo = True
                soma_intervalo = diferenca
                inicio_intervalo = i
                df.at[i, 'Em_Intervalo'] = True
                df.at[i, 'Soma_Intervalo'] = soma_intervalo
        
        # Se estamos em um intervalo
        else:
            # Se encontrar Parada com Motor Ligado = 0
            if parada_motor == 0:
                # Se a duração for > 1 minuto, fecha o intervalo
                if diferenca > 1:
                    # Se o total acumulado > 1 minuto, subtrai 1 minuto
                    if soma_intervalo > 1:
                        tempo_ocioso = soma_intervalo - 1
                        # Atribui o tempo ocioso à primeira linha do intervalo
                        # IMPORTANTE: Converter de minutos para horas antes de atribuir
                        df.at[inicio_intervalo, 'Motor Ocioso'] = tempo_ocioso / 60.0  # CORREÇÃO: Dividir por 60 para converter minutos em horas
                    
                    # Reseta o intervalo
                    em_intervalo = False
                    soma_intervalo = 0
                    inicio_intervalo = None
                else:
                    # Se <= 1 minuto, soma ao intervalo atual
                    soma_intervalo += diferenca
                    df.at[i, 'Em_Intervalo'] = True
                    df.at[i, 'Soma_Intervalo'] = soma_intervalo
            
            # Se encontrar Parada com Motor Ligado = 1
            else:
                soma_intervalo += diferenca
                df.at[i, 'Em_Intervalo'] = True
                df.at[i, 'Soma_Intervalo'] = soma_intervalo
    
    # Tratar último intervalo aberto, se houver
    if em_intervalo and soma_intervalo > 1:
        tempo_ocioso = soma_intervalo - 1
        # CORREÇÃO: Converter de minutos para horas antes de atribuir
        df.at[inicio_intervalo, 'Motor Ocioso'] = tempo_ocioso / 60.0  # Dividir por 60 para converter minutos em horas
    
        # Remover colunas auxiliares
    df = df.drop(['Diferença_Minutos', 'Em_Intervalo', 'Soma_Intervalo'], axis=1)
    
    # Garantir que o tempo ocioso nunca seja maior que o tempo de motor ligado para cada registro
    for i in range(len(df)):
        if df.iloc[i]['Motor Ocioso'] > 0:
            # Se o motor estiver ligado, limitar o tempo ocioso ao tempo de motor ligado
            if df.iloc[i]['Motor Ligado'] == 1:
                tempo_hora = df.iloc[i]['Diferença_Hora']
                if df.iloc[i]['Motor Ocioso'] > tempo_hora:
                    df.at[i, 'Motor Ocioso'] = tempo_hora
            else:
                # Se o motor não estiver ligado, o tempo ocioso deve ser zero
                df.at[i, 'Motor Ocioso'] = 0
    
    return df

def processar_arquivo_base(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo (TXT ou CSV são tratados da mesma forma se usarem separador ';')
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio (apenas cabeçalhos sem dados)
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                # Retornar o DataFrame vazio mas com as colunas, em vez de None
                # Garantir que todas as colunas desejadas existam
                for col in COLUNAS_DESEJADAS:
                    if col not in df.columns:
                        df[col] = np.nan
                # Reorganizar as colunas na ordem desejada
                colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Verificar se 'Data/Hora' existe e processá-la
            if 'Data/Hora' in df.columns:
                # Processamento de data e hora
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
                
                # Conversão e cálculo de diferenças de hora
                df_hora_temp = pd.to_datetime(df['Hora'], format='%H:%M:%S')
                df['Diferença_Hora'] = df_hora_temp.diff().dt.total_seconds() / 3600
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: max(x, 0))
                
                # Nova regra: se Diferença_Hora > 0.50, então 0
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else x)
                
                # Filtro adicional: remover registros com Diferença_Hora muito pequena
                # Usando constante global para tempo mínimo válido
                registros_antes_filtro = len(df)
                
                df_filtrado = df[
                    (df['Diferença_Hora'] >= TEMPO_MINIMO_VALIDO) |
                    (df['Diferença_Hora'] == 0)  # Manter zeros criados pela regra dos 0.50
                ]
                
                registros_depois_filtro = len(df_filtrado)
                registros_removidos_filtro = registros_antes_filtro - registros_depois_filtro
                
                if registros_removidos_filtro > 0:
                    print(f"Filtro de validação aplicado:")
                    print(f"  Registros originais: {registros_antes_filtro}")
                    print(f"  Registros após filtro: {registros_depois_filtro}")
                    print(f"  Registros removidos: {registros_removidos_filtro} ({registros_removidos_filtro/registros_antes_filtro*100:.1f}%)")
                    print(f"  Critério: Diferença_Hora >= {TEMPO_MINIMO_VALIDO} horas ({TEMPO_MINIMO_VALIDO*3600:.1f} segundos) ou = 0")
                    df = df_filtrado
                
                # Manter apenas a parte da hora como string no formato HH:MM:SS
                df['Hora'] = df_hora_temp.dt.strftime('%H:%M:%S')
            
            # Cálculos adicionais
            RPM_MINIMO = 300  # Definindo constante para RPM mínimo

            # Carregar configurações
            config = carregar_config_calculos()
            operacoes_excluidas = config['CD']['motor_ocioso']['operacoes_excluidas']
            grupos_operacao_excluidos = config['CD']['motor_ocioso']['grupos_operacao_excluidos']

            # Filtrar dados para cálculo de motor ocioso
            df_motor_ocioso = df[
                ~df['Operacao'].isin(operacoes_excluidas) & 
                ~df['Grupo Operacao'].isin(grupos_operacao_excluidos)
            ]

            # Calcular Parada com Motor Ligado usando dados filtrados
            df['Parada com Motor Ligado'] = 0  # Inicializa com 0
            df.loc[df_motor_ocioso.index, 'Parada com Motor Ligado'] = (
                (df_motor_ocioso['Velocidade'] == 0) & 
                (df_motor_ocioso['RPM Motor'] >= RPM_MINIMO)
            ).astype(int)
            
            # Calcular Motor Ocioso usando método correto (filtros + intervalos sequenciais)
            df = calcular_motor_ocioso_correto(df)
            
            # Horas Produtivas não são mais utilizadas neste fluxo; remover se existir
            if 'Horas Produtivas' in df.columns:
                df = df.drop(columns=['Horas Produtivas'])
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    # Se chegou aqui, todas as tentativas de codificação falharam
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado.
    Calcula médias diárias considerando os dias efetivos de trabalho de cada operador.
    
    Cálculos principais:
    - Horas totais: CORRIGIDO - usa Diferença_Hora (tempo de apontamento)
    - Motor Ligado: CORRIGIDO - usa horímetro inicial e final quando disponível
    - Horas elevador: CORRIGIDO - usa horímetro inicial e final quando disponível
      Condições: Grupo Operacao == "Produtiva" AND Pressao de Corte >= 400 AND Velocidade > 0
    - RTK: CORRIGIDO - usa horímetro inicial e final quando disponível
      Condições: Grupo Operacao == "Produtiva" AND Pressao de Corte >= 400 AND Velocidade > 0 AND RTK = 1
    - Horas Produtivas: CORRIGIDO - usa horímetro inicial e final quando disponível
    - Parado Com Motor Ligado: MÉTODO AVANÇADO - soma da coluna Motor Ocioso, que usa o cálculo com intervalos
    - Manutenção e outras operações: usa Diferença_Hora (tempo de apontamento)
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Detectar número de dias totais nos dados (apenas para informação)
    dias_unicos_total = df['Data'].nunique() if 'Data' in df.columns else 1
    print(f"Detectados {dias_unicos_total} dias distintos na base de dados.")
    
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    def calcular_tempo_por_horimetro(dados_filtrados, condicao_filtro=None):
        """
        Calcula tempo baseado no horímetro inicial e final.
        Se horímetro não estiver disponível, usa soma de Diferença_Hora como fallback.
        
        Args:
            dados_filtrados: DataFrame com os dados filtrados
            condicao_filtro: Condição adicional para filtrar os dados (opcional)
        
        Returns:
            float: Tempo calculado em horas
        """
        # Se não há dados, retornar 0
        if len(dados_filtrados) == 0:
            return 0.0
        
        # Aplicar condição de filtro se fornecida
        if condicao_filtro is not None:
            dados_para_calculo = dados_filtrados[condicao_filtro]
        else:
            dados_para_calculo = dados_filtrados
        
        # Se não há dados após filtro, retornar 0
        if len(dados_para_calculo) == 0:
            return 0.0
        
        # Verificar se coluna Horimetro existe e tem dados válidos
        if 'Horimetro' in dados_para_calculo.columns and not dados_para_calculo['Horimetro'].isna().all():
            # Usar método do horímetro
            try:
                # Converter horímetro para numérico se necessário
                horimetro_values = pd.to_numeric(dados_para_calculo['Horimetro'], errors='coerce')
                
                # Remover valores nulos
                horimetro_values = horimetro_values.dropna()
                
                if len(horimetro_values) >= 2:
                    # Calcular diferença entre último e primeiro horímetro
                    horimetro_inicial = horimetro_values.iloc[0]
                    horimetro_final = horimetro_values.iloc[-1]
                    tempo_horimetro = horimetro_final - horimetro_inicial
                    
                    # Validar se o resultado é positivo e razoável
                    if tempo_horimetro > 0 and tempo_horimetro < 24:  # Máximo 24 horas por dia
                        print(f"Usando horímetro: {horimetro_inicial:.2f} -> {horimetro_final:.2f} = {tempo_horimetro:.4f}h")
                        return tempo_horimetro
                    else:
                        print(f"Horímetro inválido ({tempo_horimetro:.4f}h), usando fallback")
                else:
                    print("Dados insuficientes no horímetro, usando fallback")
            except Exception as e:
                print(f"Erro ao processar horímetro: {e}, usando fallback")
        
        # Fallback: usar soma de Diferença_Hora
        tempo_diferenca = dados_para_calculo['Diferença_Hora'].sum()
        print(f"Usando soma de Diferença_Hora: {tempo_diferenca:.4f}h")
        return tempo_diferenca
    
    def calcular_tempo_por_diferenca_hora(dados_filtrados, condicao_filtro=None):
        """
        Calcula tempo baseado APENAS na soma de Diferença_Hora (tempo de apontamento).
        Usado para horas totais, manutenção e outras operações.
        
        Args:
            dados_filtrados: DataFrame com os dados filtrados
            condicao_filtro: Condição adicional para filtrar os dados (opcional)
        
        Returns:
            float: Tempo calculado em horas
        """
        # Se não há dados, retornar 0
        if len(dados_filtrados) == 0:
            return 0.0
        
        # Aplicar condição de filtro se fornecida
        if condicao_filtro is not None:
            dados_para_calculo = dados_filtrados[condicao_filtro]
        else:
            dados_para_calculo = dados_filtrados
        
        # Se não há dados após filtro, retornar 0
        if len(dados_para_calculo) == 0:
            return 0.0
        
        # Sempre usar soma de Diferença_Hora
        tempo_diferenca = dados_para_calculo['Diferença_Hora'].sum()
        print(f"Usando tempo de apontamento (Diferença_Hora): {tempo_diferenca:.4f}h")
        return tempo_diferenca
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Determinar o número de dias efetivos para este operador
        dias_operador = dados_filtrados['Data'].nunique() if 'Data' in dados_filtrados.columns else 1
        
        print(f"\n=== Calculando métricas para {operador} em {equipamento} ===")
        
        # CORRIGIDO: Horas totais - usar Diferença_Hora (tempo de apontamento)
        horas_totais = calcular_tempo_por_diferenca_hora(dados_filtrados)
        if dias_operador > 1:
            horas_totais = horas_totais / dias_operador
        
        # Motor Ligado - usar horímetro quando disponível
        motor_ligado = calcular_tempo_por_horimetro(
            dados_filtrados, 
            dados_filtrados['Motor Ligado'] == 1
        )
        if dias_operador > 1:
            motor_ligado = motor_ligado / dias_operador
        
        # Horas elevador - usar horímetro quando disponível
        horas_elevador = calcular_tempo_por_horimetro(
            dados_filtrados,
            (dados_filtrados['Grupo Operacao'] == 'Produtiva') & 
            (dados_filtrados['Pressao de Corte'] >= 400) &
            (dados_filtrados['Velocidade'] > 0)
        )
        if dias_operador > 1:
            horas_elevador = horas_elevador / dias_operador
        
        # Percentual horas elevador (em decimal 0-1)
        percent_elevador = calcular_porcentagem(horas_elevador, horas_totais)
        
        # RTK - usar horímetro quando disponível
        rtk = calcular_tempo_por_horimetro(
            dados_filtrados,
            (dados_filtrados['Grupo Operacao'] == 'Produtiva') &
            (dados_filtrados['Pressao de Corte'] >= 400) &
            (dados_filtrados['Velocidade'] > 0) &
            (dados_filtrados['RTK (Piloto Automatico)'] == 1)
        )
        if dias_operador > 1:
            rtk = rtk / dias_operador
        
        # Horas Produtivas - usar horímetro quando disponível (mesmos filtros das horas elevador)
        horas_produtivas = calcular_tempo_por_horimetro(
            dados_filtrados,
            (dados_filtrados['Grupo Operacao'] == 'Produtiva') &
            (dados_filtrados['Pressao de Corte'] >= 400) &
            (dados_filtrados['Velocidade'] > 0)
        )
        if dias_operador > 1:
            horas_produtivas = horas_produtivas / dias_operador
        
        # % Utilização RTK (em decimal 0-1)
        utilizacao_rtk = calcular_porcentagem(rtk, horas_produtivas)
        
        # % Eficiência Elevador (em decimal 0-1)
        eficiencia_elevador = calcular_porcentagem(horas_elevador, motor_ligado)
        
        # NOVO MÉTODO: Parado com Motor Ligado - usando o valor calculado pela função calcular_motor_ocioso_novo
        # A coluna 'Motor Ocioso' contém o tempo ocioso após aplicar a lógica de intervalos e tolerância
        parado_motor_ligado = dados_filtrados['Motor Ocioso'].sum()
        if dias_operador > 1:
            parado_motor_ligado = parado_motor_ligado / dias_operador
        
        # % Parado com motor ligado (em decimal 0-1)
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        # Debug para verificar os valores
        print(f"Horas totais (apontamento): {horas_totais:.6f}")
        print(f"Motor Ligado (horímetro): {motor_ligado:.6f}")
        print(f"Horas elevador (horímetro): {horas_elevador:.6f}")
        print(f"Horas Produtivas (horímetro): {horas_produtivas:.6f}")
        print(f"RTK (horímetro): {rtk:.6f}")
        print(f"Parado com Motor Ligado (método avançado): {parado_motor_ligado:.6f}")
        print(f"% Parado com motor ligado: {percent_parado_motor:.6f}")
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas elevador': horas_elevador,
            '%': percent_elevador,
            'RTK': rtk,
            'Horas Produtivas': horas_produtivas,
            '% Utilização RTK': utilizacao_rtk,
            'Motor Ligado': motor_ligado,
            '% Eficiência Elevador': eficiencia_elevador,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor
        })
    
    # Criar DataFrame com os resultados
    base_calculo = pd.DataFrame(resultados)
    
    # Adicionar coluna Frente extraída
    if not base_calculo.empty:
        base_calculo['Frente'] = base_calculo['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    return base_calculo

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento.
    Fórmula: (Total Geral - Manutenção) / Total Geral
    CORRIGIDO: 
    - Total Geral: usa Diferença_Hora (tempo de apontamento)
    - Manutenção: usa Diferença_Hora (tempo de apontamento)
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento
    """
    # Filtramos os dados excluindo os operadores da lista
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    def calcular_tempo_por_diferenca_hora_equipamento(dados_equip, condicao_filtro=None):
        """
        Calcula tempo baseado APENAS na soma de Diferença_Hora (tempo de apontamento).
        Usado para total geral e manutenção.
        """
        # Se não há dados, retornar 0
        if len(dados_equip) == 0:
            return 0.0
        
        # Aplicar condição de filtro se fornecida
        if condicao_filtro is not None:
            dados_para_calculo = dados_equip[condicao_filtro]
        else:
            dados_para_calculo = dados_equip
        
        # Se não há dados após filtro, retornar 0
        if len(dados_para_calculo) == 0:
            return 0.0
        
        # Sempre usar soma de Diferença_Hora
        return dados_para_calculo['Diferença_Hora'].sum()
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df_filtrado['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        
        print(f"\n=== Calculando disponibilidade mecânica para {equipamento} ===")
        
        # CORREÇÃO: Usar Diferença_Hora (tempo de apontamento) para total geral
        total_geral = calcular_tempo_por_diferenca_hora_equipamento(dados_equip)
        
        # CORREÇÃO: Usar Diferença_Hora (tempo de apontamento) para manutenção
        horas_manutencao = calcular_tempo_por_diferenca_hora_equipamento(
            dados_equip, 
            dados_equip['Grupo Operacao'] == 'Manutenção'
        )
        
        # CORREÇÃO: Fórmula exata como no Excel: (Total Geral - Manutenção) / Total Geral
        # A disponibilidade mecânica é: (Total - Manutenção) / Total
        if total_geral > 0:
            disp_mecanica = (total_geral - horas_manutencao) / total_geral
        else:
            disp_mecanica = 0.0
        
        # Debug: mostrar valores para verificação
        print(f"Total Geral (apontamento): {total_geral:.6f}")
        print(f"Manutenção (apontamento): {horas_manutencao:.6f}")
        print(f"Disponibilidade: {disp_mecanica:.6f} ({disp_mecanica*100:.2f}%)")
        print(f"Fórmula: ({total_geral:.6f} - {horas_manutencao:.6f}) / {total_geral:.6f} = {disp_mecanica:.6f}")
        
        resultados.append({
            'Frota': equipamento,
            'Disponibilidade': disp_mecanica,
            'Tempo Manutenção': horas_manutencao
        })
    
    return pd.DataFrame(resultados)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Calcula médias diárias considerando os dias efetivos de cada frota.
    Esta função NÃO aplica qualquer filtro de operador.
    Também identifica as faltas de horário por dia específico.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota com detalhamento por dia
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    resultados = []
    
    # Obter todos os dias únicos no dataset (filtrar valores NaN)
    if 'Data' in df.columns:
        dias_unicos = sorted([d for d in df['Data'].unique() if pd.notna(d)])
    else:
        dias_unicos = []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        
        # Determinar número de dias efetivos para este equipamento
        dias_equip = dados_equip['Data'].nunique() if 'Data' in dados_equip.columns else 1
        
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 2)
        
        # Se houver múltiplos dias, usar média diária
        if dias_equip > 1:
            total_horas = round(total_horas / dias_equip, 2)
        
        # Calcular a diferença para 24 horas
        diferenca_24h = round(max(24 - total_horas, 0), 2)
        
        # Criar o resultado básico (colunas originais mantidas)
        resultado = {
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        }
        
        # Adicionar detalhamento por dia (novas colunas)
        if len(dias_unicos) > 0:
            for dia in dias_unicos:
                dados_dia = dados_equip[dados_equip['Data'] == dia]
                
                # Se não houver dados para este dia e equipamento, a diferença é 24h
                if len(dados_dia) == 0:
                    resultado[f'Falta {dia}'] = 24.0
                    continue
                
                # Calcular horas registradas neste dia
                horas_dia = round(dados_dia['Diferença_Hora'].sum(), 2)
                
                # Calcular a diferença para 24 horas neste dia
                diferenca_dia = round(max(24 - horas_dia, 0), 2)
                
                # Adicionar ao resultado apenas se houver falta (diferença > 0)
                if diferenca_dia > 0:
                    resultado[f'Falta {dia}'] = diferenca_dia
                else:
                    resultado[f'Falta {dia}'] = 0.0
        
        resultados.append(resultado)
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso(base_calculo, df_base):
    """
    Calcula o percentual de motor ocioso por operador usando os dados da Base Calculo.
    Agrega os dados por operador, calculando a média quando um operador aparece em múltiplas frotas.
    
    Regras de cálculo (Método Avançado):
    1. Existe uma tolerância de 1 minuto para operações de "Parada com motor ligado" (valor 1)
    2. Se uma operação tem "Parada com motor ligado = 1" e dura menos de 1 minuto, ela é desconsiderada se:
       - A próxima operação tem "Parada com motor ligado = 0" e dura mais de 1 minuto
    3. Se existem duas ou mais operações com "Parada com motor ligado = 1" em sequência (ou com outras operações entre elas que duram menos de 1 minuto):
       - Somamos o tempo total dessas operações
       - Subtraímos 1 minuto do total
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
        df_base (DataFrame): DataFrame base para aplicar filtros de operações
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador (agregado)
    """
    # Agrupar por operador e usar os valores já calculados na Base Calculo
    agrupado = base_calculo.groupby('Operador').agg({
        'Motor Ligado': 'sum',
        'Parado Com Motor Ligado': 'sum'
    }).reset_index()
    
    resultados = []
    print("\n=== DETALHAMENTO DO CÁLCULO DE MOTOR OCIOSO (MÉTODO AVANÇADO) ===")
    print("Utilizando valores da coluna Motor Ocioso calculados com a lógica de intervalos e tolerância")
    print("=" * 60)

    for _, row in agrupado.iterrows():
        operador = row['Operador']
        tempo_ligado = row['Motor Ligado']
        tempo_ocioso = row['Parado Com Motor Ligado']

        # Determinar frotas associadas a este operador
        if df_base is not None and 'Equipamento' in df_base.columns:
            frotas = sorted(df_base[df_base['Operador'] == operador]['Equipamento'].astype(str).unique())
        else:
            frotas = sorted(base_calculo[base_calculo['Operador'] == operador]['Equipamento'].astype(str).unique())
        operador_nome = f"{operador} ({', '.join(frotas)})" if frotas else operador
        
        # Calcular porcentagem de tempo ocioso
        porcentagem = tempo_ocioso / tempo_ligado if tempo_ligado > 0 else 0
        
        print(f"\nOperador: {operador_nome}")
        print(f"Tempo Ocioso (método avançado) = {tempo_ocioso:.6f} horas")
        print(f"Horas Motor = {tempo_ligado:.6f} horas")
        print(f"Porcentagem = {porcentagem:.6f} ({porcentagem*100:.2f}%)")
        print("-" * 60)
        
        resultados.append({
            'Operador': operador_nome,
            'Porcentagem': porcentagem,
            'Horas Motor': tempo_ligado,
            'Tempo Ocioso': tempo_ocioso
        })
    
    return pd.DataFrame(resultados)

def calcular_velocidade_media_produtiva(df_base):
    """
    Calcula a velocidade média produtiva por máquina usando os mesmos filtros das horas de elevador.
    Filtros: Grupo Operacao='Produtiva', Pressao de Corte > 400, Velocidade > 0
    
    Args:
        df_base (DataFrame): DataFrame com os dados base
    
    Returns:
        DataFrame: Velocidade média produtiva por máquina
    """
    resultados = []
    equipamentos = df_base['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    
    for equipamento in equipamentos:
        dados_equip = df_base[df_base['Equipamento'] == equipamento]
        
        # Filtrar dados produtivos (mesmos filtros das horas de elevador)
        dados_produtivos = dados_equip[
            (dados_equip['Grupo Operacao'] == 'Produtiva') & 
            (dados_equip['Pressao de Corte'] > 400) &
            (dados_equip['Velocidade'] > 0)
        ]
        
        print(f"\n=== Calculando velocidade média produtiva para {equipamento} ===")
        print(f"Registros que atendem critérios produtivos: {len(dados_produtivos)}")
        
        # Calcular velocidade média ponderada pelo tempo (Diferença_Hora)
        if len(dados_produtivos) > 0 and dados_produtivos['Diferença_Hora'].sum() > 0:
            velocidade_media = (dados_produtivos['Velocidade'] * dados_produtivos['Diferença_Hora']).sum() / dados_produtivos['Diferença_Hora'].sum()
        else:
            velocidade_media = 0
        
        print(f"Velocidade média produtiva: {velocidade_media:.2f} km/h")
        
        resultados.append({
            'Frota': equipamento,
            'Velocidade Média Produtiva': velocidade_media
        })
    
    # Ordenar por velocidade média (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values('Velocidade Média Produtiva', ascending=False)

def calcular_hora_elevador(df_base, base_calculo):
    """
    Calcula as horas de elevador por máquina usando Diferença_Hora.
    CORREÇÃO: Horas motor calculadas usando Diferença_Hora APENAS onde Motor Ligado == 1
    para manter consistência com a planilha Motor Ocioso.
    
    Args:
        df_base: DataFrame base para calcular
        base_calculo (DataFrame): Não usado mais, mantido para compatibilidade
    
    Returns:
        DataFrame: Horas de elevador, horas motor por máquina
    """
    resultados = []
    equipamentos = df_base['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    
    for equipamento in equipamentos:
        dados_equip = df_base[df_base['Equipamento'] == equipamento]
        
        # Filtrar dados para condição de elevador
        dados_elevador = dados_equip[
            (dados_equip['Grupo Operacao'] == 'Produtiva') & 
            (dados_equip['Pressao de Corte'] > 400) &
            (dados_equip['Velocidade'] > 0)
        ]
        
        print(f"\n=== Calculando horas elevador para {equipamento} ===")
        print(f"Registros que atendem critérios elevador: {len(dados_elevador)}")
        
        # Calcular horas elevador usando Diferença_Hora
        horas_elevador = dados_elevador['Diferença_Hora'].sum()
        print(f"Horas elevador (Diferença_Hora): {horas_elevador:.4f}h")
        
        # CORREÇÃO: Calcular horas motor usando Diferença_Hora APENAS onde Motor Ligado == 1
        # Para manter consistência com a planilha Motor Ocioso
        horas_motor = dados_equip[dados_equip['Motor Ligado'] == 1]['Diferença_Hora'].sum()
        print(f"Horas motor (Motor Ligado == 1): {horas_motor:.4f}h")
        
        # Calcular percentual de eficiência do elevador
        percentual_eficiencia = (horas_elevador / horas_motor) if horas_motor > 0 else 0
        
        print(f"Horas Motor: {horas_motor:.4f}h")
        print(f"Eficiência Elevador: {percentual_eficiencia:.4f} ({percentual_eficiencia*100:.2f}%)")
        
        resultados.append({
            'Frota': equipamento,
            'Horas Elevador': horas_elevador,
            'Horas Motor': horas_motor,
            'Eficiência Elevador': percentual_eficiencia
        })
    
    # Ordenar por horas elevador (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values('Horas Elevador', ascending=False)

def calcular_uso_gps(df_base, base_calculo):
    """
    Extrai o uso de GPS da Base Calculo.
    Agrega os dados por operador, calculando a média ponderada quando um operador aparece em múltiplas frotas.
    
    Args:
        df_base: Não usado mais, mantido para compatibilidade
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador (agregado)
    """
    # Agrupar por operador e calcular a média ponderada
    agrupado = base_calculo.groupby('Operador').agg({
        'RTK': 'sum',
        'Horas Produtivas': 'sum'
    }).reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        operador = row['Operador']
        porcentagem = row['RTK'] / row['Horas Produtivas'] if row['Horas Produtivas'] > 0 else 0
        frotas = sorted(base_calculo[base_calculo['Operador'] == operador]['Equipamento'].astype(str).unique())
        operador_nome = f"{operador} ({', '.join(frotas)})" if frotas else operador
        resultados.append({
            'Operador': operador_nome,
            'Porcentagem': porcentagem
        })
    
    return pd.DataFrame(resultados)

def calcular_media_velocidade(df):
    """
    Calcula a média de velocidade para cada operador considerando apenas registros produtivos e deslocamentos.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: DataFrame com a média de velocidade por operador
    """
    # Filtramos os dados excluindo os operadores da lista de exclusão
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Identificar registros produtivos e deslocamentos
    # Produtivos: onde Grupo Operacao é 'Produtiva'
    # Deslocamentos: onde Estado é 'DESLOCAMENTO' ou velocidade > 0 em operações não produtivas
    registros_validos = (
        (df_filtrado['Grupo Operacao'] == 'Produtiva') | 
        (df_filtrado['Estado'] == 'DESLOCAMENTO') |
        ((df_filtrado['Velocidade'] > 0) & (df_filtrado['Estado'] != 'PARADO'))
    )
    
    df_velocidade = df_filtrado[registros_validos]
    
    # Obter lista de operadores únicos do DataFrame original
    operadores = df_filtrado['Operador'].unique()
    resultados = []
    
    for operador in operadores:
        # Filtrar dados para este operador
        dados_op = df_velocidade[df_velocidade['Operador'] == operador]
        
        # Se não houver dados válidos para este operador, adicionar com velocidade zero
        if len(dados_op) == 0:
            resultados.append({
                'Operador': operador,
                'Velocidade': 0
            })
            continue
        
        # Calcular a média de velocidade (ponderada pelo tempo, se houver a coluna Diferença_Hora)
        if 'Diferença_Hora' in dados_op.columns:
            # Média ponderada pelo tempo 
            velocidade_media = round(
                (dados_op['Velocidade'] * dados_op['Diferença_Hora']).sum() / dados_op['Diferença_Hora'].sum()
                if dados_op['Diferença_Hora'].sum() > 0 else 0,
                2
            )
        else:
            # Média simples
            velocidade_media = round(dados_op['Velocidade'].mean(), 2)
        
        resultados.append({
            'Operador': operador,
            'Velocidade': velocidade_media
        })
    
    return pd.DataFrame(resultados)

def identificar_operadores_duplicados(df):
    """
    Identifica operadores que aparecem com IDs diferentes no mesmo conjunto de dados.
    Detecta principalmente IDs que começam com '133' e têm 7 dígitos, verificando se
    existe outra ID com o mesmo nome mas com menos dígitos.
    
    Args:
        df (DataFrame): DataFrame com os dados dos operadores
    
    Returns:
        dict: Dicionário com mapeamento {id_incorreta: id_correta}
        DataFrame: DataFrame com as duplicidades encontradas para relatório
    """
    if 'Operador' not in df.columns or len(df) == 0:
        return {}, pd.DataFrame(columns=['ID Incorreta', 'ID Correta', 'Nome'])
    
    # Extrair operadores únicos
    operadores = df['Operador'].unique()
    
    # Mapear nomes para IDs
    nomes_para_ids = {}
    for op in operadores:
        if ' - ' in op:
            try:
                id_parte, nome_parte = op.split(' - ', 1)
                # Normalizar nome para comparação (maiúsculo e sem espaços extras)
                nome_normalizado = nome_parte.upper().strip()
                if nome_normalizado not in nomes_para_ids:
                    nomes_para_ids[nome_normalizado] = []
                nomes_para_ids[nome_normalizado].append(op)
            except:
                continue
    
    # Encontrar nomes com múltiplas IDs
    duplicidades = []
    mapeamento = {}
    
    for nome, ids in nomes_para_ids.items():
        if len(ids) > 1:
            print(f"Encontrado operador duplicado: {nome} com {len(ids)} IDs diferentes")
            
            # Separar IDs que começam com '133' e têm 7 dígitos
            ids_suspeitas = [id_op for id_op in ids if ' - ' in id_op and id_op.split(' - ')[0].startswith('133') and len(id_op.split(' - ')[0]) == 7]
            ids_normais = [id_op for id_op in ids if id_op not in ids_suspeitas]
            
            # Se temos IDs suspeitas e normais, considerar a suspeita como incorreta
            if ids_suspeitas and ids_normais:
                for id_suspeita in ids_suspeitas:
                    # Usar a ID normal mais curta como destino (geralmente a correta)
                    id_correta = min(ids_normais, key=lambda x: len(x.split(' - ')[0]) if ' - ' in x else float('inf'))
                    
                    # Adicionar ao mapeamento
                    mapeamento[id_suspeita] = id_correta
                    
                    # Extrair as partes para o relatório
                    id_incorreta_parte = id_suspeita.split(' - ')[0]
                    id_correta_parte = id_correta.split(' - ')[0]
                    
                    print(f"  - Mapeando: {id_suspeita} -> {id_correta}")
                    
                    # Adicionar à lista de duplicidades para o relatório
                    duplicidades.append({
                        'ID Incorreta': id_suspeita,
                        'ID Correta': id_correta,
                        'Nome': nome
                    })
            
            # Caso especial: todas as IDs são suspeitas ou todas são normais
            # Neste caso, apresentamos no relatório mas não fazemos substituição automática
            else:
                print(f"  - Múltiplas IDs do mesmo tipo encontradas para {nome}, sem ação automática")
                # Ainda assim, adicionar ao relatório para conhecimento
                for i, id1 in enumerate(ids):
                    for id2 in ids[i+1:]:
                        duplicidades.append({
                            'ID Incorreta': id1, 
                            'ID Correta': id2,
                            'Nome': nome,
                            'Observação': "Duplicidade detectada, verificar manualmente"
                        })
    
    print(f"Encontrados {len(duplicidades)} operadores com IDs duplicadas.")
    return mapeamento, pd.DataFrame(duplicidades)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Calcula médias diárias considerando os dias efetivos de cada frota.
    ATUALIZADO: Inclui todos os detalhes dos horímetros iniciais e finais.
    Esta função NÃO aplica qualquer filtro de operador.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota com detalhamento completo incluindo horímetros
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    resultados = []
    
    # Obter todos os dias únicos no dataset (filtrar valores NaN)
    if 'Data' in df.columns:
        dias_unicos = sorted([d for d in df['Data'].unique() if pd.notna(d)])
    else:
        dias_unicos = []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        
        # Determinar número de dias efetivos para este equipamento
        dias_equip = dados_equip['Data'].nunique() if 'Data' in dados_equip.columns else 1
        
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 2)
        
        # Se houver múltiplos dias, usar média diária
        if dias_equip > 1:
            total_horas = round(total_horas / dias_equip, 2)
        
        # Calcular a diferença para 24 horas
        diferenca_24h = round(max(24 - total_horas, 0), 2)
        
        # Calcular apenas horímetro geral (último menos o primeiro, sem filtros)
        horimetro_geral_inicio = None
        horimetro_geral_fim = None
        horimetro_geral_total = None
        
        # Horímetros gerais (todos os registros, sem filtros)
        if 'Horimetro' in dados_equip.columns and not dados_equip['Horimetro'].isna().all():
            try:
                horimetro_values = pd.to_numeric(dados_equip['Horimetro'], errors='coerce').dropna()
                if len(horimetro_values) >= 2:
                    horimetro_geral_inicio = horimetro_values.iloc[0]
                    horimetro_geral_fim = horimetro_values.iloc[-1]
                    horimetro_geral_total = horimetro_geral_fim - horimetro_geral_inicio
            except Exception as e:
                print(f"Erro ao processar horímetro geral para {equipamento}: {e}")
        
        # Criar o resultado básico (apenas horímetro geral)
        resultado = {
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h,
            'Horimetro Inicio': horimetro_geral_inicio,
            'Horimetro Fim': horimetro_geral_fim,
            'Horimetro Total': horimetro_geral_total
        }
        
        # Adicionar detalhamento por dia (novas colunas)
        if len(dias_unicos) > 0:
            for dia in dias_unicos:
                dados_dia = dados_equip[dados_equip['Data'] == dia]
                
                # Se não houver dados para este dia e equipamento, a diferença é 24h
                if len(dados_dia) == 0:
                    resultado[f'Falta {dia}'] = 24.0
                    continue
                
                # Calcular horas registradas neste dia
                horas_dia = round(dados_dia['Diferença_Hora'].sum(), 2)
                
                # Calcular a diferença para 24 horas neste dia
                diferenca_dia = round(max(24 - horas_dia, 0), 2)
                
                # Adicionar ao resultado apenas se houver falta (diferença > 0)
                if diferenca_dia > 0:
                    resultado[f'Falta {dia}'] = diferenca_dia
                else:
                    resultado[f'Falta {dia}'] = 0.0
        
        resultados.append(resultado)
    
    return pd.DataFrame(resultados)



def calcular_ofensores(df):
    """
    Calcula os top 5 ofensores gerais.
    Agrupa apenas por Operacao onde Estado Operacional é 'PARADA',
    soma a Diferença_Hora, classifica do maior para o menor e seleciona os top 5.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame com os top 5 ofensores gerais
    """
    # Filtrar apenas os registros com Estado Operacional PARADA
    df_paradas = df[df['Estado Operacional'] == 'PARADA'].copy()
    
    # Se não houver dados de parada, retornar DataFrame vazio
    if len(df_paradas) == 0:
        return pd.DataFrame(columns=['Operação', 'Tempo', 'Porcentagem'])
    
    # Agrupar apenas por Operacao, somar o tempo
    paradas_agrupadas = df_paradas.groupby('Operacao')['Diferença_Hora'].sum().reset_index()
    
    # Calcular o tempo total de todas as paradas
    tempo_total = paradas_agrupadas['Diferença_Hora'].sum()
    
    # Adicionar coluna de porcentagem
    paradas_agrupadas['Porcentagem'] = paradas_agrupadas.apply(
        lambda row: row['Diferença_Hora'] / tempo_total if tempo_total > 0 else 0,
        axis=1
    )
    
    # Ordenar por tempo (decrescente)
    paradas_agrupadas = paradas_agrupadas.sort_values(by='Diferença_Hora', ascending=False)
    
    # Selecionar os top 5 gerais
    resultado = paradas_agrupadas.head(5)
    
    # Renomear colunas para melhor visualização
    resultado = resultado.rename(columns={
        'Operacao': 'Operação',
        'Diferença_Hora': 'Tempo'
    })
    
    return resultado

def calcular_lavagem(df):
    """
    Calcula os intervalos de lavagem para cada equipamento.
    Identifica início, fim e duração de cada intervalo de lavagem,
    e calcula o tempo total por dia e equipamento.
    
    Se não houver registros de lavagem, retorna um DataFrame com uma linha informativa.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame com os intervalos de lavagem detalhados
    """
    print("Calculando intervalos de lavagem...")
    
    # Filtrar apenas registros de lavagem
    df_lavagem = df[df['Operacao'] == '8490 - LAVAGEM'].copy()
    
    if len(df_lavagem) == 0:
        print("Nenhum registro de lavagem encontrado.")
        print("A planilha de Lavagem será criada com uma linha informativa.")
        
        # Retornar DataFrame com linha informativa
        return pd.DataFrame([{
            'Data': 'N/A',
            'Equipamento': 'NÃO FORAM ENCONTRADOS DADOS DE LAVAGEM PARA A DATA INFORMADA',
            'Intervalo': 'N/A',
            'Início': 'N/A',
            'Fim': 'N/A',
            'Duração (horas)': 0,
            'Tempo Total do Dia': 0
        }])
    
    print(f"Encontrados {len(df_lavagem)} registros de lavagem.")
    
    # Garantir que as colunas necessárias existam
    colunas_necessarias = ['Equipamento', 'Data', 'Hora', 'Diferença_Hora']
    for coluna in colunas_necessarias:
        if coluna not in df_lavagem.columns:
            print(f"Coluna '{coluna}' não encontrada nos dados de lavagem.")
            return pd.DataFrame(columns=[
                'Data', 'Equipamento', 'Intervalo', 'Início', 'Fim', 
                'Duração (horas)', 'Tempo Total do Dia'
            ])
    
    # Ordenar por equipamento, data e hora
    df_lavagem = df_lavagem.sort_values(['Equipamento', 'Data', 'Hora'])
    
    resultados = []
    
    # Agrupar por equipamento e data
    for (equipamento, data), grupo in df_lavagem.groupby(['Equipamento', 'Data']):
        print(f"Processando lavagem: {equipamento} - {data}")
        
        # Resetar índices para facilitar a iteração
        grupo = grupo.reset_index(drop=True)
        
        # Identificar intervalos contínuos de lavagem
        intervalos = []
        inicio_intervalo = None
        fim_intervalo = None
        duracao_intervalo = 0
        
        for i in range(len(grupo)):
            registro = grupo.iloc[i]
            
            # Converter hora string para datetime para comparação
            hora_atual = pd.to_datetime(registro['Hora'], format='%H:%M:%S')
            
            # Se é o primeiro registro ou se houve uma pausa longa (> 30 minutos)
            if (inicio_intervalo is None or 
                (i > 0 and (hora_atual - pd.to_datetime(grupo.iloc[i-1]['Hora'], format='%H:%M:%S')).total_seconds() > 1800)):
                
                # Finalizar intervalo anterior se existir
                if inicio_intervalo is not None:
                    intervalos.append({
                        'inicio': inicio_intervalo,
                        'fim': fim_intervalo,
                        'duracao': duracao_intervalo
                    })
                
                # Iniciar novo intervalo
                inicio_intervalo = registro['Hora']
                duracao_intervalo = registro['Diferença_Hora']
            else:
                # Continuar intervalo atual
                duracao_intervalo += registro['Diferença_Hora']
            
            # Atualizar fim do intervalo
            fim_intervalo = registro['Hora']
        
        # Adicionar o último intervalo
        if inicio_intervalo is not None:
            intervalos.append({
                'inicio': inicio_intervalo,
                'fim': fim_intervalo,
                'duracao': duracao_intervalo
            })
        
        # Calcular tempo total do dia
        tempo_total_dia = sum(intervalo['duracao'] for intervalo in intervalos)
        
        # Adicionar cada intervalo aos resultados
        for i, intervalo in enumerate(intervalos, 1):
            # Preencher "Tempo Total do Dia" apenas no último intervalo
            tempo_total_mostrar = tempo_total_dia if i == len(intervalos) else None
            
            resultados.append({
                'Data': data,
                'Equipamento': equipamento,
                'Intervalo': f"Intervalo {i}",
                'Início': intervalo['inicio'],
                'Fim': intervalo['fim'],
                'Duração (horas)': intervalo['duracao'],
                'Tempo Total do Dia': tempo_total_mostrar
            })
    
    # Criar DataFrame com os resultados
    df_resultado = pd.DataFrame(resultados)
    
    # Ordenar por data, equipamento e intervalo
    if not df_resultado.empty:
        df_resultado = df_resultado.sort_values(['Data', 'Equipamento', 'Intervalo'])
    
    print(f"Processamento de lavagem concluído. {len(df_resultado)} intervalos identificados.")
    
    return df_resultado

def criar_planilha_coordenadas(df_base):
    """
    Cria uma planilha com coordenadas de TODAS as operações (sem filtro em Grupo Operacao),
    ordenada por hora e por frota. Adiciona as colunas "Velocidade" e "RTK" para
    posterior análise de deslocamento.
    
    Args:
        df_base (DataFrame): DataFrame com os dados base COMPLETOS
        
    Returns:
        DataFrame: DataFrame com as colunas Equipamento, Hora, Latitude, Longitude, Velocidade e RTK
    """
    # Verificar se as colunas necessárias existem
    colunas_necessarias = ['Equipamento', 'Hora', 'Latitude', 'Longitude', 'Velocidade', 'Pressao de Corte', 'RTK (Piloto Automatico)']
    for coluna in colunas_necessarias:
        if coluna not in df_base.columns:
            print(f"Aviso: Coluna '{coluna}' não encontrada para criar planilha de coordenadas.")
            # Criar um DataFrame vazio com as colunas necessárias
            return pd.DataFrame(columns=['Equipamento', 'Hora', 'Latitude', 'Longitude', 'Velocidade', 'RTK'])
    
    print(f"Total de coordenadas antes dos filtros: {len(df_base)} registros")
    
    # Usar todos os registros (sem filtro). Copiamos as colunas desejadas.
    colunas_saida = ['Equipamento', 'Hora', 'Latitude', 'Longitude', 'Velocidade', 'Pressao de Corte', 'RTK (Piloto Automatico)']
    df_coordenadas = df_base[colunas_saida].copy()
    
    # Garantir que as colunas sejam numéricas
    df_coordenadas['Latitude'] = pd.to_numeric(df_coordenadas['Latitude'], errors='coerce')
    df_coordenadas['Longitude'] = pd.to_numeric(df_coordenadas['Longitude'], errors='coerce')
    df_coordenadas['Velocidade'] = pd.to_numeric(df_coordenadas['Velocidade'], errors='coerce')
    df_coordenadas['Pressao de Corte'] = pd.to_numeric(df_coordenadas['Pressao de Corte'], errors='coerce')
    df_coordenadas['RTK (Piloto Automatico)'] = pd.to_numeric(df_coordenadas['RTK (Piloto Automatico)'], errors='coerce')
    
    # Filtrar apenas coordenadas válidas (não zero e não nulas)
    # MANTER este filtro pois coordenadas 0,0 são dados inválidos
    df_coordenadas = df_coordenadas[
        (df_coordenadas['Latitude'] != 0) & 
        (df_coordenadas['Longitude'] != 0) &
        (df_coordenadas['Latitude'].notna()) &
        (df_coordenadas['Longitude'].notna())
    ]
    
    print(f"Coordenadas APÓS filtro de GPS válidos: {len(df_coordenadas)} registros")
    
    # Criar coluna RTK com valores "Sim"/"Não" baseado nos critérios
    # Critério para "Sim": Velocidade > 0 AND Pressão de Corte >= 400 AND RTK (Piloto Automatico) = 1
    df_coordenadas['RTK'] = df_coordenadas.apply(
        lambda row: 'Sim' if (
            row['Velocidade'] > 0 and 
            row['Pressao de Corte'] >= 400 and 
            row['RTK (Piloto Automatico)'] == 1
        ) else 'Não', 
        axis=1
    )
    
    # Selecionar apenas as colunas finais
    colunas_finais = ['Equipamento', 'Hora', 'Latitude', 'Longitude', 'Velocidade', 'RTK']
    df_coordenadas = df_coordenadas[colunas_finais]
    
    # Formatar as coordenadas como strings com ponto decimal (mantém Velocidade numérica)
    df_coordenadas['Latitude'] = df_coordenadas['Latitude'].apply(lambda x: f"{x:.9f}" if pd.notnull(x) else '')
    df_coordenadas['Longitude'] = df_coordenadas['Longitude'].apply(lambda x: f"{x:.9f}" if pd.notnull(x) else '')
    
    # Remover duplicatas completas para reduzir tamanho da planilha
    df_coordenadas = df_coordenadas.drop_duplicates()
    
    print(f"Coordenadas FINAIS (sem duplicatas): {len(df_coordenadas)} registros")
    print(f"Registros com RTK = 'Sim': {len(df_coordenadas[df_coordenadas['RTK'] == 'Sim'])} registros")
    print(f"Registros com RTK = 'Não': {len(df_coordenadas[df_coordenadas['RTK'] == 'Não'])} registros")
    
    return df_coordenadas

def criar_excel_com_planilhas(df_base, disp_mecanica, eficiencia_energetica, velocidade_media_produtiva,
                            hora_elevador, motor_ocioso, uso_gps, horas_por_frota, caminho_saida,
                            caminho_arquivo, media_velocidade=None, 
                            df_lavagem=None, df_ofensores=None):
    """
    Cria um arquivo Excel com todas as planilhas necessárias.
    Também gera um arquivo CSV da planilha Coordenadas.
    """
    # Definir função de ajuste de largura de colunas
    def ajustar_largura_colunas(worksheet):
        """Ajusta a largura das colunas da planilha"""
        for col in worksheet.columns:
            max_length = 10
            column = col[0].column_letter
            header_text = str(col[0].value)
            if header_text:
                max_length = max(max_length, len(header_text) + 2)
            for cell in col[1:min(20, len(col))]:
                if cell.value:
                    cell_text = str(cell.value)
                    max_length = max(max_length, len(cell_text) + 2)
            max_length = min(max_length, 40)
            worksheet.column_dimensions[column].width = max_length
    
    # Calcular horas por frota (com detalhes de horímetros)
    horas_por_frota = calcular_horas_por_frota(df_base)
    
    # Criar planilha de coordenadas
    df_coordenadas = criar_planilha_coordenadas(df_base)
    
    # Gerar arquivo CSV das coordenadas
    nome_base_original = os.path.splitext(os.path.basename(caminho_arquivo))[0]  # Nome original do arquivo sem extensão
    diretorio_saida = os.path.dirname(caminho_saida)
    caminho_csv_coordenadas = os.path.join(diretorio_saida, f"{nome_base_original}_Coordenadas.csv")
    
    try:
        df_coordenadas.to_csv(caminho_csv_coordenadas, index=False, encoding='utf-8-sig', sep=';')
        print(f"Arquivo CSV de coordenadas gerado: {os.path.basename(caminho_csv_coordenadas)}")
    except Exception as e:
        print(f"Erro ao gerar arquivo CSV de coordenadas: {str(e)}")
    
        # ===== CÁLCULO DE MANOBRAS (por intervalos sequenciais) =====
    df_manobras_frota, df_manobras_operador = calcular_manobras_por_intervalos(df_base)
    # ===== FIM CÁLCULO DE MANOBRAS =====
    
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        # Planilha BASE (sempre primeira)
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        
        # Planilhas principais
        disp_mecanica.to_excel(writer, sheet_name='Disponibilidade Mecânica', index=False)
        eficiencia_energetica.to_excel(writer, sheet_name='Eficiência Energética', index=False)
        velocidade_media_produtiva.to_excel(writer, sheet_name='Velocidade Média Produtiva', index=False)
        hora_elevador.to_excel(writer, sheet_name='Eficiência Energética', index=False)
        uso_gps.to_excel(writer, sheet_name='Uso GPS', index=False)
        
        # Planilhas auxiliares
        horas_por_frota.to_excel(writer, sheet_name='Horas por Frota', index=False)
        
        if media_velocidade is None:
            media_velocidade = pd.DataFrame(columns=['Operador', 'Velocidade'])
        media_velocidade.to_excel(writer, sheet_name='Média Velocidade', index=False)
        
        # Planilha de coordenadas
        df_coordenadas.to_excel(writer, sheet_name='Coordenadas', index=False)
        
        # Planilhas de análise de problemas
        # Garantir que os valores numéricos do motor_ocioso sejam mantidos como números
        motor_ocioso['Horas Motor'] = pd.to_numeric(motor_ocioso['Horas Motor'], errors='coerce')
        motor_ocioso['Tempo Ocioso'] = pd.to_numeric(motor_ocioso['Tempo Ocioso'], errors='coerce')
        motor_ocioso['Porcentagem'] = pd.to_numeric(motor_ocioso['Porcentagem'], errors='coerce')
        motor_ocioso.to_excel(writer, sheet_name='Motor Ocioso', index=False)
        
        # Adicionar planilha de ofensores
        if df_ofensores is not None and not df_ofensores.empty:
            df_ofensores.to_excel(writer, sheet_name='Ofensores', index=False)
        
        # Adicionar planilha de lavagem (sempre incluir, mesmo se não houver registros)
        if df_lavagem is not None:
            df_lavagem.to_excel(writer, sheet_name='Lavagem', index=False)
        
        # Adicionar planilhas de manobras
        if not df_manobras_frota.empty:
            df_manobras_frota.to_excel(writer, sheet_name='Manobras Frotas', index=False)
        if not df_manobras_operador.empty:
            df_manobras_operador.to_excel(writer, sheet_name='Manobras Operador', index=False)
        
        # Formatar cada planilha
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            ajustar_largura_colunas(worksheet)
            
            if sheet_name == 'BASE':
                # Formatação da planilha BASE
                # Encontrar a coluna "Hora" e formatá-la como texto (já está no formato HH:MM:SS)
                for col in range(1, worksheet.max_column + 1):
                    header = worksheet.cell(row=1, column=col).value
                    if header == 'Hora':
                        # Formatar toda a coluna "Hora" como texto
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.number_format = '@'  # Formato de texto
                        break
            
            elif sheet_name == 'Disponibilidade Mecânica':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Disponibilidade)
                    cell.number_format = '0.00%'
            
            elif sheet_name == 'Eficiência Energética':
                for row in range(2, worksheet.max_row + 1):
                    # Coluna B (Horas Motor)
                    cell = worksheet.cell(row=row, column=2)
                    cell.number_format = '0.00'
                    # Coluna C (Horas Produtivas)
                    cell = worksheet.cell(row=row, column=3)
                    cell.number_format = '0.00'
                    # Coluna D (Eficiência Elevador)
                    cell = worksheet.cell(row=row, column=4)
                    cell.number_format = '0.00%'
            
            elif sheet_name == 'Velocidade Média Produtiva':
                for row in range(2, worksheet.max_row + 1):
                    # Coluna B (Velocidade Média Produtiva)
                    cell = worksheet.cell(row=row, column=2)
                    cell.number_format = '0.00'
            
            elif sheet_name == 'Eficiência Energética':
                for row in range(2, worksheet.max_row + 1):
                    # Coluna B (Eficiência)
                    cell = worksheet.cell(row=row, column=2)
                    cell.number_format = '0.00%'
            
            elif sheet_name == 'Uso GPS':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
                    cell = worksheet.cell(row=row, column=3)  # Coluna C (Porcentagem Sem Pressão)
                    cell.number_format = '0.00%'
            
            elif sheet_name == 'Horas por Frota':
                for row in range(2, worksheet.max_row + 1):
                    # Formatar apenas algumas colunas iniciais para não quebrar com colunas dinâmicas
                    for col in range(2, min(worksheet.max_column + 1, 12)):  # Até coluna 12 (L)
                        cell = worksheet.cell(row=row, column=col)
                        cell.number_format = '0.00'
            
            elif sheet_name == 'Motor Ocioso':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
                    cell = worksheet.cell(row=row, column=3)  # Coluna C (Horas Motor)
                    cell.number_format = '0.00'
                    cell = worksheet.cell(row=row, column=4)  # Coluna D (Tempo Ocioso)
                    cell.number_format = '0.00'
            
            elif sheet_name == 'Média Velocidade':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Velocidade)
                    cell.number_format = '0.00'
            
            elif sheet_name == 'Horas por Frota':
                for row in range(2, worksheet.max_row + 1):
                    for col in range(2, worksheet.max_column + 1):  # Todas as colunas de tempo
                        cell = worksheet.cell(row=row, column=col)
                        cell.number_format = '0.00'
            
            elif sheet_name == 'Ofensores':
                if df_ofensores is not None and not df_ofensores.empty:
                    for row in range(2, worksheet.max_row + 1):
                        # Coluna B (Tempo)
                        cell = worksheet.cell(row=row, column=2)
                        cell.number_format = '0.00'  # Formato decimal
                        
                        # Coluna C (Porcentagem)
                        cell = worksheet.cell(row=row, column=3)
                        cell.number_format = '0.00%'  # Formato percentual
            

            

            
            elif sheet_name == 'Lavagem':
                if df_lavagem is not None:
                    for row in range(2, worksheet.max_row + 1):
                        # Verificar se não é a linha informativa (para não formatar como hora)
                        equipamento_cell = worksheet.cell(row=row, column=2)
                        if equipamento_cell.value != 'NÃO FORAM ENCONTRADOS DADOS DE LAVAGEM PARA A DATA INFORMADA':
                            # Coluna D (Início)
                            cell = worksheet.cell(row=row, column=4)
                            cell.number_format = 'hh:mm:ss'
                            
                            # Coluna E (Fim)
                            cell = worksheet.cell(row=row, column=5)
                            cell.number_format = 'hh:mm:ss'
                            
                            # Coluna F (Duração)
                            cell = worksheet.cell(row=row, column=6)
                            cell.number_format = '0.00'
                            
                            # Coluna G (Tempo Total do Dia) - só formatar se não for None/vazio
                            cell = worksheet.cell(row=row, column=7)
                            if cell.value is not None and cell.value != "":
                                cell.number_format = '0.00'
            
            elif sheet_name == 'Coordenadas':
                # Formatar coluna Hora como hora
                for row in range(2, worksheet.max_row + 1):
                    # Hora
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Hora)
                    cell.number_format = 'hh:mm:ss'
                    # Latitude
                    cell = worksheet.cell(row=row, column=3)  # Coluna C (Latitude)
                    cell.number_format = '0.000000'
                    # Longitude
                    cell = worksheet.cell(row=row, column=4)  # Coluna D (Longitude)
                    cell.number_format = '0.000000'
                    # Velocidade
                    cell = worksheet.cell(row=row, column=5)  # Coluna E (Velocidade)
                    cell.number_format = '0.00'
            


            elif sheet_name in ['Manobras Operador', 'Manobras Frotas']:
                # Formatação das novas colunas de manobras por intervalos
                for row in range(2, worksheet.max_row + 1):
                    # Coluna 2: Intervalos Válidos
                    cell = worksheet.cell(row=row, column=2)
                    cell.number_format = '0'
                    
                    # Coluna 3: Tempo Total (horas)
                    cell = worksheet.cell(row=row, column=3)
                    cell.number_format = '0.0000'
                    
                    # Coluna 4: Tempo Médio (horas)
                    cell = worksheet.cell(row=row, column=4)
                    cell.number_format = '0.0000'
                
                # Adicionar colunas formatadas como hh:mm:ss
                if worksheet.max_row > 1:  # Só se houver dados
                    # Adicionar cabeçalhos para colunas de tempo formatado
                    worksheet.cell(row=1, column=worksheet.max_column + 1).value = 'Tempo Total (hh:mm)'
                    worksheet.cell(row=1, column=worksheet.max_column + 1).value = 'Tempo Médio (hh:mm)'
                    
                    for row in range(2, worksheet.max_row + 1):
                        # Tempo Total em formato hh:mm:ss
                        tempo_total = worksheet.cell(row=row, column=3).value
                        worksheet.cell(row=row, column=worksheet.max_column - 1).value = tempo_total / 24 if tempo_total else 0
                        worksheet.cell(row=row, column=worksheet.max_column - 1).number_format = 'h:mm:ss'
                        
                        # Tempo Médio em formato hh:mm:ss
                        tempo_medio = worksheet.cell(row=row, column=4).value
                        worksheet.cell(row=row, column=worksheet.max_column).value = tempo_medio / 24 if tempo_medio else 0
                        worksheet.cell(row=row, column=worksheet.max_column).number_format = 'h:mm:ss'
                
                # Reajustar largura das colunas
                ajustar_largura_colunas(worksheet)

def processar_arquivo_maquina(caminho_arquivo, diretorio_saida):
    """
    Versão simplificada de `processar_arquivo` trabalhando apenas por MÁQUINA.
    Gera planilhas BASE, Lavagem, Disponibilidade Mecânica e Uso GPS por máquina.
    """
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}_processado.xlsx")
    
    print(f"\nProcessando arquivo (simplificado): {os.path.basename(caminho_arquivo)}")
    print(f"Arquivo de saída: {os.path.basename(arquivo_saida)}")
    
    df_base = processar_arquivo_base(caminho_arquivo)
    if df_base is None or df_base.empty:
        print("Arquivo sem dados válidos. Pulando.")
        return
    
    # Ordenar por máquina/data/hora para facilitar leitura
    if 'Data' in df_base.columns:
        df_base = df_base.sort_values(by=['Equipamento', 'Data', 'Hora'])
    else:
        df_base = df_base.sort_values(by=['Equipamento', 'Hora'])
    
    # Criar coluna adicional formatada em horas (HH:MM:SS) a partir de Diferença_Hora
    if 'Diferença_Hora' in df_base.columns:
        idx_diff = df_base.columns.get_loc('Diferença_Hora')
        df_base.insert(idx_diff + 1, 'Diferença_Hora_hhmm', df_base['Diferença_Hora'] / 24)

    disp_mecanica = calcular_disponibilidade_mecanica(df_base)
    velocidade_media_produtiva = calcular_velocidade_media_produtiva(df_base)
    df_lavagem = calcular_lavagem(df_base)
    df_roletes = calcular_roletes(df_base)
    df_ofensores = calcular_ofensores(df_base)
    df_intervalos = calcular_intervalos_operacionais(df_base)
    uso_gps_maquina = calcular_uso_gps_maquina(df_base)
    motor_ocioso_maquina = calcular_motor_ocioso_maquina_correto(df_base)
    
    # Calcular horas por frota para verificação
    horas_por_frota = calcular_horas_por_frota(df_base)
    
    # ===== GERAR CSV DE COORDENADAS PARA MAPAS =====
    print("\n=== GERANDO CSV DE COORDENADAS ===")
    df_coordenadas = criar_planilha_coordenadas(df_base)
    
    # Nome do arquivo CSV baseado no arquivo original
    nome_base_original = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    caminho_csv_coordenadas = os.path.join(diretorio_saida, f"{nome_base_original}_Coordenadas.csv")
    
    try:
        df_coordenadas.to_csv(caminho_csv_coordenadas, index=False, encoding='utf-8-sig', sep=';')
        print(f"Arquivo CSV de coordenadas gerado: {os.path.basename(caminho_csv_coordenadas)}")
    except Exception as e:
        print(f"Erro ao gerar arquivo CSV de coordenadas: {str(e)}")
    # ===== FIM CSV DE COORDENADAS =====
    
    # Calcular Base Calculo para obter hora elevador
    equipamentos = df_base['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    base_calculo_data = []
    
    for equipamento in equipamentos:
        dados_equip = df_base[df_base['Equipamento'] == equipamento]
        
        # Calcular horas elevador usando horímetro
        def calcular_tempo_por_horimetro_simples(dados, condicao_filtro):
            if len(dados) == 0:
                return 0.0
            dados_filtrados = dados[condicao_filtro] if condicao_filtro is not None else dados
            if len(dados_filtrados) == 0:
                return 0.0
            
            if 'Horimetro' in dados_filtrados.columns and not dados_filtrados['Horimetro'].isna().all():
                try:
                    horimetro_values = pd.to_numeric(dados_filtrados['Horimetro'], errors='coerce').dropna()
                    if len(horimetro_values) >= 2:
                        tempo_horimetro = horimetro_values.iloc[-1] - horimetro_values.iloc[0]
                        if tempo_horimetro > 0 and tempo_horimetro < 48:
                            return tempo_horimetro
                except:
                    pass
            return dados_filtrados['Diferença_Hora'].sum()
        
        # Horas elevador
        horas_elevador = calcular_tempo_por_horimetro_simples(
            dados_equip,
            (dados_equip['Grupo Operacao'] == 'Produtiva') & 
            (dados_equip['Pressao de Corte'] > 400) &
            (dados_equip['Velocidade'] > 0)
        )
        
        # Motor ligado
        motor_ligado = calcular_tempo_por_horimetro_simples(
            dados_equip,
            dados_equip['Motor Ligado'] == 1
        )
        
        base_calculo_data.append({
            'Equipamento': equipamento,
            'Horas elevador': horas_elevador,
            'Motor Ligado': motor_ligado
        })
    
    base_calculo_simples = pd.DataFrame(base_calculo_data)
    hora_elevador_maquina = calcular_hora_elevador(df_base, base_calculo_simples)
    
    # ===== CÁLCULO DE MANOBRAS (por intervalos sequenciais) =====
    df_manobras_frota, df_manobras_operador = calcular_manobras_por_intervalos(df_base)
    # ===== FIM CÁLCULO DE MANOBRAS =====
    
    # Criar planilha de coordenadas
    df_coordenadas = criar_planilha_coordenadas(df_base)
    
    # Calcular operadores por frota
    try:
        df_operadores = calcular_operadores_por_frota(df_base)
        if df_operadores.empty:
            print("⚠️ Nenhum operador encontrado, criando planilha vazia")
            df_operadores = pd.DataFrame(columns=['Frota', 'Operador', 'Horas Elevador'])
    except Exception as e:
        print(f"⚠️ Erro ao calcular operadores, criando planilha vazia: {e}")
        df_operadores = pd.DataFrame(columns=['Frota', 'Operador', 'Horas Elevador'])

    # Calcular parâmetros médios técnicos
    try:
        df_parametros_medios = calcular_parametros_medios(df_base, uso_gps_maquina, hora_elevador_maquina, velocidade_media_produtiva)
        if df_parametros_medios.empty:
            print("⚠️ Nenhum parâmetro médio calculado, criando planilha vazia")
            df_parametros_medios = pd.DataFrame(columns=[
                'Frota', 'Horimetro', 'Uso RTK (%)', 'Horas Elevador', 'Horas Motor',
                'Velocidade Media (km/h)', 'RPM Motor Media', 'RPM Extrator Media',
                'Pressao Corte Media (psi)', 'Corte Base Auto (%)'
            ])
    except Exception as e:
        print(f"⚠️ Erro ao calcular parâmetros médios, criando planilha vazia: {e}")
        df_parametros_medios = pd.DataFrame(columns=[
            'Frota', 'Horimetro', 'Uso RTK (%)', 'Horas Elevador', 'Horas Motor',
            'Velocidade Media (km/h)', 'RPM Motor Media', 'RPM Extrator Media',
            'Pressao Corte Media (psi)', 'Corte Base Auto (%)'
        ])

    # Calcular produção por frota (usando toneladas específicas da frente)
    try:
        df_producao = calcular_producao_por_frota(hora_elevador_maquina, caminho_arquivo=caminho_arquivo)
        if df_producao.empty:
            print("⚠️ Nenhuma produção calculada, criando planilha vazia")
            df_producao = pd.DataFrame(columns=['Frota', 'Toneladas', 'Horas Elevador', 'Ton/h'])
    except Exception as e:
        print(f"⚠️ Erro ao calcular produção, criando planilha vazia: {e}")
        df_producao = pd.DataFrame(columns=['Frota', 'Toneladas', 'Horas Elevador', 'Ton/h'])

    # Calcular dados do painel esquerdo
    try:
        df_painel_esquerdo = calcular_painel_esquerdo(
            df_base, horas_por_frota, hora_elevador_maquina, 
            df_manobras_frota, disp_mecanica, df_operadores, df_producao
        )
        if df_painel_esquerdo.empty:
            print("⚠️ Nenhum dado de painel esquerdo calculado")
    except Exception as e:
        print(f"⚠️ Erro ao calcular painel esquerdo: {e}")
        df_painel_esquerdo = pd.DataFrame()

    # Calcular dados do painel direito (será calculado por frota individualmente no envio)
    try:
        # Manter dados globais para referência, mas calcular por frota no envio
        dados_painel_direito_global = calcular_painel_direito_por_frota(df_lavagem, df_ofensores, frota_especifica=None)
    except Exception as e:
        print(f"⚠️ Erro ao calcular painel direito global: {e}")
        dados_painel_direito_global = {
            "lavagem": {"tem_dados": False, "total_intervalos": 0, "tempo_total_horas": 0, "equipamentos": []},
            "ofensores": []
        }

    criar_excel_planilhas_reduzidas(
        df_base=df_base,
        disp_mecanica=disp_mecanica,
        velocidade_media_produtiva=velocidade_media_produtiva,
        uso_gps=uso_gps_maquina,
        motor_ocioso=motor_ocioso_maquina,
        hora_elevador=hora_elevador_maquina,
        df_lavagem=df_lavagem,
        df_roletes=df_roletes,
        df_ofensores=df_ofensores,
        df_intervalos=df_intervalos,
        horas_por_frota=horas_por_frota,
        df_operadores=df_operadores,
        df_coordenadas=df_coordenadas,
        df_manobras_frota=df_manobras_frota,
        df_manobras_operador=df_manobras_operador,
        df_parametros_medios=df_parametros_medios,
        caminho_saida=arquivo_saida,
        df_producao=df_producao,
        df_painel_esquerdo=df_painel_esquerdo
    )
    
    # Enviar parâmetros médios para Supabase
    try:
        print("\n" + "="*50)
        print("📡 ENVIANDO DADOS PARA SUPABASE")
        print("="*50)
        enviar_dados_supabase(df_parametros_medios, df_painel_esquerdo, df_lavagem, df_ofensores, caminho_arquivo)
    except Exception as e:
        print(f"⚠️ Erro ao enviar dados para Supabase: {e}")
        print("   Continuando processamento normalmente...")
    
    # Gerar gráfico de linha do tempo operacional
    try:
        import importlib.util, pathlib
        visual_path = os.path.join(os.path.dirname(__file__), '3_GerarVisualizacaoRelatorios.py')
        spec = importlib.util.spec_from_file_location('vis_op', visual_path)
        vis_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(vis_module)  # type: ignore
        if hasattr(vis_module, 'processar_arquivo_excel'):
            vis_module.processar_arquivo_excel(arquivo_saida, exibir=False)  # type: ignore
    except Exception as e:
        print(f"Erro ao gerar visualização do arquivo {arquivo_saida}: {e}")
    
    print(f"Arquivo {arquivo_saida} gerado com sucesso! (fluxo simplificado)")
    return arquivo_saida

# Sobrescreve a referência anterior para usar a nova implementação
processar_arquivo = processar_arquivo_maquina

def extrair_arquivo_zip(caminho_zip, pasta_destino=None):
    """
    Extrai o conteúdo de um arquivo ZIP para uma pasta temporária ou destino especificado.
    Renomeia os arquivos extraídos para terem o mesmo nome do arquivo ZIP original.
    
    Args:
        caminho_zip (str): Caminho para o arquivo ZIP
        pasta_destino (str, optional): Pasta onde os arquivos serão extraídos.
                                       Se None, usa uma pasta temporária.
    
    Returns:
        list: Lista de caminhos dos arquivos extraídos e renomeados (apenas TXT e CSV)
        str: Caminho da pasta temporária (se criada) ou None
    """
    # Se pasta_destino não foi especificada, criar uma pasta temporária
    pasta_temp = None
    if pasta_destino is None:
        pasta_temp = tempfile.mkdtemp()
        pasta_destino = pasta_temp
    
    arquivos_extraidos = []
    nome_zip_sem_extensao = os.path.splitext(os.path.basename(caminho_zip))[0]
    
    try:
        with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
            # Extrair todos os arquivos do ZIP
            zip_ref.extractall(pasta_destino)
            
            # Processar cada arquivo extraído (apenas TXT e CSV)
            for arquivo in zip_ref.namelist():
                caminho_completo = os.path.join(pasta_destino, arquivo)
                # Verificar se é um arquivo e não uma pasta
                if os.path.isfile(caminho_completo):
                    # Verificar extensão
                    extensao = os.path.splitext(arquivo)[1].lower()
                    if extensao in ['.txt', '.csv']:
                        # Criar novo nome: nome do ZIP + extensão original
                        novo_nome = f"{nome_zip_sem_extensao}{extensao}"
                        novo_caminho = os.path.join(pasta_destino, novo_nome)
                        
                        # Renomear o arquivo extraído
                        try:
                            # Se já existe um arquivo com esse nome, remover primeiro
                            if os.path.exists(novo_caminho):
                                os.remove(novo_caminho)
                            # Renomear o arquivo
                            os.rename(caminho_completo, novo_caminho)
                            arquivos_extraidos.append(novo_caminho)
                            print(f"Arquivo extraído renomeado: {novo_nome}")
                        except Exception as e:
                            print(f"Erro ao renomear arquivo {arquivo} para {novo_nome}: {str(e)}")
                            arquivos_extraidos.append(caminho_completo)  # Adicionar o caminho original em caso de erro
        
        return arquivos_extraidos, pasta_temp
    
    except Exception as e:
        print(f"Erro ao extrair o arquivo ZIP {caminho_zip}: {str(e)}")
        # Se houve erro e criamos uma pasta temporária, tentar limpá-la
        if pasta_temp:
            try:
                print(f"Removendo pasta temporária: {pasta_temp}")
                shutil.rmtree(pasta_temp)
                print(f"Pasta temporária removida com sucesso")
            except Exception as e:
                print(f"Erro ao remover pasta temporária {pasta_temp}: {str(e)}")
        return [], None

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT, CSV ou ZIP de colhedoras na pasta dados.
    Ignora arquivos que contenham "transbordo" no nome.
    """
    print("\nIniciando processamento de arquivos...")
    
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    print(f"Diretório do script: {diretorio_script}")
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    print(f"Diretório raiz: {diretorio_raiz}")
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    print(f"Diretório de dados: {diretorio_dados}")
    print(f"Diretório de saída: {diretorio_saida}")
    
    # Verificar se os diretórios existem, caso contrário criar
    if not os.path.exists(diretorio_dados):
        print(f"Criando diretório de dados: {diretorio_dados}")
        os.makedirs(diretorio_dados)
    if not os.path.exists(diretorio_saida):
        print(f"Criando diretório de saída: {diretorio_saida}")
        os.makedirs(diretorio_saida)
    
    # Encontrar todos os arquivos TXT/CSV/ZIP de colhedoras
    arquivos = []
    arquivos_zip = []
    
    # Adicionar arquivos TXT sempre
    arquivos += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.txt"))
    arquivos += glob.glob(os.path.join(diretorio_dados, "*colhedora*.txt"))
    arquivos += glob.glob(os.path.join(diretorio_dados, "colhedora*.txt"))
    
    # Adicionar arquivos CSV apenas se processCsv for True
    if processCsv:
        arquivos += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.csv"))
        arquivos += glob.glob(os.path.join(diretorio_dados, "*colhedora*.csv"))
        arquivos += glob.glob(os.path.join(diretorio_dados, "colhedora*.csv"))
    
    # Adicionar arquivos ZIP
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.zip"))
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "*colhedora*.zip"))
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "colhedora*.zip"))
    
    print("\nArquivos encontrados antes da filtragem:")
    print(f"TXT/CSV: {[os.path.basename(a) for a in arquivos]}")
    print(f"ZIP: {[os.path.basename(a) for a in arquivos_zip]}")
    
    # Filtrar arquivos que contenham "transbordo" no nome (case insensitive)
    arquivos = [arquivo for arquivo in arquivos if "transbordo" not in os.path.basename(arquivo).lower()]
    arquivos_zip = [arquivo for arquivo in arquivos_zip if "transbordo" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos = list(set(arquivos))
    arquivos_zip = list(set(arquivos_zip))
    
    print("\nArquivos encontrados após a filtragem:")
    print(f"TXT/CSV: {[os.path.basename(a) for a in arquivos]}")
    print(f"ZIP: {[os.path.basename(a) for a in arquivos_zip]}")
    
    if not arquivos and not arquivos_zip:
        print("Nenhum arquivo de colhedoras encontrado na pasta dados!")
        return
    
    print(f"\nEncontrados {len(arquivos)} arquivos de colhedoras (TXT/CSV) para processar.")
    print(f"Encontrados {len(arquivos_zip)} arquivos ZIP de colhedoras para processar.")
    
    # Processar cada arquivo TXT/CSV
    for arquivo in arquivos:
        print(f"\nProcessando arquivo TXT/CSV: {os.path.basename(arquivo)}")
        processar_arquivo(arquivo, diretorio_saida)
    
    # Processar cada arquivo ZIP
    for arquivo_zip in arquivos_zip:
        print(f"\nProcessando arquivo ZIP: {os.path.basename(arquivo_zip)}")
        
        # Extrair arquivo ZIP para pasta temporária
        print(f"Extraindo arquivo ZIP: {arquivo_zip}")
        arquivos_extraidos, pasta_temp = extrair_arquivo_zip(arquivo_zip)
        
        if not arquivos_extraidos:
            print(f"Nenhum arquivo TXT ou CSV encontrado no ZIP {os.path.basename(arquivo_zip)}")
            continue
        
        print(f"Extraídos {len(arquivos_extraidos)} arquivos do ZIP:")
        for arquivo in arquivos_extraidos:
            print(f"  - {os.path.basename(arquivo)}")
        
        # Processar cada arquivo extraído
        for arquivo_extraido in arquivos_extraidos:
            # Filtrar arquivos que contenham "transbordo" no nome
            if "transbordo" not in os.path.basename(arquivo_extraido).lower():
                print(f"\nProcessando arquivo extraído: {os.path.basename(arquivo_extraido)}")
                processar_arquivo(arquivo_extraido, diretorio_saida)
        
        # Limpar pasta temporária se foi criada
        if pasta_temp:
            try:
                print(f"Removendo pasta temporária: {pasta_temp}")
                shutil.rmtree(pasta_temp)
                print(f"Pasta temporária removida com sucesso")
            except Exception as e:
                print(f"Erro ao remover pasta temporária {pasta_temp}: {str(e)}")
    
    print("\nProcessamento de todos os arquivos concluído!")

def carregar_substituicoes_operadores():
    """
    Carrega o arquivo substituiroperadores.json que contém os mapeamentos 
    de substituição de operadores.
    
    Returns:
        dict: Dicionário com mapeamento {operador_origem: operador_destino}
        ou dicionário vazio se o arquivo não existir ou for inválido
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Caminho para o arquivo de substituição
    arquivo_substituicao = os.path.join(diretorio_raiz, "config", "substituiroperadores.json")
    
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo_substituicao):
        print(f"Arquivo de substituição de operadores não encontrado: {arquivo_substituicao}")
        return {}
    
    try:
        # Carregar o arquivo JSON
        with open(arquivo_substituicao, 'r', encoding='utf-8') as f:
            substituicoes = json.load(f)
        
        # Criar dicionário de substituições
        mapeamento = {item['operador_origem']: item['operador_destino'] for item in substituicoes}
        
        print(f"Carregadas {len(mapeamento)} substituições de operadores.")
        return mapeamento
        
    except Exception as e:
        print(f"Erro ao carregar arquivo de substituição de operadores: {str(e)}")
        return {}

def carregar_substituicoes_operadores_horario():
    """
    Carrega o arquivo substituiroperadores_horario.json que contém os mapeamentos 
    de substituição de operadores com intervalos de horário.
    
    Returns:
        list: Lista de dicionários com mapeamentos 
              {operador_origem, operador_destino, hora_inicio, hora_fim, frota_origem}
        ou lista vazia se o arquivo não existir ou for inválido.
        O campo frota_origem é opcional:
        - Se presente e não vazio, a substituição é aplicada apenas a registros daquela frota específica
        - Se ausente ou vazio, a substituição é aplicada a todos os registros do operador
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Caminho para o arquivo de substituição
    arquivo_substituicao = os.path.join(diretorio_raiz, "config", "substituiroperadores_horario.json")
    
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo_substituicao):
        print(f"Arquivo de substituição de operadores com horário não encontrado: {arquivo_substituicao}")
        return []
    
    try:
        # Carregar o arquivo JSON
        with open(arquivo_substituicao, 'r', encoding='utf-8') as f:
            substituicoes = json.load(f)
        
        # Converter strings de hora para objetos datetime.time
        for item in substituicoes:
            if 'hora_inicio' in item and isinstance(item['hora_inicio'], str):
                hora_str = item['hora_inicio']
                # Adicionar segundos se não estiverem presentes
                if len(hora_str.split(':')) == 2:
                    hora_str += ':00'
                item['hora_inicio_obj'] = datetime.strptime(hora_str, '%H:%M:%S').time()
            
            if 'hora_fim' in item and isinstance(item['hora_fim'], str):
                hora_str = item['hora_fim']
                # Adicionar segundos se não estiverem presentes
                if len(hora_str.split(':')) == 2:
                    hora_str += ':00'
                item['hora_fim_obj'] = datetime.strptime(hora_str, '%H:%M:%S').time()
        
        print(f"Carregadas {len(substituicoes)} substituições de operadores com intervalos de horário.")
        return substituicoes
        
    except Exception as e:
        print(f"Erro ao carregar arquivo de substituição de operadores com horário: {str(e)}")
        return []

def aplicar_substituicao_operadores(df, mapeamento_substituicoes, mapeamento_horario=None):
    """
    Aplica as substituições de operadores no DataFrame.
    
    Args:
        df (DataFrame): DataFrame a ser processado
        mapeamento_substituicoes (dict): Dicionário com mapeamento {operador_origem: operador_destino}
        mapeamento_horario (list, optional): Lista de dicionários com mapeamentos
            {operador_origem, operador_destino, hora_inicio, hora_fim, frota_origem}.
            O campo frota_origem é opcional:
            - Se presente e não vazio, a substituição é aplicada apenas a registros daquela frota específica
            - Se ausente ou vazio, a substituição é aplicada a todos os registros do operador
    
    Returns:
        tuple: (DataFrame com substituições aplicadas, DataFrame com registro das substituições)
    """
    if (not mapeamento_substituicoes and not mapeamento_horario) or 'Operador' not in df.columns:
        return df, pd.DataFrame(columns=['ID Original', 'Nome Original', 'ID Nova', 'Nome Novo', 'Registros Afetados'])
    
    # Criar uma cópia para não alterar o DataFrame original
    df_modificado = df.copy()
    
    # Lista para armazenar as substituições realizadas
    substituicoes_realizadas = []
    total_registros_substituidos = 0
    
    # Verificar operadores antes da substituição para relatório
    operadores_antes = df_modificado['Operador'].unique()
    print(f"\nOperadores antes da substituição: {len(operadores_antes)}")
    
    # Aplicar as substituições por horário se disponíveis e se o DataFrame tiver coluna de data/hora
    if mapeamento_horario and 'Data' in df_modificado.columns and 'Hora' in df_modificado.columns:
        # Criar uma cópia backup dos operadores originais
        df_modificado['Operador_Original'] = df_modificado['Operador'].copy()
        
        # Para cada linha no DataFrame
        for idx, row in df_modificado.iterrows():
            # Tenta extrair a hora do registro
            try:
                if isinstance(row['Hora'], str):
                    hora_registro = datetime.strptime(row['Hora'], '%H:%M:%S').time()
                elif hasattr(row['Hora'], 'time'):
                    hora_registro = row['Hora'].time()
                else:
                    continue  # Pula se não conseguir converter
                
                operador = row['Operador']
                
                # Verificar todas as regras de substituição por horário
                for regra in mapeamento_horario:
                    # Verificar se a condição básica é atendida: operador de origem correto e horário dentro do intervalo
                    condicao_basica = (operador == regra['operador_origem'] and 
                                      regra['hora_inicio_obj'] <= hora_registro <= regra['hora_fim_obj'])
                    
                    # Verificar condição de frota se ela existir na regra
                    condicao_frota = True
                    if 'frota_origem' in regra and regra['frota_origem'] and 'Equipamento' in df_modificado.columns:
                        condicao_frota = (row['Equipamento'] == regra['frota_origem'])
                    # Caso contrário, aplica a todos os registros do operador
                    
                    # Aplicar substituição apenas se ambas condições forem atendidas
                    if condicao_basica and condicao_frota:
                        # Aplicar a substituição
                        df_modificado.at[idx, 'Operador'] = regra['operador_destino']
                        break
            except Exception as e:
                print(f"Erro ao processar substituição com horário para linha {idx}: {str(e)}")
        
        # Contar substituições realizadas por horário
        substituicoes_contagem = {}
        for idx, row in df_modificado[df_modificado['Operador'] != df_modificado['Operador_Original']].iterrows():
            origem = row['Operador_Original']
            destino = row['Operador']
            chave = (origem, destino)
            if chave in substituicoes_contagem:
                substituicoes_contagem[chave] += 1
            else:
                substituicoes_contagem[chave] = 1
        
        # Adicionar as substituições com horário à lista de substituições realizadas
        for (origem, destino), count in substituicoes_contagem.items():
            total_registros_substituidos += count
            id_original = origem.split(' - ')[0] if ' - ' in origem else origem
            nome_original = origem.split(' - ')[1] if ' - ' in origem else ''
            id_nova = destino.split(' - ')[0] if ' - ' in destino else destino
            nome_novo = destino.split(' - ')[1] if ' - ' in destino else ''
            
            substituicoes_realizadas.append({
                'ID Original': id_original,
                'Nome Original': nome_original,
                'ID Nova': id_nova,
                'Nome Novo': nome_novo,
                'Registros Afetados': count,
                'Por Horário': True
            })
            print(f"Operador '{origem}' substituído por '{destino}' em {count} registros (por horário)")
        
        # Remover a coluna temporária
        df_modificado.drop('Operador_Original', axis=1, inplace=True)
    
    # Contar operadores e registros antes da substituição padrão
    contagem_antes = df_modificado['Operador'].value_counts().to_dict()
    
    # Aplicar as substituições padrão (sem horário)
    for origem, destino in mapeamento_substituicoes.items():
        # Verificar se o operador de origem existe no DataFrame
        registros_afetados = df_modificado[df_modificado['Operador'] == origem].shape[0]
        
        if registros_afetados > 0:
            # Substituir o operador
            df_modificado.loc[df_modificado['Operador'] == origem, 'Operador'] = destino
            
            total_registros_substituidos += registros_afetados
            
            # Extrair IDs e nomes
            id_original = origem.split(' - ')[0] if ' - ' in origem else origem
            nome_original = origem.split(' - ')[1] if ' - ' in origem else ''
            id_nova = destino.split(' - ')[0] if ' - ' in destino else destino
            nome_novo = destino.split(' - ')[1] if ' - ' in destino else ''
            
            substituicoes_realizadas.append({
                'ID Original': id_original,
                'Nome Original': nome_original,
                'ID Nova': id_nova, 
                'Nome Novo': nome_novo,
                'Registros Afetados': registros_afetados,
                'Por Horário': False
            })
            print(f"Operador '{origem}' substituído por '{destino}' em {registros_afetados} registros")
    
    # Verificar operadores após substituição
    operadores_depois = df_modificado['Operador'].unique()
    print(f"Operadores após substituição: {len(operadores_depois)}")
    print(f"Total de registros substituídos: {total_registros_substituidos}")
    
    # Criar DataFrame com as substituições realizadas
    df_substituicoes = pd.DataFrame(substituicoes_realizadas)
    
    return df_modificado, df_substituicoes

# === NOVA FUNÇÃO: Uso de GPS por máquina ===

def calcular_uso_gps_maquina(df):
    """
    Calcula o percentual de uso de GPS (RTK) por máquina (Equipamento).
    Usa Diferença_Hora para todos os cálculos.
    Aplica filtros: Produtiva + Pressão ≥ 400 + Velocidade > 0

    Args:
        df (DataFrame): DataFrame processado

    Returns:
        DataFrame: Colunas 'Frota', 'Porcentagem' com o percentual de uso de GPS por máquina
    """
    resultados = []
    equipamentos = df['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    
    for equipamento in equipamentos:
        dados = df[df['Equipamento'] == equipamento]
        
        print(f"\n=== Calculando uso GPS para {equipamento} ===")
        
        # CÁLCULO ATUAL: Usar Diferença_Hora para horas produtivas (mesmos filtros das horas elevador)
        horas_prod = dados[
            (dados['Grupo Operacao'] == 'Produtiva') & 
            (dados['Pressao de Corte'] >= 400) &
            (dados['Velocidade'] > 0)
        ]['Diferença_Hora'].sum()
        
        # CÁLCULO ATUAL: Usar Diferença_Hora para RTK
        rtk = dados[
            (dados['Grupo Operacao'] == 'Produtiva') &
            (dados['Pressao de Corte'] >= 400) &
            (dados['Velocidade'] > 0) &
            (dados['RTK (Piloto Automatico)'] == 1)
        ]['Diferença_Hora'].sum()
        
        porcentagem = rtk / horas_prod if horas_prod > 0 else 0
        
        print(f"CÁLCULO USO GPS (Produtiva + Pressão ≥ 400 + Velocidade > 0):")
        print(f"  Horas Produtivas: {horas_prod:.4f} h")
        print(f"  RTK: {rtk:.4f} h")
        print(f"  % Uso GPS: {porcentagem:.4f} ({porcentagem*100:.2f}%)")
        
        resultados.append({
            'Frota': equipamento, 
            'Porcentagem': porcentagem
        })
    return pd.DataFrame(resultados)

def criar_excel_planilhas_reduzidas(df_base, disp_mecanica, velocidade_media_produtiva, uso_gps, motor_ocioso, hora_elevador, df_lavagem, df_roletes, df_ofensores, df_intervalos, horas_por_frota, df_operadores, df_coordenadas, df_manobras_frota, df_manobras_operador, df_parametros_medios, caminho_saida, df_producao=None, df_painel_esquerdo=None):
    """
    Gera arquivo Excel contendo planilhas BASE, Lavagem, Ofensores, Intervalos, Horas por Frota, Disponibilidade Mecânica, Velocidade Média Produtiva, Uso GPS e Hora Elevador por máquina.

    A planilha de Lavagem será sempre criada, mesmo quando não houver registros de lavagem 
    (neste caso, uma linha informativa será exibida).

    Args:
        df_base (DataFrame): Dados base processados
        disp_mecanica (DataFrame): Disponibilidade mecânica por máquina
        velocidade_media_produtiva (DataFrame): Velocidade média produtiva por máquina
        uso_gps (DataFrame): Uso de GPS por máquina
        motor_ocioso (DataFrame): Motor ocioso por máquina
        hora_elevador (DataFrame): Horas elevador por máquina
        df_lavagem (DataFrame): Intervalos de lavagem (sempre presente)
        df_roletes (DataFrame): Intervalos de aferição de roletes (sempre presente)
        df_ofensores (DataFrame): Top 5 ofensores gerais
        df_intervalos (DataFrame): Intervalos operacionais
        horas_por_frota (DataFrame): Horas registradas por frota
        caminho_saida (str): Caminho do arquivo Excel a ser criado
    """
    def _ajustar_largura_colunas(worksheet, max_preview_rows: int = 20):
        """Ajusta a largura das colunas com base no conteúdo (pré-visualiza até max_preview_rows linhas)."""
        for col in worksheet.columns:
            max_length = 10  # largura mínima
            column_letter = col[0].column_letter
            header_val = str(col[0].value) if col[0].value is not None else ""
            max_length = max(max_length, len(header_val) + 2)

            # Limitar amostra para performance
            for cell in col[1:max_preview_rows]:
                if cell.value is not None:
                    cell_len = len(str(cell.value)) + 2
                    if cell_len > max_length:
                        max_length = cell_len

            # Limitar largura extrema
            max_length = min(max_length, 40)
            worksheet.column_dimensions[column_letter].width = max_length
 
    # Gerar arquivo CSV das coordenadas
    nome_base_original = os.path.splitext(os.path.basename(caminho_saida))[0]
    nome_base_original = nome_base_original.replace('_processado', '')  # Remover sufixo se existir
    diretorio_saida = os.path.dirname(caminho_saida)
    caminho_csv_coordenadas = os.path.join(diretorio_saida, f"{nome_base_original}_Coordenadas.csv")
    
    try:
        df_coordenadas.to_csv(caminho_csv_coordenadas, index=False, encoding='utf-8-sig', sep=';')
        print(f"Arquivo CSV de coordenadas gerado: {os.path.basename(caminho_csv_coordenadas)}")
    except Exception as e:
        print(f"Erro ao gerar arquivo CSV de coordenadas: {str(e)}")
    
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        # Planilha BASE
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        
        # Parâmetros Médios (logo após BASE)
        if df_parametros_medios is not None:
            df_parametros_medios.to_excel(writer, sheet_name='Parâmetros Médios', index=False)
        
        # Disponibilidade Mecânica
        disp_mecanica.to_excel(writer, sheet_name='Disponibilidade Mecânica', index=False)
        # Velocidade Média Produtiva
        velocidade_media_produtiva.to_excel(writer, sheet_name='Velocidade Média Produtiva', index=False)
        # Uso GPS por máquina
        uso_gps.to_excel(writer, sheet_name='Uso GPS', index=False)
        # Motor Ocioso por máquina
        motor_ocioso.to_excel(writer, sheet_name='Motor Ocioso', index=False)
        # Hora Elevador por máquina
        hora_elevador.to_excel(writer, sheet_name='Eficiência Energética', index=False)
        # Lavagem (sempre incluir, mesmo se não houver registros)
        if df_lavagem is not None:
            df_lavagem.to_excel(writer, sheet_name='Lavagem', index=False)
        
        # Roletes
        if df_roletes is not None:
            df_roletes.to_excel(writer, sheet_name='Roletes', index=False)
        
        # Intervalos (caso exista)
        if df_intervalos is not None and not df_intervalos.empty:
            df_intervalos.to_excel(writer, sheet_name='Intervalos', index=False)
        
        # Horas por Frota
        if horas_por_frota is not None and not horas_por_frota.empty:
            horas_por_frota.to_excel(writer, sheet_name='Horas por Frota', index=False)
        
        # Ofensores (caso exista)
        if df_ofensores is not None and not df_ofensores.empty:
            df_ofensores.to_excel(writer, sheet_name='Ofensores', index=False)
        
        # Coordenadas
        df_coordenadas.to_excel(writer, sheet_name='Coordenadas', index=False)
        
        # Operadores por Frota (sempre incluir, mesmo se vazio)
        if df_operadores is not None:
            df_operadores.to_excel(writer, sheet_name='Operadores', index=False)
        
        # Manobras por Frota
        if df_manobras_frota is not None and not df_manobras_frota.empty:
            df_manobras_frota.to_excel(writer, sheet_name='Manobras', index=False)
        
        # Produção por Frota (nova planilha)
        if df_producao is not None and not df_producao.empty:
            df_producao.to_excel(writer, sheet_name='Produção', index=False)

        # Ajustar largura das colunas e aplicar formatação específica
        wb = writer.book
        for sh_name in wb.sheetnames:
            ws = wb[sh_name]
            # Formato de hora para nova coluna, se existir
            if sh_name == 'BASE':
                # Encontrar índice da nova coluna
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == 'Diferença_Hora_hhmm':
                        for row in range(2, ws.max_row + 1):
                            ws.cell(row=row, column=col).number_format = 'h:mm:ss'
                        break
            _ajustar_largura_colunas(ws)
            # Congelar a primeira linha (cabeçalho)
            ws.freeze_panes = ws['A2']
            # Formatação específica
            if sh_name == 'Disponibilidade Mecânica':
                for row in range(2, ws.max_row + 1):
                    # Coluna B (Disponibilidade)
                    ws.cell(row=row, column=2).number_format = '0.00%'
            
            elif sh_name == 'Velocidade Média Produtiva':
                for row in range(2, ws.max_row + 1):
                    # Coluna B (Velocidade Média Produtiva)
                    ws.cell(row=row, column=2).number_format = '0.00'
            
            elif sh_name == 'Uso GPS':
                for row in range(2, ws.max_row + 1):
                    # Coluna B (Porcentagem)
                    ws.cell(row=row, column=2).number_format = '0.00%'

            
            elif sh_name == 'Motor Ocioso':
                for row in range(2, ws.max_row + 1):
                    # Coluna B Porcentagem
                    ws.cell(row=row, column=2).number_format = '0.00%'
                    # Coluna C (Horas Motor)
                    ws.cell(row=row, column=3).number_format = '0.00'
                    # Coluna D (Tempo Ocioso)
                    ws.cell(row=row, column=4).number_format = '0.00'
            
            elif sh_name == 'Eficiência Energética':
                for row in range(2, ws.max_row + 1):
                    # Coluna B (Eficiência)
                    ws.cell(row=row, column=2).number_format = '0.00%'
            
            elif sh_name == 'Parâmetros Médios':
                for row in range(2, ws.max_row + 1):
                    # Coluna B (Horimetro) - 2 casas decimais
                    ws.cell(row=row, column=2).number_format = '0.00'
                    # Coluna C (Uso RTK %) - 2 casas decimais
                    ws.cell(row=row, column=3).number_format = '0.00'
                    # Coluna D (Horas Elevador) - 2 casas decimais
                    ws.cell(row=row, column=4).number_format = '0.00'
                    # Coluna E (Horas Motor) - 2 casas decimais
                    ws.cell(row=row, column=5).number_format = '0.00'
                    # Coluna F (Velocidade Media) - 2 casas decimais
                    ws.cell(row=row, column=6).number_format = '0.00'
                    # Coluna G (RPM Motor Media) - 2 casas decimais
                    ws.cell(row=row, column=7).number_format = '0.00'
                    # Coluna H (RPM Extrator Media) - 2 casas decimais
                    ws.cell(row=row, column=8).number_format = '0.00'
                    # Coluna I (Pressao Corte Media - psi) - 2 casas decimais
                    ws.cell(row=row, column=9).number_format = '0.00'
                    # Coluna J (Corte Base Auto %) - 2 casas decimais
                    ws.cell(row=row, column=10).number_format = '0.00'
            
            elif sh_name == 'Ofensores':
                if df_ofensores is not None and not df_ofensores.empty:
                    for row in range(2, ws.max_row + 1):
                        # Coluna B (Tempo)
                        ws.cell(row=row, column=2).number_format = '0.00'  # Formato decimal
                        # Coluna C (Porcentagem)
                        ws.cell(row=row, column=3).number_format = '0.00%'  # Formato percentual
            
            elif sh_name == 'Lavagem':
                if df_lavagem is not None:
                    for row in range(2, ws.max_row + 1):
                        # Verificar se não é a linha informativa (para não formatar como hora)
                        equipamento_cell = ws.cell(row=row, column=2)
                        if equipamento_cell.value != 'NÃO FORAM ENCONTRADOS DADOS DE LAVAGEM PARA A DATA INFORMADA':
                            # Coluna D (Início)
                            cell = ws.cell(row=row, column=4)
                            cell.number_format = 'hh:mm:ss'
                            
                            # Coluna E (Fim)
                            cell = ws.cell(row=row, column=5)
                            cell.number_format = 'hh:mm:ss'
                            
                            # Coluna F (Duração)
                            cell = ws.cell(row=row, column=6)
                            cell.number_format = '0.00'
                            
                            # Coluna G (Tempo Total do Dia) - só formatar se não for None/vazio
                            cell = ws.cell(row=row, column=7)
                            if cell.value is not None and cell.value != "":
                                cell.number_format = '0.00'
            
            elif sh_name == 'Intervalos':
                if df_intervalos is not None and not df_intervalos.empty:
                    for row in range(2, ws.max_row + 1):
                        # Identificar posição das colunas Início, Fim e Duração
                        header_row = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
                        
                        # Formatar Início (hh:mm:ss)
                        if 'Início' in header_row:
                            col_inicio = header_row.index('Início') + 1
                            ws.cell(row=row, column=col_inicio).number_format = 'hh:mm:ss'
                        
                        # Formatar Fim (hh:mm:ss)
                        if 'Fim' in header_row:
                            col_fim = header_row.index('Fim') + 1
                            ws.cell(row=row, column=col_fim).number_format = 'hh:mm:ss'
                        
                        # Formatar Duração (0.00)
                        if 'Duração (horas)' in header_row:
                            col_duracao = header_row.index('Duração (horas)') + 1
                            ws.cell(row=row, column=col_duracao).number_format = '0.00'
                        
                        # Formatar Duração (hh:mm)
                        if 'Duração (hh:mm)' in header_row:
                            col_duracao_hhmm = header_row.index('Duração (hh:mm)') + 1
                            ws.cell(row=row, column=col_duracao_hhmm).number_format = 'h:mm:ss'
            
            elif sh_name == 'Horas por Frota':
                if horas_por_frota is not None and not horas_por_frota.empty:
                    for row in range(2, ws.max_row + 1):
                        # Formatar as colunas de horas (numéricas) como decimal
                        for col in range(2, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                cell.number_format = '0.00'
            
            elif sh_name == 'Coordenadas':
                # Formatação da planilha de coordenadas
                for row in range(2, ws.max_row + 1):
                    # Hora
                    ws.cell(row=row, column=2).number_format = 'hh:mm:ss'
                    # Latitude
                    ws.cell(row=row, column=3).number_format = '0.000000'
                    # Longitude
                    ws.cell(row=row, column=4).number_format = '0.000000'
                    # Velocidade
                    ws.cell(row=row, column=5).number_format = '0.00'
            
            elif sh_name in ['Manobras Frotas', 'Manobras Operador']:
                # Formatação das novas colunas de manobras por intervalos
                for row in range(2, ws.max_row + 1):
                    # Coluna 2: Intervalos Válidos
                    ws.cell(row=row, column=2).number_format = '0'
                    # Coluna 3: Tempo Total (horas)
                    ws.cell(row=row, column=3).number_format = '0.0000'
                    # Coluna 4: Tempo Médio (horas)
                    ws.cell(row=row, column=4).number_format = '0.0000'
                
                # Adicionar colunas formatadas como hh:mm:ss
                if ws.max_row > 1:  # Só se houver dados
                    # Adicionar cabeçalhos para colunas de tempo formatado
                    ws.cell(row=1, column=ws.max_column + 1).value = 'Tempo Total (hh:mm)'
                    ws.cell(row=1, column=ws.max_column + 1).value = 'Tempo Médio (hh:mm)'
                    
                    for row in range(2, ws.max_row + 1):
                        # Tempo Total em formato hh:mm:ss
                        tempo_total = ws.cell(row=row, column=3).value
                        ws.cell(row=row, column=ws.max_column - 1).value = tempo_total / 24 if tempo_total else 0
                        ws.cell(row=row, column=ws.max_column - 1).number_format = 'h:mm:ss'
                        
                        # Tempo Médio em formato hh:mm:ss
                        tempo_medio = ws.cell(row=row, column=4).value
                        ws.cell(row=row, column=ws.max_column).value = tempo_medio / 24 if tempo_medio else 0
                        ws.cell(row=row, column=ws.max_column).number_format = 'h:mm:ss'
            
            _ajustar_largura_colunas(ws)

# === NOVA FUNÇÃO: motor ocioso por máquina ===

def calcular_motor_ocioso_maquina(df: pd.DataFrame) -> pd.DataFrame:
    """Calcula o percentual de motor ocioso por máquina.

    A coluna `Motor Ocioso Correto` deve estar presente no DataFrame (calculada por
    `calcular_motor_ocioso_correto`). O cálculo é baseado em:

    horas_motor   = soma de Diferença_Hora onde Motor Ligado == 1
    tempo_ocioso  = soma da coluna Motor Ocioso Correto
    porcentagem   = tempo_ocioso / horas_motor
    """
    resultados = []
    equipamentos = df['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    
    for equipamento in equipamentos:
        dados = df[df['Equipamento'] == equipamento]
        
        print(f"\n=== Calculando motor ocioso para {equipamento} ===")
        
        # Usar Diferença_Hora para horas motor
        horas_motor = dados[dados['Motor Ligado'] == 1]['Diferença_Hora'].sum()
        
        # Tempo ocioso correto
        tempo_ocioso = dados['Motor Ocioso Correto'].sum()
        
        porcentagem = tempo_ocioso / horas_motor if horas_motor > 0 else 0

        print(f"Horas Motor (Diferença_Hora): {horas_motor:.4f} h")
        print(f"Tempo Ocioso (método correto): {tempo_ocioso:.4f} h")
        print(f"% Ocioso: {porcentagem:.4f} ({porcentagem*100:.2f}%)")

        resultados.append({
            'Frota': equipamento,
            'Porcentagem': porcentagem,
            'Horas Motor': horas_motor,
            'Tempo Ocioso': tempo_ocioso
        })

    return pd.DataFrame(resultados)

def calcular_intervalos_operacionais(df_base):
    """
    Calcula intervalos operacionais sequenciais baseados no Grupo Operacao.
    
    Lógica:
    1. Analisa registros sequencialmente (por data/hora)
    2. Agrupa registros consecutivos do mesmo Grupo Operacao
    3. Classifica intervalos: Manutenção, Produtivo, Disponível
    4. Calcula início, fim e duração de cada intervalo
    5. Agrupa por equipamento
    
    Args:
        df_base (DataFrame): DataFrame com os dados base
    
    Returns:
        DataFrame: DataFrame com intervalos operacionais detalhados
    """
    print(f"\n=== CALCULANDO INTERVALOS OPERACIONAIS ===")
    print("Agrupando registros sequenciais por Grupo Operacao")
    print("="*60)
    
    # Verificar se as colunas necessárias existem
    if not {'Equipamento', 'Grupo Operacao', 'Diferença_Hora', 'Hora'}.issubset(df_base.columns):
        print("Colunas necessárias não encontradas para cálculo de intervalos")
        return pd.DataFrame(columns=['Data', 'Equipamento', 'Intervalo', 'Tipo', 'Início', 'Fim', 'Duração (horas)'])
    
    # Filtrar registros inválidos antes do cálculo
    print("Filtrando registros inválidos...")
    registros_antes = len(df_base)
    
    # Remover registros com Diferença_Hora muito pequena ou zero
    # Usando constante global para tempo mínimo válido
    df_filtrado = df_base[
        (df_base['Diferença_Hora'] >= TEMPO_MINIMO_VALIDO) &
        (df_base['Diferença_Hora'].notna())
    ].copy()
    
    registros_depois = len(df_filtrado)
    registros_removidos = registros_antes - registros_depois
    
    print(f"Registros originais: {registros_antes}")
    print(f"Registros após filtro: {registros_depois}")
    print(f"Registros removidos: {registros_removidos} ({registros_removidos/registros_antes*100:.1f}%)")
    
    if registros_removidos > 0:
        print(f"Filtro aplicado: Diferença_Hora >= {TEMPO_MINIMO_VALIDO} horas ({TEMPO_MINIMO_VALIDO*3600:.1f} segundos)")
    
    # Usar o DataFrame filtrado para o restante do cálculo
    df_para_calculo = df_filtrado
    
    # Ordenar por equipamento, data e hora para análise sequencial
    colunas_ordenacao = ['Equipamento']
    if 'Data' in df_para_calculo.columns:
        colunas_ordenacao.append('Data')
    if 'Hora' in df_para_calculo.columns:
        colunas_ordenacao.append('Hora')
    
    df_ordenado = df_para_calculo.sort_values(colunas_ordenacao).reset_index(drop=True)
    
    resultados = []
    
    # Processar cada equipamento separadamente
    equipamentos = df_ordenado['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    
    for equipamento in equipamentos:
        dados_equip = df_ordenado[df_ordenado['Equipamento'] == equipamento].reset_index(drop=True)
        
        print(f"\n--- Processando {equipamento} ---")
        print(f"Total de registros: {len(dados_equip)}")
        
        if len(dados_equip) == 0:
            continue
        
        # Agrupar por data se existir
        if 'Data' in dados_equip.columns:
            for data in dados_equip['Data'].unique():
                dados_dia = dados_equip[dados_equip['Data'] == data].reset_index(drop=True)
                intervalos_dia = processar_intervalos_dia(dados_dia, equipamento, data)
                resultados.extend(intervalos_dia)
        else:
            intervalos_equip = processar_intervalos_dia(dados_equip, equipamento, None)
            resultados.extend(intervalos_equip)
    
    # Criar DataFrame final
    df_intervalos = pd.DataFrame(resultados)
    
    print(f"\n=== RESUMO GERAL ===")
    print(f"Total de intervalos identificados: {len(df_intervalos)}")
    if not df_intervalos.empty:
        print(f"Intervalos por tipo:")
        for tipo in df_intervalos['Tipo'].unique():
            count = len(df_intervalos[df_intervalos['Tipo'] == tipo])
            duracao_total = df_intervalos[df_intervalos['Tipo'] == tipo]['Duração (horas)'].sum()
            print(f"  {tipo}: {count} intervalos, {duracao_total:.2f}h total")
        
        # Verificar se a soma das durações bate com as horas registradas
        print(f"\n=== VERIFICAÇÃO DE CONSISTÊNCIA ===")
        print("Soma das durações dos intervalos por equipamento:")
        for equipamento in df_intervalos['Equipamento'].unique():
            soma_duracao = df_intervalos[df_intervalos['Equipamento'] == equipamento]['Duração (horas)'].sum()
            # Calcular horas registradas do equipamento (usando dados filtrados)
            horas_registradas = df_para_calculo[df_para_calculo['Equipamento'] == equipamento]['Diferença_Hora'].sum()
            diferenca = abs(soma_duracao - horas_registradas)
            print(f"  {equipamento}: Intervalos={soma_duracao:.4f}h, Registros={horas_registradas:.4f}h, Diferença={diferenca:.4f}h")
    print("="*60)
    
    return df_intervalos

def processar_intervalos_dia(dados_dia, equipamento, data):
    """
    Processa intervalos operacionais para um equipamento em um dia específico.
    
    Args:
        dados_dia (DataFrame): Dados do equipamento para um dia
        equipamento (str): Nome do equipamento
        data (str): Data (pode ser None)
    
    Returns:
        list: Lista de intervalos encontrados
    """
    if len(dados_dia) == 0:
        return []
    
    intervalos = []
    
    # Variáveis para controle do intervalo atual
    intervalo_atual = None
    inicio_idx = 0
    
    for idx, row in dados_dia.iterrows():
        grupo_operacao = row['Grupo Operacao']
        
        # Classificar o tipo de intervalo
        if grupo_operacao == 'Manutenção':
            tipo_intervalo = 'Manutenção'
        elif grupo_operacao == 'Produtiva':
            tipo_intervalo = 'Produtivo'
        else:
            tipo_intervalo = 'Disponível'
        
        # Se é o primeiro registro ou mudou o tipo de intervalo
        if intervalo_atual is None or intervalo_atual != tipo_intervalo:
            # Finalizar intervalo anterior se existir
            if intervalo_atual is not None:
                # Para intervalos de um único registro, passar o índice do próximo
                proximo_inicio = idx if inicio_idx == idx-1 else None
                intervalo_info = finalizar_intervalo(dados_dia, inicio_idx, idx-1, intervalo_atual, equipamento, data, len(intervalos)+1, proximo_inicio)
                if intervalo_info:
                    intervalos.append(intervalo_info)
            
            # Iniciar novo intervalo
            intervalo_atual = tipo_intervalo
            inicio_idx = idx
    
    # Finalizar último intervalo
    if intervalo_atual is not None:
        intervalo_info = finalizar_intervalo(dados_dia, inicio_idx, len(dados_dia)-1, intervalo_atual, equipamento, data, len(intervalos)+1)
        if intervalo_info:
            intervalos.append(intervalo_info)
    
    return intervalos

def finalizar_intervalo(dados_dia, inicio_idx, fim_idx, tipo_intervalo, equipamento, data, numero_intervalo, proximo_inicio_idx=None):
    """
    Finaliza um intervalo operacional calculando início, fim e duração.
    
    Args:
        dados_dia (DataFrame): Dados do dia
        inicio_idx (int): Índice de início do intervalo
        fim_idx (int): Índice de fim do intervalo
        tipo_intervalo (str): Tipo do intervalo (Manutenção, Produtivo, Disponível)
        equipamento (str): Nome do equipamento
        data (str): Data (pode ser None)
        numero_intervalo (int): Número sequencial do intervalo
        proximo_inicio_idx (int, optional): Índice do próximo intervalo para intervalos únicos
    
    Returns:
        dict: Informações do intervalo ou None se inválido
    """
    if inicio_idx > fim_idx:
        return None
    
    # Calcular duração total do intervalo
    duracao_total = dados_dia.iloc[inicio_idx:fim_idx+1]['Diferença_Hora'].sum()
    
    # Obter hora de início
    hora_inicio = dados_dia.iloc[inicio_idx]['Hora']
    
    # Obter hora de fim
    if inicio_idx == fim_idx and proximo_inicio_idx is not None:
        # Para intervalos de um único registro, usar hora do próximo intervalo
        hora_fim = dados_dia.iloc[proximo_inicio_idx]['Hora']
    else:
        # Para intervalos múltiplos, usar hora do último registro
        hora_fim = dados_dia.iloc[fim_idx]['Hora']
    
    # ------------------------------------------------------------------
    # VALIDAR INTERVALO – descartar se duração muito pequena ou se hora
    # de início e fim coincidem (possível erro de coleta).
    # ------------------------------------------------------------------
    if duracao_total < TEMPO_MINIMO_VALIDO or hora_inicio == hora_fim:
        # Intervalo considerado inválido
        return None
    
    # Criar informações do intervalo
    intervalo_info = {
        'Equipamento': equipamento,
        'Intervalo': f"Intervalo {numero_intervalo}",
        'Tipo': tipo_intervalo,
        'Início': hora_inicio,
        'Fim': hora_fim,
        'Duração (horas)': duracao_total,
        'Duração (hh:mm)': duracao_total / 24  # Para formatação como tempo
    }
    
    # Adicionar data se disponível
    if data is not None:
        intervalo_info['Data'] = data
    
    return intervalo_info

def calcular_manobras_por_intervalos(df_base):
    """
    Calcula manobras por intervalos sequenciais, agrupando por frota/equipamento.
    
    Lógica:
    1. Analisa registros sequencialmente (por data/hora)
    2. Agrupa manobras consecutivas em intervalos
    3. Interrompe intervalo se há registro NÃO-MANOBRA ≥ 30 segundos
    4. Continua intervalo se há registro NÃO-MANOBRA < 30 segundos (SEM somar sua duração)
    5. Soma apenas a duração de registros com Estado = 'MANOBRA'
    6. Filtra intervalos pelo tempo mínimo configurado
    7. Calcula métricas por frota: quantidade, tempo total, tempo médio
    
    Args:
        df_base (DataFrame): DataFrame com os dados base
    
    Returns:
        tuple: (df_manobras_frota, df_manobras_operador_vazio)
    """
    print(f"\n=== CALCULANDO MANOBRAS POR INTERVALOS ===")
    print(f"Tempo mínimo por intervalo: {tempoMinimoManobras} segundos")
    print(f"Tolerância para interrupção: 30 segundos")
    print(f"REGRA: Apenas tempo de operações MANOBRA é somado ao intervalo")
    print(f"       Operações NÃO-MANOBRA < 30s não interrompem mas não são somadas")
    print("="*60)
    
    # Converter tempo mínimo para horas
    tempo_minimo_horas = tempoMinimoManobras / 3600
    tolerancia_interrupcao = 30 / 3600  # 30 segundos em horas
    
    # Verificar se as colunas necessárias existem
    if not {'Equipamento', 'Estado', 'Diferença_Hora'}.issubset(df_base.columns):
        print("Colunas necessárias não encontradas para cálculo de manobras")
        return pd.DataFrame(columns=['Frota', 'Intervalos Válidos', 'Tempo Total', 'Tempo Médio']), pd.DataFrame()
    
    # Filtrar registros inválidos antes do cálculo
    print("Filtrando registros inválidos...")
    registros_antes = len(df_base)
    
    # Remover registros com Diferença_Hora muito pequena ou zero
    # Usando constante global para tempo mínimo válido
    df_filtrado = df_base[
        (df_base['Diferença_Hora'] >= TEMPO_MINIMO_VALIDO) &
        (df_base['Diferença_Hora'].notna())
    ].copy()
    
    registros_depois = len(df_filtrado)
    registros_removidos = registros_antes - registros_depois
    
    print(f"Registros originais: {registros_antes}")
    print(f"Registros após filtro: {registros_depois}")
    print(f"Registros removidos: {registros_removidos} ({registros_removidos/registros_antes*100:.1f}%)")
    
    if registros_removidos > 0:
        print(f"Filtro aplicado: Diferença_Hora >= {TEMPO_MINIMO_VALIDO} horas ({TEMPO_MINIMO_VALIDO*3600:.1f} segundos)")
    
    # Usar o DataFrame filtrado para o restante do cálculo
    df_para_calculo = df_filtrado
    
    # Ordenar por equipamento, data e hora para análise sequencial
    colunas_ordenacao = ['Equipamento']
    if 'Data' in df_para_calculo.columns:
        colunas_ordenacao.append('Data')
    if 'Hora' in df_para_calculo.columns:
        colunas_ordenacao.append('Hora')
    
    df_ordenado = df_para_calculo.sort_values(colunas_ordenacao).reset_index(drop=True)
    
    resultados_frota = []
    
    # Processar cada equipamento separadamente
    equipamentos = df_ordenado['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    
    for equipamento in equipamentos:
        dados_equip = df_ordenado[df_ordenado['Equipamento'] == equipamento].reset_index(drop=True)
        
        print(f"\n--- Processando {equipamento} ---")
        print(f"Total de registros: {len(dados_equip)}")
        
        intervalos_validos = []
        intervalo_atual = {
            'tempo_total': 0,
            'inicio_idx': None,
            'em_intervalo': False
        }
        
        for idx, row in dados_equip.iterrows():
            estado = row['Estado']
            tempo_registro = row['Diferença_Hora']
            
            if estado == 'MANOBRA':
                # Registro de manobra
                if not intervalo_atual['em_intervalo']:
                    # Inicia novo intervalo
                    intervalo_atual['tempo_total'] = tempo_registro
                    intervalo_atual['inicio_idx'] = idx
                    intervalo_atual['em_intervalo'] = True
                else:
                    # Continua intervalo existente
                    intervalo_atual['tempo_total'] += tempo_registro
            
            else:
                # Registro NÃO-MANOBRA
                if intervalo_atual['em_intervalo']:
                    if tempo_registro >= tolerancia_interrupcao:
                        # Interrupção >= 30 segundos: fecha o intervalo
                        if intervalo_atual['tempo_total'] >= tempo_minimo_horas:
                            intervalos_validos.append({
                                'tempo_total': intervalo_atual['tempo_total'],
                                'inicio_idx': intervalo_atual['inicio_idx'],
                                'fim_idx': idx - 1
                            })
                            print(f"  Intervalo válido: {intervalo_atual['tempo_total']*3600:.1f}s (registros {intervalo_atual['inicio_idx']}-{idx-1})")
                        else:
                            print(f"  Intervalo descartado: {intervalo_atual['tempo_total']*3600:.1f}s < {tempoMinimoManobras}s")
                        
                        # Reset do intervalo
                        intervalo_atual = {'tempo_total': 0, 'inicio_idx': None, 'em_intervalo': False}
                    else:
                        # Interrupção < 30 segundos: continua o intervalo (NÃO soma a duração da pausa)
                        print(f"  Pausa curta ({tempo_registro*3600:.1f}s) ignorada, intervalo continua")
        
        # Processar último intervalo se ainda estiver aberto
        if intervalo_atual['em_intervalo']:
            if intervalo_atual['tempo_total'] >= tempo_minimo_horas:
                intervalos_validos.append({
                    'tempo_total': intervalo_atual['tempo_total'],
                    'inicio_idx': intervalo_atual['inicio_idx'],
                    'fim_idx': len(dados_equip) - 1
                })
                print(f"  Último intervalo válido: {intervalo_atual['tempo_total']*3600:.1f}s")
            else:
                print(f"  Último intervalo descartado: {intervalo_atual['tempo_total']*3600:.1f}s < {tempoMinimoManobras}s")
        
        # Calcular métricas para este equipamento
        num_intervalos = len(intervalos_validos)
        tempo_total = sum(intervalo['tempo_total'] for intervalo in intervalos_validos)
        tempo_medio = tempo_total / num_intervalos if num_intervalos > 0 else 0
        
        print(f"  Resultado: {num_intervalos} intervalos válidos, {tempo_total*3600:.1f}s total, {tempo_medio*3600:.1f}s médio")
        
        resultados_frota.append({
            'Frota': equipamento,
            'Intervalos Válidos': num_intervalos,
            'Tempo Total': tempo_total,
            'Tempo Médio': tempo_medio
        })
    
    # Criar DataFrame final ordenado por tempo total (decrescente)
    df_manobras_frota = pd.DataFrame(resultados_frota)
    df_manobras_frota = df_manobras_frota.sort_values('Tempo Total', ascending=False)
    
    # Retornar DataFrame vazio para operador (não usado neste arquivo)
    df_manobras_operador_vazio = pd.DataFrame(columns=['Operador', 'Intervalos Válidos', 'Tempo Total', 'Tempo Médio'])
    
    print(f"\n=== RESUMO GERAL ===")
    print(f"Total de equipamentos processados: {len(df_manobras_frota)}")
    print(f"Total de intervalos válidos: {df_manobras_frota['Intervalos Válidos'].sum()}")
    print(f"Tempo total de manobras: {df_manobras_frota['Tempo Total'].sum()*3600:.1f}s")
    print("="*60)
    
    return df_manobras_frota, df_manobras_operador_vazio

def calcular_motor_ocioso_simples(df):
    """
    Calcula motor ocioso de forma simples e direta:
    Soma os valores da coluna Diferença_Hora onde:
    - Motor Ligado == 1 
    - Estado Operacional == PARADA
    
    Args:
        df (DataFrame): DataFrame com os dados de operação
        
    Returns:
        DataFrame: DataFrame com a coluna 'Motor Ocioso Simples' adicionada
    """
    # Inicializar coluna
    df['Motor Ocioso Simples'] = 0
    
    # Identificar registros onde motor está ligado mas máquina está parada
    condicao_motor_ocioso = (df['Motor Ligado'] == 1) & (df['Estado Operacional'] == 'PARADA')
    
    # Atribuir diretamente o tempo da Diferença_Hora para esses registros
    df.loc[condicao_motor_ocioso, 'Motor Ocioso Simples'] = df.loc[condicao_motor_ocioso, 'Diferença_Hora']
    
    return df

def calcular_motor_ocioso_maquina_simples(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula o percentual de motor ocioso por máquina usando a lógica simples.
    
    Motor Ocioso = soma de Diferença_Hora onde Motor Ligado == 1 e Estado Operacional == PARADA
    Horas Motor = soma de Diferença_Hora onde Motor Ligado == 1
    Porcentagem = tempo_ocioso / horas_motor
    """
    resultados = []
    equipamentos = df['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    
    print("\n=== CÁLCULO DE MOTOR OCIOSO (MÉTODO SIMPLES) ===")
    print("Motor Ocioso = Diferença_Hora onde Motor Ligado == 1 e Estado == PARADA")
    print("=" * 60)
    
    for equipamento in equipamentos:
        dados = df[df['Equipamento'] == equipamento]
        
        print(f"\n=== Calculando motor ocioso para {equipamento} ===")
        
        # Horas motor: soma de Diferença_Hora onde Motor Ligado == 1
        horas_motor = dados[dados['Motor Ligado'] == 1]['Diferença_Hora'].sum()
        
        # Motor ocioso: soma de Diferença_Hora onde Motor Ligado == 1 E Estado == PARADA
        tempo_ocioso = dados[(dados['Motor Ligado'] == 1) & (dados['Estado Operacional'] == 'PARADA')]['Diferença_Hora'].sum()
        
        porcentagem = tempo_ocioso / horas_motor if horas_motor > 0 else 0

        print(f"Horas Motor (Motor Ligado == 1): {horas_motor:.4f} h")
        print(f"Tempo Ocioso (Motor Ligado == 1 e Estado == PARADA): {tempo_ocioso:.4f} h")
        print(f"% Ocioso: {porcentagem:.4f} ({porcentagem*100:.2f}%)")

        resultados.append({
            'Frota': equipamento,
            'Porcentagem': porcentagem,
            'Horas Motor': horas_motor,
            'Tempo Ocioso': tempo_ocioso
        })

    return pd.DataFrame(resultados)

def calcular_motor_ocioso_correto(df):
    """
    Calcula motor ocioso com a lógica sequencial correta:
    
    FILTROS APLICADOS ANTES DO CÁLCULO:
    1. Excluir registros onde Grupo Operacao == 'Manutenção'
    2. Considerar APENAS registros onde Motor Ligado == 1 E Estado Operacional == 'PARADA'
    
    LÓGICA SEQUENCIAL:
    1. Leitura linha por linha (sequencial, sem agrupamento)
    2. Identificar intervalos contínuos de registros que atendem aos critérios
    3. Para cada intervalo:
       - Somar todo o tempo (Diferença_Hora) do intervalo
       - Subtrair 1 minuto do total
       - Se resultado >= 1 minuto → incluir no cálculo
       - Se resultado < 1 minuto → descartar intervalo
    
    Args:
        df (DataFrame): DataFrame com os dados de operação
        
    Returns:
        DataFrame: DataFrame com a coluna 'Motor Ocioso Correto' adicionada
    """
    print("\n=== INICIANDO CÁLCULO DE MOTOR OCIOSO CORRETO ===")
    
    # PASSO 1: Aplicar filtros ANTES do cálculo
    print("PASSO 1: Aplicando filtros...")
    
    # Filtro 1: Excluir Manutenção
    df_filtrado = df[df['Grupo Operacao'] != 'Manutenção'].copy()
    registros_manutencao = len(df) - len(df_filtrado)
    if registros_manutencao > 0:
        print(f"  • Excluídos {registros_manutencao} registros de Manutenção")
    
    # Filtro 2: Considerar apenas Motor Ligado == 1 E Estado == PARADA
    condicao_motor_ocioso = (df_filtrado['Motor Ligado'] == 1) & (df_filtrado['Estado Operacional'] == 'PARADA')
    registros_validos = df_filtrado[condicao_motor_ocioso].copy()
    
    print(f"  • Total registros após filtros: {len(registros_validos)}")
    print(f"  • Critério: Motor Ligado == 1 E Estado == PARADA (exceto Manutenção)")
    
    if len(registros_validos) == 0:
        print("  • Nenhum registro atende aos critérios. Motor ocioso = 0")
        df['Motor Ocioso Correto'] = 0
        return df
    
    # PASSO 2: Lógica sequencial para identificar intervalos
    print("\nPASSO 2: Identificando intervalos sequenciais...")
    
    # Inicializar resultado
    df['Motor Ocioso Correto'] = 0
    
    # Resetar índice para facilitar iteração sequencial
    registros_validos = registros_validos.reset_index()
    
    # Variáveis para controle de intervalos
    intervalo_atual = []
    intervalos_encontrados = []
    
    # Iterar pelos registros válidos de forma sequencial
    for i in range(len(registros_validos)):
        registro_atual = registros_validos.iloc[i]
        
        # Se é o primeiro registro ou é sequencial ao anterior
        if i == 0 or registros_validos.iloc[i]['index'] == registros_validos.iloc[i-1]['index'] + 1:
            # Adicionar ao intervalo atual
            intervalo_atual.append(i)
        else:
            # Fim do intervalo atual, processar se não estiver vazio
            if intervalo_atual:
                intervalos_encontrados.append(intervalo_atual)
            # Iniciar novo intervalo
            intervalo_atual = [i]
    
    # Processar último intervalo se existir
    if intervalo_atual:
        intervalos_encontrados.append(intervalo_atual)
    
    print(f"  • Encontrados {len(intervalos_encontrados)} intervalos sequenciais")
    
    # PASSO 3: Processar cada intervalo
    print("\nPASSO 3: Processando intervalos...")
    
    total_tempo_ocioso = 0
    intervalos_validos = 0
    
    for idx_intervalo, indices_intervalo in enumerate(intervalos_encontrados):
        # Somar tempo do intervalo
        tempo_intervalo_horas = 0
        for idx in indices_intervalo:
            tempo_intervalo_horas += registros_validos.iloc[idx]['Diferença_Hora']
        
        # Converter para minutos para aplicar a regra
        tempo_intervalo_minutos = tempo_intervalo_horas * 60
        
        # Aplicar regra: subtrair 1 minuto
        tempo_ocioso_minutos = tempo_intervalo_minutos - 1
        
        # Verificar se é válido (>= 1 minuto)
        if tempo_ocioso_minutos >= 1:
            # Converter de volta para horas
            tempo_ocioso_horas = tempo_ocioso_minutos / 60
            
            # Atribuir o tempo ocioso ao primeiro registro do intervalo
            indice_original = registros_validos.iloc[indices_intervalo[0]]['index']
            df.at[indice_original, 'Motor Ocioso Correto'] = tempo_ocioso_horas
            
            total_tempo_ocioso += tempo_ocioso_horas
            intervalos_validos += 1
            
            print(f"  • Intervalo {idx_intervalo + 1}: {len(indices_intervalo)} registros, {tempo_intervalo_minutos:.1f}min → {tempo_ocioso_minutos:.1f}min ocioso ✓")
        else:
            print(f"  • Intervalo {idx_intervalo + 1}: {len(indices_intervalo)} registros, {tempo_intervalo_minutos:.1f}min → descartado (< 1min) ✗")
    
    print(f"\nRESULTADO:")
    print(f"  • Intervalos válidos: {intervalos_validos}/{len(intervalos_encontrados)}")
    print(f"  • Tempo total ocioso: {total_tempo_ocioso:.4f} horas ({total_tempo_ocioso*60:.1f} minutos)")
    
    return df

def calcular_motor_ocioso_maquina_correto(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula o percentual de motor ocioso por máquina usando o método correto.
    
    Método Correto:
    - Filtros aplicados: excluir Manutenção, considerar apenas Motor Ligado == 1 e Estado == PARADA
    - Lógica sequencial com intervalos e subtração de 1 minuto
    """
    resultados = []
    equipamentos = df['Equipamento'].unique()
    # Filtrar equipamentos NaN
    equipamentos = [equip for equip in equipamentos if pd.notna(equip)]
    
    print("\n=== CÁLCULO DE MOTOR OCIOSO POR MÁQUINA (MÉTODO CORRETO) ===")
    print("Filtros: excluir Manutenção + Motor Ligado == 1 + Estado == PARADA")
    print("Lógica: intervalos sequenciais - 1 minuto")
    print("=" * 70)
    
    for equipamento in equipamentos:
        dados = df[df['Equipamento'] == equipamento]
        
        print(f"\n=== Calculando motor ocioso para {equipamento} ===")
        
        # Horas motor: soma de Diferença_Hora onde Motor Ligado == 1
        horas_motor = dados[dados['Motor Ligado'] == 1]['Diferença_Hora'].sum()
        
        # Motor ocioso: soma da coluna Motor Ocioso Correto
        tempo_ocioso = dados['Motor Ocioso Correto'].sum()
        
        porcentagem = tempo_ocioso / horas_motor if horas_motor > 0 else 0

        print(f"Horas Motor (Motor Ligado == 1): {horas_motor:.4f} h")
        print(f"Tempo Ocioso (método correto): {tempo_ocioso:.4f} h")
        print(f"% Ocioso: {porcentagem:.4f} ({porcentagem*100:.2f}%)")

        resultados.append({
            'Frota': equipamento,
            'Porcentagem': porcentagem,
            'Horas Motor': horas_motor,
            'Tempo Ocioso': tempo_ocioso
        })

    return pd.DataFrame(resultados)

# === NOVO: cálculo de Aferição de Roletes ===
def calcular_roletes(df):
    """
    Calcula os intervalos de aferição de roletes para cada equipamento, filtrando
    a operação "9029 - MEDIR TEMPERATURA ROLETES". Lógica espelhada de
    calcular_lavagem.

    Se não houver registros, retorna DataFrame com linha informativa.
    """
    print("Calculando intervalos de aferição de roletes…")

    df_roletes = df[df['Operacao'] == '9029 - MEDIR TEMPERATURA ROLETES'].copy()

    if len(df_roletes) == 0:
        print("Nenhum registro de roletes encontrado. Criando linha informativa.")
        return pd.DataFrame([{
            'Data': 'N/A',
            'Equipamento': 'NÃO FORAM ENCONTRADOS DADOS DE ROLETES PARA A DATA INFORMADA',
            'Intervalo': 'N/A',
            'Início': 'N/A',
            'Fim': 'N/A',
            'Duração (horas)': 0,
            'Tempo Total do Dia': 0
        }])

    # Mesma lógica de cálculo de intervalos contínuos
    colunas_necessarias = ['Equipamento', 'Data', 'Hora', 'Diferença_Hora']
    for coluna in colunas_necessarias:
        if coluna not in df_roletes.columns:
            print(f"Coluna '{coluna}' não encontrada nos dados de roletes.")
            return pd.DataFrame(columns=[
                'Data', 'Equipamento', 'Intervalo', 'Início', 'Fim',
                'Duração (horas)', 'Tempo Total do Dia'
            ])

    df_roletes = df_roletes.sort_values(['Equipamento', 'Data', 'Hora'])

    resultados = []
    for (equipamento, data), grupo in df_roletes.groupby(['Equipamento', 'Data']):
        grupo = grupo.reset_index(drop=True)
        inicio_intervalo = None
        fim_intervalo = None
        duracao_intervalo = 0
        for i in range(len(grupo)):
            registro = grupo.iloc[i]
            hora_atual = pd.to_datetime(registro['Hora'], format='%H:%M:%S')
            if (inicio_intervalo is None or (i > 0 and (hora_atual - pd.to_datetime(grupo.iloc[i-1]['Hora'], format='%H:%M:%S')).total_seconds() > 1800)):
                if inicio_intervalo is not None:
                    resultados.append({
                        'Data': data,
                        'Equipamento': equipamento,
                        'Intervalo': f"Intervalo {len([r for r in resultados if r['Equipamento']==equipamento and r['Data']==data])+1}",
                        'Início': inicio_intervalo,
                        'Fim': fim_intervalo,
                        'Duração (horas)': round(duracao_intervalo,4),
                        'Tempo Total do Dia': 0  # preencher depois
                    })
                inicio_intervalo = registro['Hora']
                duracao_intervalo = 0
            fim_intervalo = registro['Hora']
            duracao_intervalo += registro['Diferença_Hora']

        # push último intervalo
        if inicio_intervalo is not None:
            resultados.append({
                'Data': data,
                'Equipamento': equipamento,
                'Intervalo': f"Intervalo {len([r for r in resultados if r['Equipamento']==equipamento and r['Data']==data])+1}",
                'Início': inicio_intervalo,
                'Fim': fim_intervalo,
                'Duração (horas)': round(duracao_intervalo,4),
                'Tempo Total do Dia': 0
            })

    # Calcular tempo total por equipamento/dia
    df_result = pd.DataFrame(resultados)
    if not df_result.empty:
        df_result['Tempo Total do Dia'] = df_result.groupby(['Equipamento','Data'])['Duração (horas)'].transform('sum')
    print(f"Processamento de roletes concluído. {len(df_result)} intervalos identificados.")
    return df_result


def calcular_parametros_medios(df_base, uso_gps, hora_elevador, velocidade_media_produtiva):
    """
    Calcula parâmetros médios técnicos por equipamento para exibição no painel.
    
    Args:
        df_base (DataFrame): DataFrame com dados base
        uso_gps (DataFrame): DataFrame com dados de uso GPS já calculados
        hora_elevador (DataFrame): DataFrame com horas de elevador já calculadas
        velocidade_media_produtiva (DataFrame): DataFrame com velocidade média produtiva já calculada
    
    Returns:
        DataFrame: Parâmetros médios por equipamento
    """
    try:
        print("=== CALCULANDO PARÂMETROS MÉDIOS ===")
        
        if df_base.empty:
            print("DataFrame base vazio")
            return pd.DataFrame()
        
        # Filtrar dados produtivos para cálculos de médias
        df_produtivo = df_base[df_base['Grupo Operacao'] == 'Produtiva'].copy()
        
        if df_produtivo.empty:
            print("Nenhum registro produtivo encontrado")
            return pd.DataFrame()
        
        print(f"Registros produtivos para cálculo de médias: {len(df_produtivo)}")
        
        equipamentos = df_base['Equipamento'].unique()
        equipamentos = [equip for equip in equipamentos if pd.notna(equip) and str(equip).strip() != '']
        
        resultados = []
        
        for equipamento in equipamentos:
            print(f"\n--- Calculando parâmetros para equipamento {equipamento} ---")
            
            # Dados gerais do equipamento
            dados_equip = df_base[df_base['Equipamento'] == equipamento]
            dados_prod = df_produtivo[df_produtivo['Equipamento'] == equipamento]
            
            # 1. Horímetro - último registro da frota
            horimetro_final = 0
            if 'Horimetro' in dados_equip.columns:
                horimetros_validos = pd.to_numeric(dados_equip['Horimetro'], errors='coerce').dropna()
                if not horimetros_validos.empty:
                    horimetro_final = horimetros_validos.iloc[-1]  # Último registro
            
            # 2. Uso RTK - buscar na planilha já calculada
            uso_rtk_pct = 0
            try:
                if not uso_gps.empty and 'Frota' in uso_gps.columns:
                    rtk_row = uso_gps[uso_gps['Frota'] == int(equipamento)]
                    if not rtk_row.empty and 'Porcentagem' in rtk_row.columns:
                        uso_rtk_pct = float(rtk_row['Porcentagem'].iloc[0]) * 100  # Converter decimal para número (0.484 -> 48.4)
                        print(f"  Uso RTK encontrado: {uso_rtk_pct:.1f}")
            except (ValueError, TypeError, KeyError) as e:
                print(f"  Erro ao buscar uso RTK: {e}")
                uso_rtk_pct = 0
            
            # 3. Horas Elevador e Horas Motor - buscar na planilha já calculada
            horas_elevador_val = 0
            horas_motor_val = 0
            try:
                if not hora_elevador.empty and 'Frota' in hora_elevador.columns:
                    elevador_row = hora_elevador[hora_elevador['Frota'] == int(equipamento)]
                    if not elevador_row.empty:
                        if 'Horas Elevador' in elevador_row.columns:
                            horas_elevador_val = float(elevador_row['Horas Elevador'].iloc[0])
                        if 'Horas Motor' in elevador_row.columns:
                            horas_motor_val = float(elevador_row['Horas Motor'].iloc[0])
            except (ValueError, TypeError, KeyError) as e:
                print(f"  Erro ao buscar horas elevador/motor: {e}")
                horas_elevador_val = 0
                horas_motor_val = 0
            
            # 4. Velocidade Média - BUSCAR da planilha já calculada "Velocidade Média Produtiva"
            velocidade_media = 0
            try:
                # Buscar na planilha velocidade_media_produtiva já calculada
                if not velocidade_media_produtiva.empty and 'Frota' in velocidade_media_produtiva.columns:
                    vel_row = velocidade_media_produtiva[velocidade_media_produtiva['Frota'] == equipamento]
                    if not vel_row.empty and 'Velocidade Média Produtiva' in vel_row.columns:
                        velocidade_media = float(vel_row['Velocidade Média Produtiva'].iloc[0])
                        print(f"  Velocidade média (da planilha): {velocidade_media:.2f} km/h")
                    else:
                        print(f"  Velocidade média não encontrada para equipamento {equipamento}")
                else:
                    print(f"  Planilha velocidade_media_produtiva não disponível")
            except (ValueError, TypeError, KeyError) as e:
                print(f"  Erro ao buscar velocidade média: {e}")
                velocidade_media = 0
            
            # 5. RPM Motor - média em operação produtiva
            rpm_motor_media = 0
            try:
                if 'RPM Motor' in dados_prod.columns and not dados_prod.empty:
                    rpm_motor = pd.to_numeric(dados_prod['RPM Motor'], errors='coerce').dropna()
                    if not rpm_motor.empty and len(rpm_motor) > 0:
                        rpm_motor_media = float(rpm_motor.mean())
            except (ValueError, TypeError) as e:
                print(f"  Erro ao calcular RPM Motor média: {e}")
                rpm_motor_media = 0
            
            # 6. RPM Extrator - média em operação produtiva
            rpm_extrator_media = 0
            try:
                if 'RPM Extrator' in dados_prod.columns and not dados_prod.empty:
                    rpm_extrator = pd.to_numeric(dados_prod['RPM Extrator'], errors='coerce').dropna()
                    if not rpm_extrator.empty and len(rpm_extrator) > 0:
                        rpm_extrator_media = float(rpm_extrator.mean())
            except (ValueError, TypeError) as e:
                print(f"  Erro ao calcular RPM Extrator média: {e}")
                rpm_extrator_media = 0
            
            # 7. Pressão de Corte - média em operação produtiva
            pressao_corte_media = 0
            try:
                if 'Pressao de Corte' in dados_prod.columns and not dados_prod.empty:
                    pressao_corte = pd.to_numeric(dados_prod['Pressao de Corte'], errors='coerce').dropna()
                    if not pressao_corte.empty and len(pressao_corte) > 0:
                        pressao_corte_media = float(pressao_corte.mean())
            except (ValueError, TypeError) as e:
                print(f"  Erro ao calcular pressão de corte média: {e}")
                pressao_corte_media = 0
            
            # 8. Corte Base Automático - apenas % de automático
            corte_base_auto_pct = 0
            try:
                # PRIMEIRO filtrar por EQUIPAMENTO, depois por 'Grupo Operacao' = 'Produtiva'
                if 'Grupo Operacao' in df_base.columns and 'Equipamento' in df_base.columns:
                    # Filtrar por equipamento específico E por produtiva
                    dados_equipamento_produtivo = df_base[
                        (df_base['Equipamento'] == equipamento) & 
                        (df_base['Grupo Operacao'] == 'Produtiva')
                    ].copy()
                    print(f"  Registros da frota {equipamento} + 'Produtiva': {len(dados_equipamento_produtivo)}")
                    
                    if 'Corte Base Automatico/Manual' in dados_equipamento_produtivo.columns and not dados_equipamento_produtivo.empty:
                        # SOLUÇÃO SIMPLES: Remover colunas duplicadas antes de processar
                        print(f"  Colunas antes da limpeza: {len(dados_equipamento_produtivo.columns)}")
                        
                        # Remove colunas duplicadas mantendo apenas a primeira
                        dados_limpos = dados_equipamento_produtivo.loc[:, ~dados_equipamento_produtivo.columns.duplicated()]
                        print(f"  Colunas após limpeza: {len(dados_limpos.columns)}")
                        
                        # Agora pode acessar a coluna normalmente
                        corte_base_raw = dados_limpos['Corte Base Automatico/Manual']
                        
                        # Debug: mostrar alguns valores da coluna
                        if len(corte_base_raw) > 0:
                            sample_values = corte_base_raw.head(10).values.tolist()
                            print(f"  Amostra de valores da coluna Corte Base: {sample_values}")
                        
                        # Converter para numérico e remover NaN
                        corte_base_numericos = pd.to_numeric(corte_base_raw, errors='coerce').dropna()
                        print(f"  Registros válidos de Corte Base após conversão numérica: {len(corte_base_numericos)}")
                        
                        if len(corte_base_numericos) > 0:
                            # Filtrar apenas valores 0 e 1
                            corte_base_validos = corte_base_numericos[(corte_base_numericos == 0) | (corte_base_numericos == 1)]
                            print(f"  Registros com valores 0 ou 1: {len(corte_base_validos)}")
                            
                            if len(corte_base_validos) > 0:
                                # Contar registros automáticos (1) e calcular percentual
                                total_registros = len(corte_base_validos)
                                registros_auto = int((corte_base_validos == 1).sum())
                                registros_manual = int((corte_base_validos == 0).sum())
                                
                                # Calcular percentual de automático
                                corte_base_auto_pct = round((registros_auto / total_registros) * 100) if total_registros > 0 else 0
                                
                                print(f"  Total registros válidos: {total_registros}")
                                print(f"  Manual (0): {registros_manual}, Automático (1): {registros_auto}")
                                print(f"  % Automático: {corte_base_auto_pct}%")
                            else:
                                print(f"  Nenhum registro com valores 0 ou 1 encontrado")
                        else:
                            print(f"  Nenhum registro numérico válido encontrado")
                    else:
                        print(f"  Coluna 'Corte Base Automatico/Manual' não encontrada para frota {equipamento}")
                else:
                    print(f"  Colunas 'Grupo Operacao' ou 'Equipamento' não encontradas nos dados base")
            except Exception as e:
                print(f"  Erro ao calcular corte base automático: {e}")
                corte_base_auto_pct = 0
            
            # Debug dos valores calculados
            print(f"  Horímetro final: {horimetro_final:.1f} h")
            print(f"  Uso RTK: {uso_rtk_pct:.1f}")
            print(f"  Horas Elevador: {horas_elevador_val:.2f} h")
            print(f"  Corte Base Auto final: {corte_base_auto_pct}%")
            print(f"  Horas Motor: {horas_motor_val:.2f} h")
            print(f"  Velocidade Média: {velocidade_media:.1f} km/h")
            print(f"  RPM Motor Média: {rpm_motor_media:.0f} rpm")
            print(f"  RPM Extrator Média: {rpm_extrator_media:.0f} rpm")
            print(f"  Pressão Corte Média: {pressao_corte_media:.0f} psi")
            print(f"  Corte Base Automático: {corte_base_auto_pct}%")
            
            resultados.append({
                'Frota': int(equipamento),
                'Horimetro': round(horimetro_final, 2),
                'Uso RTK (%)': round(uso_rtk_pct, 2),
                'Horas Elevador': round(horas_elevador_val, 2),
                'Horas Motor': round(horas_motor_val, 2),
                'Velocidade Media (km/h)': round(velocidade_media, 2),
                'RPM Motor Media': round(rpm_motor_media, 2),
                'RPM Extrator Media': round(rpm_extrator_media, 2),
                'Pressao Corte Media (psi)': round(pressao_corte_media, 2),
                'Corte Base Auto (%)': round(corte_base_auto_pct, 2)
            })
        
        if not resultados:
            print("Nenhum parâmetro médio calculado")
            return pd.DataFrame()
        
        df_resultado = pd.DataFrame(resultados)
        print(f"✅ {len(df_resultado)} equipamentos com parâmetros médios calculados")
        
        return df_resultado
        
    except Exception as e:
        print(f"❌ Erro ao calcular parâmetros médios: {e}")
        return pd.DataFrame()


def calcular_operadores_por_frota(df_base):
    """
    Calcula horas elevador por operador por frota
    Baseado na mesma lógica de calcular_hora_elevador mas agrupado por operador
    """
    try:
        print("=== CALCULANDO OPERADORES POR FROTA ===")
        
        if df_base.empty:
            print("DataFrame base vazio")
            return pd.DataFrame()
        
        # Filtros para horas elevador (mesma lógica do calcular_hora_elevador)
        filtros = (
            (df_base['Grupo Operacao'] == 'Produtiva') &
            (df_base['Pressao de Corte'] > 400) &
            (df_base['Velocidade'] > 0)
        )
        
        df_filtrado = df_base[filtros].copy()
        
        if df_filtrado.empty:
            print("Nenhum registro atende aos critérios de horas elevador")
            return pd.DataFrame()
        
        print(f"Registros filtrados para horas elevador: {len(df_filtrado)}")
        
        # Agrupar por equipamento e operador
        resultado_operadores = []
        
        for equipamento in df_filtrado['Equipamento'].unique():
            dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
            
            # Agrupar por operador dentro do equipamento
            for operador in dados_equip['Operador'].unique():
                if pd.isna(operador) or operador == 'N/A' or operador == '':
                    continue
                    
                dados_operador = dados_equip[dados_equip['Operador'] == operador]
                
                # Usar Diferença_Hora diretamente (método mais simples e confiável)
                if 'Diferença_Hora' in dados_operador.columns:
                    tempo_total = dados_operador['Diferença_Hora'].sum()
                else:
                    # Fallback: calcular tempo usando diferença de horário
                    dados_operador = dados_operador.sort_values('Hora').reset_index(drop=True)
                    
                    tempo_total = 0
                    if len(dados_operador) > 1:
                        try:
                            for i in range(len(dados_operador) - 1):
                                hora_atual = pd.to_datetime(dados_operador.iloc[i]['Hora'])
                                hora_proxima = pd.to_datetime(dados_operador.iloc[i + 1]['Hora'])
                                diferenca = (hora_proxima - hora_atual).total_seconds() / 3600
                                
                                # Limitar diferenças muito grandes (provavelmente quebras)
                                if diferenca <= 0.1:  # máximo 6 minutos entre registros
                                    tempo_total += diferenca
                        except (KeyError, IndexError, ValueError) as e:
                            print(f"  Erro ao calcular tempo para {operador}: {e}")
                            tempo_total = 0
                
                if tempo_total > 0:
                    resultado_operadores.append({
                        'Frota': int(equipamento),
                        'Operador': str(operador),
                        'Horas Elevador': round(tempo_total, 4)
                    })
        
        if not resultado_operadores:
            print("Nenhum operador com horas elevador encontrado")
            return pd.DataFrame()
        
        df_resultado = pd.DataFrame(resultado_operadores)
        print(f"✅ {len(df_resultado)} registros de operadores calculados")
        
        return df_resultado
        
    except Exception as e:
        print(f"❌ Erro ao calcular operadores por frota: {e}")
        return pd.DataFrame()

def obter_toneladas_por_frente(caminho_arquivo):
    """
    Obtém as toneladas configuradas para a frente baseado no nome do arquivo.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo processado
    
    Returns:
        tuple: (toneladas, nome_frente) ou (None, None) se não encontrado
    """
    try:
        nome_arquivo = os.path.basename(caminho_arquivo).lower()
        print(f"🔍 Analisando arquivo: {nome_arquivo}")
        
        # Padrões para identificar frentes
        padroes_frente = {
            'frente03': 'Frente03',
            'frente04': 'Frente04', 
            'frente08': 'Frente08',
            'zirleno': 'Zirleno'
        }
        
        # Buscar padrão no nome do arquivo
        for padrao, nome_frente in padroes_frente.items():
            if padrao in nome_arquivo:
                toneladas = TONELADAS_POR_FRENTE.get(nome_frente)
                if toneladas:
                    print(f"✅ Frente identificada: {nome_frente} = {toneladas} toneladas")
                    return toneladas, nome_frente
                else:
                    print(f"⚠️ Frente {nome_frente} identificada mas sem configuração de toneladas")
                    return None, nome_frente
        
        # Se não encontrou padrão específico, usar primeira frente como padrão
        print(f"⚠️ Frente não identificada no arquivo, usando Frente03 como padrão")
        return TONELADAS_FRENTE_03, 'Frente03'
        
    except Exception as e:
        print(f"❌ Erro ao identificar frente: {e}")
        return TONELADAS_FRENTE_03, 'Frente03'

def calcular_producao_por_frota(hora_elevador_df, toneladas_totais=None, caminho_arquivo=None):
    """
    Calcula a produção (toneladas) por frota baseada na proporção de horas elevador.
    
    Args:
        hora_elevador_df (DataFrame): DataFrame com dados de eficiência energética (horas elevador)
        toneladas_totais (float): Total de toneladas (opcional, será calculado por frente se não informado)
        caminho_arquivo (str): Caminho do arquivo para identificar a frente
    
    Returns:
        DataFrame: Produção por frota com colunas ['Frota', 'Toneladas', 'Horas Elevador', 'Ton/h']
    """
    try:
        print("=== CALCULANDO PRODUÇÃO POR FROTA ===")
        
        if hora_elevador_df.empty:
            print("DataFrame de horas elevador vazio")
            return pd.DataFrame(columns=['Frota', 'Toneladas', 'Horas Elevador', 'Ton/h'])
        
        # Obter toneladas específicas da frente se não informado
        if toneladas_totais is None and caminho_arquivo:
            toneladas_totais, nome_frente = obter_toneladas_por_frente(caminho_arquivo)
            if toneladas_totais is None:
                toneladas_totais = TONELADAS_FRENTE_03  # Fallback
                nome_frente = 'Frente03'
        elif toneladas_totais is None:
            toneladas_totais = TONELADAS_FRENTE_03  # Fallback se não tiver arquivo
            nome_frente = 'Padrão'
        else:
            nome_frente = 'Manual'
        
        print(f"Frente: {nome_frente}")
        print(f"Total de toneladas a distribuir: {toneladas_totais}")
        
        # Verificar se temos a coluna de horas elevador
        coluna_horas = None
        for col in hora_elevador_df.columns:
            if 'elevador' in col.lower() or 'horas' in col.lower():
                coluna_horas = col
                break
        
        if coluna_horas is None:
            print("⚠️ Coluna de horas elevador não encontrada")
            # Tentar usar a segunda coluna (assumindo que a primeira é Frota/Equipamento)
            if len(hora_elevador_df.columns) >= 2:
                coluna_horas = hora_elevador_df.columns[1]
                print(f"Usando coluna: {coluna_horas}")
            else:
                return pd.DataFrame(columns=['Frota', 'Toneladas', 'Horas Elevador', 'Ton/h'])
        
        # Criar DataFrame de trabalho
        df_trabalho = hora_elevador_df.copy()
        
        # Identificar coluna de frota/equipamento
        coluna_frota = None
        for col in df_trabalho.columns:
            if 'frota' in col.lower() or 'equipamento' in col.lower():
                coluna_frota = col
                break
        
        if coluna_frota is None:
            coluna_frota = df_trabalho.columns[0]  # Primeira coluna
        
        print(f"Usando coluna frota: {coluna_frota}")
        print(f"Usando coluna horas: {coluna_horas}")
        
        # Garantir que horas elevador são numéricas
        df_trabalho[coluna_horas] = pd.to_numeric(df_trabalho[coluna_horas], errors='coerce').fillna(0)
        
        # Calcular total de horas elevador
        total_horas_elevador = df_trabalho[coluna_horas].sum()
        
        if total_horas_elevador == 0:
            print("⚠️ Total de horas elevador é zero")
            return pd.DataFrame(columns=['Frota', 'Toneladas', 'Horas Elevador', 'Ton/h'])
        
        print(f"Total de horas elevador: {total_horas_elevador:.2f}h")
        
        resultados = []
        
        for _, linha in df_trabalho.iterrows():
            frota = linha[coluna_frota]
            horas_elevador = linha[coluna_horas]
            
            # Calcular proporção da frota
            proporcao = horas_elevador / total_horas_elevador if total_horas_elevador > 0 else 0
            
            # Calcular toneladas proporcionais
            toneladas_frota = toneladas_totais * proporcao
            
            # Calcular toneladas por hora
            ton_por_hora = toneladas_frota / horas_elevador if horas_elevador > 0 else 0
            
            resultados.append({
                'Frota': int(frota) if pd.notna(frota) else 0,
                'Toneladas': round(toneladas_frota, 2),
                'Horas Elevador': round(horas_elevador, 2),
                'Ton/h': round(ton_por_hora, 2)
            })
            
            print(f"  Frota {frota}: {horas_elevador:.2f}h ({proporcao*100:.1f}%) = {toneladas_frota:.2f}t ({ton_por_hora:.2f}t/h)")
        
        df_resultado = pd.DataFrame(resultados)
        print(f"✅ {len(df_resultado)} registros de produção calculados")
        
        return df_resultado
        
    except Exception as e:
        print(f"❌ Erro ao calcular produção por frota: {e}")
        return pd.DataFrame(columns=['Frota', 'Toneladas', 'Horas Elevador', 'Ton/h'])

def calcular_painel_esquerdo(df_base, horas_por_frota, hora_elevador, df_manobras_frota, disp_mecanica, df_operadores, df_producao=None):
    """
    Calcula todos os dados necessários para o painel esquerdo do dashboard.
    
    Args:
        df_base (DataFrame): Dados base
        horas_por_frota (DataFrame): Horas registradas por frota
        hora_elevador (DataFrame): Horas elevador e motor por frota
        df_manobras_frota (DataFrame): Dados de manobras por frota
        disp_mecanica (DataFrame): Disponibilidade mecânica por frota
        df_operadores (DataFrame): Operadores por frota
        df_producao (DataFrame): Produção por frota (opcional, será calculada se não fornecida)
    
    Returns:
        DataFrame: Dados consolidados para o painel esquerdo
    """
    try:
        print("=== CALCULANDO DADOS DO PAINEL ESQUERDO ===")
        
        # Se produção não foi fornecida, calcular
        if df_producao is None or df_producao.empty:
            df_producao = calcular_producao_por_frota(hora_elevador)
        
        # Obter lista de frotas únicas
        frotas_unicas = set()
        
        for df in [horas_por_frota, hora_elevador, df_manobras_frota, disp_mecanica]:
            if not df.empty:
                # Tentar encontrar coluna de frota
                coluna_frota = None
                for col in df.columns:
                    if any(palavra in col.lower() for palavra in ['frota', 'equipamento']):
                        coluna_frota = col
                        break
                
                if coluna_frota:
                    frotas_df = pd.to_numeric(df[coluna_frota], errors='coerce').dropna()
                    frotas_unicas.update(frotas_df.astype(int).tolist())
        
        frotas_unicas = sorted(list(frotas_unicas))
        print(f"Frotas encontradas: {frotas_unicas}")
        
        resultados = []
        
        for frota in frotas_unicas:
            print(f"\n--- Consolidando dados para frota {frota} ---")
            
            dados_frota = {"frota": frota}
            
            # 1. Horas registradas (da planilha "Horas por Frota")
            horas_registradas = 0
            if not horas_por_frota.empty:
                linha_frota = horas_por_frota[horas_por_frota.iloc[:, 0] == frota]
                if not linha_frota.empty:
                    # Procurar por coluna que contenha "registradas" ou similar
                    for col in linha_frota.columns:
                        if any(palavra in col.lower() for palavra in ['registradas', 'total', 'geral']):
                            horas_registradas = pd.to_numeric(linha_frota[col].iloc[0], errors='coerce') or 0
                            break
                    if horas_registradas == 0 and len(linha_frota.columns) >= 2:
                        horas_registradas = pd.to_numeric(linha_frota.iloc[0, 1], errors='coerce') or 0
            
            dados_frota["horas_registradas"] = horas_registradas
            
            # 2. Horas motor e horas elevador (da planilha "Eficiência Energética")
            horas_motor = 0
            horas_elevador_val = 0
            if not hora_elevador.empty:
                linha_frota = hora_elevador[hora_elevador.iloc[:, 0] == frota]
                if not linha_frota.empty:
                    # Horas elevador
                    for col in linha_frota.columns:
                        if 'elevador' in col.lower():
                            horas_elevador_val = pd.to_numeric(linha_frota[col].iloc[0], errors='coerce') or 0
                            break
                    
                    # Horas motor
                    for col in linha_frota.columns:
                        if 'motor' in col.lower():
                            horas_motor = pd.to_numeric(linha_frota[col].iloc[0], errors='coerce') or 0
                            break
            
            dados_frota["horas_motor"] = horas_motor
            dados_frota["horas_elevador"] = horas_elevador_val
            
            # 3. Toneladas (da planilha "Produção")
            toneladas = 0
            ton_por_hora = 0
            if not df_producao.empty:
                linha_frota = df_producao[df_producao['Frota'] == frota]
                if not linha_frota.empty:
                    toneladas = linha_frota['Toneladas'].iloc[0]
                    ton_por_hora = linha_frota['Ton/h'].iloc[0]
            
            dados_frota["toneladas"] = toneladas
            dados_frota["ton_por_hora"] = ton_por_hora
            
            # 4. Eficiências calculadas
            # Eficiência operacional = horas elevador / horas registradas
            eficiencia_operacional = (horas_elevador_val / horas_registradas * 100) if horas_registradas > 0 else 0
            dados_frota["eficiencia_operacional"] = eficiencia_operacional
            
            # Eficiência energética = horas elevador / horas motor  
            eficiencia_energetica = (horas_elevador_val / horas_motor * 100) if horas_motor > 0 else 0
            dados_frota["eficiencia_energetica"] = eficiencia_energetica
            
            # 5. Manobras (da planilha "Manobras") - enviar dados brutos sem conversão
            intervalos_manobras = 0
            tempo_total_manobras = 0
            tempo_medio_manobras = 0
            if not df_manobras_frota.empty:
                linha_frota = df_manobras_frota[df_manobras_frota.iloc[:, 0] == frota]
                if not linha_frota.empty:
                    for col in linha_frota.columns:
                        if any(palavra in col.lower() for palavra in ['intervalos', 'válidos']):
                            # Usar fillna(0) ao invés de 'or 0' para preservar valores pequenos
                            intervalos_manobras = pd.to_numeric(linha_frota[col].iloc[0], errors='coerce')
                            intervalos_manobras = 0 if pd.isna(intervalos_manobras) else intervalos_manobras
                        elif any(palavra in col.lower() for palavra in ['total', 'tempo']) and 'médio' not in col.lower():
                            tempo_total_manobras = pd.to_numeric(linha_frota[col].iloc[0], errors='coerce')
                            tempo_total_manobras = 0 if pd.isna(tempo_total_manobras) else tempo_total_manobras
                        elif any(palavra in col.lower() for palavra in ['médio', 'medio', 'média']):
                            tempo_medio_manobras = pd.to_numeric(linha_frota[col].iloc[0], errors='coerce')
                            tempo_medio_manobras = 0 if pd.isna(tempo_medio_manobras) else tempo_medio_manobras
            
            # Enviar dados exatos como estão na planilha (em horas)
            dados_frota["manobras_intervalos"] = int(intervalos_manobras)
            dados_frota["manobras_tempo_total"] = tempo_total_manobras
            dados_frota["manobras_tempo_medio"] = tempo_medio_manobras
            
            # 6. Disponibilidade mecânica
            disponibilidade_pct = 0
            tempo_manutencao = 0
            if not disp_mecanica.empty:
                linha_frota = disp_mecanica[disp_mecanica.iloc[:, 0] == frota]
                if not linha_frota.empty:
                    for col in linha_frota.columns:
                        if any(palavra in col.lower() for palavra in ['disponibilidade', '%']):
                            valor = linha_frota[col].iloc[0]
                            if isinstance(valor, str) and '%' in valor:
                                disponibilidade_pct = float(valor.replace('%', ''))
                            else:
                                disponibilidade_pct = pd.to_numeric(valor, errors='coerce') or 0
                                # Se valor está entre 0 e 1, converter para porcentagem
                                if 0 <= disponibilidade_pct <= 1:
                                    disponibilidade_pct *= 100
                        elif any(palavra in col.lower() for palavra in ['manutencao', 'manutenção']):
                            tempo_manutencao = pd.to_numeric(linha_frota[col].iloc[0], errors='coerce') or 0
            
            dados_frota["disponibilidade_mecanica"] = disponibilidade_pct
            dados_frota["tempo_manutencao"] = tempo_manutencao
            
            # 7. Operadores
            operadores_lista = []
            if not df_operadores.empty:
                linhas_frota = df_operadores[df_operadores.iloc[:, 0] == frota]
                for _, linha_op in linhas_frota.iterrows():
                    if len(linha_op) >= 2:
                        operador = str(linha_op.iloc[1]) if pd.notna(linha_op.iloc[1]) else "N/A"
                        horas_op = pd.to_numeric(linha_op.iloc[2], errors='coerce') if len(linha_op) >= 3 else 0
                        operadores_lista.append({
                            "nome": operador,
                            "horas": horas_op
                        })
            
            dados_frota["operadores"] = operadores_lista
            
            resultados.append(dados_frota)
            
            print(f"  ✅ Frota {frota}: {horas_registradas:.1f}h reg., {horas_elevador_val:.1f}h elev., {toneladas:.1f}t")
        
        df_painel = pd.DataFrame(resultados)
        print(f"\n✅ {len(df_painel)} registros consolidados para painel esquerdo")
        
        return df_painel
        
    except Exception as e:
        print(f"❌ Erro ao calcular painel esquerdo: {e}")
        return pd.DataFrame()

def calcular_painel_direito_por_frota(df_lavagem, df_ofensores, frota_especifica=None):
    """
    Calcula dados para o painel direito do dashboard (Lavagem e Ofensores) para uma frota específica.
    
    Args:
        df_lavagem (DataFrame): Dados de intervalos de lavagem
        df_ofensores (DataFrame): Dados dos principais ofensores
        frota_especifica (int, optional): ID da frota para filtrar os dados. Se None, retorna dados globais.
    
    Returns:
        dict: Dados consolidados para o painel direito da frota específica
    """
    try:
        print("=== CALCULANDO DADOS DO PAINEL DIREITO ===")
        
        painel_direito = {
            "lavagem": {},
            "ofensores": []
        }
        
        # 1. Processar dados de LAVAGEM (filtrados por frota se especificada)
        if df_lavagem is not None and not df_lavagem.empty:
            # Verificar se é a mensagem informativa (sem dados)
            primeira_linha = df_lavagem.iloc[0]
            if 'NÃO FORAM ENCONTRADOS DADOS' in str(primeira_linha.get('Equipamento', '')):
                print(f"  📋 Nenhum dado de lavagem encontrado{' para frota ' + str(frota_especifica) if frota_especifica else ''}")
                painel_direito["lavagem"] = {
                    "tem_dados": False,
                    "total_intervalos": 0,
                    "tempo_total_horas": 0,
                    "equipamentos": []
                }
            else:
                # Filtrar por frota específica se informada
                df_lavagem_filtrado = df_lavagem
                if frota_especifica is not None:
                    df_lavagem_filtrado = df_lavagem[df_lavagem['Equipamento'] == frota_especifica]
                    print(f"  🧽 Processando lavagem para frota {frota_especifica}: {len(df_lavagem_filtrado)} registros")
                else:
                    print(f"  🧽 Processando {len(df_lavagem)} registros de lavagem (todos)")
                
                if df_lavagem_filtrado.empty:
                    print(f"  📋 Nenhum dado de lavagem para frota {frota_especifica}")
                    painel_direito["lavagem"] = {
                        "tem_dados": False,
                        "total_intervalos": 0,
                        "tempo_total_horas": 0,
                        "equipamentos": []
                    }
                else:
                    # Agrupar por equipamento
                    equipamentos_lavagem = []
                    total_intervalos = 0
                    tempo_total_global = 0
                    
                    # Verificar se as colunas esperadas existem
                    colunas_lavagem = df_lavagem_filtrado.columns.tolist()
                    print(f"  📊 Colunas de lavagem: {colunas_lavagem}")
                    
                    for equipamento in df_lavagem_filtrado['Equipamento'].unique():
                        dados_equip = df_lavagem_filtrado[df_lavagem_filtrado['Equipamento'] == equipamento]
                        
                        # Calcular métricas por equipamento
                        intervalos_equip = len(dados_equip)
                        
                        # Buscar tempo total do dia (última coluna geralmente)
                        tempo_total_equip = 0
                        for col in dados_equip.columns:
                            if any(palavra in col.lower() for palavra in ['total', 'dia']):
                                tempo_total_equip = pd.to_numeric(dados_equip[col].iloc[0], errors='coerce')
                                tempo_total_equip = 0 if pd.isna(tempo_total_equip) else tempo_total_equip
                                break
                        
                        # Se não encontrou, somar as durações
                        if tempo_total_equip == 0:
                            for col in dados_equip.columns:
                                if any(palavra in col.lower() for palavra in ['duração', 'duracao', 'horas']):
                                    duracao_col = pd.to_numeric(dados_equip[col], errors='coerce').fillna(0)
                                    tempo_total_equip = duracao_col.sum()
                                    break
                        
                        equipamentos_lavagem.append({
                            "equipamento": int(equipamento) if pd.notna(equipamento) else 0,
                            "intervalos": intervalos_equip,
                            "tempo_total_horas": tempo_total_equip,
                            "detalhes": dados_equip.to_dict('records')
                        })
                        
                        total_intervalos += intervalos_equip
                        tempo_total_global += tempo_total_equip
                        
                        print(f"    🚜 Equipamento {equipamento}: {intervalos_equip} intervalos, {tempo_total_equip:.2f}h total")
                    
                    painel_direito["lavagem"] = {
                        "tem_dados": True,
                        "total_intervalos": total_intervalos,
                        "tempo_total_horas": tempo_total_global,
                        "equipamentos": equipamentos_lavagem
                    }
        else:
            print("  📋 DataFrame de lavagem vazio")
            painel_direito["lavagem"] = {
                "tem_dados": False,
                "total_intervalos": 0,
                "tempo_total_horas": 0,
                "equipamentos": []
            }
        
        # 2. Processar dados de OFENSORES (filtrados por frota se especificada)
        if df_ofensores is not None and not df_ofensores.empty:
            # Filtrar por frota específica se informada
            df_ofensores_filtrado = df_ofensores
            if frota_especifica is not None:
                # Filtrar ofensores que tenham relação com a frota (primeira coluna contém o número da frota)
                def contem_frota(linha_operacao):
                    if pd.isna(linha_operacao):
                        return False
                    if isinstance(linha_operacao, str):
                        match = re.match(r'^(\d+)', str(linha_operacao))
                        if match:
                            return int(match.group(1)) == frota_especifica
                    elif pd.api.types.is_numeric_dtype(type(linha_operacao)):
                        return int(linha_operacao) == frota_especifica
                    return False
                
                mask_frota = df_ofensores[df_ofensores.columns[0]].apply(contem_frota)
                df_ofensores_filtrado = df_ofensores[mask_frota]
                print(f"  ⚠️ Processando ofensores para frota {frota_especifica}: {len(df_ofensores_filtrado)} registros")
            else:
                print(f"  ⚠️ Processando {len(df_ofensores)} registros de ofensores (todos)")
            
            colunas_ofensores = df_ofensores_filtrado.columns.tolist()
            print(f"  📊 Colunas de ofensores: {colunas_ofensores}")
            
            # Converter dados de ofensores
            ofensores_lista = []
            for idx, linha in df_ofensores_filtrado.iterrows():
                ofensor = {}
                
                # Processar primeira coluna (pode ser equipamento ou operação)
                primeira_coluna = linha.iloc[0]
                primeira_col_nome = df_ofensores_filtrado.columns[0].lower().replace(' ', '_').replace('(', '').replace(')', '').replace('%', 'pct')
                
                if pd.notna(primeira_coluna):
                    # Tentar extrair número da primeira coluna se for string
                    if isinstance(primeira_coluna, str):
                        # Procurar por números no início da string
                        match = re.match(r'^(\d+)', str(primeira_coluna))
                        if match:
                            ofensor["equipamento"] = int(match.group(1))
                        else:
                            ofensor["equipamento"] = 0
                        ofensor[primeira_col_nome] = str(primeira_coluna)
                    elif pd.api.types.is_numeric_dtype(type(primeira_coluna)):
                        ofensor["equipamento"] = int(primeira_coluna)
                        ofensor[primeira_col_nome] = primeira_coluna
                    else:
                        ofensor["equipamento"] = 0
                        ofensor[primeira_col_nome] = str(primeira_coluna)
                else:
                    ofensor["equipamento"] = 0
                    ofensor[primeira_col_nome] = ""
                
                # Mapear outras colunas dinamicamente
                for i, col in enumerate(df_ofensores_filtrado.columns[1:], 1):
                    if i < len(linha):
                        valor = linha.iloc[i]
                        # Converter nome da coluna para snake_case
                        nome_campo = col.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('%', 'pct')
                        
                        # Converter valor
                        if pd.notna(valor):
                            if isinstance(valor, str) and '%' in valor:
                                # Converter porcentagem
                                try:
                                    ofensor[nome_campo] = float(valor.replace('%', ''))
                                except:
                                    ofensor[nome_campo] = str(valor)
                            elif pd.api.types.is_numeric_dtype(type(valor)):
                                ofensor[nome_campo] = valor
                            else:
                                ofensor[nome_campo] = str(valor)
                        else:
                            ofensor[nome_campo] = 0
                
                ofensores_lista.append(ofensor)
                print(f"    ⚠️ Ofensor {len(ofensores_lista)}: Equipamento {ofensor['equipamento']}, Dados: {primeira_coluna}")
            
            painel_direito["ofensores"] = ofensores_lista
        else:
            print("  📋 DataFrame de ofensores vazio")
            painel_direito["ofensores"] = []
        
        frota_info = f" (frota {frota_especifica})" if frota_especifica else " (global)"
        print(f"✅ Painel direito calculado{frota_info}:")
        print(f"   🧽 Lavagem: {painel_direito['lavagem']['total_intervalos']} intervalos, {painel_direito['lavagem']['tempo_total_horas']:.2f}h")
        print(f"   ⚠️ Ofensores: {len(painel_direito['ofensores'])} registros")
        
        return painel_direito
        
    except Exception as e:
        print(f"❌ Erro ao calcular painel direito: {e}")
        return {
            "lavagem": {"tem_dados": False, "total_intervalos": 0, "tempo_total_horas": 0, "equipamentos": []},
            "ofensores": []
        }

def extrair_info_arquivo(caminho_arquivo):
    """
    Extrai informações do nome do arquivo para determinar data, frente e máquina.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo
        
    Returns:
        tuple: (data_dia, frente_id, maquina_id) ou (None, None, None) se não conseguir extrair
    """
    try:
        nome_arquivo = os.path.basename(caminho_arquivo)
        
        # Padrão: colhedorasFrente03_05082025.txt ou similar
        # Extrair frente (ex: Frente03, Zirleno)
        frente_match = re.search(r'colhedoras([A-Za-z0-9]+)_', nome_arquivo)
        if not frente_match:
            return None, None, None
        
        frente_id = frente_match.group(1)
        
        # Extrair data (ex: 05082025)
        data_match = re.search(r'_(\d{8})', nome_arquivo)
        if not data_match:
            return None, None, None
        
        data_str = data_match.group(1)
        # Converter formato ddmmyyyy para yyyy-mm-dd
        if len(data_str) == 8:
            dia = data_str[:2]
            mes = data_str[2:4]
            ano = data_str[4:]
            data_dia = f"{ano}-{mes}-{dia}"
        else:
            return None, None, None
        
        # Para máquina ID, usar 0 como padrão (pode ser ajustado conforme necessário)
        maquina_id = 0
        
        return data_dia, frente_id, maquina_id
        
    except Exception as e:
        print(f"❌ Erro ao extrair info do arquivo {caminho_arquivo}: {e}")
        return None, None, None

def converter_chaves_snake_case(dados_dict):
    """
    Converte as chaves do dicionário para snake_case, removendo espaços, 
    parênteses e caracteres especiais para facilitar uso em código Python.
    
    Args:
        dados_dict (dict): Dicionário com chaves originais
        
    Returns:
        dict: Dicionário com chaves convertidas para snake_case
    """
    mapeamento_chaves = {
        'Frota': 'frota',
        'Horimetro': 'horimetro',
        'Uso RTK (%)': 'uso_rtk',
        'Horas Elevador': 'horas_elevador',
        'Horas Motor': 'horas_motor',
        'Velocidade Media (km/h)': 'vel_media',
        'RPM Motor Media': 'rpm_motor_media',
        'RPM Extrator Media': 'rpm_extrator_media',
        'Pressao Corte Media (psi)': 'pressao_corte_media',
        'Corte Base Auto (%)': 'corte_base_auto'
    }
    
    dados_convertidos = {}
    for chave_original, valor in dados_dict.items():
        chave_nova = mapeamento_chaves.get(chave_original, chave_original.lower().replace(' ', '_'))
        dados_convertidos[chave_nova] = valor
    
    return dados_convertidos

def enviar_dados_supabase(df_parametros, df_painel_esquerdo, df_lavagem, df_ofensores, caminho_arquivo):
    """
    Envia dados completos (parâmetros médios, painel esquerdo e painel direito) para a tabela do Supabase.
    Cria um registro separado para cada frota com UUID único e chaves em snake_case.
    Calcula dados do painel direito individualmente por frota.
    
    Args:
        df_parametros (DataFrame): DataFrame com os parâmetros médios
        df_painel_esquerdo (DataFrame): DataFrame com dados do painel esquerdo
        df_lavagem (DataFrame): DataFrame com dados de lavagem
        df_ofensores (DataFrame): DataFrame com dados de ofensores
        caminho_arquivo (str): Caminho do arquivo processado para extrair metadata
    """
    try:
        if df_parametros.empty:
            print("⚠️ DataFrame de parâmetros vazio, não enviando para Supabase")
            return False
        
        # Extrair informações do arquivo
        data_dia, frente_id, _ = extrair_info_arquivo(caminho_arquivo)
        
        if not all([data_dia, frente_id]):
            print("⚠️ Não foi possível extrair informações do arquivo para Supabase")
            return False
        
        print(f"📡 Enviando parâmetros para Supabase - Data: {data_dia}, Frente: {frente_id}")
        print(f"📊 Total de frotas a processar: {len(df_parametros)}")
        
        # Headers para a requisição com UPSERT baseado na chave primária composta
        headers = {
            "apikey": SUPABASE_ANON_KEY,
            "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates"
        }
        
        # URL da API Supabase
        url = f"{SUPABASE_URL}/rest/v1/registros_painelmaq"
        
        sucessos = 0
        erros = 0
        
        # Processar cada frota individualmente
        for index, linha in df_parametros.iterrows():
            try:
                frota = int(linha['Frota'])
                
                # Usar o número da frota como maquina_id
                maquina_id = frota
                
                # Converter a linha em um registro JSON com chaves snake_case
                parametros_frota_original = linha.to_dict()
                parametros_frota = converter_chaves_snake_case(parametros_frota_original)
                
                # Buscar dados do painel esquerdo para esta frota
                painel_esquerdo_frota = None
                if df_painel_esquerdo is not None and not df_painel_esquerdo.empty:
                    linha_painel = df_painel_esquerdo[df_painel_esquerdo['frota'] == frota]
                    if not linha_painel.empty:
                        painel_esquerdo_frota = linha_painel.iloc[0].to_dict()
                
                # Calcular dados do painel direito específicos para esta frota
                try:
                    dados_painel_direito_frota = calcular_painel_direito_por_frota(df_lavagem, df_ofensores, frota_especifica=frota)
                    print(f"      📊 Painel direito calculado para frota {frota}")
                except Exception as e:
                    print(f"      ⚠️ Erro ao calcular painel direito para frota {frota}: {e}")
                    dados_painel_direito_frota = {
                        "lavagem": {"tem_dados": False, "total_intervalos": 0, "tempo_total_horas": 0, "equipamentos": []},
                        "ofensores": []
                    }
                
                # Dados para UPSERT (inserir ou atualizar) baseado na chave primária
                dados_registro = {
                    "data_dia": data_dia,
                    "frente_id": frente_id,
                    "maquina_id": maquina_id,
                    "parametros_medios": [parametros_frota],  # Array com um registro da frota
                    "painel_esquerdo": painel_esquerdo_frota,  # Dados do painel esquerdo
                    "painel_direito": dados_painel_direito_frota,  # Dados do painel direito específicos da frota
                    "updated_at": datetime.now().isoformat()
                }
                
                # Verificar se registro já existe para log
                check_url = f"{url}?data_dia=eq.{data_dia}&frente_id=eq.{frente_id}&maquina_id=eq.{maquina_id}"
                check_response = requests.get(check_url, headers={
                    "apikey": SUPABASE_ANON_KEY,
                    "Authorization": f"Bearer {SUPABASE_ANON_KEY}"
                })
                
                registro_existe = check_response.status_code == 200 and len(check_response.json()) > 0
                
                print(f"   🚜 {'Atualizando' if registro_existe else 'Criando'} frota {frota}...")
                
                # Estratégia UPSERT correta para Supabase
                if registro_existe:
                    # UPDATE: Usar PATCH com filtro específico
                    update_headers = {
                        "apikey": SUPABASE_ANON_KEY,
                        "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
                        "Content-Type": "application/json"
                    }
                    
                    # Dados apenas para atualização (sem incluir chave primária)
                    dados_update = {
                        "parametros_medios": dados_registro["parametros_medios"],
                        "painel_esquerdo": dados_registro["painel_esquerdo"],
                        "painel_direito": dados_registro["painel_direito"],  # Agora específico da frota
                        "updated_at": dados_registro["updated_at"]
                    }
                    
                    response = requests.patch(
                        check_url,
                        headers=update_headers,
                        json=dados_update
                    )
                else:
                    # INSERT: Usar POST normal
                    insert_headers = {
                        "apikey": SUPABASE_ANON_KEY,
                        "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
                        "Content-Type": "application/json"
                    }
                    
                    response = requests.post(
                        url,
                        headers=insert_headers,
                        json=dados_registro
                    )
                
                if response.status_code in [200, 201, 204]:
                    action = "atualizada" if registro_existe else "criada"
                    print(f"      ✅ Frota {frota} {action} com sucesso")
                    sucessos += 1
                else:
                    print(f"      ❌ Erro ao processar frota {frota}: {response.status_code}")
                    print(f"         Resposta: {response.text[:200]}")
                    erros += 1
                    
            except Exception as e:
                print(f"      ❌ Erro ao processar frota {linha.get('Frota', 'N/A')}: {e}")
                erros += 1
        
        # Resumo do envio
        print(f"\n📋 RESUMO DO ENVIO:")
        print(f"   ✅ Sucessos: {sucessos}")
        print(f"   ❌ Erros: {erros}")
        print(f"   📅 Data: {data_dia}")
        print(f"   🏭 Frente: {frente_id}")
        print(f"   🆔 Cada registro tem UUID único gerado automaticamente")
        
        return sucessos > 0
            
    except Exception as e:
        print(f"❌ Erro geral ao enviar parâmetros para Supabase: {e}")
        return False

if __name__ == "__main__":
    print("="*80)
    print("Iniciando processamento de arquivos de colhedoras...")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Este script processa arquivos de colhedoras e gera planilhas Excel com métricas")
    print("Suporta arquivos TXT, CSV e ZIP")
    print("Ignorando arquivos que contenham 'transbordo' no nome")
    print("="*50)
    print("USANDO MÉTODO AVANÇADO DE CÁLCULO DE MOTOR OCIOSO")
    print("- Tolerância de 1 minuto é aplicada a cada sequência de paradas")
    print("- Sequências de paradas com menos de 1 minuto são ignoradas")
    print("- Paradas de motor com velocidade zero são agrupadas em intervalos")
    print("="*80)
    
    try:
        processar_todos_arquivos()
        print("\nProcessamento concluído com sucesso!")
    except Exception as e:
        print(f"\nErro durante o processamento: {str(e)}")
        print("Detalhes do erro:")
        import traceback
        traceback.print_exc() 