import pandas as pd
import openpyxl
import os

# Criar um DataFrame com valores decimais (0-1) para porcentagens
df = pd.DataFrame({
    'Frota': ['Frota1', 'Frota2', 'Frota3'],
    'Disponibilidade': [0.25, 0.5, 0.75],  # Valores decimais (0-1)
    'Uso GPS': [0.15, 0.35, 0.65],         # Valores decimais (0-1)
    'Horas': [12.5, 15.3, 18.7]            # Valores numéricos normais
})

print("DataFrame original:")
print(df)

# Salvar para Excel
arquivo = 'teste_formato_porcentagem.xlsx'
with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Teste', index=False)
    
    # Aplicar formatação
    workbook = writer.book
    worksheet = workbook['Teste']
    
    # Formatar colunas de porcentagem
    for row in range(2, len(df) + 2):  # +2 porque o Excel é 1-indexado e temos um cabeçalho
        # Coluna B - Disponibilidade (formato porcentagem)
        cell = worksheet.cell(row=row, column=2)
        cell.number_format = '0.00%'
        
        # Coluna C - Uso GPS (formato porcentagem)
        cell = worksheet.cell(row=row, column=3)
        cell.number_format = '0.00%'
        
        # Coluna D - Horas (formato decimal)
        cell = worksheet.cell(row=row, column=4)
        cell.number_format = '0.00'

print(f"Arquivo '{arquivo}' criado com sucesso!")
print("Abra o arquivo no Excel e verifique que os valores são mostrados como porcentagens")
print("mas na barra de fórmulas aparecem como decimais (0-1).")

# Agora vamos ler o arquivo novamente para verificar os valores
input("Pressione Enter para continuar e verificar os valores no arquivo Excel...")

# Ler o arquivo Excel
wb = openpyxl.load_workbook(arquivo)
ws = wb.active

print("\nValores lidos do arquivo Excel:")
print(f"{'Frota':<10} | {'Disponibilidade':<15} | {'Uso GPS':<15} | {'Horas':<10}")
print("-" * 60)

for row in range(2, len(df) + 2):
    frota = ws.cell(row=row, column=1).value
    disponibilidade = ws.cell(row=row, column=2).value
    uso_gps = ws.cell(row=row, column=3).value
    horas = ws.cell(row=row, column=4).value
    
    print(f"{frota:<10} | {disponibilidade:<15.6f} | {uso_gps:<15.6f} | {horas:<10.2f}")

print("\nComo você pode ver, os valores são armazenados como decimais (0-1) no Excel,")
print("mas são exibidos como porcentagens na visualização devido ao formato aplicado.")
print("Quando você seleciona uma célula, o valor decimal é mostrado na barra de fórmulas.") 