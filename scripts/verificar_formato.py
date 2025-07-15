import openpyxl
import os

# Caminho para o arquivo Excel gerado
arquivo = '../output/colhedorasFrente08_14072025_processado.xlsx'

# Verificar se o arquivo existe
if not os.path.exists(arquivo):
    print(f"Arquivo {arquivo} não encontrado!")
    exit(1)

print(f"Verificando formato das células no arquivo {os.path.basename(arquivo)}...")

# Ler o arquivo Excel usando openpyxl para verificar os formatos
wb = openpyxl.load_workbook(arquivo)

# Verificar formatos na planilha de Disponibilidade Mecânica
if '1_Disponibilidade Mecânica' in wb.sheetnames:
    ws = wb['1_Disponibilidade Mecânica']
    print("\n=== Planilha: 1_Disponibilidade Mecânica ===")
    
    # Verificar formato da célula
    for row in range(2, min(ws.max_row + 1, 6)):  # Limitar a 5 linhas para não sobrecarregar a saída
        frota = ws.cell(row=row, column=1).value
        disponibilidade = ws.cell(row=row, column=2).value
        formato = ws.cell(row=row, column=2).number_format
        
        print(f"Frota: {frota}")
        print(f"  Valor: {disponibilidade}")
        print(f"  Formato: '{formato}'")
        print(f"  Tipo: {type(disponibilidade)}")
        print()

# Corrigir o formato para garantir que seja '0.00%' (com ponto decimal)
print("\nCorrigindo formato das células para garantir '0.00%'...")

# Planilha de Disponibilidade Mecânica
if '1_Disponibilidade Mecânica' in wb.sheetnames:
    ws = wb['1_Disponibilidade Mecânica']
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        cell.number_format = '0.00%'  # Garantir formato correto

# Planilha de Uso GPS
if '2_Uso GPS' in wb.sheetnames:
    ws = wb['2_Uso GPS']
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        cell.number_format = '0.00%'  # Garantir formato correto

# Planilha de Motor Ocioso
if '3_Motor Ocioso' in wb.sheetnames:
    ws = wb['3_Motor Ocioso']
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        cell.number_format = '0.00%'  # Garantir formato correto

# Salvar as alterações
novo_arquivo = '../output/colhedorasFrente08_14072025_corrigido.xlsx'
wb.save(novo_arquivo)

print(f"\nArquivo corrigido salvo como {os.path.basename(novo_arquivo)}")
print("Abra o arquivo no Excel e verifique que os valores são mostrados como porcentagens")
print("com o formato '0.00%' e na barra de fórmulas aparecem como decimais (0-1).") 