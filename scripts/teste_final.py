import pandas as pd
import openpyxl
import os

# Criar um DataFrame simples com valores decimais (0-1)
df = pd.DataFrame({
    'Frota': ['Frota1', 'Frota2', 'Frota3'],
    'Disponibilidade': [0.25, 0.5, 0.75]  # Valores decimais (0-1)
})

print("DataFrame criado:")
print(df)

# Salvar para Excel
arquivo = 'teste_final.xlsx'
with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Teste', index=False)
    
    # Aplicar formatação
    workbook = writer.book
    worksheet = workbook['Teste']
    
    # Formatar coluna de porcentagem
    for row in range(2, len(df) + 2):  # +2 porque o Excel é 1-indexado e temos um cabeçalho
        cell = worksheet.cell(row=row, column=2)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas decimais

print(f"Arquivo '{arquivo}' criado com sucesso!")

# Verificar o formato
wb = openpyxl.load_workbook(arquivo)
ws = wb.active

print("\nVerificando formato das células:")
for row in range(2, len(df) + 2):
    frota = ws.cell(row=row, column=1).value
    valor = ws.cell(row=row, column=2).value
    formato = ws.cell(row=row, column=2).number_format
    
    print(f"Frota: {frota}")
    print(f"  Valor: {valor}")
    print(f"  Formato: '{formato}'")
    print(f"  Tipo: {type(valor)}")
    print()

print("Abra o arquivo no Excel e verifique que os valores são mostrados como porcentagens")
print("com o formato '0.00%' e na barra de fórmulas aparecem como decimais (0-1).") 