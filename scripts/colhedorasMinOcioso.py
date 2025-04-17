"""
Script para processamento completo de dados de monitoramento de colhedoras.
Lê arquivos TXT ou CSV na pasta raiz, processa-os e gera arquivos Excel com planilhas auxiliares prontas.
Também processa arquivos ZIP contendo TXT ou CSV.
"""

import pandas as pd
import numpy as np
import os
import glob
import zipfile
import tempfile
import shutil
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurações
processCsv = False  # Altere para True quando quiser processar arquivos CSV

# Constantes
COLUNAS_REMOVER = [
    'Justificativa Corte Base Desligado',
    'Latitude',
    'Longitude',
    'Regional',
    'Tipo de Equipamento',
    'Unidade',
    'Centro de Custo',
    'Trabalhando em File',
    'Trabalhando Frente Dividida',
    'Trabalhando em Fila'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento', 'Apertura do Rolo', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo', 'Corte Base Automatico/Manual',
    'Descricao Equipamento', 'Estado', 'Estado Operacional', 'Esteira Ligada',
    'Field Cruiser', 'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Implemento Ligado', 'Motor Ligado', 'Operacao', 'Operador', 'Pressao de Corte',
    'RPM Extrator', 'RPM Motor', 'RTK (Piloto Automatico)', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parada com Motor Ligado',
    'Horas Produtivas'
]

# Valores a serem filtrados
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO"]

# Mapeamento de valores booleanos para 1/0
MAPEAMENTO_BOOLEANO = {
    'VERDADEIRO': 1, 'FALSO': 0,
    'TRUE': 1, 'FALSE': 0,
    'LIGADO': 1, 'DESLIGADO': 0,
    True: 1, False: 0,
    1: 1, 0: 0
}

def carregar_config_calculos():
    """
    Carrega as configurações de cálculos do arquivo JSON.
    Se o arquivo não existir, retorna configurações padrão.
    """
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config", "calculos_config.json")
    
    # Configuração padrão
    config_padrao = {
        "CD": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [],
                "grupos_operacao_excluidos": []
            },
            "equipamentos_excluidos": []
        },
        "TR": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [],
                "grupos_operacao_excluidos": []
            },
            "equipamentos_excluidos": []
        }
    }
    
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                
                # Garantir que os equipamentos excluídos sejam tratados como texto
                for tipo in ["CD", "TR"]:
                    if tipo in config and "equipamentos_excluidos" in config[tipo]:
                        config[tipo]["equipamentos_excluidos"] = [str(eq).replace('.0', '') for eq in config[tipo]["equipamentos_excluidos"]]
                
                return config
        else:
            # Criar diretório config se não existir
            config_dir = os.path.dirname(config_path)
            if not os.path.exists(config_dir):
                os.makedirs(config_dir)
                
            # Criar arquivo de configuração padrão
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config_padrao, f, indent=4, ensure_ascii=False)
                
            print(f"Arquivo de configuração criado em {config_path} com valores padrão.")
            return config_padrao
    except Exception as e:
        print(f"Erro ao carregar configurações: {str(e)}. Usando configuração padrão.")
        return config_padrao

def calcular_motor_ocioso_novo(df):
    """
    Calcula o tempo de motor ocioso de acordo com as novas regras:
    1. Intervalo é fechado quando encontra 'Parada com Motor Ligado = 0' com duração > 1 minuto
    2. Soma sequências de 'Parada com Motor Ligado = 1' com 'Parada com Motor Ligado = 0' < 1 minuto
    3. Se o total > 1 minuto, subtrai 1 minuto; se menor, descarta o intervalo
    
    Args:
        df (DataFrame): DataFrame com os dados de operação
        
    Returns:
        DataFrame: DataFrame com a coluna 'Motor Ocioso' atualizada
    """
    # Converter a coluna de diferença para timedelta em minutos
    df['Diferença_Minutos'] = df['Diferença_Hora'].dt.total_seconds() / 60
    
    # Inicializar colunas
    df['Motor Ocioso'] = 0
    df['Em_Intervalo'] = False
    df['Soma_Intervalo'] = 0
    
    # Variáveis para controle do intervalo atual
    em_intervalo = False
    soma_intervalo = 0
    inicio_intervalo = None
    
    # Iterar sobre as linhas do DataFrame
    for i in range(len(df)):
        parada_motor = df.iloc[i]['Parada com Motor Ligado']
        diferenca = df.iloc[i]['Diferença_Minutos']
        
        # Se não estamos em um intervalo
        if not em_intervalo:
            # Se encontrar Parada com Motor Ligado = 1, inicia novo intervalo
            if parada_motor == 1:
                em_intervalo = True
                soma_intervalo = diferenca
                inicio_intervalo = i
                df.at[i, 'Em_Intervalo'] = True
                df.at[i, 'Soma_Intervalo'] = soma_intervalo
        
        # Se estamos em um intervalo
        else:
            # Se encontrar Parada com Motor Ligado = 0
            if parada_motor == 0:
                # Se a duração for > 1 minuto, fecha o intervalo
                if diferenca > 1:
                    # Se o total acumulado > 1 minuto, subtrai 1 minuto
                    if soma_intervalo > 1:
                        tempo_ocioso = soma_intervalo - 1
                        # Atribui o tempo ocioso à primeira linha do intervalo
                        df.at[inicio_intervalo, 'Motor Ocioso'] = tempo_ocioso
                    
                    # Reseta o intervalo
                    em_intervalo = False
                    soma_intervalo = 0
                    inicio_intervalo = None
                else:
                    # Se <= 1 minuto, soma ao intervalo atual
                    soma_intervalo += diferenca
                    df.at[i, 'Em_Intervalo'] = True
                    df.at[i, 'Soma_Intervalo'] = soma_intervalo
            
            # Se encontrar Parada com Motor Ligado = 1
            else:
                soma_intervalo += diferenca
                df.at[i, 'Em_Intervalo'] = True
                df.at[i, 'Soma_Intervalo'] = soma_intervalo
    
    # Tratar último intervalo aberto, se houver
    if em_intervalo and soma_intervalo > 1:
        tempo_ocioso = soma_intervalo - 1
        df.at[inicio_intervalo, 'Motor Ocioso'] = tempo_ocioso
    
    # Remover colunas auxiliares
    df = df.drop(['Diferença_Minutos', 'Em_Intervalo', 'Soma_Intervalo'], axis=1)
    
    return df

def processar_arquivo_base(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo (TXT ou CSV são tratados da mesma forma se usarem separador ';')
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio (apenas cabeçalhos sem dados)
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                # Retornar o DataFrame vazio mas com as colunas, em vez de None
                # Garantir que todas as colunas desejadas existam
                for col in COLUNAS_DESEJADAS:
                    if col not in df.columns:
                        df[col] = np.nan
                # Reorganizar as colunas na ordem desejada
                df = df[COLUNAS_DESEJADAS]
                return df
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            df = df[COLUNAS_DESEJADAS]
            
            # Remover linhas onde Data ou Hora são nulos
            df = df.dropna(subset=['Data', 'Hora'])
            
            # Converter colunas de data e hora para string e limpar espaços
            df['Data'] = df['Data'].astype(str).str.strip()
            df['Hora'] = df['Hora'].astype(str).str.strip()
            
            # Criar coluna de data/hora combinada
            df['Data_Hora'] = pd.to_datetime(df['Data'] + ' ' + df['Hora'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
            
            # Remover linhas onde a conversão de data/hora falhou
            df = df.dropna(subset=['Data_Hora'])
            
            # Ordenar por data/hora
            df = df.sort_values('Data_Hora')
            
            # Calcular diferença de tempo entre linhas consecutivas
            df['Diferença_Hora'] = df['Data_Hora'].diff()
            
            # Converter 'Parada com Motor Ligado' para numérico
            df['Parada com Motor Ligado'] = pd.to_numeric(df['Parada com Motor Ligado'], errors='coerce').fillna(0)
            
            # Aplicar o novo cálculo de motor ocioso
            df = calcular_motor_ocioso_novo(df)
            
            return df
            
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado.
    Calcula médias diárias considerando os dias efetivos de trabalho de cada operador.
    
    Cálculos principais:
    - Horas totais: soma de Diferença_Hora
    - Horas elevador: soma de Diferença_Hora onde Esteira Ligada = 1 E Pressão de Corte > 400
    - Motor Ligado: soma de Diferença_Hora onde Motor Ligado = 1
    - Parado Com Motor Ligado: soma de Diferença_Hora onde Motor Ligado = 1 E Velocidade = 0
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Detectar número de dias totais nos dados (apenas para informação)
    dias_unicos_total = df['Data'].nunique() if 'Data' in df.columns else 1
    print(f"Detectados {dias_unicos_total} dias distintos na base de dados.")
    
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Determinar o número de dias efetivos para este operador
        dias_operador = dados_filtrados['Data'].nunique() if 'Data' in dados_filtrados.columns else 1
        
        # Horas totais - soma de Diferença_Hora
        horas_totais = dados_filtrados['Diferença_Hora'].sum()
        if dias_operador > 1:
            horas_totais = horas_totais / dias_operador
        
        # Motor Ligado - soma de Diferença_Hora onde Motor Ligado = 1
        motor_ligado = dados_filtrados[
            dados_filtrados['Motor Ligado'] == 1
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            motor_ligado = motor_ligado / dias_operador
        
        # Horas elevador - soma de Diferença_Hora onde Esteira Ligada = 1 E Pressão de Corte > 400
        horas_elevador = dados_filtrados[
            (dados_filtrados['Esteira Ligada'] == 1) & 
            (dados_filtrados['Pressao de Corte'] > 400)
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            horas_elevador = horas_elevador / dias_operador
        
        # Percentual horas elevador (em decimal 0-1)
        percent_elevador = calcular_porcentagem(horas_elevador, horas_totais)
        
        # RTK - soma de Diferença_Hora onde todas as condições são atendidas
        rtk = dados_filtrados[
            (dados_filtrados['Operacao'] == '7290 - COLHEITA CANA MECANIZADA') &
            (dados_filtrados['Pressao de Corte'] > 300) &
            (dados_filtrados['RTK (Piloto Automatico)'] == 1) &
            (dados_filtrados['Esteira Ligada'] == 1)
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            rtk = rtk / dias_operador
        
        # Horas Produtivas
        horas_produtivas = dados_filtrados[
            dados_filtrados['Grupo Operacao'] == 'Produtiva'
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            horas_produtivas = horas_produtivas / dias_operador
        
        # % Utilização RTK (em decimal 0-1)
        utilizacao_rtk = calcular_porcentagem(rtk, horas_produtivas)
        
        # % Eficiência Elevador (em decimal 0-1)
        eficiencia_elevador = calcular_porcentagem(horas_elevador, motor_ligado)
        
        # Parado com Motor Ligado - soma de Diferença_Hora onde Motor Ligado = 1 E Velocidade = 0
        parado_motor_ligado = dados_filtrados[
            (dados_filtrados['Motor Ligado'] == 1) & 
            (dados_filtrados['Velocidade'] == 0)
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            parado_motor_ligado = parado_motor_ligado / dias_operador
        
        # % Parado com motor ligado (em decimal 0-1)
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        # Debug para verificar os valores
        print(f"\nOperador: {operador} em {equipamento}")
        print(f"Motor Ligado: {motor_ligado:.6f}")
        print(f"Parado com Motor Ligado: {parado_motor_ligado:.6f}")
        print(f"% Parado com motor ligado: {percent_parado_motor:.6f}")
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas elevador': horas_elevador,
            '%': percent_elevador,
            'RTK': rtk,
            'Horas Produtivas': horas_produtivas,
            '% Utilização RTK': utilizacao_rtk,
            'Motor Ligado': motor_ligado,
            '% Eficiência Elevador': eficiencia_elevador,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor
        })
    
    return pd.DataFrame(resultados)

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento.
    Calcula médias diárias considerando os dias efetivos de cada equipamento.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento
    """
    # Filtramos os dados excluindo os operadores da lista
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df_filtrado['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        
        # Determinar número de dias efetivos para este equipamento
        dias_equip = dados_equip['Data'].nunique() if 'Data' in dados_equip.columns else 1
        
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 4)
        
        # Calcular horas de manutenção
        manutencao = round(dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum(), 4)
        
        # Se houver múltiplos dias, usar médias diárias
        if dias_equip > 1:
            total_horas = round(total_horas / dias_equip, 4)
            manutencao = round(manutencao / dias_equip, 4)
            print(f"Equipamento: {equipamento}, Dias efetivos: {dias_equip}, Média diária: {total_horas:.2f} horas")
        
        # A disponibilidade mecânica é o percentual de tempo fora de manutenção
        disp_mecanica = calcular_porcentagem(total_horas - manutencao, total_horas)
        
        resultados.append({
            'Frota': equipamento,
            'Disponibilidade': disp_mecanica
        })
    
    return pd.DataFrame(resultados)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Calcula médias diárias considerando os dias efetivos de cada frota.
    Esta função NÃO aplica qualquer filtro de operador.
    Também identifica as faltas de horário por dia específico.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota com detalhamento por dia
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    # Obter todos os dias únicos no dataset
    dias_unicos = sorted(df['Data'].unique()) if 'Data' in df.columns else []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        
        # Determinar número de dias efetivos para este equipamento
        dias_equip = dados_equip['Data'].nunique() if 'Data' in dados_equip.columns else 1
        
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 2)
        
        # Se houver múltiplos dias, usar média diária
        if dias_equip > 1:
            total_horas = round(total_horas / dias_equip, 2)
        
        # Calcular a diferença para 24 horas
        diferenca_24h = round(max(24 - total_horas, 0), 2)
        
        # Criar o resultado básico (colunas originais mantidas)
        resultado = {
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        }
        
        # Adicionar detalhamento por dia (novas colunas)
        if len(dias_unicos) > 0:
            for dia in dias_unicos:
                dados_dia = dados_equip[dados_equip['Data'] == dia]
                
                # Se não houver dados para este dia e equipamento, a diferença é 24h
                if len(dados_dia) == 0:
                    resultado[f'Falta {dia}'] = 24.0
                    continue
                
                # Calcular horas registradas neste dia
                horas_dia = round(dados_dia['Diferença_Hora'].sum(), 2)
                
                # Calcular a diferença para 24 horas neste dia
                diferenca_dia = round(max(24 - horas_dia, 0), 2)
                
                # Adicionar ao resultado apenas se houver falta (diferença > 0)
                if diferenca_dia > 0:
                    resultado[f'Falta {dia}'] = diferenca_dia
                else:
                    resultado[f'Falta {dia}'] = 0.0
        
        resultados.append(resultado)
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso(base_calculo, df_base):
    """
    Calcula o percentual de motor ocioso por operador usando os dados da Base Calculo.
    Agrega os dados por operador, calculando a média quando um operador aparece em múltiplas frotas.
    
    Regras de cálculo:
    1. Existe uma tolerância de 1 minuto para operações de "Parada com motor ligado" (valor 1)
    2. Se uma operação tem "Parada com motor ligado = 1" e dura menos de 1 minuto, ela é desconsiderada se:
       - A próxima operação tem "Parada com motor ligado = 0" e dura mais de 1 minuto
    3. Se existem duas ou mais operações com "Parada com motor ligado = 1" em sequência (ou com outras operações entre elas que duram menos de 1 minuto):
       - Somamos o tempo total dessas operações
       - Subtraímos 1 minuto do total
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
        df_base (DataFrame): DataFrame base para aplicar filtros de operações
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador (agregado)
    """
    # Carregar configurações
    config = carregar_config_calculos()
    operacoes_excluidas = config['CD']['motor_ocioso']['operacoes_excluidas']
    grupos_operacao_excluidos = config['CD']['motor_ocioso']['grupos_operacao_excluidos']

    # Filtrar dados base excluindo operações configuradas
    df_filtrado = df_base[
        ~df_base['Operacao'].isin(operacoes_excluidas) & 
        ~df_base['Grupo Operacao'].isin(grupos_operacao_excluidos)
    ]

    # Agrupar por operador e calcular as somas apenas dos dados filtrados
    resultados = []
    print("\n=== DETALHAMENTO DO CÁLCULO DE MOTOR OCIOSO (AGREGADO) ===")
    print("Usando dados filtrados (excluindo operações configuradas)")
    print("=" * 50)

    # Converter a coluna de diferença para minutos
    df_filtrado['Diferença_Minutos'] = df_filtrado['Diferença_Hora'] * 60

    # Agrupar por operador
    for operador in df_filtrado['Operador'].unique():
        if operador in OPERADORES_EXCLUIR:
            continue

        # Filtrar dados do operador
        dados_operador = df_filtrado[df_filtrado['Operador'] == operador].copy()
        
        # Ordenar por Data e Hora para garantir sequência temporal
        if 'Data' in dados_operador.columns and 'Hora' in dados_operador.columns:
            dados_operador = dados_operador.sort_values(by=['Data', 'Hora'])
        
        # Calcular tempo total com motor ligado
        tempo_ligado = dados_operador[dados_operador['Motor Ligado'] == 1]['Diferença_Hora'].sum()
        
        # Calcular tempo ocioso com a nova lógica de tolerância
        tempo_ocioso = 0
        em_intervalo = False
        soma_intervalo = 0
        
        for i in range(len(dados_operador)):
            parada_motor = dados_operador.iloc[i]['Parada com Motor Ligado']
            diferenca = dados_operador.iloc[i]['Diferença_Minutos']
            
            # Se não estamos em um intervalo
            if not em_intervalo:
                # Se encontrar Parada com Motor Ligado = 1, inicia novo intervalo
                if parada_motor == 1:
                    em_intervalo = True
                    soma_intervalo = diferenca
            
            # Se estamos em um intervalo
            else:
                # Se encontrar Parada com Motor Ligado = 0
                if parada_motor == 0:
                    # Se a duração for > 1 minuto, fecha o intervalo
                    if diferenca > 1:
                        # Se o total acumulado > 1 minuto, subtrai 1 minuto
                        if soma_intervalo > 1:
                            tempo_ocioso += (soma_intervalo - 1) / 60  # Converter de volta para horas
                        
                        # Reseta o intervalo
                        em_intervalo = False
                        soma_intervalo = 0
                    else:
                        # Se <= 1 minuto, soma ao intervalo atual
                        soma_intervalo += diferenca
                
                # Se encontrar Parada com Motor Ligado = 1
                else:
                    soma_intervalo += diferenca
        
        # Tratar último intervalo aberto, se houver
        if em_intervalo and soma_intervalo > 1:
            tempo_ocioso += (soma_intervalo - 1) / 60  # Converter de volta para horas
        
        # Calcular porcentagem de tempo ocioso
        porcentagem = tempo_ocioso / tempo_ligado if tempo_ligado > 0 else 0
        
        print(f"\nOperador: {operador}")
        print(f"Tempo Ocioso (soma) = {tempo_ocioso:.6f}")
        print(f"Tempo Ligado (soma) = {tempo_ligado:.6f}")
        print(f"Porcentagem = {porcentagem:.6f}")
        print("-" * 50)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': porcentagem,
            'Tempo Ligado': tempo_ligado,
            'Tempo Ocioso': tempo_ocioso
        })
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica(base_calculo):
    """
    Calcula a eficiência energética por operador usando os dados da Base Calculo.
    Agrega os dados por operador, calculando a média ponderada quando um operador aparece em múltiplas frotas.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador (agregado)
    """
    # Agrupar por operador e calcular a média ponderada
    agrupado = base_calculo.groupby('Operador').agg({
        'Horas elevador': 'sum',
        'Motor Ligado': 'sum'
    }).reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        eficiencia = row['Horas elevador'] / row['Motor Ligado'] if row['Motor Ligado'] > 0 else 0
        resultados.append({
            'Operador': row['Operador'],
            'Eficiência': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_hora_elevador(df_base, base_calculo):
    """
    Extrai as horas de elevador da Base Calculo.
    Agrega os dados por operador, somando quando um operador aparece em múltiplas frotas.
    
    Args:
        df_base: Não usado mais, mantido para compatibilidade
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Horas de elevador por operador (agregado)
    """
    # Agrupar por operador e somar as horas
    agrupado = base_calculo.groupby('Operador')['Horas elevador'].sum().reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        resultados.append({
            'Operador': row['Operador'],
            'Horas': row['Horas elevador']
        })
    
    return pd.DataFrame(resultados)

def calcular_uso_gps(df_base, base_calculo):
    """
    Extrai o uso de GPS da Base Calculo.
    Agrega os dados por operador, calculando a média ponderada quando um operador aparece em múltiplas frotas.
    
    Args:
        df_base: Não usado mais, mantido para compatibilidade
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador (agregado)
    """
    # Agrupar por operador e calcular a média ponderada
    agrupado = base_calculo.groupby('Operador').agg({
        'RTK': 'sum',
        'Horas Produtivas': 'sum'
    }).reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        porcentagem = row['RTK'] / row['Horas Produtivas'] if row['Horas Produtivas'] > 0 else 0
        resultados.append({
            'Operador': row['Operador'],
            'Porcentagem': porcentagem
        })
    
    return pd.DataFrame(resultados)

def calcular_media_velocidade(df):
    """
    Calcula a média de velocidade para cada operador considerando apenas registros produtivos e deslocamentos.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: DataFrame com a média de velocidade por operador
    """
    # Filtramos os dados excluindo os operadores da lista de exclusão
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Identificar registros produtivos e deslocamentos
    # Produtivos: onde Grupo Operacao é 'Produtiva'
    # Deslocamentos: onde Estado é 'DESLOCAMENTO' ou velocidade > 0 em operações não produtivas
    registros_validos = (
        (df_filtrado['Grupo Operacao'] == 'Produtiva') | 
        (df_filtrado['Estado'] == 'DESLOCAMENTO') |
        ((df_filtrado['Velocidade'] > 0) & (df_filtrado['Estado'] != 'PARADO'))
    )
    
    df_velocidade = df_filtrado[registros_validos]
    
    # Obter lista de operadores únicos do DataFrame original
    operadores = df_filtrado['Operador'].unique()
    resultados = []
    
    for operador in operadores:
        # Filtrar dados para este operador
        dados_op = df_velocidade[df_velocidade['Operador'] == operador]
        
        # Se não houver dados válidos para este operador, adicionar com velocidade zero
        if len(dados_op) == 0:
            resultados.append({
                'Operador': operador,
                'Velocidade': 0
            })
            continue
        
        # Calcular a média de velocidade (ponderada pelo tempo, se houver a coluna Diferença_Hora)
        if 'Diferença_Hora' in dados_op.columns:
            # Média ponderada pelo tempo 
            velocidade_media = round(
                (dados_op['Velocidade'] * dados_op['Diferença_Hora']).sum() / dados_op['Diferença_Hora'].sum()
                if dados_op['Diferença_Hora'].sum() > 0 else 0,
                2
            )
        else:
            # Média simples
            velocidade_media = round(dados_op['Velocidade'].mean(), 2)
        
        resultados.append({
            'Operador': operador,
            'Velocidade': velocidade_media
        })
    
    return pd.DataFrame(resultados)

def identificar_operadores_duplicados(df):
    """
    Identifica operadores que aparecem com IDs diferentes no mesmo conjunto de dados.
    Detecta principalmente IDs que começam com '133' e têm 7 dígitos, verificando se
    existe outra ID com o mesmo nome mas com menos dígitos.
    
    Args:
        df (DataFrame): DataFrame com os dados dos operadores
    
    Returns:
        dict: Dicionário com mapeamento {id_incorreta: id_correta}
        DataFrame: DataFrame com as duplicidades encontradas para relatório
    """
    if 'Operador' not in df.columns or len(df) == 0:
        return {}, pd.DataFrame(columns=['ID Incorreta', 'ID Correta', 'Nome'])
    
    # Extrair operadores únicos
    operadores = df['Operador'].unique()
    
    # Mapear nomes para IDs
    nomes_para_ids = {}
    for op in operadores:
        if ' - ' in op:
            try:
                id_parte, nome_parte = op.split(' - ', 1)
                if nome_parte not in nomes_para_ids:
                    nomes_para_ids[nome_parte] = []
                nomes_para_ids[nome_parte].append(op)
            except:
                continue
    
    # Encontrar nomes com múltiplas IDs
    duplicidades = []
    mapeamento = {}
    
    for nome, ids in nomes_para_ids.items():
        if len(ids) > 1:
            # Verificar se uma das IDs inicia com 133 e tem 7 dígitos
            ids_longas = [id_op for id_op in ids if ' - ' in id_op and id_op.split(' - ')[0].startswith('133') and len(id_op.split(' - ')[0]) == 7]
            ids_curtas = [id_op for id_op in ids if id_op not in ids_longas]
            
            if ids_longas and ids_curtas:
                for id_longa in ids_longas:
                    # Encontrar a ID correta (a mais curta)
                    id_correta = min(ids_curtas, key=lambda x: len(x.split(' - ')[0]) if ' - ' in x else float('inf'))
                    
                    # Adicionar ao mapeamento
                    mapeamento[id_longa] = id_correta
                    
                    # Adicionar à lista de duplicidades para o relatório
                    duplicidades.append({
                        'ID Incorreta': id_longa,
                        'ID Correta': id_correta,
                        'Nome': nome
                    })
    
    print(f"Encontrados {len(duplicidades)} operadores com IDs duplicadas.")
    return mapeamento, pd.DataFrame(duplicidades)

def criar_excel_com_planilhas(df_base_calculo, df_motor_ocioso, df_horas_por_frota, caminho_saida):
    """
    Cria um arquivo Excel com três planilhas: BASE_CALCULO, MOTOR_OCIOSO e HORAS_POR_FROTA.
    
    Args:
        df_base_calculo (DataFrame): DataFrame com os dados da base de cálculo
        df_motor_ocioso (DataFrame): DataFrame com os dados de motor ocioso
        df_horas_por_frota (DataFrame): DataFrame com as horas por frota
        caminho_saida (str): Caminho do arquivo Excel de saída
    """
    try:
        # Criar cópias dos DataFrames para não modificar os originais
        df_base = df_base_calculo.copy()
        df_motor = df_motor_ocioso.copy()
        df_horas = df_horas_por_frota.copy()
        
        # Identificar colunas de tempo em cada DataFrame
        colunas_tempo_base = ['Diferença_Hora', 'Horas Produtivas']
        colunas_tempo_motor = ['Tempo Ocioso', 'Tempo Ligado']
        colunas_tempo_horas = ['Horas Produtivas', 'Horas Improdutivas', 'Horas Totais']
        
        # Ordenar df_motor_ocioso pela porcentagem (do menor para o maior)
        if 'Porcentagem' in df_motor.columns:
            df_motor = df_motor.sort_values('Porcentagem')
            
        # Ordenar df_horas_por_frota pelas horas totais (do maior para o menor)
        if 'Horas Totais' in df_horas.columns:
            df_horas = df_horas.sort_values('Horas Totais', ascending=False)
        
        # Criar um novo arquivo Excel
        writer = pd.ExcelWriter(caminho_saida, engine='openpyxl')
        
        # Salvar os DataFrames nas respectivas abas
        df_base.to_excel(writer, sheet_name='BASE_CALCULO', index=False)
        df_motor.to_excel(writer, sheet_name='MOTOR_OCIOSO', index=False)
        df_horas.to_excel(writer, sheet_name='HORAS_POR_FROTA', index=False)
        
        # Ajustar largura das colunas e aplicar formatação
        workbook = writer.book
        
        # Configurar cada planilha
        sheets_config = {
            'BASE_CALCULO': {'df': df_base, 'time_cols': colunas_tempo_base},
            'MOTOR_OCIOSO': {'df': df_motor, 'time_cols': colunas_tempo_motor},
            'HORAS_POR_FROTA': {'df': df_horas, 'time_cols': colunas_tempo_horas}
        }
        
        for sheet_name, config in sheets_config.items():
            worksheet = writer.sheets[sheet_name]
            df = config['df']
            time_cols = config['time_cols']
            
            # Ajustar largura das colunas
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                )
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width
                
                # Aplicar formatação de tempo sem dividir por 24
                if col in time_cols:
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=idx + 1)
                        cell.number_format = '[h]:mm:ss'  # Formato que permite mais de 24 horas
                
                # Aplicar formatação de porcentagem
                if col == 'Porcentagem':
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=idx + 1)
                        cell.number_format = '0.00%'
        
        # Salvar o arquivo
        writer.close()
        print(f"Arquivo {caminho_saida} gerado com sucesso!")
        
    except Exception as e:
        print(f"Erro ao gerar arquivo {caminho_saida}: {str(e)}")
        print(f"Arquivo {caminho_saida} gerado com apenas a planilha BASE_CALCULO (sem dados).")

def extrair_arquivo_zip(caminho_zip, pasta_destino=None):
    """
    Extrai o conteúdo de um arquivo ZIP para uma pasta temporária ou destino especificado.
    Renomeia os arquivos extraídos para terem o mesmo nome do arquivo ZIP original.
    
    Args:
        caminho_zip (str): Caminho para o arquivo ZIP
        pasta_destino (str, optional): Pasta onde os arquivos serão extraídos.
                                       Se None, usa uma pasta temporária.
    
    Returns:
        list: Lista de caminhos dos arquivos extraídos e renomeados (apenas TXT e CSV)
        str: Caminho da pasta temporária (se criada) ou None
    """
    # Se pasta_destino não foi especificada, criar uma pasta temporária
    pasta_temp = None
    if pasta_destino is None:
        pasta_temp = tempfile.mkdtemp()
        pasta_destino = pasta_temp
    
    arquivos_extraidos = []
    nome_zip_sem_extensao = os.path.splitext(os.path.basename(caminho_zip))[0]
    
    try:
        with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
            # Extrair todos os arquivos do ZIP
            zip_ref.extractall(pasta_destino)
            
            # Processar cada arquivo extraído (apenas TXT e CSV)
            for arquivo in zip_ref.namelist():
                caminho_completo = os.path.join(pasta_destino, arquivo)
                # Verificar se é um arquivo e não uma pasta
                if os.path.isfile(caminho_completo):
                    # Verificar extensão
                    extensao = os.path.splitext(arquivo)[1].lower()
                    if extensao in ['.txt', '.csv']:
                        # Criar novo nome: nome do ZIP + extensão original
                        novo_nome = f"{nome_zip_sem_extensao}{extensao}"
                        novo_caminho = os.path.join(pasta_destino, novo_nome)
                        
                        # Renomear o arquivo extraído
                        try:
                            # Se já existe um arquivo com esse nome, remover primeiro
                            if os.path.exists(novo_caminho):
                                os.remove(novo_caminho)
                            # Renomear o arquivo
                            os.rename(caminho_completo, novo_caminho)
                            arquivos_extraidos.append(novo_caminho)
                            print(f"Arquivo extraído renomeado: {novo_nome}")
                        except Exception as e:
                            print(f"Erro ao renomear arquivo {arquivo} para {novo_nome}: {str(e)}")
                            arquivos_extraidos.append(caminho_completo)  # Adicionar o caminho original em caso de erro
        
        return arquivos_extraidos, pasta_temp
    
    except Exception as e:
        print(f"Erro ao extrair o arquivo ZIP {caminho_zip}: {str(e)}")
        # Se houve erro e criamos uma pasta temporária, tentar limpá-la
        if pasta_temp:
            try:
                shutil.rmtree(pasta_temp)
            except:
                pass
        return [], None

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT, CSV ou ZIP de colhedoras na pasta dados.
    Ignora arquivos que contenham "transbordo" no nome.
    """
    print("\nIniciando processamento de arquivos...")
    
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    print(f"Diretório do script: {diretorio_script}")
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    print(f"Diretório raiz: {diretorio_raiz}")
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    print(f"Diretório de dados: {diretorio_dados}")
    print(f"Diretório de saída: {diretorio_saida}")
    
    # Verificar se os diretórios existem, caso contrário criar
    if not os.path.exists(diretorio_dados):
        print(f"Criando diretório de dados: {diretorio_dados}")
        os.makedirs(diretorio_dados)
    if not os.path.exists(diretorio_saida):
        print(f"Criando diretório de saída: {diretorio_saida}")
        os.makedirs(diretorio_saida)
    
    # Encontrar todos os arquivos TXT/CSV/ZIP de colhedoras
    arquivos = []
    arquivos_zip = []
    
    # Adicionar arquivos TXT sempre
    arquivos += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.txt"))
    arquivos += glob.glob(os.path.join(diretorio_dados, "*colhedora*.txt"))
    arquivos += glob.glob(os.path.join(diretorio_dados, "colhedora*.txt"))
    
    # Adicionar arquivos CSV apenas se processCsv for True
    if processCsv:
        arquivos += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.csv"))
        arquivos += glob.glob(os.path.join(diretorio_dados, "*colhedora*.csv"))
        arquivos += glob.glob(os.path.join(diretorio_dados, "colhedora*.csv"))
    
    # Adicionar arquivos ZIP
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.zip"))
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "*colhedora*.zip"))
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "colhedora*.zip"))
    
    print("\nArquivos encontrados antes da filtragem:")
    print(f"TXT/CSV: {[os.path.basename(a) for a in arquivos]}")
    print(f"ZIP: {[os.path.basename(a) for a in arquivos_zip]}")
    
    # Filtrar arquivos que contenham "transbordo" no nome (case insensitive)
    arquivos = [arquivo for arquivo in arquivos if "transbordo" not in os.path.basename(arquivo).lower()]
    arquivos_zip = [arquivo for arquivo in arquivos_zip if "transbordo" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos = list(set(arquivos))
    arquivos_zip = list(set(arquivos_zip))
    
    print("\nArquivos encontrados após a filtragem:")
    print(f"TXT/CSV: {[os.path.basename(a) for a in arquivos]}")
    print(f"ZIP: {[os.path.basename(a) for a in arquivos_zip]}")
    
    if not arquivos and not arquivos_zip:
        print("Nenhum arquivo de colhedoras encontrado na pasta dados!")
        return
    
    print(f"\nEncontrados {len(arquivos)} arquivos de colhedoras (TXT/CSV) para processar.")
    print(f"Encontrados {len(arquivos_zip)} arquivos ZIP de colhedoras para processar.")
    
    # Processar cada arquivo TXT/CSV
    for arquivo in arquivos:
        print(f"\nProcessando arquivo TXT/CSV: {os.path.basename(arquivo)}")
        processar_arquivo(arquivo, diretorio_saida)
    
    # Processar cada arquivo ZIP
    for arquivo_zip in arquivos_zip:
        print(f"\nProcessando arquivo ZIP: {os.path.basename(arquivo_zip)}")
        
        # Extrair arquivo ZIP para pasta temporária
        print(f"Extraindo arquivo ZIP: {arquivo_zip}")
        arquivos_extraidos, pasta_temp = extrair_arquivo_zip(arquivo_zip)
        
        if not arquivos_extraidos:
            print(f"Nenhum arquivo TXT ou CSV encontrado no ZIP {os.path.basename(arquivo_zip)}")
            continue
        
        print(f"Extraídos {len(arquivos_extraidos)} arquivos do ZIP:")
        for arquivo in arquivos_extraidos:
            print(f"  - {os.path.basename(arquivo)}")
        
        # Processar cada arquivo extraído
        for arquivo_extraido in arquivos_extraidos:
            # Filtrar arquivos que contenham "transbordo" no nome
            if "transbordo" not in os.path.basename(arquivo_extraido).lower():
                print(f"\nProcessando arquivo extraído: {os.path.basename(arquivo_extraido)}")
                processar_arquivo(arquivo_extraido, diretorio_saida)
        
        # Limpar pasta temporária se foi criada
        if pasta_temp:
            try:
                print(f"Removendo pasta temporária: {pasta_temp}")
                shutil.rmtree(pasta_temp)
                print(f"Pasta temporária removida com sucesso")
            except Exception as e:
                print(f"Erro ao remover pasta temporária {pasta_temp}: {str(e)}")
    
    print("\nProcessamento de todos os arquivos concluído!")

def carregar_substituicoes_operadores():
    """
    Carrega o arquivo substituiroperadores.json que contém os mapeamentos 
    de substituição de operadores.
    
    Returns:
        dict: Dicionário com mapeamento {operador_origem: operador_destino}
        ou dicionário vazio se o arquivo não existir ou for inválido
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Caminho para o arquivo de substituição
    arquivo_substituicao = os.path.join(diretorio_raiz, "config", "substituiroperadores.json")
    
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo_substituicao):
        print(f"Arquivo de substituição de operadores não encontrado: {arquivo_substituicao}")
        return {}
    
    try:
        # Carregar o arquivo JSON
        with open(arquivo_substituicao, 'r', encoding='utf-8') as f:
            substituicoes = json.load(f)
        
        # Criar dicionário de substituições
        mapeamento = {item['operador_origem']: item['operador_destino'] for item in substituicoes}
        
        print(f"Carregadas {len(mapeamento)} substituições de operadores.")
        return mapeamento
        
    except Exception as e:
        print(f"Erro ao carregar arquivo de substituição de operadores: {str(e)}")
        return {}

def aplicar_substituicao_operadores(df, mapeamento_substituicoes):
    """
    Aplica as substituições de operadores no DataFrame.
    
    Args:
        df (DataFrame): DataFrame a ser processado
        mapeamento_substituicoes (dict): Dicionário com mapeamento {operador_origem: operador_destino}
    
    Returns:
        tuple: (DataFrame com substituições aplicadas, DataFrame com registro das substituições)
    """
    if not mapeamento_substituicoes or 'Operador' not in df.columns:
        return df, pd.DataFrame(columns=['ID Original', 'Nome Original', 'ID Nova', 'Nome Novo', 'Registros Afetados'])
    
    # Criar uma cópia para não alterar o DataFrame original
    df_modificado = df.copy()
    
    # Lista para armazenar as substituições realizadas
    substituicoes_realizadas = []
    
    # Contar operadores antes da substituição
    contagem_antes = df_modificado['Operador'].value_counts()
    
    # Aplicar as substituições
    df_modificado['Operador'] = df_modificado['Operador'].replace(mapeamento_substituicoes)
    
    # Contar operadores depois da substituição
    contagem_depois = df_modificado['Operador'].value_counts()
    
    # Verificar quais operadores foram substituídos
    for operador_origem, operador_destino in mapeamento_substituicoes.items():
        if operador_origem in contagem_antes:
            registros_afetados = contagem_antes.get(operador_origem, 0)
            if registros_afetados > 0:
                # Extrair IDs e nomes
                id_original = operador_origem.split(' - ')[0] if ' - ' in operador_origem else operador_origem
                nome_original = operador_origem.split(' - ')[1] if ' - ' in operador_origem else ''
                id_nova = operador_destino.split(' - ')[0] if ' - ' in operador_destino else operador_destino
                nome_novo = operador_destino.split(' - ')[1] if ' - ' in operador_destino else ''
                
                substituicoes_realizadas.append({
                    'ID Original': id_original,
                    'Nome Original': nome_original,
                    'ID Nova': id_nova,
                    'Nome Novo': nome_novo,
                    'Registros Afetados': registros_afetados
                })
                print(f"Operador '{operador_origem}' substituído por '{operador_destino}' em {registros_afetados} registros")
    
    # Criar DataFrame com as substituições realizadas
    df_substituicoes = pd.DataFrame(substituicoes_realizadas)
    
    return df_modificado, df_substituicoes

def processar_arquivo(caminho_arquivo, diretorio_saida):
    """
    Processa um único arquivo e gera o Excel de saída.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo a ser processado
        diretorio_saida (str): Diretório onde o arquivo de saída será salvo
    """
    try:
        # Processar o arquivo base
        df_base = processar_arquivo_base(caminho_arquivo)
        if df_base is None:
            print(f"Erro ao processar o arquivo {caminho_arquivo}. Pulando para o próximo arquivo.")
            return
        
        # Carregar substituições de operadores
        substituicoes = carregar_substituicoes_operadores()
        print(f"Carregadas {len(substituicoes)} substituições de operadores.")
        
        # Aplicar substituições de operadores
        df_base = aplicar_substituicao_operadores(df_base, substituicoes)
        
        # Calcular motor ocioso
        df_base = calcular_motor_ocioso_novo(df_base)
        
        # Salvar o arquivo Excel
        salvar_planilha_base(df_base, diretorio_saida)
        
    except Exception as e:
        print(f"Erro ao processar arquivo {caminho_arquivo}: {str(e)}")
        print(f"Pulando para o próximo arquivo.")

def salvar_planilha_base(df, caminho_saida):
    """
    Salva o DataFrame em uma planilha Excel com a aba BASE.
    
    Args:
        df (DataFrame): DataFrame com os dados processados
        caminho_saida (str): Caminho do arquivo Excel de saída
    """
    try:
        # Criar uma cópia do DataFrame para não modificar o original
        df_excel = df.copy()
        
        # Identificar colunas de tempo
        colunas_tempo = ['Diferença_Hora', 'Horas Produtivas']
        
        # Criar um novo arquivo Excel
        writer = pd.ExcelWriter(caminho_saida, engine='openpyxl')
        
        # Salvar o DataFrame na aba BASE
        df_excel.to_excel(writer, sheet_name='BASE', index=False)
        
        # Ajustar largura das colunas
        workbook = writer.book
        worksheet = writer.sheets['BASE']
        
        for idx, col in enumerate(df_excel.columns):
            # Encontrar a largura máxima na coluna
            max_length = max(
                df_excel[col].astype(str).apply(len).max(),  # Comprimento máximo dos dados
                len(str(col))  # Comprimento do cabeçalho
            )
            # Adicionar um pouco de espaço extra
            adjusted_width = (max_length + 2)
            # Definir a largura da coluna (converter para unidades do Excel)
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width
            
            # Formatar colunas de tempo
            if col in colunas_tempo:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=idx + 1)
                    cell.number_format = '[h]:mm:ss'  # Formato que permite mais de 24 horas
        
        # Salvar o arquivo
        writer.close()
        print(f"Arquivo {caminho_saida} gerado com sucesso!")
        
    except Exception as e:
        print(f"Erro ao gerar arquivo {caminho_saida}: {str(e)}")
        print(f"Arquivo {caminho_saida} gerado com apenas a planilha BASE (sem dados).")

if __name__ == "__main__":
    print("="*80)
    print("Iniciando processamento de arquivos de colhedoras...")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Este script processa arquivos de colhedoras e gera planilhas Excel com métricas")
    print("Suporta arquivos TXT, CSV e ZIP")
    print("Ignorando arquivos que contenham 'transbordo' no nome")
    print("="*80)
    
    try:
        processar_todos_arquivos()
        print("\nProcessamento concluído com sucesso!")
    except Exception as e:
        print(f"\nErro durante o processamento: {str(e)}")
        print("Detalhes do erro:")
        import traceback
        traceback.print_exc() 