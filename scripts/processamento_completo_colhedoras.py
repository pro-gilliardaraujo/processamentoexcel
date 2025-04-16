"""
Script para processamento completo de dados de monitoramento de colhedoras.
Lê arquivos TXT ou CSV na pasta raiz, processa-os e gera arquivos Excel com planilhas auxiliares prontas.
Também processa arquivos ZIP contendo TXT ou CSV.
"""

import pandas as pd
import numpy as np
import os
import glob
import zipfile
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import json

# Configurações
processCsv = False  # Altere para True quando quiser processar arquivos CSV

# Constantes
COLUNAS_REMOVER = [
    'Justificativa Corte Base Desligado',
    'Latitude',
    'Longitude',
    'Regional',
    'Tipo de Equipamento',
    'Unidade',
    'Centro de Custo'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento', 'Apertura do Rolo', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo', 'Corte Base Automatico/Manual',
    'Descricao Equipamento', 'Estado', 'Estado Operacional', 'Esteira Ligada',
    'Field Cruiser', 'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Implemento Ligado', 'Motor Ligado', 'Operacao', 'Operador', 'Pressao de Corte',
    'RPM Extrator', 'RPM Motor', 'RTK (Piloto Automatico)', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parada com Motor Ligado',
    'Horas Produtivas'
]

# Valores a serem filtrados
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO"]

def carregar_config_calculos():
    """
    Carrega as configurações de cálculos do arquivo JSON.
    Se o arquivo não existir, retorna configurações padrão.
    """
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config", "calculos_config.json")
    
    # Configuração padrão
    config_padrao = {
        "CD": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": []
            },
            "equipamentos_excluidos": []
        },
        "TR": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": []
            },
            "equipamentos_excluidos": []
        }
    }
    
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                
                # Garantir que os equipamentos excluídos sejam tratados como texto
                for tipo in ["CD", "TR"]:
                    if tipo in config and "equipamentos_excluidos" in config[tipo]:
                        config[tipo]["equipamentos_excluidos"] = [str(eq).replace('.0', '') for eq in config[tipo]["equipamentos_excluidos"]]
                
                return config
        else:
            # Criar diretório config se não existir
            config_dir = os.path.dirname(config_path)
            if not os.path.exists(config_dir):
                os.makedirs(config_dir)
                
            # Criar arquivo de configuração padrão
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config_padrao, f, indent=4, ensure_ascii=False)
                
            print(f"Arquivo de configuração criado em {config_path} com valores padrão.")
            return config_padrao
    except Exception as e:
        print(f"Erro ao carregar configurações: {str(e)}. Usando configuração padrão.")
        return config_padrao

def carregar_substituicao_operadores():
    """
    Carrega as configurações de substituição de operadores do arquivo JSON.
    Se o arquivo não existir, cria um novo com configuração padrão.
    Este arquivo serve para padronizar nomes de operadores que podem 
    aparecer com variações nos registros de entrada.
    
    Returns:
        dict: Dicionário com mapeamento de substituição de operadores
    """
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                             "config", "substituicao_operadores.json")
    
    # Configuração padrão com alguns exemplos
    config_padrao = {
        "substituicoes": {
            "JOAO SILVA": "João Silva",
            "JOAO S.": "João Silva",
            "MARIA SANTOS": "Maria Santos",
            "PEDRO OLIVEIRA": "Pedro Oliveira"
        }
    }
    
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config
        else:
            # Criar diretório config se não existir
            config_dir = os.path.dirname(config_path)
            if not os.path.exists(config_dir):
                os.makedirs(config_dir)
                
            # Criar arquivo de configuração padrão
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config_padrao, f, indent=4, ensure_ascii=False)
                
            print(f"Arquivo de substituição de operadores criado em {config_path} com valores de exemplo.")
            return config_padrao
    except Exception as e:
        print(f"Erro ao carregar substituições de operadores: {str(e)}. Usando configuração padrão.")
        return config_padrao

def processar_arquivo_base(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo (TXT ou CSV são tratados da mesma forma se usarem separador ';')
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio (apenas cabeçalhos sem dados)
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                # Retornar o DataFrame vazio mas com as colunas, em vez de None
                # Garantir que todas as colunas desejadas existam
                for col in COLUNAS_DESEJADAS:
                    if col not in df.columns:
                        df[col] = np.nan
                # Reorganizar as colunas na ordem desejada
                colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Verificar se 'Data/Hora' existe e processá-la
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Conversão e cálculo de diferenças de hora
            if isinstance(df['Hora'].iloc[0], str):  # Se ainda for string, converter para datetime
                df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular a diferença de hora se ainda não existir
            if 'Diferença_Hora' not in df.columns or df['Diferença_Hora'].isna().any():
                df['Diferença_Hora'] = df['Hora'].diff().dt.total_seconds() / 3600
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if pd.isna(x) or x < 0 else x)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            else:
                # Certifica-se de que a coluna 'Diferença_Hora' esteja limpa e como número
                df['Diferença_Hora'] = pd.to_numeric(df['Diferença_Hora'].astype(str).str.strip(), errors='coerce')
                df['Diferença_Hora'] = df['Diferença_Hora'].fillna(0)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            
            # Conversão de colunas binárias para valores numéricos
            for col in ['Esteira Ligada', 'Motor Ligado', 'Field Cruiser', 'RTK (Piloto Automatico)', 'Implemento Ligado', 'Corte Base Automatico/Manual']:
                if col in df.columns:
                    if df[col].dtype == 'object':
                        df[col] = df[col].replace({'LIGADO': 1, 'DESLIGADO': 0, 'AUTOMATICO': 1, 'MANUAL': 0})
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
            # Tratar coluna Equipamento como texto e remover sufixo ".0"
            if 'Equipamento' in df.columns:
                df['Equipamento'] = df['Equipamento'].astype(str)
                df['Equipamento'] = df['Equipamento'].str.replace('.0$', '', regex=True)
            
            # Verificar e calcular "Parada com Motor Ligado" se necessário
            if 'Parada com Motor Ligado' not in df.columns:
                RPM_MINIMO = 300  # RPM mínimo considerado como motor ligado
                df['Parada com Motor Ligado'] = ((df['Velocidade'] == 0) & 
                                                (df['RPM Motor'] >= RPM_MINIMO) &
                                                (df['Motor Ligado'] == 1)).astype(int)
            
            # Verificar se Horas Produtivas já existe
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                # Para colhedoras, horas produtivas são registros onde Estado = 'TRABALHANDO' ou 'COLHEITA'
                condicao_produtiva = (df['Estado'] == 'TRABALHANDO') | (df['Estado'] == 'COLHEITA')
                df['Horas Produtivas'] = df['Diferença_Hora'] * condicao_produtiva.astype(int)
            else:
                # Limpa e converte para número
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # Aplicar substituição de operadores se a coluna existir
            if 'Operador' in df.columns:
                # Carregar configuração de substituição
                config_substituicao = carregar_substituicao_operadores()
                substituicoes = config_substituicao.get("substituicoes", {})
                
                # Garantir que a coluna de operador seja string e remover espaços extras
                df['Operador'] = df['Operador'].astype(str).str.strip()
                
                # Criar um dicionário onde as chaves são versões normalizadas dos operadores originais
                substituicoes_normalizadas = {
                    op_original.strip().upper(): op_novo
                    for op_original, op_novo in substituicoes.items()
                }
                
                # Criar uma coluna temporária com versões normalizadas dos operadores atuais
                df['_operador_normalizado'] = df['Operador'].str.strip().str.upper()
                
                # Aplicar as substituições usando a coluna normalizada
                for operador_original, operador_novo in substituicoes_normalizadas.items():
                    df.loc[df['_operador_normalizado'] == operador_original, 'Operador'] = operador_novo
                
                # Remover a coluna temporária
                df = df.drop(columns=['_operador_normalizado'])
                
                print(f"Substituição de operadores aplicada. Total de mapeamentos: {len(substituicoes)}")
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Carregar config para saber quais equipamentos excluir
    config = carregar_config_calculos()
    equipamentos_excluidos = config.get("CD", {}).get("equipamentos_excluidos", [])
    
    # Filtramos os dados excluindo os operadores da lista e os equipamentos excluídos
    df = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Aplicar filtro de equipamentos excluídos
    if equipamentos_excluidos:
        df = df[~df['Equipamento'].isin(equipamentos_excluidos)]
    
    # Extrair combinações únicas de Equipamento, Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round(numerador / denominador, precisao)
        return 0.0
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Horas totais
        horas_totais = round(dados_filtrados['Diferença_Hora'].sum(), 4)
        
        # Horas Produtivas - registros com Estado 'TRABALHANDO' ou 'COLHEITA'
        horas_produtivas = round(dados_filtrados[
            (dados_filtrados['Estado'] == 'TRABALHANDO') | 
            (dados_filtrados['Estado'] == 'COLHEITA')
        ]['Diferença_Hora'].sum(), 4)
        
        # Esteira Ligada com pressão de corte adequada > 400
        esteira_com_pressao = round(dados_filtrados[
            (dados_filtrados['Esteira Ligada'] == 1) & 
            (dados_filtrados['Pressao de Corte'] > 400)
        ]['Diferença_Hora'].sum(), 4)
        
        # Cálculo de GPS - Quando RTK está ligado, velocidade > 0 e Estado = TRABALHANDO ou COLHEITA
        gps = round(dados_filtrados[
            (dados_filtrados['RTK (Piloto Automatico)'] == 1) & 
            (dados_filtrados['Velocidade'] > 0) &
            ((dados_filtrados['Estado'] == 'TRABALHANDO') | (dados_filtrados['Estado'] == 'COLHEITA'))
        ]['Diferença_Hora'].sum(), 4)
        
        # % Utilização GPS
        utilizacao_gps = calcular_porcentagem(gps, horas_produtivas)
        
        # Verificar se o GPS está maior que as horas produtivas (ajuste necessário)
        if gps > horas_produtivas and horas_produtivas > 0:
            gps = horas_produtivas  # Corrigir para não ultrapassar as horas produtivas
            utilizacao_gps = 1.0  # 100% de utilização
        
        # Motor Ligado
        motor_ligado = round(dados_filtrados[dados_filtrados['Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # Parado com Motor Ligado
        parado_motor_ligado = round(dados_filtrados[dados_filtrados['Parada com Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # % Parado com motor ligado
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        # Tempo de Horas elevador
        horas_elevador = esteira_com_pressao
        
        # % Hora Elevador (em relação ao Motor Ligado)
        percent_hora_elevador = calcular_porcentagem(horas_elevador, motor_ligado)
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas Produtivas': horas_produtivas,
            'GPS': gps,
            '% Utilização GPS': utilizacao_gps,
            'Motor Ligado': motor_ligado,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor,
            'Horas elevador': horas_elevador,
            '% Hora Elevador': percent_hora_elevador
        })
    
    return pd.DataFrame(resultados)

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento
    """
    # Carregar config para saber quais equipamentos excluir
    config = carregar_config_calculos()
    equipamentos_excluidos = config.get("CD", {}).get("equipamentos_excluidos", [])
    
    # Filtramos os dados excluindo os operadores da lista e os equipamentos excluídos
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Aplicar filtro de equipamentos excluídos
    if equipamentos_excluidos:
        df_filtrado = df_filtrado[~df_filtrado['Equipamento'].isin(equipamentos_excluidos)]
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df_filtrado['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        # Garantir que o equipamento seja string e sem o sufixo ".0"
        equipamento_str = str(equipamento).replace('.0', '')
        
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 4)
        
        # Calcular horas de manutenção
        manutencao = round(dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum(), 4)
        
        # A disponibilidade mecânica é o percentual de tempo fora de manutenção
        disp_mecanica = calcular_porcentagem(total_horas - manutencao, total_horas)
        
        resultados.append({
            'Frota': equipamento_str,
            'Disponibilidade': disp_mecanica
        })
    
    return pd.DataFrame(resultados)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Esta função NÃO aplica qualquer filtro de operador (análise total da frota).
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota
    """
    # Carregar config para saber quais equipamentos excluir
    config = carregar_config_calculos()
    equipamentos_excluidos = config.get("CD", {}).get("equipamentos_excluidos", [])
    
    # Aplicar filtro de equipamentos excluídos
    if equipamentos_excluidos:
        df = df[~df['Equipamento'].isin(equipamentos_excluidos)]
    
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        # Garantir que o equipamento seja string e sem o sufixo ".0"
        equipamento_str = str(equipamento).replace('.0', '')
        
        dados_equip = df[df['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 2)
        
        # Calcular a diferença para 24 horas
        diferenca_24h = round(max(24 - total_horas, 0), 2)
        
        resultados.append({
            'Frota': equipamento_str,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        })
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica(base_calculo):
    """
    Calcula a eficiência energética por operador.
    Eficiência energética = Horas elevador / Horas motor ligado
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador
    """
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Eficiência Energética = horas elevador / motor ligado
        horas_elevador_sum = round(dados_op['Horas elevador'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        # Calcular eficiência - já está em decimal, não precisa multiplicar por 100
        eficiencia = calcular_porcentagem(horas_elevador_sum, motor_ligado_sum)
        
        # Garantir que não ultrapasse 100%
        eficiencia = min(eficiencia, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Eficiência': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_hora_elevador(df, base_calculo):
    """
    Calcula as horas de elevador por operador.
    Considera-se horas de elevador quando:
    - Esteira Ligada = 1
    - Pressão de Corte > 400
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Horas de elevador por operador
    """
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Somar horas de elevador da base de cálculo (já filtradas corretamente)
        horas_elevador_sum = round(dados_op['Horas elevador'].sum(), 2)
        
        resultados.append({
            'Operador': operador,
            'Horas': horas_elevador_sum
        })
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso(base_calculo, df_base=None):
    """
    Calcula o percentual de motor ocioso por operador.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
        df_base (DataFrame, optional): DataFrame base completo para filtrar operações excluídas
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador, tempo operação e tempo ocioso
    """
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Carregar configurações para exclusões
    config = carregar_config_calculos()
    tipo_equipamento = "CD"  # Colhedoras
    
    operacoes_excluidas = []
    tipo_calculo = "Remover do cálculo"
    
    if tipo_equipamento in config and "motor_ocioso" in config[tipo_equipamento]:
        motor_ocioso_config = config[tipo_equipamento]["motor_ocioso"]
        if "operacoes_excluidas" in motor_ocioso_config:
            operacoes_excluidas = motor_ocioso_config["operacoes_excluidas"]
        if "tipo_calculo" in motor_ocioso_config:
            tipo_calculo = motor_ocioso_config["tipo_calculo"]
    
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Motor Ocioso = Parado Com Motor Ligado / Motor Ligado
        parado_motor_sum = round(dados_op['Parado Com Motor Ligado'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        # Se temos o df_base e há operações a excluir, ajustar os tempos
        if df_base is not None and operacoes_excluidas and tipo_calculo == "Remover do cálculo":
            # Filtrar registros do operador/grupo no df_base
            filtro_base = (df_base['Operador'] == operador) & (df_base['Grupo Equipamento/Frente'] == grupo)
            dados_base = df_base[filtro_base]
            
            # Calcular tempo em operações excluídas
            tempo_op_excluidas = round(dados_base[dados_base['Operacao'].isin(operacoes_excluidas)]['Diferença_Hora'].sum(), 4)
            
            # Ajustar o tempo de motor ligado (remover o tempo em operações excluídas)
            motor_ligado_sum = max(0, motor_ligado_sum - tempo_op_excluidas)
        
        percentual = calcular_porcentagem(parado_motor_sum, motor_ligado_sum)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual,
            'Tempo Operação': motor_ligado_sum,
            'Tempo Ocioso': parado_motor_sum
        })
    
    # Criar DataFrame e organizar colunas na ordem desejada
    df_resultado = pd.DataFrame(resultados)
    if not df_resultado.empty:
        df_resultado = df_resultado[['Operador', 'Porcentagem', 'Tempo Operação', 'Tempo Ocioso']]
    
    return df_resultado

def calcular_uso_gps(df, base_calculo):
    """
    Calcula o percentual de uso de GPS por operador.
    O uso de GPS é definido como o tempo em que o equipamento está:
    - Estado é "TRABALHANDO" ou "COLHEITA" 
    - Piloto Automático (RTK) = 1
    - Velocidade > 0
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador
    """
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados base para este operador e grupo
        filtro_base = (df['Operador'] == operador) & \
                      (df['Grupo Equipamento/Frente'] == grupo)
        dados_op_base = df[filtro_base]
        
        # Calcular tempo total trabalhando
        tempo_trabalhando = round(dados_op_base[
            (dados_op_base['Estado'].isin(['TRABALHANDO', 'COLHEITA']))
        ]['Diferença_Hora'].sum(), 4)
        
        # Calcular tempo com GPS ativo (condições combinadas)
        tempo_gps_ativo = round(dados_op_base[
            (dados_op_base['Estado'].isin(['TRABALHANDO', 'COLHEITA'])) &
            (dados_op_base['RTK (Piloto Automatico)'] == 1) &
            (dados_op_base['Velocidade'] > 0)
        ]['Diferença_Hora'].sum(), 4)
        
        # Calcular percentual em formato decimal (0-1)
        percentual = calcular_porcentagem(tempo_gps_ativo, tempo_trabalhando)
        
        # Garantir que não ultrapasse 100% (1.0)
        percentual = min(percentual, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def calcular_media_velocidade(df):
    """
    Calcula a velocidade média por operador durante períodos produtivos.
    Considera apenas registros onde:
    - Estado é "TRABALHANDO" ou "COLHEITA"
    - Velocidade > 0
    
    Args:
        df (DataFrame): DataFrame base processado
    
    Returns:
        DataFrame: Média de velocidade por operador
    """
    # Carregar config para saber quais equipamentos excluir
    config = carregar_config_calculos()
    equipamentos_excluidos = config.get("CD", {}).get("equipamentos_excluidos", [])
    
    # Filtramos os dados excluindo os operadores da lista e os equipamentos excluídos
    df = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Aplicar filtro de equipamentos excluídos
    if equipamentos_excluidos:
        df = df[~df['Equipamento'].isin(equipamentos_excluidos)]
    
    # Filtrar registros produtivos com velocidade > 0
    df_filtrado = df[
        (df['Estado'].isin(['TRABALHANDO', 'COLHEITA'])) &
        (df['Velocidade'] > 0)
    ]
    
    # Agrupar por operador e grupo equipamento/frente
    resultados = []
    operadores = df_filtrado[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates().dropna()
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (df_filtrado['Operador'] == operador) & (df_filtrado['Grupo Equipamento/Frente'] == grupo)
        dados_op = df_filtrado[filtro]
        
        if len(dados_op) > 0:
            # Calcular média ponderada pela diferença de hora
            velocidade_total = (dados_op['Velocidade'] * dados_op['Diferença_Hora']).sum()
            horas_total = dados_op['Diferença_Hora'].sum()
            
            if horas_total > 0:
                velocidade_media = round(velocidade_total / horas_total, 2)
            else:
                velocidade_media = 0
            
            resultados.append({
                'Operador': operador,
                'Grupo Equipamento/Frente': grupo,
                'Velocidade Média (km/h)': velocidade_media
            })
    
    return pd.DataFrame(resultados)

def criar_excel_com_planilhas(df_base, base_calculo, disp_mecanica, eficiencia_energetica, 
                             hora_elevador, motor_ocioso, uso_gps, horas_por_frota, media_velocidade, caminho_saida):
    """
    Cria um arquivo Excel com todas as planilhas auxiliares.
    
    Args:
        df_base (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
        disp_mecanica (DataFrame): Disponibilidade mecânica
        eficiencia_energetica (DataFrame): Eficiência energética
        hora_elevador (DataFrame): Horas de elevador
        motor_ocioso (DataFrame): Motor ocioso
        uso_gps (DataFrame): Uso GPS
        horas_por_frota (DataFrame): Horas totais registradas por frota
        media_velocidade (DataFrame): Média de velocidade por operador
        caminho_saida (str): Caminho do arquivo Excel de saída
    """
    # Garantir que a coluna Equipamento esteja como string sem ".0"
    if 'Equipamento' in df_base.columns:
        df_base['Equipamento'] = df_base['Equipamento'].astype(str).str.replace('.0$', '', regex=True)
    
    if 'Equipamento' in base_calculo.columns:
        base_calculo['Equipamento'] = base_calculo['Equipamento'].astype(str).str.replace('.0$', '', regex=True)
    
    # Garantir que a coluna Frota esteja como string sem ".0"
    if 'Frota' in disp_mecanica.columns:
        disp_mecanica['Frota'] = disp_mecanica['Frota'].astype(str).str.replace('.0$', '', regex=True)
    
    if 'Frota' in horas_por_frota.columns:
        horas_por_frota['Frota'] = horas_por_frota['Frota'].astype(str).str.replace('.0$', '', regex=True)
    
    writer = pd.ExcelWriter(caminho_saida, engine='openpyxl')
    
    # Arredondamento fixo para 2 casas decimais em todas as colunas numéricas antes de exportar
    # Base Calculo - garantir que todas as colunas numéricas tenham 2 casas decimais
    colunas_numericas = ['Horas totais', 'Horas elevador', '%', 'RTK', 'Horas Produtivas', 
                         '% Utilização RTK', 'Motor Ligado', '% Eficiência Elevador', 
                         'Parado Com Motor Ligado', '% Parado com motor ligado']
    
    for col in colunas_numericas:
        if col in base_calculo.columns:
            base_calculo[col] = base_calculo[col].apply(lambda x: round(x, 2))
    
    # Arredondar valores nas outras planilhas
    disp_mecanica['Disponibilidade'] = disp_mecanica['Disponibilidade'].apply(lambda x: round(x, 4))
    eficiencia_energetica['Eficiência'] = eficiencia_energetica['Eficiência'].apply(lambda x: round(x, 4))
    hora_elevador['Horas'] = hora_elevador['Horas'].apply(lambda x: round(x, 2))
    motor_ocioso['Tempo Operação'] = motor_ocioso['Tempo Operação'].apply(lambda x: round(x, 2))
    motor_ocioso['Tempo Ocioso'] = motor_ocioso['Tempo Ocioso'].apply(lambda x: round(x, 2))
    motor_ocioso['Porcentagem'] = motor_ocioso['Porcentagem'].apply(lambda x: round(x, 4))
    uso_gps['Porcentagem'] = uso_gps['Porcentagem'].apply(lambda x: round(x, 4))
    horas_por_frota['Horas Registradas'] = horas_por_frota['Horas Registradas'].apply(lambda x: round(x, 2))
    horas_por_frota['Diferença para 24h'] = horas_por_frota['Diferença para 24h'].apply(lambda x: round(x, 2))
    if 'Velocidade Média (km/h)' in media_velocidade.columns:
        media_velocidade['Velocidade Média (km/h)'] = media_velocidade['Velocidade Média (km/h)'].apply(lambda x: round(x, 2))
    
    # Salvar cada DataFrame em uma planilha separada
    df_base.to_excel(writer, sheet_name='BASE', index=False)
    base_calculo.to_excel(writer, sheet_name='Base Calculo', index=False)
    
    # Planilhas auxiliares (formatadas conforme necessário)
    disp_mecanica.to_excel(writer, sheet_name='1_Disponibilidade Mecânica', index=False)
    eficiencia_energetica.to_excel(writer, sheet_name='2_Eficiência Energética', index=False)
    hora_elevador.to_excel(writer, sheet_name='3_Hora Elevador', index=False)
    motor_ocioso.to_excel(writer, sheet_name='4_Motor Ocioso', index=False)
    uso_gps.to_excel(writer, sheet_name='5_Uso GPS', index=False)
    media_velocidade.to_excel(writer, sheet_name='6_Média Velocidade', index=False)
    horas_por_frota.to_excel(writer, sheet_name='Horas por Frota', index=False)
    
    # Aplicar formatação nas planilhas
    workbook = writer.book
    
    # Formatar planilha de Disponibilidade Mecânica
    worksheet = workbook['1_Disponibilidade Mecânica']
    for row in range(2, worksheet.max_row + 1):  # Começando da linha 2 (ignorando cabeçalho)
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Disponibilidade)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Eficiência Energética
    worksheet = workbook['2_Eficiência Energética']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Eficiência)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Hora Elevador
    worksheet = workbook['3_Hora Elevador']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Horas)
        cell.number_format = '0.00'  # Formato decimal com 2 casas
    
    # Formatar planilha de Motor Ocioso
    worksheet = workbook['4_Motor Ocioso']
    for row in range(2, worksheet.max_row + 1):
        # Porcentagem (coluna B)
        cell_porc = worksheet.cell(row=row, column=2)
        cell_porc.number_format = '0.00%'  # Formato de porcentagem com 2 casas
        
        # Tempo Operação (coluna C)
        cell_tempo_op = worksheet.cell(row=row, column=3)
        cell_tempo_op.number_format = '0.00'  # Formato decimal com 2 casas
        
        # Tempo Ocioso (coluna D)
        cell_tempo_oc = worksheet.cell(row=row, column=4)
        cell_tempo_oc.number_format = '0.00'  # Formato decimal com 2 casas
    
    # Formatar planilha de Uso GPS
    worksheet = workbook['5_Uso GPS']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Média Velocidade
    worksheet = workbook['6_Média Velocidade']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=3)  # Coluna C (Velocidade Média)
        cell.number_format = '0.00'  # Formato decimal com 2 casas
    
    # Formatar planilha de Base Calculo
    worksheet = workbook['Base Calculo']
    for row in range(2, worksheet.max_row + 1):
        # Formatar colunas decimais
        columns_decimal = [4, 5, 7, 8, 10, 12]  # Colunas D, E, G, H, J, L (Horas totais, Horas elevador, etc.)
        for col in columns_decimal:
            if col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                cell.number_format = '0.00'  # Formato decimal com 2 casas
        
        # Formatar colunas de porcentagem
        columns_percent = [6, 9, 11, 13]  # Colunas F, I, K, M (%, % Utilização RTK, etc.)
        for col in columns_percent:
            if col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Horas por Frota
    worksheet = workbook['Horas por Frota']
    for row in range(2, worksheet.max_row + 1):
        # Coluna B (Horas Registradas)
        cell_b = worksheet.cell(row=row, column=2)
        cell_b.number_format = '0.00'  # Formato decimal com 2 casas
        
        # Coluna C (Diferença para 24h)
        cell_c = worksheet.cell(row=row, column=3)
        cell_c.number_format = '0.00'  # Formato decimal com 2 casas
    
    writer.close()
    print(f"Arquivo Excel salvo com sucesso em {caminho_saida}")

def extrair_arquivo_zip(caminho_zip, pasta_destino=None):
    """
    Extrai o conteúdo de um arquivo ZIP para uma pasta temporária ou destino especificado.
    Renomeia os arquivos extraídos para terem o mesmo nome do arquivo ZIP original.
    
    Args:
        caminho_zip (str): Caminho para o arquivo ZIP
        pasta_destino (str, optional): Pasta onde os arquivos serão extraídos.
                                       Se None, usa uma pasta temporária.
    
    Returns:
        list: Lista de caminhos dos arquivos extraídos e renomeados (apenas TXT e CSV)
        str: Caminho da pasta temporária (se criada) ou None
    """
    # Se pasta_destino não foi especificada, criar uma pasta temporária
    pasta_temp = None
    if pasta_destino is None:
        pasta_temp = tempfile.mkdtemp()
        pasta_destino = pasta_temp
    
    arquivos_extraidos = []
    nome_zip_sem_extensao = os.path.splitext(os.path.basename(caminho_zip))[0]
    
    try:
        with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
            # Extrair todos os arquivos do ZIP
            zip_ref.extractall(pasta_destino)
            
            # Processar cada arquivo extraído (apenas TXT e CSV)
            for arquivo in zip_ref.namelist():
                caminho_completo = os.path.join(pasta_destino, arquivo)
                # Verificar se é um arquivo e não uma pasta
                if os.path.isfile(caminho_completo):
                    # Verificar extensão
                    extensao = os.path.splitext(arquivo)[1].lower()
                    if extensao in ['.txt', '.csv']:
                        # Criar novo nome: nome do ZIP + extensão original
                        novo_nome = f"{nome_zip_sem_extensao}{extensao}"
                        novo_caminho = os.path.join(pasta_destino, novo_nome)
                        
                        # Renomear o arquivo extraído
                        try:
                            # Se já existe um arquivo com esse nome, remover primeiro
                            if os.path.exists(novo_caminho):
                                os.remove(novo_caminho)
                            # Renomear o arquivo
                            os.rename(caminho_completo, novo_caminho)
                            arquivos_extraidos.append(novo_caminho)
                            print(f"Arquivo extraído renomeado: {novo_nome}")
                        except Exception as e:
                            print(f"Erro ao renomear arquivo {arquivo} para {novo_nome}: {str(e)}")
                            arquivos_extraidos.append(caminho_completo)  # Adicionar o caminho original em caso de erro
        
        return arquivos_extraidos, pasta_temp
    
    except Exception as e:
        print(f"Erro ao extrair o arquivo ZIP {caminho_zip}: {str(e)}")
        # Se houve erro e criamos uma pasta temporária, tentar limpá-la
        if pasta_temp:
            try:
                shutil.rmtree(pasta_temp)
            except:
                pass
        return [], None

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT, CSV ou ZIP de colhedoras na pasta dados.
    Ignora arquivos que contenham "transbordo" no nome.
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    
    # Verificar se os diretórios existem, caso contrário criar
    if not os.path.exists(diretorio_dados):
        os.makedirs(diretorio_dados)
    if not os.path.exists(diretorio_saida):
        os.makedirs(diretorio_saida)
    
    # Encontrar todos os arquivos TXT/CSV/ZIP de colhedoras em ambos os diretórios
    arquivos = []
    arquivos_zip = []
    
    # Adicionar arquivos TXT sempre
    arquivos += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.txt"))
    arquivos += glob.glob(os.path.join(diretorio_dados, "*colhedora*.txt"))
    arquivos += glob.glob(os.path.join(diretorio_dados, "colhedora*.txt"))
    
    # Adicionar arquivos CSV apenas se processCsv for True
    if processCsv:
        arquivos += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.csv"))
        arquivos += glob.glob(os.path.join(diretorio_dados, "*colhedora*.csv"))
        arquivos += glob.glob(os.path.join(diretorio_dados, "colhedora*.csv"))
    
    # Adicionar arquivos ZIP
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.zip"))
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "*colhedora*.zip"))
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "colhedora*.zip"))
    
    # Filtrar arquivos que contenham "transbordo" no nome (case insensitive)
    arquivos = [arquivo for arquivo in arquivos if "transbordo" not in os.path.basename(arquivo).lower()]
    arquivos_zip = [arquivo for arquivo in arquivos_zip if "transbordo" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos = list(set(arquivos))
    arquivos_zip = list(set(arquivos_zip))
    
    if not arquivos and not arquivos_zip:
        print("Nenhum arquivo de colhedoras encontrado na pasta dados!")
        return
    
    print(f"Encontrados {len(arquivos)} arquivos de colhedoras (TXT/CSV) para processar.")
    print(f"Encontrados {len(arquivos_zip)} arquivos ZIP de colhedoras para processar.")
    
    # Processar cada arquivo TXT/CSV
    for arquivo in arquivos:
        processar_arquivo(arquivo, diretorio_saida)
    
    # Processar cada arquivo ZIP
    for arquivo_zip in arquivos_zip:
        print(f"\nProcessando arquivo ZIP: {os.path.basename(arquivo_zip)}")
        
        # Extrair arquivo ZIP para pasta temporária
        arquivos_extraidos, pasta_temp = extrair_arquivo_zip(arquivo_zip)
        
        if not arquivos_extraidos:
            print(f"Nenhum arquivo TXT ou CSV encontrado no ZIP {os.path.basename(arquivo_zip)}")
            continue
        
        print(f"Extraídos {len(arquivos_extraidos)} arquivos do ZIP.")
        
        # Processar cada arquivo extraído
        for arquivo_extraido in arquivos_extraidos:
            # Filtrar arquivos que contenham "transbordo" no nome
            if "transbordo" not in os.path.basename(arquivo_extraido).lower():
                processar_arquivo(arquivo_extraido, diretorio_saida)
        
        # Limpar pasta temporária se foi criada
        if pasta_temp:
            try:
                shutil.rmtree(pasta_temp)
                print(f"Pasta temporária removida: {pasta_temp}")
            except Exception as e:
                print(f"Erro ao remover pasta temporária {pasta_temp}: {str(e)}")

def processar_arquivo(caminho_arquivo, diretorio_saida):
    """
    Processa um único arquivo e gera o Excel de saída.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo a ser processado
        diretorio_saida (str): Diretório onde o arquivo de saída será salvo
    """
    # Obter apenas o nome do arquivo (sem caminho e sem extensão)
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    
    # Nome de saída igual ao original, mas com extensão .xlsx na pasta output
    arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}.xlsx")
    
    print(f"\nProcessando arquivo: {os.path.basename(caminho_arquivo)}")
    print(f"Arquivo de saída: {os.path.basename(arquivo_saida)}")
    
    # Processar o arquivo base
    df_base = processar_arquivo_base(caminho_arquivo)
    if df_base is None:
        print(f"Erro ao processar {os.path.basename(caminho_arquivo)}. Pulando para o próximo arquivo.")
        return
    
    # Se o DataFrame estiver vazio, gerar apenas a planilha BASE
    if len(df_base) == 0:
        writer = pd.ExcelWriter(arquivo_saida, engine='openpyxl')
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        writer.close()
        print(f"Arquivo {arquivo_saida} gerado com apenas a planilha BASE (sem dados).")
        return
    
    # Calcular a Base Calculo
    base_calculo = calcular_base_calculo(df_base)
    
    # Calcular as métricas auxiliares
    disp_mecanica = calcular_disponibilidade_mecanica(df_base)
    eficiencia_energetica = calcular_eficiencia_energetica(base_calculo)
    hora_elevador = calcular_hora_elevador(df_base, base_calculo)
    motor_ocioso = calcular_motor_ocioso(base_calculo, df_base)
    uso_gps = calcular_uso_gps(df_base, base_calculo)
    horas_por_frota = calcular_horas_por_frota(df_base)
    media_velocidade = calcular_media_velocidade(df_base)
    
    # Criar o arquivo Excel com todas as planilhas
    criar_excel_com_planilhas(
        df_base, base_calculo, disp_mecanica, eficiencia_energetica,
        hora_elevador, motor_ocioso, uso_gps, horas_por_frota, media_velocidade, arquivo_saida
    )
    
    print(f"Arquivo {arquivo_saida} gerado com sucesso!")

if __name__ == "__main__":
    print("="*80)
    print("Iniciando processamento de arquivos de colhedoras...")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Este script processa arquivos de colhedoras e gera planilhas Excel com métricas")
    print("Suporta arquivos TXT, CSV e ZIP")
    print("Ignorando arquivos que contenham 'transbordo' no nome")
    print("="*80)
    processar_todos_arquivos()
    print("\nProcessamento concluído!") 