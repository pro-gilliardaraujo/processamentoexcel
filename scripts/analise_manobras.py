"""
Script para análise de manobras de equipamentos agrícolas.
Processa dados de monitoramento para identificar e analisar padrões de manobra.
Suporta arquivos CSV na pasta dados/manobras.
Salva resultados na pasta output/manobras.
"""

import os
import sys
import json
import pandas as pd
import numpy as np
import zipfile
import tempfile
import shutil
from datetime import datetime, timedelta
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import dash
from dash import dcc, html, Output, Input
import dash_bootstrap_components as dbc
import glob
import plotly.express as px
import openpyxl

# Configurações
CONFIG_FILE = 'config/config_calculos.json'
INPUT_DIR = 'dados'
OUTPUT_DIR = 'output'

# Garante que os diretórios existem
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

def processar_arquivo_base(caminho_arquivo):
    """
    Processa um arquivo base (TXT, CSV ou ZIP) e retorna um DataFrame.
    """
    print(f"\nProcessando arquivo: {caminho_arquivo}")
    
    # Verifica se o arquivo existe
    if not os.path.exists(caminho_arquivo):
        print(f"Arquivo não encontrado: {caminho_arquivo}")
        return None
        
    # Processa arquivo ZIP
    if caminho_arquivo.lower().endswith('.zip'):
        print("Arquivo ZIP detectado")
        return processar_arquivo_zip(caminho_arquivo)
        
    # Processa arquivo TXT ou CSV
    if caminho_arquivo.lower().endswith(('.txt', '.csv')):
        print("Arquivo TXT/CSV detectado")
        return ler_arquivo_com_encoding(caminho_arquivo)
        
    print(f"Formato de arquivo não suportado: {caminho_arquivo}")
    return None

def processar_arquivo_zip(caminho_arquivo):
    """
    Processa um arquivo ZIP contendo arquivos TXT ou CSV.
    Extrai o conteúdo para um diretório temporário e processa cada arquivo.
    """
    try:
        print(f"Processando arquivo ZIP: {caminho_arquivo}")
        temp_dir = tempfile.mkdtemp()
        
        try:
            # Extrai arquivos para o diretório temporário
            with zipfile.ZipFile(caminho_arquivo, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
                
            # Lista todos os arquivos extraídos
            arquivos_extraidos = []
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    if file.lower().endswith(('.txt', '.csv')):
                        arquivos_extraidos.append(os.path.join(root, file))
            
            print(f"Extraídos {len(arquivos_extraidos)} arquivos")
            
            if not arquivos_extraidos:
                print("Nenhum arquivo TXT ou CSV encontrado no ZIP")
                return None
                
            # Processa todos os arquivos encontrados
            dfs = []
            for arquivo in arquivos_extraidos:
                print(f"Processando arquivo extraído: {os.path.basename(arquivo)}")
                df = ler_arquivo_com_encoding(arquivo)
                if df is not None:
                    dfs.append(df)
            
            if not dfs:
                print("Nenhum arquivo foi processado com sucesso")
                return None
                
            # Concatena todos os DataFrames
            df_final = pd.concat(dfs, ignore_index=True)
            print(f"Total de registros após concatenação: {len(df_final)}")
            return df_final
            
        finally:
            # Limpa o diretório temporário
            shutil.rmtree(temp_dir)
            
    except Exception as e:
        print(f"Erro ao processar arquivo ZIP: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def ler_arquivo_com_encoding(caminho_arquivo, encodings=['utf-8', 'latin1', 'ISO-8859-1']):
    """
    Tenta ler um arquivo com diferentes encodings.
    """
    for encoding in encodings:
        try:
            print(f"Tentando ler com encoding: {encoding}")
            
            # Detecta o separador (vírgula ou ponto e vírgula)
            with open(caminho_arquivo, 'r', encoding=encoding) as f:
                primeira_linha = f.readline().strip()
                
            if ';' in primeira_linha:
                separador = ';'
            else:
                separador = ','
                
            print(f"Separador detectado: '{separador}'")
            
            # Lê o arquivo com pandas
            df = pd.read_csv(caminho_arquivo, encoding=encoding, sep=separador, low_memory=False)
            
            # Verifica se o DataFrame foi lido corretamente
            if df.empty:
                print(f"Arquivo vazio ou inválido: {caminho_arquivo}")
                continue
            
            # Filtra apenas as colunas necessárias para o processamento
            colunas_necessarias = [
                'Equipamento', 'Data/Hora', 'Grupo', 'Motor Ligado', 
                'Operacao', 'Operador', 'RPM Motor', 'Tipo de Equipamento', 'Velocidade'
            ]
            
            # Verifica se as colunas necessárias existem
            colunas_faltantes = [col for col in colunas_necessarias if col not in df.columns]
            if colunas_faltantes:
                print(f"AVISO: Colunas faltantes: {colunas_faltantes}")
                
            # Mantém apenas as colunas necessárias (se existirem)
            colunas_para_manter = [col for col in colunas_necessarias if col in df.columns]
            df = df[colunas_para_manter]
                
            # Renomeia colunas para padrão interno
            if 'Data/Hora' in df.columns:
                df['Data_Hora'] = pd.to_datetime(df['Data/Hora'], errors='coerce', dayfirst=True)
            elif 'Tempo' in df.columns:
                df['Data_Hora'] = pd.to_datetime(df['Tempo'], errors='coerce', dayfirst=True)
                
            # Converte coluna de data/hora para datetime
            if 'Data_Hora' in df.columns:
                df['Data_Hora'] = pd.to_datetime(df['Data_Hora'], errors='coerce')
                
                # Ordena o DataFrame por equipamento e data/hora
                df = df.sort_values(['Equipamento', 'Data_Hora'])
                
                # Extrai informações adicionais da data/hora
                df['Data'] = df['Data_Hora'].dt.date
                df['Hora'] = df['Data_Hora'].dt.hour
                
                # Calcula a diferença de tempo entre registros consecutivos em horas
                df['Diferença_Hora'] = df.groupby('Equipamento')['Data_Hora'].diff().dt.total_seconds() / 3600
                
                # Preenche valores NaN com um valor razoável (15 minutos = 0.25 horas)
                # Isso ocorre no primeiro registro de cada equipamento
                df['Diferença_Hora'] = df['Diferença_Hora'].fillna(0.25)
                
                # Limita valores muito grandes (gaps maiores que 1 hora são provavelmente pausas)
                df.loc[df['Diferença_Hora'] > 1, 'Diferença_Hora'] = 1.0
                
                print(f"Coluna 'Diferença_Hora' criada com sucesso.")
                
            # Converte 'Motor Ligado' para valores numéricos
            if 'Motor Ligado' in df.columns:
                df['Motor_Ligado_Num'] = df['Motor Ligado'].apply(lambda x: 1 if str(x).lower() == 'sim' else 0)
                print(f"Coluna 'Motor_Ligado_Num' criada com sucesso.")
            
            # Converte 'RPM Motor' para numérico
            if 'RPM Motor' in df.columns:
                df['RPM_Motor'] = pd.to_numeric(df['RPM Motor'], errors='coerce').fillna(0)
                
            # Converte 'Velocidade' para numérico
            if 'Velocidade' in df.columns:
                df['Velocidade'] = pd.to_numeric(df['Velocidade'], errors='coerce').fillna(0)
                
            # Renomeia 'Grupo' para 'Frente' para clareza
            if 'Grupo' in df.columns:
                df.rename(columns={'Grupo': 'Frente'}, inplace=True)
                
            print(f"Arquivo lido com sucesso usando encoding '{encoding}'")
            print(f"Total de registros: {len(df)}")
            print(f"Colunas: {df.columns.tolist()}")
            
            return df
            
        except Exception as e:
            print(f"Erro com encoding {encoding}: {str(e)}")
            continue
            
    print(f"Não foi possível ler o arquivo com nenhum dos encodings tentados: {caminho_arquivo}")
    return None

def identificar_manobras(df):
    """
    Identifica manobras no DataFrame baseado na velocidade e direção.
    """
    try:
        print("\nIniciando identificação de manobras...")
        
        # Verifica se as colunas necessárias existem
        colunas_necessarias = ['Data/Hora', 'Equipamento', 'Velocidade']
        for coluna in colunas_necessarias:
            if coluna not in df.columns:
                print(f"Erro: Coluna '{coluna}' não encontrada no DataFrame")
                return None
        
        # Cria cópia do DataFrame para não modificar o original
        df_manobras = df.copy()
        
        # Adiciona coluna de direção se não existir
        if 'Direção' not in df_manobras.columns:
            print("Adicionando coluna de direção...")
            df_manobras['Direção'] = df_manobras['Velocidade'].apply(lambda x: 1 if x > 0 else -1 if x < 0 else 0)
        
        # Identifica mudanças de direção
        print("Identificando mudanças de direção...")
        df_manobras['Mudança_Direção'] = df_manobras['Direção'].diff().fillna(0) != 0
        
        # Agrupa por equipamento
        print("Agrupando por equipamento...")
        manobras_por_equipamento = []
        
        for equipamento, grupo in df_manobras.groupby('Equipamento'):
            print(f"\nProcessando equipamento: {equipamento}")
            
            # Ordena por data/hora
            grupo = grupo.sort_values('Data/Hora')
            
            # Identifica início e fim das manobras
            inicio_manobras = grupo[grupo['Mudança_Direção']].index
            if len(inicio_manobras) == 0:
                print(f"Nenhuma manobra encontrada para o equipamento {equipamento}")
                continue
                
            print(f"Encontradas {len(inicio_manobras)} manobras para o equipamento {equipamento}")
            
            # Processa cada manobra
            for i in range(len(inicio_manobras)):
                inicio_idx = inicio_manobras[i]
                fim_idx = inicio_manobras[i + 1] if i < len(inicio_manobras) - 1 else grupo.index[-1]
                
                manobra = grupo.loc[inicio_idx:fim_idx].copy()
                
                # Calcula duração da manobra
                duracao = (manobra['Data/Hora'].iloc[-1] - manobra['Data/Hora'].iloc[0]).total_seconds() / 3600
                
                # Adiciona informações da manobra
                manobra['Duração_Horas'] = duracao
                manobra['Tipo_Manobra'] = 'Reversão' if manobra['Direção'].iloc[0] != manobra['Direção'].iloc[-1] else 'Mudança'
                
                manobras_por_equipamento.append(manobra)
        
        if not manobras_por_equipamento:
            print("Nenhuma manobra foi identificada")
            return None
            
        # Concatena todas as manobras
        df_manobras_final = pd.concat(manobras_por_equipamento)
        print(f"\nTotal de manobras identificadas: {len(df_manobras_final)}")
        
        return df_manobras_final
        
    except Exception as e:
        print(f"Erro ao identificar manobras: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def calcular_horas_por_frota(df, df_manobras):
    """
    Calcula as horas totais, horas em manobra e horas sem manobra por frota.
    """
    try:
        # Cria uma cópia do DataFrame para não modificar o original
        df_horas = df.copy()
        
        # Calcula horas totais por frota (Equipamento)
        horas_totais = df_horas.groupby('Equipamento')['Diferença_Hora'].sum() / 24
        
        # Calcula horas em manobra por frota (Equipamento)
        horas_manobra = df_manobras.groupby('Frota')['Tempo_Manobra'].sum()
        
        # Cria DataFrame com as métricas
        df_horas_por_frota = pd.DataFrame({
            'Horas_Totais': horas_totais,
            'Horas_Manobra': horas_manobra
        }).fillna(0)
        
        # Calcula horas sem manobra
        df_horas_por_frota['Horas_Sem_Manobra'] = (
            df_horas_por_frota['Horas_Totais'] - df_horas_por_frota['Horas_Manobra']
        )
        
        # Calcula porcentagem de manobra
        df_horas_por_frota['Porcentagem_Manobra'] = (
            df_horas_por_frota['Horas_Manobra'] / df_horas_por_frota['Horas_Totais']
        ).fillna(0)
        
        # Arredonda valores
        df_horas_por_frota = df_horas_por_frota.round(4)
        
        return df_horas_por_frota
        
    except Exception as e:
        print(f"Erro ao calcular horas por frota: {str(e)}")
        return None

def criar_dashboard(df_manobras, metricas):
    """
    Cria uma dashboard interativa com Plotly Dash para visualização das métricas de manobra.
    
    Args:
        df_manobras (DataFrame): DataFrame com as manobras identificadas
        metricas (dict): Dicionário com as métricas calculadas
    """
    app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])
    
    # Extrair opções únicas para os filtros
    frentes = sorted(df_manobras['Frente'].unique()) if 'Frente' in df_manobras.columns else []
    equipamentos = sorted(df_manobras['Equipamento'].unique())
    tipos_equipamento = sorted(df_manobras['Tipo_Equipamento'].unique()) if 'Tipo_Equipamento' in df_manobras.columns else []
    operadores = sorted(df_manobras['Operador'].unique()) if 'Operador' in df_manobras.columns else []
    
    # Layout da dashboard
    app.layout = dbc.Container([
        dbc.Row([
            dbc.Col(html.H1("Dashboard de Análise de Manobras", className="text-center mb-4"), width=12)
        ]),
        
        # Filtros
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Filtros", className="card-title"),
                        dbc.Row([
                            dbc.Col([
                                html.Label("Frente:"),
                                dcc.Dropdown(
                                    id='filtro-frente',
                                    options=[{'label': f, 'value': f} for f in frentes],
                                    value=None,
                                    clearable=True,
                                    placeholder="Selecione a Frente"
                                )
                            ], width=3) if frentes else None,
                            dbc.Col([
                                html.Label("Equipamento:"),
                                dcc.Dropdown(
                                    id='filtro-equipamento',
                                    options=[{'label': e, 'value': e} for e in equipamentos],
                                    value=None,
                                    clearable=True,
                                    placeholder="Selecione o Equipamento"
                                )
                            ], width=3),
                            dbc.Col([
                                html.Label("Tipo de Equipamento:"),
                                dcc.Dropdown(
                                    id='filtro-tipo-equipamento',
                                    options=[{'label': t, 'value': t} for t in tipos_equipamento],
                                    value=None,
                                    clearable=True,
                                    placeholder="Selecione o Tipo"
                                )
                            ], width=3) if tipos_equipamento else None,
                            dbc.Col([
                                html.Label("Operador:"),
                                dcc.Dropdown(
                                    id='filtro-operador',
                                    options=[{'label': o, 'value': o} for o in operadores],
                                    value=None,
                                    clearable=True,
                                    placeholder="Selecione o Operador"
                                )
                            ], width=3) if operadores else None
                        ])
                    ])
                ])
            ], width=12)
        ], className="mb-4"),
        
        # Cards com métricas principais
        dbc.Row([
            dbc.Col(
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Total de Manobras", className="card-title"),
                        html.H2(id="total-manobras", className="card-text text-center")
                    ])
                ]), width=3
            ),
            dbc.Col(
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Tempo Médio (min)", className="card-title"),
                        html.H2(id="tempo-medio", className="card-text text-center")
                    ])
                ]), width=3
            ),
            dbc.Col(
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Tempo Total (h)", className="card-title"),
                        html.H2(id="tempo-total", className="card-text text-center")
                    ])
                ]), width=3
            ),
            dbc.Col(
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Manobras por Hora", className="card-title"),
                        html.H2(id="manobras-por-hora", className="card-text text-center")
                    ])
                ]), width=3
            ),
        ], className="mb-4"),
        
        # Gráficos - Primeira linha
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Manobras por Hora do Dia", className="card-title"),
                        dcc.Graph(id='grafico-manobras-hora')
                    ])
                ])
            ], width=6),
            
            dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Evolução das Manobras", className="card-title"),
                        dcc.Graph(id='grafico-evolucao-manobras')
                    ])
                ])
            ], width=6)
        ], className="mb-4"),
        
        # Gráficos - Segunda linha
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Distribuição por Tipo de Equipamento", className="card-title"),
                        dcc.Graph(id='grafico-tipo-equipamento')
                    ])
                ])
            ], width=6) if tipos_equipamento else None,
            
            dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Distribuição por Equipamento", className="card-title"),
                        dcc.Graph(id='grafico-equipamento')
                    ])
                ])
            ], width=6)
        ], className="mb-4"),
        
        # Tabela de detalhes por operador
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Detalhes por Operador", className="card-title"),
                        html.Div(id='tabela-operadores')
                    ])
                ])
            ], width=12)
        ], className="mb-4") if operadores else None
    ], fluid=True)
    
    # Callbacks para atualização dos gráficos e tabelas
    @app.callback(
        [Output('total-manobras', 'children'),
         Output('tempo-medio', 'children'),
         Output('tempo-total', 'children'),
         Output('manobras-por-hora', 'children'),
         Output('grafico-manobras-hora', 'figure'),
         Output('grafico-evolucao-manobras', 'figure'),
         Output('grafico-tipo-equipamento', 'figure'),
         Output('grafico-equipamento', 'figure'),
         Output('tabela-operadores', 'children')],
        [Input('filtro-frente', 'value'),
         Input('filtro-equipamento', 'value'),
         Input('filtro-tipo-equipamento', 'value'),
         Input('filtro-operador', 'value')]
    )
    def atualizar_dashboard(frente, equipamento, tipo_equipamento, operador):
        # Aplicar filtros
        df_filtrado = df_manobras.copy()
        if frente and 'Frente' in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado['Frente'] == frente]
        if equipamento:
            df_filtrado = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        if tipo_equipamento and 'Tipo_Equipamento' in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado['Tipo_Equipamento'] == tipo_equipamento]
        if operador and 'Operador' in df_filtrado.columns:
            df_filtrado = df_filtrado[df_filtrado['Operador'] == operador]
        
        # Calcular métricas atualizadas
        total_manobras = len(df_filtrado)
        tempo_medio = df_filtrado['Duração (h)'].mean() * 60 if not df_filtrado.empty else 0
        tempo_total = df_filtrado['Duração (h)'].sum() if not df_filtrado.empty else 0
        manobras_por_hora = total_manobras / tempo_total if tempo_total > 0 else 0
        
        # Gráfico de manobras por hora
        manobras_hora = df_filtrado.groupby('Hora_Inicio').size()
        fig_manobras_hora = go.Figure(data=[
            go.Scatter(
                x=manobras_hora.index,
                y=manobras_hora.values,
                mode='lines+markers',
                line=dict(shape='spline', smoothing=0.3)
            )
        ])
        fig_manobras_hora.update_layout(
            title="Manobras por Hora do Dia",
            xaxis_title="Hora",
            yaxis_title="Quantidade de Manobras",
            margin=dict(l=40, r=40, t=40, b=40)
        )
        
        # Gráfico de evolução
        df_filtrado['Data'] = pd.to_datetime(df_filtrado['Inicio'])
        evolucao = df_filtrado.groupby('Data').size()
        fig_evolucao = go.Figure(data=[
            go.Scatter(
                x=evolucao.index,
                y=evolucao.values,
                mode='lines+markers',
                line=dict(shape='spline', smoothing=0.3)
            )
        ])
        fig_evolucao.update_layout(
            title="Evolução das Manobras",
            xaxis_title="Data",
            yaxis_title="Quantidade de Manobras",
            margin=dict(l=40, r=40, t=40, b=40)
        )
        
        # Gráfico por tipo de equipamento
        if 'Tipo_Equipamento' in df_filtrado.columns:
        tipo_equip = df_filtrado.groupby('Tipo_Equipamento').size()
        fig_tipo_equip = go.Figure(data=[
            go.Bar(
                x=tipo_equip.index,
                y=tipo_equip.values,
                text=tipo_equip.values,
                textposition='auto'
            )
        ])
        fig_tipo_equip.update_layout(
            title="Distribuição por Tipo de Equipamento",
            xaxis_title="Tipo de Equipamento",
            yaxis_title="Quantidade de Manobras",
            margin=dict(l=40, r=40, t=40, b=40)
        )
        else:
            fig_tipo_equip = go.Figure()
        
        # Gráfico por equipamento
        equip_dist = df_filtrado.groupby('Equipamento').size()
        fig_equip = go.Figure(data=[
            go.Bar(
                x=equip_dist.index,
                y=equip_dist.values,
                text=equip_dist.values,
                textposition='auto'
            )
        ])
        fig_equip.update_layout(
            title="Distribuição por Equipamento",
            xaxis_title="Equipamento",
            yaxis_title="Quantidade de Manobras",
            margin=dict(l=40, r=40, t=40, b=40)
        )
        
        # Tabela de operadores
        if 'Operador' in df_filtrado.columns:
            df_operadores = df_filtrado.groupby(['Equipamento', 'Operador']).agg({
            'Equipamento': 'count',
                'Duração (h)': 'sum'
        }).round(4)
        
        df_operadores.columns = ['Quantidade', 'Tempo_Total (h)']
        df_operadores = df_operadores.reset_index()
        
        tabela = dbc.Table.from_dataframe(
            df_operadores,
            striped=True,
            bordered=True,
            hover=True,
            responsive=True
        )
        else:
            tabela = html.Div("Dados de operadores não disponíveis")
        
        return (
            f"{total_manobras:,}",
            f"{tempo_medio:.1f}",
            f"{tempo_total:.1f}",
            f"{manobras_por_hora:.1f}",
            fig_manobras_hora,
            fig_evolucao,
            fig_tipo_equip,
            fig_equip,
            tabela
        )
    
    return app

def criar_excel_com_metricas_manobras(df_base, df_manobras, df_metricas_gerais, metricas_agregadas, caminho_saida):
    """
    Cria arquivo Excel com métricas de manobras.
    Gera planilha BASE + planilhas auxiliares: Por Equipamento, Por Operador e Turno.
    """
    try:
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            
            # 0. Planilha BASE - Dados brutos processados do arquivo original
            df_base.to_excel(writer, sheet_name='BASE', index=False)
            
            # 1. Planilha auxiliar: Tempo de MANOBRA por Equipamento
            manobras_por_equipamento = df_manobras.groupby('Equipamento').agg({
                'Duração (h)': ['count', 'sum'],
                'RPM_Medio': 'mean',
                'Velocidade_Media': 'mean'
            }).round(4)
            
            # Renomear colunas para clareza
            manobras_por_equipamento.columns = [
                'Total de Manobras',
                'Tempo Total MANOBRA (h)',
                'RPM Médio',
                'Velocidade Média (km/h)'
            ]
            
            # Resetar índice para incluir Equipamento como coluna
            manobras_por_equipamento = manobras_por_equipamento.reset_index()
            
            # Calcular tempo médio de manobra
            manobras_por_equipamento['Tempo Médio MANOBRA (h)'] = (
                manobras_por_equipamento['Tempo Total MANOBRA (h)'] / manobras_por_equipamento['Total de Manobras']
            ).round(4)
            
            # Converter tempos para formato de hora (decimal / 24)
            manobras_por_equipamento['Tempo Total MANOBRA (formato)'] = manobras_por_equipamento['Tempo Total MANOBRA (h)'] / 24
            manobras_por_equipamento['Tempo Médio MANOBRA (formato)'] = manobras_por_equipamento['Tempo Médio MANOBRA (h)'] / 24
            
            # Reordenar colunas
            manobras_por_equipamento = manobras_por_equipamento[[
                'Equipamento', 'Total de Manobras', 'Tempo Total MANOBRA (h)', 'Tempo Total MANOBRA (formato)',
                'Tempo Médio MANOBRA (h)', 'Tempo Médio MANOBRA (formato)', 'RPM Médio', 'Velocidade Média (km/h)'
            ]]
            
            # Adicionar linha de totais
            total_row = {
                'Equipamento': 'TOTAL GERAL',
                'Total de Manobras': manobras_por_equipamento['Total de Manobras'].sum(),
                'Tempo Total MANOBRA (h)': manobras_por_equipamento['Tempo Total MANOBRA (h)'].sum(),
                'Tempo Total MANOBRA (formato)': manobras_por_equipamento['Tempo Total MANOBRA (h)'].sum() / 24,
                'Tempo Médio MANOBRA (h)': (manobras_por_equipamento['Tempo Total MANOBRA (h)'].sum() / manobras_por_equipamento['Total de Manobras'].sum()).round(4) if manobras_por_equipamento['Total de Manobras'].sum() > 0 else 0,
                'Tempo Médio MANOBRA (formato)': ((manobras_por_equipamento['Tempo Total MANOBRA (h)'].sum() / manobras_por_equipamento['Total de Manobras'].sum()) / 24) if manobras_por_equipamento['Total de Manobras'].sum() > 0 else 0,
                'RPM Médio': manobras_por_equipamento['RPM Médio'].mean(),
                'Velocidade Média (km/h)': manobras_por_equipamento['Velocidade Média (km/h)'].mean()
            }
            manobras_por_equipamento = pd.concat([manobras_por_equipamento, pd.DataFrame([total_row])], ignore_index=True)
            
            # Salvar planilha
            manobras_por_equipamento.to_excel(writer, sheet_name='Por Equipamento', index=False)
            
            # 2. Planilha auxiliar: Tempo de MANOBRA por Operador
            manobras_por_operador = df_manobras.groupby('Operador').agg({
                'Duração (h)': ['count', 'sum'],
                'RPM_Medio': 'mean',
                'Velocidade_Media': 'mean'
            }).round(4)
            
            # Renomear colunas para clareza
            manobras_por_operador.columns = [
                'Total de Manobras',
                'Tempo Total MANOBRA (h)',
                'RPM Médio',
                'Velocidade Média (km/h)'
            ]
            
            # Resetar índice para incluir Operador como coluna
            manobras_por_operador = manobras_por_operador.reset_index()
            
            # Calcular tempo médio de manobra
            manobras_por_operador['Tempo Médio MANOBRA (h)'] = (
                manobras_por_operador['Tempo Total MANOBRA (h)'] / manobras_por_operador['Total de Manobras']
            ).round(4)
            
            # Converter tempos para formato de hora (decimal / 24)
            manobras_por_operador['Tempo Total MANOBRA (formato)'] = manobras_por_operador['Tempo Total MANOBRA (h)'] / 24
            manobras_por_operador['Tempo Médio MANOBRA (formato)'] = manobras_por_operador['Tempo Médio MANOBRA (h)'] / 24
            
            # Reordenar colunas
            manobras_por_operador = manobras_por_operador[[
                'Operador', 'Total de Manobras', 'Tempo Total MANOBRA (h)', 'Tempo Total MANOBRA (formato)',
                'Tempo Médio MANOBRA (h)', 'Tempo Médio MANOBRA (formato)', 'RPM Médio', 'Velocidade Média (km/h)'
            ]]
            
            # Adicionar linha de totais
            total_row_op = {
                'Operador': 'TOTAL GERAL',
                'Total de Manobras': manobras_por_operador['Total de Manobras'].sum(),
                'Tempo Total MANOBRA (h)': manobras_por_operador['Tempo Total MANOBRA (h)'].sum(),
                'Tempo Total MANOBRA (formato)': manobras_por_operador['Tempo Total MANOBRA (h)'].sum() / 24,
                'Tempo Médio MANOBRA (h)': (manobras_por_operador['Tempo Total MANOBRA (h)'].sum() / manobras_por_operador['Total de Manobras'].sum()).round(4) if manobras_por_operador['Total de Manobras'].sum() > 0 else 0,
                'Tempo Médio MANOBRA (formato)': ((manobras_por_operador['Tempo Total MANOBRA (h)'].sum() / manobras_por_operador['Total de Manobras'].sum()) / 24) if manobras_por_operador['Total de Manobras'].sum() > 0 else 0,
                'RPM Médio': manobras_por_operador['RPM Médio'].mean(),
                'Velocidade Média (km/h)': manobras_por_operador['Velocidade Média (km/h)'].mean()
            }
            manobras_por_operador = pd.concat([manobras_por_operador, pd.DataFrame([total_row_op])], ignore_index=True)
            
            # Salvar planilha
            manobras_por_operador.to_excel(writer, sheet_name='Por Operador', index=False)
            
            # 3. Planilha Turno - Análise de trocas de operadores por turno
            df_turno = analisar_turnos(df_base)
            if df_turno is not None and not df_turno.empty:
                df_turno.to_excel(writer, sheet_name='Turno', index=False)
            
            # 4. Ajustar aparência das planilhas
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                
                # Ajustar largura das colunas automaticamente
                    for idx, col in enumerate(worksheet.columns, 1):
                        max_length = 0
                        column = col[0].column_letter
                        
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        
                    adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column].width = adjusted_width
                    
                # Congelar primeira linha (cabeçalhos)
                    worksheet.freeze_panes = worksheet.cell(row=2, column=1)
                    
                # Formatar cabeçalhos em negrito
                    for cell in worksheet[1]:
                    cell.font = openpyxl.styles.Font(bold=True)
                
                # Formatar colunas de tempo como formato de hora
                if sheet_name in ['Por Equipamento', 'Por Operador']:
                    # Encontrar colunas com "(formato)" no nome
                    for col_idx, col_name in enumerate(worksheet[1]):
                        if col_name.value and '(formato)' in str(col_name.value):
                            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                            # Aplicar formato de hora para toda a coluna
                            for row in range(2, worksheet.max_row + 1):
                                cell = worksheet[f'{col_letter}{row}']
                                cell.number_format = '[h]:mm:ss'
                
                # Destacar linha de totais em negrito (apenas para planilhas auxiliares)
                if sheet_name in ['Por Equipamento', 'Por Operador']:
                    last_row = worksheet.max_row
                    for cell in worksheet[last_row]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        print(f"Arquivo Excel criado com sucesso: {caminho_saida}")
        print(f"Planilhas criadas:")
        print(f"- BASE: {len(df_base)} registros originais")
        print(f"- Por Equipamento: {len(manobras_por_equipamento)-1} equipamentos")
        print(f"- Por Operador: {len(manobras_por_operador)-1} operadores")
        if df_turno is not None and not df_turno.empty:
            print(f"- Turno: {len(df_turno)} trocas de operadores identificadas")
        
        return True
        
    except PermissionError as e:
        print(f"ERRO: Não foi possível salvar o arquivo. Verifique se ele não está aberto: {str(e)}")
        return False
    except Exception as e:
        print(f"ERRO ao criar arquivo Excel: {str(e)}")
        import traceback
        traceback.print_exc()
            return False

def analisar_turnos(df_base):
    """
    Analisa trocas de operadores na planilha BASE para identificar turnos.
    Turno A: próximo às 7h (6:40 - 7:20)
    Turno B: próximo às 15h (14:40 - 15:20) 
    Turno C: próximo às 23h (22:40 - 23:20)
    """
    try:
        if 'Data_Hora' not in df_base.columns or 'Operador' not in df_base.columns or 'Equipamento' not in df_base.columns:
            print("AVISO: Colunas necessárias para análise de turnos não encontradas")
            return None
        
        # Criar cópia e ordenar por equipamento e data/hora
        df = df_base.copy()
        df = df.sort_values(['Equipamento', 'Data_Hora'])
        
        # Detectar trocas de operador por equipamento
        trocas_operador = []
        
        for equipamento, grupo in df.groupby('Equipamento'):
            grupo = grupo.reset_index(drop=True)
            
            # Identificar onde houve troca de operador
            for i in range(1, len(grupo)):
                operador_anterior = grupo.loc[i-1, 'Operador']
                operador_atual = grupo.loc[i, 'Operador']
                
                if operador_anterior != operador_atual:
                    data_hora = grupo.loc[i, 'Data_Hora']
                    hora = data_hora.hour
                    minuto = data_hora.minute
                    hora_decimal = hora + minuto/60
                    
                    # Classificar turno baseado no horário
                    turno = classificar_turno(hora_decimal)
                    
                    trocas_operador.append({
                        'Equipamento': equipamento,
                        'Data': data_hora.date(),
                        'Hora_Troca': data_hora.strftime('%H:%M'),
                        'Operador_Anterior': operador_anterior,
                        'Operador_Novo': operador_atual,
                        'Turno': turno,
                        'Hora_Decimal': round(hora_decimal, 2)
                    })
        
        if not trocas_operador:
            print("Nenhuma troca de operador detectada")
            return None
        
        df_turnos = pd.DataFrame(trocas_operador)
        
        # Agrupar por turno para análise
        resumo_turnos = df_turnos.groupby('Turno').agg({
            'Equipamento': 'nunique',
            'Operador_Novo': 'nunique',
            'Data': 'nunique'
        }).rename(columns={
            'Equipamento': 'Equipamentos_Afetados',
            'Operador_Novo': 'Operadores_Únicos',
            'Data': 'Dias_Com_Trocas'
        })
        
        print(f"\nResumo de trocas por turno:")
        for turno in resumo_turnos.index:
            print(f"- {turno}: {len(df_turnos[df_turnos['Turno'] == turno])} trocas")
        
        return df_turnos
    
    except Exception as e:
        print(f"ERRO ao analisar turnos: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def classificar_turno(hora_decimal):
    """
    Classifica o turno baseado no horário da troca de operador.
    """
    # Turno A: 6:40 às 7:20 (6.67 às 7.33)
    if 6.67 <= hora_decimal <= 7.33:
        return "Turno A"
    # Turno B: 14:40 às 15:20 (14.67 às 15.33)
    elif 14.67 <= hora_decimal <= 15.33:
        return "Turno B"
    # Turno C: 22:40 às 23:20 (22.67 às 23.33)
    elif 22.67 <= hora_decimal <= 23.33:
        return "Turno C"
    else:
        return "Horário Irregular"

def processar_todos_arquivos():
    """
    Processa todos os arquivos CSV, TXT e ZIP na pasta de dados e gera um arquivo Excel com as métricas.
    """
    try:
        # Lista todos os arquivos na pasta de dados
        arquivos = []
        for ext in ['.csv', '.txt', '.zip']:
            arquivos.extend([f for f in os.listdir(INPUT_DIR) if f.lower().endswith(ext)])
        
        if not arquivos:
            print("Nenhum arquivo encontrado na pasta dados/manobras")
            return
        
        # Processa cada arquivo
        for arquivo in arquivos:
            print(f"\nProcessando arquivo: {arquivo}")
            
            # Caminho completo do arquivo
            caminho_arquivo = os.path.join(INPUT_DIR, arquivo)
            
            # Processa o arquivo base
            df = processar_arquivo_base(caminho_arquivo)
            if df is None or df.empty:
                print(f"Erro ao processar arquivo {arquivo}")
                continue
            
            # Identifica manobras
            df_manobras = identificar_manobras(df)
            
            # Calcula horas por frota
            df_horas_por_frota = calcular_horas_por_frota(df, df_manobras)
            
            # Nome do arquivo de saída
            nome_base = os.path.splitext(arquivo)[0]
            arquivo_saida = os.path.join(OUTPUT_DIR, f"{nome_base}.xlsx")
            
            # Cria diretório de saída se não existir
            os.makedirs(os.path.dirname(arquivo_saida), exist_ok=True)
            
            # Criar arquivo Excel com as planilhas formatadas
            if criar_excel_com_metricas_manobras(df, df_manobras, df_horas_por_frota, arquivo_saida):
                print(f"Arquivo de métricas salvo em {arquivo_saida}")
            else:
                print(f"Erro ao salvar arquivo {arquivo_saida}")
                
    except Exception as e:
        print(f"Erro ao processar arquivos: {str(e)}")
        import traceback
        traceback.print_exc()

def processar_arquivo_manobras(caminho_arquivo):
    """
    Processa um arquivo de monitoramento e gera análise de manobras com dashboard.
    """
    try:
        print(f"\nIniciando processamento de manobras para o arquivo: {caminho_arquivo}")
        
        # Carrega e processa o arquivo base
        df = processar_arquivo_base(caminho_arquivo)
        if df is None or len(df) == 0:
            print("Erro: Não foi possível processar o arquivo base ou o arquivo está vazio")
            return False
            
        print(f"\nArquivo base processado com sucesso. Total de registros: {len(df)}")
        
        # Identifica manobras
        print("\nIdentificando manobras...")
        df_manobras = identificar_manobras(df)
        if df_manobras is None or len(df_manobras) == 0:
            print("Erro: Não foi possível identificar manobras no arquivo")
            return False
            
        print(f"Manobras identificadas com sucesso. Total de manobras: {len(df_manobras)}")
        
        # Calcula métricas
        print("\nCalculando métricas...")
        df_metricas = calcular_horas_por_frota(df, df_manobras)
        if df_metricas is None or len(df_metricas) == 0:
            print("Erro: Não foi possível calcular métricas das manobras")
            return False
            
        print(f"Métricas calculadas com sucesso. Total de equipamentos analisados: {len(df_metricas)}")
        
        # Cria dashboard
        print("\nCriando dashboard...")
        app = criar_dashboard(df_manobras, df_metricas)
        if app is None:
            print("Erro: Não foi possível criar o dashboard")
            return False
            
        print("Dashboard criado com sucesso")
        
        # Salva métricas em Excel
        print("\nSalvando métricas em Excel...")
        nome_arquivo = os.path.splitext(os.path.basename(caminho_arquivo))[0]
        caminho_saida = os.path.join(OUTPUT_DIR, f'{nome_arquivo}.xlsx')
        
        # Cria diretório de saída se não existir
        os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
        
        # Salva arquivo Excel
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            df_manobras.to_excel(writer, sheet_name='MANOBRAS', index=False)
            df_metricas.to_excel(writer, sheet_name='MÉTRICAS', index=False)
            
            # Ajusta largura das colunas
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(df_manobras.columns if sheet_name == 'MANOBRAS' else df_metricas.columns):
                    max_length = max(
                        df_manobras[col].astype(str).apply(len).max() if sheet_name == 'MANOBRAS' else df_metricas[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2
        
        print(f"Métricas salvas com sucesso em: {caminho_saida}")
        
        # Inicia o servidor
        print("\nIniciando servidor...")
        app.run_server(debug=True, port=8050)
        
        return True
        
    except Exception as e:
        print(f"Erro ao processar arquivo de manobras: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def calcular_metricas_manobras(df, palavras_chave_manobra=None, debug=True):
    """
    Calcula métricas detalhadas sobre manobras por operador, frente e equipamento.
    
    Args:
        df (DataFrame): DataFrame com os dados de operação
        palavras_chave_manobra (list): Lista de palavras-chave para identificar operações de manobra
        debug (bool): Se True, imprime informações detalhadas durante o processamento
    
    Returns:
        tuple: (DataFrame com métricas gerais, dict com DataFrames de métricas por agrupamento)
    """
    if df is None or df.empty:
        print("DataFrame vazio ou inválido")
        return None, None, None
    
    # Se palavras_chave_manobra não for fornecido, usa padrão
    if palavras_chave_manobra is None:
        palavras_chave_manobra = ["MANOBRA"]
    
    # Verificar se as colunas necessárias existem
    colunas_basicas = ['Data_Hora', 'Operacao', 'Operador', 'Equipamento']
    
    # Garantir que a coluna 'Data_Hora' existe
    if 'Data_Hora' not in df.columns and 'Data/Hora' in df.columns:
        df['Data_Hora'] = df['Data/Hora']
    
    colunas_faltantes = [col for col in colunas_basicas if col not in df.columns]
    if colunas_faltantes:
        print(f"AVISO: Colunas faltantes: {colunas_faltantes}")
        # Se faltarem colunas críticas, retorna erro
        if 'Data_Hora' in colunas_faltantes or 'Operacao' in colunas_faltantes:
            print("ERRO: Colunas críticas faltantes. Impossível calcular métricas.")
            return None, None, None
            
    # Verificar se a coluna Diferença_Hora existe, senão criar
    if 'Diferença_Hora' not in df.columns:
        print("AVISO: Coluna 'Diferença_Hora' não encontrada. Criando...")
        # Ordena o DataFrame por equipamento e data/hora
        df = df.sort_values(['Equipamento', 'Data_Hora'])
        
        # Calcula a diferença de tempo entre registros consecutivos em horas
        df['Diferença_Hora'] = df.groupby('Equipamento')['Data_Hora'].diff().dt.total_seconds() / 3600
        
        # Preenche valores NaN com um valor razoável (15 minutos = 0.25 horas)
        # Isso ocorre no primeiro registro de cada equipamento
        df['Diferença_Hora'] = df['Diferença_Hora'].fillna(0.25)
        
        # Limita valores muito grandes (gaps maiores que 1 hora são provavelmente pausas)
        df.loc[df['Diferença_Hora'] > 1, 'Diferença_Hora'] = 1.0
        
        print(f"Coluna 'Diferença_Hora' criada com sucesso.")
    
    # Adicionar colunas extras vazias se não existirem
    colunas_extras = ['Frente', 'Tipo de Equipamento', 'Velocidade', 'RPM_Motor']
    for col in colunas_extras:
        if col not in df.columns:
            df[col] = "Não informado"
    
    if debug:
        print(f"\n=== INICIANDO CÁLCULO DE MÉTRICAS DE MANOBRAS ===")
        print(f"Parâmetros:")
        print(f"- Palavras-chave para identificar manobras: {palavras_chave_manobra}")
        print(f"- Frentes encontradas: {sorted(df['Frente'].unique())}")
    
    # Identificar operações de manobra usando correspondência exata
    df['eh_manobra'] = df['Operacao'].apply(
        lambda x: bool(x in palavras_chave_manobra) if pd.notna(x) else False
    )
    
    # Ordenar o DataFrame por Data_Hora e Equipamento para processamento sequencial
    df = df.sort_values(['Equipamento', 'Data_Hora'])
    
    # Calcular velocidade média total das manobras antes de processar
    registros_manobra = df[df['eh_manobra']]
    velocidade_media_total = registros_manobra['Velocidade'].mean() if 'Velocidade' in registros_manobra.columns else 0
    rpm_medio_total = registros_manobra['RPM_Motor'].mean() if 'RPM_Motor' in registros_manobra.columns else 0
    
    if debug:
        manobras_count = df['eh_manobra'].sum()
        total_count = len(df)
        print(f"Encontradas {manobras_count} registros de MANOBRA de um total de {total_count} ({manobras_count/total_count:.1%})")
        
        # Mostrar os tipos de operações de manobra encontradas
        if manobras_count > 0:
            operacoes_manobra = df[df['eh_manobra']]['Operacao'].unique()
            print(f"Operações identificadas como MANOBRA: {operacoes_manobra}")
    
    # Estrutura para armazenar detalhes das manobras
    manobras_detalhes = []
    
    # Variáveis para controle do processamento
    equipamento_atual = None
    operador_atual = None
    frente_atual = None
    tipo_equipamento_atual = None
    inicio_manobra = None
    ultimo_timestamp = None
    tempo_manobra = 0
    total_manobras = 0
    
    # Contadores para estatísticas
    rpm_acumulado = 0
    velocidade_acumulada = 0
    pontos_coletados = 0
    
    # Analisar cada registro
    for idx, row in df.iterrows():
        timestamp = row['Data_Hora']
        equipamento = row['Equipamento']
        operador = row['Operador'] if 'Operador' in df.columns else 'Não informado'
        frente = row['Frente'] if 'Frente' in df.columns else 'Não informado'
        tipo_equipamento = row['Tipo de Equipamento'] if 'Tipo de Equipamento' in df.columns else 'Não informado'
        operacao = row['Operacao']
        eh_manobra = row['eh_manobra']
        diferenca_hora = row['Diferença_Hora']
        velocidade = row['Velocidade'] if 'Velocidade' in df.columns else 0
        rpm = row['RPM_Motor'] if 'RPM_Motor' in df.columns else 0
        
        # Se mudou de equipamento, reseta o controle
        if equipamento != equipamento_atual:
            # Finaliza manobra anterior se existir
            if inicio_manobra is not None and tempo_manobra > 0:
                rpm_medio = rpm_acumulado / pontos_coletados if pontos_coletados > 0 else 0
                velocidade_media = velocidade_acumulada / pontos_coletados if pontos_coletados > 0 else 0
                
                manobras_detalhes.append({
                    'Equipamento': equipamento_atual,
                    'Operador': operador_atual,
                    'Frente': frente_atual,
                    'Tipo_Equipamento': tipo_equipamento_atual,
                    'Inicio': inicio_manobra,
                    'Fim': ultimo_timestamp,
                    'Duração (h)': tempo_manobra,
                    'RPM_Medio': rpm_medio,
                    'Velocidade_Media': velocidade_media,
                    'Data': inicio_manobra.date(),
                    'Hora_Inicio': inicio_manobra.hour
                })
                total_manobras += 1
            
            equipamento_atual = equipamento
            operador_atual = operador
            frente_atual = frente
            tipo_equipamento_atual = tipo_equipamento
            inicio_manobra = None
            tempo_manobra = 0
            rpm_acumulado = 0
            velocidade_acumulada = 0
            pontos_coletados = 0
        
        # Se é o primeiro registro para este equipamento
        if ultimo_timestamp is None:
            ultimo_timestamp = timestamp
            if eh_manobra:
                inicio_manobra = timestamp
                operador_atual = operador
                frente_atual = frente
                tipo_equipamento_atual = tipo_equipamento
                rpm_acumulado += rpm
                velocidade_acumulada += velocidade
                pontos_coletados += 1
                continue
                
        # Calcula o intervalo de tempo em horas
        intervalo = diferenca_hora
        
        # Se é uma manobra
        if eh_manobra:
            # Se não há manobra iniciada, inicia uma nova
            if inicio_manobra is None:
                inicio_manobra = timestamp
                operador_atual = operador
                frente_atual = frente
                tipo_equipamento_atual = tipo_equipamento
            
            # Acumula o tempo desta manobra
            tempo_manobra += intervalo
            rpm_acumulado += rpm
            velocidade_acumulada += velocidade
            pontos_coletados += 1
        
        # Se não é uma manobra, mas há uma manobra em andamento
        elif inicio_manobra is not None:
            # Finaliza a manobra
            rpm_medio = rpm_acumulado / pontos_coletados if pontos_coletados > 0 else 0
            velocidade_media = velocidade_acumulada / pontos_coletados if pontos_coletados > 0 else 0
            
            manobras_detalhes.append({
                'Equipamento': equipamento,
                'Operador': operador_atual,
                'Frente': frente_atual,
                'Tipo_Equipamento': tipo_equipamento_atual,
                'Inicio': inicio_manobra,
                'Fim': timestamp,
                'Duração (h)': tempo_manobra,
                'RPM_Medio': rpm_medio,
                'Velocidade_Media': velocidade_media,
                'Data': inicio_manobra.date(),
                'Hora_Inicio': inicio_manobra.hour
            })
            total_manobras += 1
            
            # Reseta as variáveis de controle
            inicio_manobra = None
            tempo_manobra = 0
            rpm_acumulado = 0
            velocidade_acumulada = 0
            pontos_coletados = 0
        
        ultimo_timestamp = timestamp
    
    # Criar DataFrame com os detalhes das manobras
    df_manobras = pd.DataFrame(manobras_detalhes)
    
    if df_manobras.empty:
        print("Nenhuma manobra encontrada no arquivo.")
        return None, None, None
    
    # Calcular métricas gerais
    metricas_gerais = {
        'Total de Manobras': len(df_manobras),
        'Tempo Total de Operação (h)': df['Diferença_Hora'].sum(),
        'Tempo Total em Manobras (h)': df_manobras['Duração (h)'].sum(),
        'Porcentagem do Tempo em Manobras (%)': (df_manobras['Duração (h)'].sum() / df['Diferença_Hora'].sum() * 100) if df['Diferença_Hora'].sum() > 0 else 0,
        'Tempo Médio por Manobra (h)': df_manobras['Duração (h)'].mean(),
        'RPM Médio': df_manobras['RPM_Medio'].mean(),
        'Velocidade Média (km/h)': df_manobras['Velocidade_Media'].mean()
    }
    
    # Criar DataFrame com métricas gerais
    df_metricas_gerais = pd.DataFrame([metricas_gerais])
    
    # Calcular métricas agregadas
    metricas_agregadas = {}
    
    # Métricas por Frente
    if 'Frente' in df_manobras.columns:
        metricas_por_frente = df_manobras.groupby('Frente').agg({
            'Duração (h)': ['count', 'sum', 'mean'],
            'RPM_Medio': 'mean',
            'Velocidade_Media': 'mean'
        }).round(2)
        metricas_por_frente.columns = [
            'Total de Manobras',
            'Tempo Total (h)',
            'Tempo Médio (h)',
            'RPM Médio',
            'Velocidade Média (km/h)'
        ]
        metricas_agregadas['Por Frente'] = metricas_por_frente
    
    # Métricas por Equipamento
    metricas_por_equipamento = df_manobras.groupby('Equipamento').agg({
        'Duração (h)': ['count', 'sum', 'mean'],
        'RPM_Medio': 'mean',
        'Velocidade_Media': 'mean'
    }).round(2)
    metricas_por_equipamento.columns = [
        'Total de Manobras',
        'Tempo Total (h)',
        'Tempo Médio (h)',
        'RPM Médio',
        'Velocidade Média (km/h)'
    ]
    metricas_agregadas['Por Equipamento'] = metricas_por_equipamento
    
    # Métricas por Tipo de Equipamento
    if 'Tipo_Equipamento' in df_manobras.columns:
        metricas_por_tipo = df_manobras.groupby('Tipo_Equipamento').agg({
            'Duração (h)': ['count', 'sum', 'mean'],
            'RPM_Medio': 'mean',
            'Velocidade_Media': 'mean'
        }).round(2)
        metricas_por_tipo.columns = [
            'Total de Manobras',
            'Tempo Total (h)',
            'Tempo Médio (h)',
            'RPM Médio',
            'Velocidade Média (km/h)'
        ]
        metricas_agregadas['Por Tipo de Equipamento'] = metricas_por_tipo
    
    # Métricas por Operador
    if 'Operador' in df_manobras.columns:
        metricas_por_operador = df_manobras.groupby('Operador').agg({
            'Duração (h)': ['count', 'sum', 'mean'],
            'RPM_Medio': 'mean',
            'Velocidade_Media': 'mean'
        }).round(2)
        metricas_por_operador.columns = [
            'Total de Manobras',
            'Tempo Total (h)',
            'Tempo Médio (h)',
            'RPM Médio',
            'Velocidade Média (km/h)'
        ]
        metricas_agregadas['Por Operador'] = metricas_por_operador
    
    # Métricas por Dia
    metricas_por_dia = df_manobras.groupby('Data').agg({
        'Duração (h)': ['count', 'sum', 'mean'],
        'RPM_Medio': 'mean',
        'Velocidade_Media': 'mean'
    }).round(2)
    metricas_por_dia.columns = [
        'Total de Manobras',
        'Tempo Total (h)',
        'Tempo Médio (h)',
        'RPM Médio',
        'Velocidade Média (km/h)'
    ]
    metricas_agregadas['Por Dia'] = metricas_por_dia
    
    # Métricas por Hora
    metricas_por_hora = df_manobras.groupby('Hora_Inicio').agg({
        'Duração (h)': ['count', 'sum', 'mean'],
        'RPM_Medio': 'mean',
        'Velocidade_Media': 'mean'
    }).round(2)
    metricas_por_hora.columns = [
        'Total de Manobras',
        'Tempo Total (h)',
        'Tempo Médio (h)',
        'RPM Médio',
        'Velocidade Média (km/h)'
    ]
    metricas_agregadas['Por Hora'] = metricas_por_hora
    
    if debug:
        print("\n=== MÉTRICAS CALCULADAS ===")
        print(f"Total de manobras: {len(df_manobras)}")
        print(f"Tempo total em manobras: {df_manobras['Duração (h)'].sum():.2f} horas")
        print(f"Tempo médio por manobra: {df_manobras['Duração (h)'].mean():.2f} horas")
        print(f"Velocidade média: {df_manobras['Velocidade_Media'].mean():.2f} km/h")
        print(f"RPM médio: {df_manobras['RPM_Medio'].mean():.2f}")
    
    return df_manobras, df_metricas_gerais, metricas_agregadas

def listar_arquivos_para_processar():
    """
    Lista os arquivos disponíveis para processamento.
    """
    arquivos = []
    
    # Lista arquivos TXT e CSV
    for ext in ['.txt', '.csv']:
        arquivos.extend(glob.glob(os.path.join(INPUT_DIR, f'*{ext}')))
        
    # Lista arquivos ZIP
    arquivos.extend(glob.glob(os.path.join(INPUT_DIR, f'*.zip')))
    
    if not arquivos:
        print(f"Nenhum arquivo encontrado em {INPUT_DIR}")
    else:
        print(f"Arquivos encontrados: {len(arquivos)}")
        for arq in arquivos:
            print(f"  - {os.path.basename(arq)}")
        
    return arquivos

def carregar_config_calculos(silencioso=True):
    """
    Carrega o arquivo de configuração JSON para os cálculos.
    
    Args:
        silencioso (bool): Se True, não exibe mensagens de aviso quando o arquivo não existe
    """
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            if not silencioso:
                print(f"Erro ao carregar arquivo de configuração: {str(e)}")
    else:
        if not silencioso:
            print(f"Arquivo de configuração não encontrado: {CONFIG_FILE}")
    
    return {}

def criar_dashboard_html(df_manobras, metricas_agregadas, caminho_saida=None):
    """
    Cria uma dashboard HTML com as métricas de manobras.
    
    Args:
        df_manobras (DataFrame): DataFrame com as manobras
        metricas_agregadas (dict): Dicionário com DataFrames de métricas por agrupamento
        caminho_saida (str): Caminho para salvar o arquivo HTML (opcional)
        
    Returns:
        str: Caminho do arquivo HTML gerado ou None se não for salvo
    """
    try:
        import plotly.express as px
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
        import plotly.io as pio
        
        # Extrair dados para os gráficos
        metricas_dia = metricas_agregadas.get('Por Dia', pd.DataFrame())
        metricas_hora = metricas_agregadas.get('Por Hora', pd.DataFrame())
        metricas_operador = metricas_agregadas.get('Por Operador', pd.DataFrame())
        metricas_equipamento = metricas_agregadas.get('Por Equipamento', pd.DataFrame())
        metricas_tipo = metricas_agregadas.get('Por Tipo de Equipamento', pd.DataFrame())
        
        # Criar HTML
        html_content = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Dashboard de Manobras</title>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.2.3/dist/css/bootstrap.min.css">
            <style>
                body { padding: 20px; }
                .card { margin-bottom: 20px; }
                .chart-container { height: 400px; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1 class="text-center mb-4">Dashboard de Manobras</h1>
                
                <div class="row">
                    <div class="col-md-12">
                        <div class="card">
                            <div class="card-header"><h5>Resumo</h5></div>
                            <div class="card-body">
                                <div class="row">
        """
        
        # Adicionar cards com métricas gerais
        resumo_metrics = [
            {'title': 'Total de Manobras', 'value': len(df_manobras)},
            {'title': 'Tempo Total em Manobras (h)', 'value': round(df_manobras['Duração (h)'].sum(), 2)},
            {'title': 'Velocidade Média (km/h)', 'value': round(df_manobras['Velocidade_Media'].mean(), 2)},
            {'title': 'RPM Médio', 'value': round(df_manobras['RPM_Medio'].mean(), 2)}
        ]
        
        for metric in resumo_metrics:
            html_content += f"""
                <div class="col-md-3">
                    <div class="card text-center">
                        <div class="card-body">
                            <h5 class="card-title">{metric['title']}</h5>
                            <p class="card-text display-6">{metric['value']}</p>
                        </div>
                    </div>
                </div>
            """
        
        html_content += """
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
        """
        
        # Gráficos
        if not metricas_dia.empty:
            # Gráfico de evolução por dia
            fig_dia = px.bar(
                metricas_dia, 
                x='Data', 
                y='Número de Manobras',
                title='Evolução de Manobras por Dia',
                labels={'Data': 'Data', 'Número de Manobras': 'Quantidade de Manobras'}
            )
            
            # Adicionar linha com tempo em manobras
            fig_dia.add_trace(
                go.Scatter(
                    x=metricas_dia['Data'], 
                    y=metricas_dia['Tempo Total em Manobras (h)'],
                    name='Tempo em Manobras (h)',
                    yaxis='y2'
                )
            )
            
            fig_dia.update_layout(
                yaxis2=dict(
                    title='Tempo em Manobras (h)',
                    overlaying='y',
                    side='right'
                ),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1)
            )
            
            html_content += f"""
                <div class="row">
                    <div class="col-md-12">
                        <div class="card">
                            <div class="card-header"><h5>Evolução por Dia</h5></div>
                            <div class="card-body">
                                <div class="chart-container" id="chart-dia"></div>
                            </div>
                        </div>
                    </div>
                </div>
                <script>
                    var fig_dia = {fig_dia.to_json()};
                    Plotly.newPlot('chart-dia', fig_dia.data, fig_dia.layout);
                </script>
            """
        
        if not metricas_hora.empty:
            # Gráfico por hora do dia
            fig_hora = px.bar(
                metricas_hora, 
                x='Hora', 
                y='Número de Manobras',
                title='Manobras por Hora do Dia',
                labels={'Hora': 'Hora do Dia', 'Número de Manobras': 'Quantidade de Manobras'}
            )
            
            html_content += f"""
                <div class="row">
                    <div class="col-md-12">
                        <div class="card">
                            <div class="card-header"><h5>Manobras por Hora</h5></div>
                            <div class="card-body">
                                <div class="chart-container" id="chart-hora"></div>
                            </div>
                        </div>
                    </div>
                </div>
                <script>
                    var fig_hora = {fig_hora.to_json()};
                    Plotly.newPlot('chart-hora', fig_hora.data, fig_hora.layout);
                </script>
            """
        
        if not metricas_tipo.empty:
            # Gráfico por tipo de equipamento
            fig_tipo = px.pie(
                metricas_tipo, 
                values='Número de Manobras', 
                names='Tipo de Equipamento',
                title='Manobras por Tipo de Equipamento'
            )
            
            html_content += f"""
                <div class="row">
                    <div class="col-md-6">
                        <div class="card">
                            <div class="card-header"><h5>Manobras por Tipo</h5></div>
                            <div class="card-body">
                                <div class="chart-container" id="chart-tipo"></div>
                            </div>
                        </div>
                    </div>
            """
        
        if not metricas_operador.empty:
            # Top 10 operadores
            top_operadores = metricas_operador.sort_values('Número de Manobras', ascending=False).head(10)
            fig_operador = px.bar(
                top_operadores, 
                x='Operador', 
                y='Número de Manobras',
                title='Top 10 Operadores por Número de Manobras',
                text='Número de Manobras'
            )
            
            html_content += f"""
                    <div class="col-md-6">
                        <div class="card">
                            <div class="card-header"><h5>Top Operadores</h5></div>
                            <div class="card-body">
                                <div class="chart-container" id="chart-operador"></div>
                            </div>
                        </div>
                    </div>
                </div>
                <script>
                    var fig_tipo = {fig_tipo.to_json()};
                    Plotly.newPlot('chart-tipo', fig_tipo.data, fig_tipo.layout);
                    
                    var fig_operador = {fig_operador.to_json()};
                    Plotly.newPlot('chart-operador', fig_operador.data, fig_operador.layout);
                </script>
            """
        
        html_content += """
            </div>
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.2.3/dist/js/bootstrap.bundle.min.js"></script>
        </body>
        </html>
        """
        
        # Salvar HTML se caminho fornecido
        if caminho_saida:
            with open(caminho_saida, 'w', encoding='utf-8') as f:
                f.write(html_content)
            print(f"Dashboard HTML criado: {caminho_saida}")
            return caminho_saida
        
        return html_content
        
    except Exception as e:
        print(f"Erro ao criar dashboard HTML: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def processar_arquivo(arquivo):
    """
    Processa um único arquivo de manobras e gera o Excel com métricas e dashboard interativo
    """
        if not os.path.exists(arquivo):
            print(f"Arquivo não encontrado: {arquivo}")
        return False
            
    print(f"Processando arquivo: {arquivo}")
        df = processar_arquivo_base(arquivo)
        
        if df is not None and not df.empty:
        # Carrega configurações
        config = carregar_config_calculos(silencioso=True)
        
        # Define parâmetros do cálculo de manobras
        palavras_chave_manobra = ["MANOBRA"]  # Padrão
        debug = True
        
        # Se existir configuração, usa os valores definidos
        if config and 'manobras' in config:
            palavras_chave_manobra = config.get('manobras', {}).get('palavras_chave_manobra', palavras_chave_manobra)
            debug = config.get('manobras', {}).get('debug', True)
        
        # Calcular métricas de manobras
        df_manobras, df_metricas_gerais, metricas_agregadas = calcular_metricas_manobras(
            df,
            palavras_chave_manobra=palavras_chave_manobra,
            debug=debug
        )
        
        if df_manobras is not None:
                # Salva resultados em Excel
            nome_base = os.path.splitext(os.path.basename(arquivo))[0]
            caminho_saida = os.path.join(OUTPUT_DIR, f"{nome_base}.xlsx")
                
            if criar_excel_com_metricas_manobras(df, df_manobras, df_metricas_gerais, metricas_agregadas, caminho_saida):
                    print(f"Resultados salvos em: {caminho_saida}")
                
                # Criar e iniciar o dashboard interativo
                app = criar_dashboard(df_manobras, metricas_agregadas)
                print("\nIniciando dashboard interativo...")
                print("Acesse http://127.0.0.1:8050 no seu navegador")
                app.run(debug=False)
                return True
                else:
                    print("Erro ao salvar resultados")
                return False
        else:
            print("Não foi possível identificar manobras no arquivo")
            return False
        else:
            print("Não foi possível processar o arquivo") 
        return False

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        # Processa um arquivo específico
        arquivo = sys.argv[1]
        processar_arquivo(arquivo)
    else:
        # Processa todos os arquivos na pasta dados/manobras
        print("Processando todos os arquivos na pasta dados/manobras...")
        arquivos = listar_arquivos_para_processar()
        
        if not arquivos:
            print("Nenhum arquivo encontrado para processamento.")
            print("Por favor, coloque os arquivos na pasta dados/manobras e tente novamente.")
            sys.exit(1)
            
        for arquivo in arquivos:
            print(f"\nProcessando arquivo: {arquivo}")
            processar_arquivo(arquivo)