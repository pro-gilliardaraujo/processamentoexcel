"""
Script para processamento completo de dados de monitoramento de transbordos.
Lê arquivos TXT ou CSV na pasta raiz, processa-os e gera arquivos Excel com planilhas auxiliares prontas.
Também processa arquivos ZIP contendo TXT ou CSV.
"""

import pandas as pd
import numpy as np
import os
import glob
import zipfile
import tempfile
import shutil
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# Configurações
processCsv = True  # Altere para True quando quiser processar arquivos CSV

# Constantes
COLUNAS_REMOVER = [
    'Latitude',
    'Longitude',
    'Regional',
    'Unidade',
    'Centro de Custo',
    'Fazenda', 
    'Zona', 
    'Talhao'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento', 'Descricao Equipamento', 'Estado', 'Estado Operacional',
    'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro', 'Motor Ligado', 'Operacao', 'Operador',
    'RPM Motor', 'Tipo de Equipamento', 'Velocidade', 'Parado com motor ligado',
    'Diferença_Hora', 'Horas Produtivas', 'GPS', 'Motor Ocioso'
]

# Valores a serem filtrados
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO", "1 - SEM OPERADOR", "9999 - OPERADOR TESTE"]

# Adicionar função para extrair frente antes da função processar_arquivo_base()
def extrair_frente(grupo_equipamento_frente):
    """
    Extrai a frente da coluna 'Grupo Equipamento/Frente'.
    
    Args:
        grupo_equipamento_frente (str): Valor da coluna 'Grupo Equipamento/Frente'
    
    Returns:
        str: Nome da frente extraído ou 'Não informado' se não conseguir extrair
    """
    if pd.isna(grupo_equipamento_frente) or grupo_equipamento_frente == '':
        return 'Não informado'
    
    # Tentar extrair a frente (assumindo formato "GRUPO/FRENTE" ou similar)
    try:
        # Se contém "/", pega a parte após a barra
        if '/' in str(grupo_equipamento_frente):
            return str(grupo_equipamento_frente).split('/')[-1].strip()
        # Se contém "-", pega a parte após o traço
        elif '-' in str(grupo_equipamento_frente):
            return str(grupo_equipamento_frente).split('-')[-1].strip()
        # Caso contrário, usa o valor completo
        else:
            return str(grupo_equipamento_frente).strip()
    except:
        return 'Não informado'

def processar_arquivo_base(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV de transbordos e retorna o DataFrame com as transformações necessárias.
    Usando exatamente o mesmo método do Codigo_Base_TT.py para cálculo da Diferença_Hora.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo (TXT ou CSV são tratados da mesma forma se usarem separador ';')
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio (apenas cabeçalhos sem dados)
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                # Retornar o DataFrame vazio mas com as colunas, em vez de None
                # Garantir que todas as colunas desejadas existam
                for col in COLUNAS_DESEJADAS:
                    if col not in df.columns:
                        df[col] = np.nan
                # Reorganizar as colunas na ordem desejada
                colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Padronizar valores da coluna Grupo Operacao
            if 'Grupo Operacao' in df.columns:
                df['Grupo Operacao'] = df['Grupo Operacao'].str.strip()
                # Mapear valores para garantir consistência
                mapa_grupo_operacao = {
                    'SEM APONTAMENTO': 'Sem Apontamento',
                    'PRODUTIVA': 'Produtiva',
                    'MANUTENCAO': 'Manutenção',
                    'MANUTENÇÃO': 'Manutenção'
                }
                df['Grupo Operacao'] = df['Grupo Operacao'].replace(mapa_grupo_operacao)
            
            # Verificar se 'Data/Hora' existe, caso ainda não tenha sido separado
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Remover colunas conforme solicitado no Codigo_Base_TT.py
            colunas_remover = ['Unidade', 'Centro de Custo', 'Fazenda', 'Zona', 'Talhao']
            df = df.drop(columns=colunas_remover, errors='ignore')
            
            # MÉTODO EXATO DO Codigo_Base_TT.py para cálculo da Diferença_Hora para garantir mesmos resultados
            # Conversão de Hora para datetime (apenas se ainda não for)
            if df['Hora'].dtype != 'datetime64[ns]':
                df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular Diferença_Hora sem arredondamentos usando o método EXATO do Codigo_Base_TT.py
            # NOTA: Removida a regra que zerava valores > 0.5, pois não existe no Codigo_Base_TT.py
            # e estava causando perda de aproximadamente 16 horas no total
            df['Diferença_Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S').diff()
            df['Diferença_Hora'] = pd.to_timedelta(df['Diferença_Hora'], errors='coerce')
            df['Diferença_Hora'] = df['Diferença_Hora'].dt.total_seconds() / 3600  # Conversor para horas
            df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: x if x >= 0 else 0)
            
            # Soma total para verificação de precisão (mesma lógica do Codigo_Base_TT.py)
            print(f"Diferença_Hora calculada usando método exato do Codigo_Base_TT.py. Soma total: {df['Diferença_Hora'].sum():.8f} horas")
            
            # Conversão de Motor Ligado para formato conforme Codigo_Base_TT.py
            for col in ['Motor Ligado']:
                if col in df.columns:
                    df[col] = df[col].replace({1: 'LIGADO', 0: 'DESLIGADO'})
            
            # Criar a coluna "Parado com motor ligado" exatamente como no Codigo_Base_TT.py
            df['Parado com motor ligado'] = ((df['Velocidade'] == 0) & (df['Motor Ligado'] == 'LIGADO')).astype(int)
            
            # Verifica se Horas Produtivas já existe, senão calcula usando método do Codigo_Base_TT.py
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                # Calcular horas produtivas sem arredondamento, mantendo a precisão completa
                df['Horas Produtivas'] = df.apply(
                    lambda row: row['Diferença_Hora'] if row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
                # Soma total de horas produtivas para verificação
                print(f"Total de horas produtivas: {df['Horas Produtivas'].sum():.8f}")
            else:
                # Limpa e converte para número
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # IMPORTANTE: Zerar horas produtivas dos operadores excluídos para garantir que não sejam contabilizadas
            df.loc[df['Operador'].isin(OPERADORES_EXCLUIR), 'Horas Produtivas'] = 0
            print(f"Total de horas produtivas após exclusão de operadores: {df['Horas Produtivas'].sum():.8f}")
            
            # Coluna de GPS - Para transbordos, vamos considerar GPS quando houver "RTK (Piloto Automatico)" 
            # e Velocidade > 0 (se a coluna existir)
            if 'RTK (Piloto Automatico)' in df.columns:
                # Primeiro, garantir que RTK seja convertido para valores numéricos (1/0)
                df['RTK (Piloto Automatico)'] = df['RTK (Piloto Automatico)'].apply(lambda x: 
                    1 if (isinstance(x, bool) and x) or 
                         (isinstance(x, (int, float)) and x == 1) or
                         (isinstance(x, str) and str(x).upper().strip() in ['1', 'SIM', 'S', 'VERDADEIRO', 'TRUE', 'LIGADO'])
                    else 0
                )
                
                # Verificar se a coluna GPS já existe e tem valores
                if 'GPS' not in df.columns or df['GPS'].isna().all() or (df['GPS'] == 0).all():
                    print("Calculando coluna GPS com base em RTK (Piloto Automático)...")
                    # Agora aplicar a lógica de GPS usando o RTK normalizado
                    df['GPS'] = df.apply(
                        lambda row: row['Diferença_Hora'] if row['RTK (Piloto Automatico)'] == 1 
                        and row['Velocidade'] > 0 and row['Grupo Operacao'] == 'Produtiva' else 0,
                        axis=1
                    )
                else:
                    print("Coluna GPS já existe com valores. Mantendo os valores originais.")
                
                # IMPORTANTE: Zerar GPS dos operadores excluídos para garantir que não sejam contabilizados  
                df.loc[df['Operador'].isin(OPERADORES_EXCLUIR), 'GPS'] = 0
                    
                # Verificar o total de horas com GPS ativo
                total_gps = df['GPS'].sum()
                print(f"Total de horas com GPS ativo após exclusão de operadores: {total_gps:.4f}")
            else:
                # Se não tiver a coluna RTK, criar uma coluna GPS zerada
                df['GPS'] = 0
                print("Coluna RTK (Piloto Automatico) não encontrada. GPS definido como zero.")
            
            # Conversão de colunas binárias para valores numéricos (garantindo que sejam números)
            for col in ['Esteira Ligada', 'Field Cruiser', 'Implemento Ligado']:  # RTK já foi tratado acima
                if col in df.columns and col != 'Motor Ligado':  # Motor Ligado já foi tratado acima
                    # Se a coluna for texto (LIGADO/DESLIGADO), converter para 1/0
                    if df[col].dtype == 'object':
                        df[col] = df[col].replace({'LIGADO': 1, 'DESLIGADO': 0})
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    # Se chegou aqui, todas as tentativas de codificação falharam
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_motor_ocioso_novo(df):
    """
    Extrai os dados de motor ocioso da Base Calculo.
    Esta função não faz cálculos adicionais, apenas formata os dados
    já calculados na Base Calculo para exibição na planilha auxiliar.
    Os dados são agregados por operador calculando a média quando o mesmo 
    operador aparece em diferentes equipamentos/grupos.
    
    Args:
        df (DataFrame): Base Calculo com os dados já processados
        
    Returns:
        DataFrame: DataFrame com as colunas 'Operador', 'Porcentagem', 'Tempo Ligado', 'Tempo Ocioso'
    """
    # Agrupar por operador e calcular as médias
    agrupado = df.groupby('Operador').agg({
        'Motor Ligado': 'mean',  # Média do tempo ligado
        'Parado com motor ligado': 'mean',  # Média do tempo ocioso
        'Equipamento': 'count'  # Conta quantas vezes o operador aparece
    })
    
    resultado_motor_ocioso = []
    
    print("\n=== DETALHAMENTO DO MOTOR OCIOSO (DADOS DA BASE CALCULO) ===")
    
    # Para cada operador na Base Calculo
    for operador, row in agrupado.iterrows():
        tempo_ligado = row['Motor Ligado']
        tempo_ocioso = row['Parado com motor ligado']
        ocorrencias = row['Equipamento']  # Número de vezes que o operador aparece
        
        # Calcular a porcentagem usando os valores médios
        porcentagem = tempo_ocioso / tempo_ligado if tempo_ligado > 0 else 0
        
        print(f"\nOperador: {operador} (média de {ocorrencias} registros)")
        print(f"Tempo Ligado (média): {tempo_ligado:.6f}")
        print(f"Tempo Ocioso (média): {tempo_ocioso:.6f}")
        print(f"Porcentagem: {porcentagem:.6f}")
        print("-" * 50)
        
        resultado_motor_ocioso.append({
            'Operador': operador,
            'Porcentagem': porcentagem,
            'Tempo Ligado': tempo_ligado,
            'Tempo Ocioso': tempo_ocioso
        })
    
    # Retornar DataFrame formatado para a planilha
    return pd.DataFrame(resultado_motor_ocioso)

def carregar_config_calculos():
    """
    Retorna as configurações de cálculos embutidas diretamente no código.
    As configurações são as mesmas definidas no arquivo calculos_config.json.
    """
    # Configurações embutidas diretamente no código
    config = {
        "CD": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "8490 - LAVAGEM",
                    "MANUTENCAO",
                    "LAVAGEM",
                    "INST CONFIG TECNOL EMBARCADAS",
                    "1055 - MANOBRA"
                ],
                "grupos_operacao_excluidos": [
                    "Manutenção", 
                    "Inaptidão"
                ],
                "operadores_excluidos": []
            },
            "equipamentos_excluidos": []
        },
        "TT": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "9016 - ENCH SISTEMA FREIO",
                    "6340 - BASCULANDO  TRANSBORDAGEM",
                    "9024 - DESATOLAMENTO",
                    "MANUTENCAO",
                    "MANUTENÇÃO",
                    "INST CONFIG TECNOL EMBARCADAS",
                    "1055 - MANOBRA"
                ],
                "grupos_operacao_excluidos": [
                    "Manutenção", 
                    "Inaptidão"
                ],
                "operadores_excluidos": []
            },
            "equipamentos_excluidos": []
        }
    }
    
    print("Usando configurações embutidas no código, ignorando o arquivo calculos_config.json")
    return config

def carregar_substituicoes_operadores():
    """
    Carrega o arquivo substituiroperadores.json que contém os mapeamentos 
    de substituição de operadores.
    
    Returns:
        dict: Dicionário com mapeamento {operador_origem: operador_destino}
        ou dicionário vazio se o arquivo não existir ou for inválido
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Caminho para o arquivo de substituição
    arquivo_substituicao = os.path.join(diretorio_raiz, "config", "substituiroperadores.json")
    
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo_substituicao):
        print(f"Arquivo de substituição de operadores não encontrado: {arquivo_substituicao}")
        return {}
    
    try:
        # Carregar o arquivo JSON
        with open(arquivo_substituicao, 'r', encoding='utf-8') as f:
            substituicoes = json.load(f)
        
        # Criar dicionário de substituições
        mapeamento = {item['operador_origem']: item['operador_destino'] for item in substituicoes}
        
        print(f"Carregadas {len(mapeamento)} substituições de operadores.")
        return mapeamento
        
    except Exception as e:
        print(f"Erro ao carregar arquivo de substituição de operadores: {str(e)}")
        return {}

def aplicar_substituicao_operadores(df, mapeamento_substituicoes):
    """
    Aplica as substituições de operadores no DataFrame.
    
    Args:
        df (DataFrame): DataFrame a ser processado
        mapeamento_substituicoes (dict): Dicionário com mapeamento {operador_origem: operador_destino}
    
    Returns:
        tuple: (DataFrame com substituições aplicadas, DataFrame com registro das substituições)
    """
    if not mapeamento_substituicoes or 'Operador' not in df.columns:
        return df, pd.DataFrame(columns=['ID Original', 'Nome Original', 'ID Nova', 'Nome Novo', 'Registros Afetados'])
    
    # Criar uma cópia para não alterar o DataFrame original
    df_modificado = df.copy()
    
    # Lista para armazenar as substituições realizadas
    substituicoes_realizadas = []
    total_registros_substituidos = 0
    
    # Verificar operadores antes da substituição para relatório
    operadores_antes = df_modificado['Operador'].unique()
    print(f"\nOperadores antes da substituição: {len(operadores_antes)}")
    
    # Contar operadores e registros antes da substituição
    contagem_antes = df_modificado['Operador'].value_counts().to_dict()
    
    # Aplicar as substituições
    for origem, destino in mapeamento_substituicoes.items():
        # Verificar se o operador de origem existe no DataFrame
        registros_afetados = df_modificado[df_modificado['Operador'] == origem].shape[0]
        
        if registros_afetados > 0:
            # Substituir o operador
            df_modificado.loc[df_modificado['Operador'] == origem, 'Operador'] = destino
            
            total_registros_substituidos += registros_afetados
            
            # Extrair IDs e nomes
            id_original = origem.split(' - ')[0] if ' - ' in origem else origem
            nome_original = origem.split(' - ')[1] if ' - ' in origem else ''
            id_nova = destino.split(' - ')[0] if ' - ' in destino else destino
            nome_novo = destino.split(' - ')[1] if ' - ' in destino else ''
            
            substituicoes_realizadas.append({
                'ID Original': id_original,
                'Nome Original': nome_original,
                'ID Nova': id_nova, 
                'Nome Novo': nome_novo,
                'Registros Afetados': registros_afetados
            })
            print(f"Operador '{origem}' substituído por '{destino}' em {registros_afetados} registros")
    
    # Verificar operadores após substituição
    operadores_depois = df_modificado['Operador'].unique()
    print(f"Operadores após substituição: {len(operadores_depois)}")
    print(f"Total de registros substituídos: {total_registros_substituidos}")
    
    # Criar DataFrame com as substituições realizadas
    df_substituicoes = pd.DataFrame(substituicoes_realizadas)
    
    return df_modificado, df_substituicoes

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento e frente.
    Fórmula: (Total Geral - Manutenção) / Total Geral
    Agora também inclui o percentual de uso de GPS por frota.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento e frente com coluna GPS
    """
    # Filtramos os dados excluindo os operadores da lista
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Extrair frente da coluna 'Grupo Equipamento/Frente'
    df_filtrado = df_filtrado.copy()
    df_filtrado['Frente'] = df_filtrado['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por Equipamento e Frente
    grupos = df_filtrado.groupby(['Equipamento', 'Frente'])
    resultados = []
    
    for (equipamento, frente), dados_grupo in grupos:
        # CORREÇÃO: Usar soma total direta, não média diária
        # Calcular Total Geral (soma de todas as diferenças de hora)
        total_geral = dados_grupo['Diferença_Hora'].sum()
        
        # Calcular horas de manutenção (Grupo Operacao = 'Manutenção')
        horas_manutencao = dados_grupo[dados_grupo['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum()
        
        # Calcular GPS para esta frota - usar apenas dados produtivos
        dados_produtivos = dados_grupo[dados_grupo['Grupo Operacao'] == 'Produtiva']
        total_horas_produtivas = dados_produtivos['Diferença_Hora'].sum()
        horas_gps = dados_produtivos['GPS'].sum()
        
        # CORREÇÃO: Fórmula exata como no Excel: (Total Geral - Manutenção) / Total Geral
        # A disponibilidade mecânica é: (Total - Manutenção) / Total
        if total_geral > 0:
            disp_mecanica = (total_geral - horas_manutencao) / total_geral
        else:
            disp_mecanica = 0.0
        
        # Calcular percentual de uso de GPS (GPS / tempo produtivo)
        uso_gps = calcular_porcentagem(horas_gps, total_horas_produtivas)
        
        # Debug: mostrar valores para verificação
        print(f"Equipamento: {equipamento}")
        print(f"  Total Geral: {total_geral:.6f}")
        print(f"  Manutenção: {horas_manutencao:.6f}")
        print(f"  Disponibilidade: {disp_mecanica:.6f} ({disp_mecanica*100:.2f}%)")
        print(f"  Fórmula: ({total_geral:.6f} - {horas_manutencao:.6f}) / {total_geral:.6f} = {disp_mecanica:.6f}")
        
        resultados.append({
            'Frota': equipamento,
            'Frente': frente,
            'Disponibilidade': disp_mecanica,
            'GPS': uso_gps
        })
    
    # Ordenar primeiro por frente, depois por disponibilidade (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values(['Frente', 'Disponibilidade'], ascending=[True, False])

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Calcula médias diárias considerando os dias efetivos de cada frota.
    Esta função NÃO aplica qualquer filtro de operador.
    Também identifica as faltas de horário por dia específico.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota com detalhamento por dia
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    # Obter todos os dias únicos no dataset
    dias_unicos = sorted(df['Data'].unique()) if 'Data' in df.columns else []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        
        # Determinar número de dias efetivos para este equipamento
        dias_equip = dados_equip['Data'].nunique() if 'Data' in dados_equip.columns else 1
        
        total_horas = dados_equip['Diferença_Hora'].sum()
        
        # Se houver múltiplos dias, usar média diária
        if dias_equip > 1:
            total_horas = total_horas / dias_equip
        
        # Calcular a diferença para 24 horas
        diferenca_24h = max(24 - total_horas, 0)
        
        # Criar o resultado básico (colunas originais mantidas)
        resultado = {
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        }
        
        # Adicionar detalhamento por dia (novas colunas)
        if len(dias_unicos) > 0:
            for dia in dias_unicos:
                dados_dia = dados_equip[dados_equip['Data'] == dia]
                
                # Se não houver dados para este dia e equipamento, a diferença é 24h
                if len(dados_dia) == 0:
                    resultado[f'Falta {dia}'] = 24.0
                    continue
                
                # Calcular horas registradas neste dia
                horas_dia = dados_dia['Diferença_Hora'].sum()
                
                # Calcular a diferença para 24 horas neste dia
                diferenca_dia = max(24 - horas_dia, 0)
                
                # Adicionar ao resultado apenas se houver falta (diferença > 0)
                if diferenca_dia > 0:
                    resultado[f'Falta {dia}'] = diferenca_dia
                else:
                    resultado[f'Falta {dia}'] = 0.0
        
        resultados.append(resultado)
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica(base_calculo):
    """
    Extrai e agrega a eficiência energética por operador e frente da tabela Base Calculo.
    Não realiza novos cálculos, apenas agrupa os valores já calculados por operador e frente.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador e frente (agregado)
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in base_calculo.columns:
        base_calculo = base_calculo.copy()
        base_calculo['Frente'] = base_calculo['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Extrair apenas as colunas necessárias para eficiência energética
    df_temp = base_calculo[['Operador', 'Frente', 'Motor Ligado', 'GPS', '% Utilização GPS']].copy()
    
    # Agrupar por operador e frente
    agrupado = df_temp.groupby(['Operador', 'Frente']).agg({
        'Motor Ligado': 'sum',
        'GPS': 'sum',
        '% Utilização GPS': 'mean'  # Média ponderada do percentual
    }).reset_index()
    
    # Calcular eficiência energética (que para transbordos é % Utilização GPS)
    resultados = []
    for _, row in agrupado.iterrows():
        eficiencia = row['GPS'] / row['Motor Ligado'] if row['Motor Ligado'] > 0 else 0
        resultados.append({
            'Operador': row['Operador'],
            'Frente': row['Frente'],
            'Eficiência': eficiencia
        })
    
    # Ordenar primeiro por frente, depois por eficiência (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values(['Frente', 'Eficiência'], ascending=[True, False])

def calcular_falta_apontamento(base_calculo):
    """
    Extrai o percentual de falta de apontamento por operador e frente da Base Calculo, sem realizar novos cálculos.
    Agrega os dados por operador e frente, calculando a média quando um operador aparece em múltiplas situações.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de falta de apontamento por operador e frente (agregado)
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in base_calculo.columns:
        base_calculo = base_calculo.copy()
        base_calculo['Frente'] = base_calculo['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Selecionar apenas as colunas relevantes
    df_temp = base_calculo[['Operador', 'Frente', '% Falta de Apontamento']].copy()
    
    # Agrupar por operador e frente
    agrupado = df_temp.groupby(['Operador', 'Frente'])['% Falta de Apontamento'].mean().reset_index()
    
    # Renomear a coluna para o formato esperado no relatório
    agrupado.rename(columns={'% Falta de Apontamento': 'Porcentagem'}, inplace=True)
    
    print("\n=== DETALHAMENTO DE FALTA DE APONTAMENTO POR FRENTE (EXTRAÍDO DA BASE CALCULO) ===")
    for _, row in agrupado.iterrows():
        print(f"Operador: {row['Operador']}, Frente: {row['Frente']}, Porcentagem: {row['Porcentagem']:.6f}")
    print("-" * 60)
    
    # Ordenar primeiro por frente, depois por porcentagem (decrescente)
    return agrupado.sort_values(['Frente', 'Porcentagem'], ascending=[True, False])

def calcular_uso_gps(base_calculo):
    """
    Extrai o percentual de uso de GPS por operador e frente da Base Calculo, sem realizar novos cálculos.
    Agrega os dados por operador e frente, calculando a média quando um operador aparece em múltiplas situações.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador e frente (agregado)
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in base_calculo.columns:
        base_calculo = base_calculo.copy()
        base_calculo['Frente'] = base_calculo['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Selecionar apenas as colunas relevantes
    df_temp = base_calculo[['Operador', 'Frente', '% Utilização GPS']].copy()
    
    # Agrupar por operador e frente
    agrupado = df_temp.groupby(['Operador', 'Frente'])['% Utilização GPS'].mean().reset_index()
    
    # Renomear a coluna para o formato esperado no relatório
    agrupado.rename(columns={'% Utilização GPS': 'Porcentagem'}, inplace=True)
    
    print("\n=== DETALHAMENTO DE UTILIZAÇÃO DE GPS POR FRENTE (EXTRAÍDO DA BASE CALCULO) ===")
    for _, row in agrupado.iterrows():
        print(f"Operador: {row['Operador']}, Frente: {row['Frente']}, Porcentagem: {row['Porcentagem']:.6f}")
    print("-" * 60)
    
    # Ordenar primeiro por frente, depois por porcentagem (decrescente)
    return agrupado.sort_values(['Frente', 'Porcentagem'], ascending=[True, False])

def calcular_media_velocidade(df):
    """
    Calcula a média de velocidade para cada operador e frente, separando por tipo de deslocamento:
    - Deslocamento Carregado (Estado Operacional = "DESLOCAMENTO CARREGADO")
    - Deslocamento Vazio (Estado Operacional = "DESLOCAMENTO VAZIO")
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame com a média de velocidade por operador e frente e tipo de deslocamento
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in df.columns:
        df = df.copy()
        df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Filtrar operadores excluídos
    df = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # DIAGNÓSTICO: Verificar dados antes da filtragem
    print(f"Total de registros antes da filtragem: {len(df)}")
    
    # Verificar se as colunas necessárias existem
    colunas_necessarias = ['Estado Operacional', 'Velocidade']
    for coluna in colunas_necessarias:
        if coluna not in df.columns:
            print(f"ERRO: Coluna '{coluna}' não encontrada no DataFrame!")
            return pd.DataFrame(columns=['Operador', 'Frente', 'Velocidade Geral', 'Velocidade Carregado', 'Velocidade Vazio'])
    
    # Identificar registros com velocidade > 0 (sem filtros adicionais inicialmente)
    registros_velocidade = df['Velocidade'] > 0
    
    # DIAGNÓSTICO: Mostrar estatísticas de velocidade
    print(f"Registros com velocidade > 0: {registros_velocidade.sum()}")
    print(f"Média de velocidade geral: {df[registros_velocidade]['Velocidade'].mean()}")
    
    # DIAGNÓSTICO: Verificar estados operacionais únicos disponíveis
    estados_unicos = df['Estado Operacional'].unique()
    print(f"Estados operacionais únicos disponíveis: {estados_unicos}")
    
    # Estados operacionais específicos - tratando possíveis variações
    estado_carregado = "DESLOCAMENTO CARREGADO"
    estado_vazio = "DESLOCAMENTO VAZIO"
    
    # DIAGNÓSTICO: Verificar quantos registros existem para cada estado operacional
    # antes de aplicar outros filtros
    for estado in estados_unicos:
        registros = len(df[df['Estado Operacional'] == estado])
        velocidade_media = df[df['Estado Operacional'] == estado]['Velocidade'].mean()
        print(f"Estado '{estado}': {registros} registros, velocidade média: {velocidade_media}")
    
    # DIAGNÓSTICO: Verificar especificamente para os estados que nos interessam
    print(f"Registros para '{estado_carregado}' com velocidade > 0: {len(df[(df['Estado Operacional'] == estado_carregado) & (df['Velocidade'] > 0)])}")
    print(f"Registros para '{estado_vazio}' com velocidade > 0: {len(df[(df['Estado Operacional'] == estado_vazio) & (df['Velocidade'] > 0)])}")
    
    # Aplicar filtro de velocidade > 0 (mínimo necessário)
    df_validos = df[df['Velocidade'] > 0].copy()
    
    # Estatísticas gerais para toda a população
    print("\nESTATÍSTICAS GERAIS DE VELOCIDADE:")
    print(f"Mínimo: {df_validos['Velocidade'].min()}")
    print(f"Máximo: {df_validos['Velocidade'].max()}")
    print(f"Média: {df_validos['Velocidade'].mean()}")
    print(f"Mediana: {df_validos['Velocidade'].median()}")
    print(f"Desvio Padrão: {df_validos['Velocidade'].std()}")
    print(f"Total de registros: {len(df_validos)}")
    
    # Estatísticas para cada estado operacional
    for estado in [estado_carregado, estado_vazio]:
        df_estado = df_validos[df_validos['Estado Operacional'] == estado]
        if not df_estado.empty:
            print(f"\nEstatísticas gerais para {estado}:")
            print(f"Mínimo: {df_estado['Velocidade'].min()}")
            print(f"Máximo: {df_estado['Velocidade'].max()}")
            print(f"Média: {df_estado['Velocidade'].mean()}")
            print(f"Mediana: {df_estado['Velocidade'].median()}")
            print(f"Desvio Padrão: {df_estado['Velocidade'].std()}")
            print(f"Total de registros: {len(df_estado)}")
        else:
            print(f"\nNenhum registro para {estado}")
    
    # Obter combinações únicas de operador e frente
    combinacoes = df[['Operador', 'Frente']].drop_duplicates()
    resultados = []
    
    # Calcular média geral de velocidade por operador e frente
    if not df_validos.empty:
        # Média geral (apenas filtro de velocidade > 0)
        media_geral = df_validos.groupby(['Operador', 'Frente'])['Velocidade'].mean()
        
        # Calcular média de velocidade para deslocamento carregado
        df_carregado = df_validos[df_validos['Estado Operacional'] == estado_carregado]
        if not df_carregado.empty:
            # Estatísticas detalhadas para carregado
            stats_carregado = df_carregado.groupby('Operador')['Velocidade'].agg(['min', 'max', 'mean', 'count'])
            print("\nEstatísticas detalhadas para DESLOCAMENTO CARREGADO:")
            print(stats_carregado)
            
            media_carregado = df_carregado.groupby(['Operador', 'Frente'])['Velocidade'].mean()
            print(f"Média de velocidade carregado calculada para {len(media_carregado)} combinações operador/frente.")
            print(f"Exemplo de registros carregado: {df_carregado.head(3)[['Operador', 'Estado Operacional', 'Velocidade']]}")
        else:
            media_carregado = pd.Series(dtype='float64')
            print("Nenhum registro para cálculo de média carregado.")
        
        # Calcular média de velocidade para deslocamento vazio
        df_vazio = df_validos[df_validos['Estado Operacional'] == estado_vazio]
        if not df_vazio.empty:
            # Estatísticas detalhadas para vazio
            stats_vazio = df_vazio.groupby('Operador')['Velocidade'].agg(['min', 'max', 'mean', 'count'])
            print("\nEstatísticas detalhadas para DESLOCAMENTO VAZIO:")
            print(stats_vazio)
            
            media_vazio = df_vazio.groupby(['Operador', 'Frente'])['Velocidade'].mean()
            print(f"Média de velocidade vazio calculada para {len(media_vazio)} combinações operador/frente.")
            print(f"Exemplo de registros vazio: {df_vazio.head(3)[['Operador', 'Estado Operacional', 'Velocidade']]}")
        else:
            media_vazio = pd.Series(dtype='float64')
            print("Nenhum registro para cálculo de média vazio.")
        
        # Preencher resultados para cada combinação de operador e frente
        for _, comb in combinacoes.iterrows():
            operador = comb['Operador']
            frente = comb['Frente']
            
            # Média geral
            velocidade_geral = 0
            if (operador, frente) in media_geral:
                velocidade_geral = media_geral[(operador, frente)]
            
            # Média carregado
            velocidade_carregado = 0
            if (operador, frente) in media_carregado:
                velocidade_carregado = media_carregado[(operador, frente)]
            
            # Média vazio
            velocidade_vazio = 0
            if (operador, frente) in media_vazio:
                velocidade_vazio = media_vazio[(operador, frente)]
            
            # Definir tipo de deslocamento
            tem_carregado = velocidade_carregado > 0
            tem_vazio = velocidade_vazio > 0
            if tem_carregado and tem_vazio:
                tipo = "Ambos"
            elif tem_carregado:
                tipo = "Apenas Carregado"
            elif tem_vazio:
                tipo = "Apenas Vazio"
            else:
                tipo = "Nenhum"
            
            resultados.append({
                'Operador': operador,
                'Frente': frente,
                'Velocidade Geral': velocidade_geral,
                'Velocidade Carregado': velocidade_carregado,
                'Velocidade Vazio': velocidade_vazio,
                'Tipo Deslocamento': tipo
            })
    
    # Criar DataFrame e ordenar primeiro por frente, depois por velocidade geral (decrescente)
    resultado = pd.DataFrame(resultados)
    resultado = resultado.sort_values(['Frente', 'Velocidade Geral'], ascending=[True, False])
    
    # DIAGNÓSTICO: Verificar resultado final
    print(f"\nVelocidades calculadas para {len(resultado)} combinações operador/frente.")
    print("Operadores com velocidade vazio > 0:", len(resultado[resultado['Velocidade Vazio'] > 0]))
    print("Operadores com velocidade carregado > 0:", len(resultado[resultado['Velocidade Carregado'] > 0]))
    
    # Identificar operadores com apenas um tipo de deslocamento
    apenas_carregado = resultado[(resultado['Velocidade Carregado'] > 0) & (resultado['Velocidade Vazio'] == 0)]
    apenas_vazio = resultado[(resultado['Velocidade Carregado'] == 0) & (resultado['Velocidade Vazio'] > 0)]
    ambos = resultado[(resultado['Velocidade Carregado'] > 0) & (resultado['Velocidade Vazio'] > 0)]
    nenhum = resultado[(resultado['Velocidade Carregado'] == 0) & (resultado['Velocidade Vazio'] == 0)]
    
    print(f"\nDistribuição dos operadores:")
    print(f"Apenas deslocamento carregado: {len(apenas_carregado)} operadores")
    print(f"Apenas deslocamento vazio: {len(apenas_vazio)} operadores")
    print(f"Ambos os deslocamentos: {len(ambos)} operadores")
    print(f"Nenhum deslocamento: {len(nenhum)} operadores")
    
    if len(apenas_carregado) > 0:
        print("\nExemplos de operadores apenas com deslocamento carregado:")
        print(apenas_carregado.head(3)[['Operador', 'Frente', 'Velocidade Carregado', 'Velocidade Vazio']])
        
    if len(apenas_vazio) > 0:
        print("\nExemplos de operadores apenas com deslocamento vazio:")
        print(apenas_vazio.head(3)[['Operador', 'Frente', 'Velocidade Carregado', 'Velocidade Vazio']])
    
    return resultado

def identificar_operadores_duplicados(df, substituicoes=None):
    """
    Identifica operadores que começam com '133' e têm 7 dígitos.
    Verifica se já existe uma substituição no arquivo JSON, caso contrário, registra como ID encontrada.
    
    Args:
        df (DataFrame): DataFrame com os dados dos operadores
        substituicoes (dict): Dicionário com as substituições do arquivo JSON
    
    Returns:
        dict: Dicionário com mapeamento {id_incorreta: id_correta}
        DataFrame: DataFrame com as IDs encontradas para relatório
    """
    if 'Operador' not in df.columns or len(df) == 0:
        return {}, pd.DataFrame(columns=['ID Encontrada', 'Nome', 'Status', 'ID Substituição'])
    
    # Extrair operadores únicos
    operadores = df['Operador'].unique()
    
    # Lista para armazenar as IDs encontradas
    ids_encontradas = []
    mapeamento = {}
    
    for op in operadores:
        if ' - ' in op:
            try:
                id_parte, nome_parte = op.split(' - ', 1)
                # Verificar se a ID começa com 133 e tem 7 dígitos
                if id_parte.startswith('133') and len(id_parte) == 7:
                    # Verificar se existe uma substituição no arquivo JSON
                    if substituicoes and op in substituicoes:
                        status = "Substituição encontrada"
                        id_substituicao = substituicoes[op].split(' - ')[0] if ' - ' in substituicoes[op] else substituicoes[op]
                        mapeamento[op] = substituicoes[op]
                    else:
                        status = "Sem substituição definida"
                        id_substituicao = ""
                    
                    # Adicionar à lista de IDs encontradas, mesmo se for "NAO CADASTRADO"
                    ids_encontradas.append({
                        'ID Encontrada': id_parte,
                        'Nome': nome_parte,
                        'Status': status,
                        'ID Substituição': id_substituicao
                    })
                    
                    print(f"ID encontrada: {id_parte} - {nome_parte}")
            except Exception as e:
                print(f"Erro ao processar operador {op}: {str(e)}")
                continue
    
    print(f"Encontradas {len(ids_encontradas)} IDs começando com 133 e 7 dígitos.")
    for id_enc in ids_encontradas:
        print(f"  - {id_enc['ID Encontrada']} - {id_enc['Nome']} ({id_enc['Status']})")
    
    # Criar o DataFrame e ordenar por ID Encontrada
    df_encontradas = pd.DataFrame(ids_encontradas)
    if not df_encontradas.empty:
        df_encontradas = df_encontradas.sort_values('ID Encontrada')
    
    return mapeamento, df_encontradas

def criar_planilha_tdh(df):
    """
    Cria uma planilha vazia com os equipamentos únicos e suas respectivas frotas e frentes, com coluna para TDH.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame vazio com as colunas: Frota, Frente, TDH
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in df.columns:
        df = df.copy()
        df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Extrair equipamentos únicos e suas frotas e frentes
    equipamentos_unicos = df[['Equipamento', 'Frente']].drop_duplicates()
    
    # Criar DataFrame vazio com as colunas necessárias
    planilha_tdh = pd.DataFrame({
        'Frota': equipamentos_unicos['Equipamento'].astype(str).values,  # Garantir que seja string
        'Frente': equipamentos_unicos['Frente'].values,
        'TDH': [''] * len(equipamentos_unicos)  # Strings vazias em vez de zeros
    })
    
    # Ordenar primeiro por frente, depois por frota
    return planilha_tdh.sort_values(['Frente', 'Frota'], ascending=[True, True])

def criar_planilha_diesel(df):
    """
    Cria uma planilha vazia com os equipamentos únicos e suas respectivas frotas e frentes, com coluna para Diesel.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame vazio com as colunas: Frota, Frente, Diesel
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in df.columns:
        df = df.copy()
        df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Extrair equipamentos únicos e suas frotas e frentes
    equipamentos_unicos = df[['Equipamento', 'Frente']].drop_duplicates()
    
    # Criar DataFrame vazio com as colunas necessárias
    planilha_diesel = pd.DataFrame({
        'Frota': equipamentos_unicos['Equipamento'].astype(str).values,  # Garantir que seja string
        'Frente': equipamentos_unicos['Frente'].values,
        'Diesel': [''] * len(equipamentos_unicos)  # Strings vazias em vez de zeros
    })
    
    # Ordenar primeiro por frente, depois por frota
    return planilha_diesel.sort_values(['Frente', 'Frota'], ascending=[True, True])

def criar_planilha_impureza(df):
    """
    Cria uma planilha vazia com os equipamentos únicos e suas respectivas frotas e frentes, com coluna para Impureza Vegetal.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame vazio com as colunas: Frota, Frente, Impureza
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in df.columns:
        df = df.copy()
        df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Extrair equipamentos únicos e suas frotas e frentes
    equipamentos_unicos = df[['Equipamento', 'Frente']].drop_duplicates()
    
    # Criar DataFrame vazio com as colunas necessárias
    planilha_impureza = pd.DataFrame({
        'Frota': equipamentos_unicos['Equipamento'].astype(str).values,  # Garantir que seja string
        'Frente': equipamentos_unicos['Frente'].values,
        'Impureza': [''] * len(equipamentos_unicos)  # Strings vazias em vez de zeros
    })
    
    # Ordenar primeiro por frente, depois por frota
    return planilha_impureza.sort_values(['Frente', 'Frota'], ascending=[True, True])

def calcular_ofensores(df):
    """
    Calcula os top 5 ofensores por frente.
    Agrupa por Frente e Operacao onde Estado Operacional é 'PARADA',
    soma a Diferença_Hora, classifica do maior para o menor.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame com os ofensores por frente
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in df.columns:
        df = df.copy()
        df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Filtrar apenas os registros com Estado Operacional PARADA
    df_paradas = df[df['Estado Operacional'] == 'PARADA'].copy()
    
    # Se não houver dados de parada, retornar DataFrame vazio
    if len(df_paradas) == 0:
        return pd.DataFrame(columns=['Frente', 'Operação', 'Tempo', 'Porcentagem'])
    
    # Agrupar por Frente e Operacao, somar o tempo
    paradas_agrupadas = df_paradas.groupby(['Frente', 'Operacao'])['Diferença_Hora'].sum().reset_index()
    
    # Calcular o tempo total de todas as paradas por frente
    tempo_total_por_frente = df_paradas.groupby('Frente')['Diferença_Hora'].sum().to_dict()
    
    # Adicionar coluna de porcentagem (porcentagem dentro da frente)
    paradas_agrupadas['Porcentagem'] = paradas_agrupadas.apply(
        lambda row: row['Diferença_Hora'] / tempo_total_por_frente.get(row['Frente'], 1) if tempo_total_por_frente.get(row['Frente'], 0) > 0 else 0,
        axis=1
    )
    
    # Ordenar primeiro por frente, depois por tempo (decrescente)
    paradas_agrupadas = paradas_agrupadas.sort_values(['Frente', 'Diferença_Hora'], ascending=[True, False])
    
    # Renomear colunas para melhor visualização
    resultado = paradas_agrupadas.rename(columns={
        'Operacao': 'Operação',
        'Diferença_Hora': 'Tempo'
    })
    
    return resultado

def criar_excel_com_planilhas(df_base, base_calculo, disp_mecanica, eficiencia_energetica,
                            motor_ocioso, falta_apontamento, uso_gps, horas_por_frota, caminho_saida,
                            df_duplicados=None, media_velocidade=None, df_substituicoes=None):
    """
    Cria um arquivo Excel com todas as planilhas necessárias.
    """
    # Definir função de ajuste de largura de colunas
    def ajustar_largura_colunas(worksheet):
        """Ajusta a largura das colunas da planilha"""
        for col in worksheet.columns:
            max_length = 10
            column = col[0].column_letter
            header_text = str(col[0].value)
            if header_text:
                max_length = max(max_length, len(header_text) + 2)
            for cell in col[1:min(20, len(col))]:
                if cell.value:
                    cell_text = str(cell.value)
                    max_length = max(max_length, len(cell_text) + 2)
            max_length = min(max_length, 40)
            worksheet.column_dimensions[column].width = max_length
    
    # Criar planilhas adicionais
    df_tdh = criar_planilha_tdh(df_base)
    df_diesel = criar_planilha_diesel(df_base)
    df_impureza = criar_planilha_impureza(df_base)
    
    # Calcular ofensores
    df_ofensores = calcular_ofensores(df_base)
    
    # Aplicar frotas aos operadores nos resultados antes de criar o Excel
    eficiencia_energetica = adicionar_frotas_ao_operador(df_base, eficiencia_energetica)
    motor_ocioso = adicionar_frotas_ao_operador(df_base, motor_ocioso)
    falta_apontamento = adicionar_frotas_ao_operador(df_base, falta_apontamento)
    uso_gps = adicionar_frotas_ao_operador(df_base, uso_gps)
    if media_velocidade is not None and not media_velocidade.empty:
        media_velocidade = adicionar_frotas_ao_operador(df_base, media_velocidade)
    df_ofensores = adicionar_frotas_ao_operador(df_base, df_ofensores)
    
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        # Salvar cada DataFrame em uma planilha separada
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        reordenar_colunas_frente_primeiro(base_calculo).to_excel(writer, sheet_name='Base Calculo', index=False)
        reordenar_colunas_frente_primeiro(disp_mecanica).to_excel(writer, sheet_name='1_Disponibilidade Mecânica', index=False)
        reordenar_colunas_frente_primeiro(eficiencia_energetica).to_excel(writer, sheet_name='2_Eficiência Energética', index=False)
        reordenar_colunas_frente_primeiro(motor_ocioso).to_excel(writer, sheet_name='3_Motor Ocioso', index=False)
        reordenar_colunas_frente_primeiro(falta_apontamento).to_excel(writer, sheet_name='4_Falta Apontamento', index=False)
        reordenar_colunas_frente_primeiro(uso_gps).to_excel(writer, sheet_name='5_Uso GPS', index=False)
        reordenar_colunas_frente_primeiro(horas_por_frota).to_excel(writer, sheet_name='Horas por Frota', index=False)
        
        # Adicionar nova planilha de ofensores
        reordenar_colunas_frente_primeiro(df_ofensores).to_excel(writer, sheet_name='Ofensores', index=False)
        
        # Adicionar novas planilhas
        reordenar_colunas_frente_primeiro(df_tdh).to_excel(writer, sheet_name='TDH', index=False)
        reordenar_colunas_frente_primeiro(df_diesel).to_excel(writer, sheet_name='Diesel', index=False)
        reordenar_colunas_frente_primeiro(df_impureza).to_excel(writer, sheet_name='Impureza Vegetal', index=False)
        
        if media_velocidade is None:
            media_velocidade = pd.DataFrame(columns=['Operador', 'Frente', 'Velocidade Geral', 'Velocidade Carregado', 'Velocidade Vazio'])
        reordenar_colunas_frente_primeiro(media_velocidade).to_excel(writer, sheet_name='Média Velocidade', index=False)
        
        # IDs duplicadas e substituídas
        if df_duplicados is not None and not df_duplicados.empty:
            df_duplicados.to_excel(writer, sheet_name='IDs Encontradas', index=False)
        if df_substituicoes is not None and not df_substituicoes.empty:
            df_substituicoes.to_excel(writer, sheet_name='IDs Substituídas', index=False)
        
        # Formatar cada planilha
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            ajustar_largura_colunas(worksheet)
            
            if sheet_name == '1_Disponibilidade Mecânica':
                for row in range(2, worksheet.max_row + 1):
                    # Verificar se "Frente" é a primeira coluna (reorganizada)
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        frota_cell = worksheet.cell(row=row, column=2)  # Coluna B (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        disp_cell = worksheet.cell(row=row, column=3)  # Coluna C (Disponibilidade)
                        disp_cell.number_format = '0.00%'  # Formato percentual
                        gps_cell = worksheet.cell(row=row, column=4)  # Coluna D (GPS)
                        gps_cell.number_format = '0.00%'  # Formato percentual
                    else:
                        frota_cell = worksheet.cell(row=row, column=1)  # Coluna A (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        disp_cell = worksheet.cell(row=row, column=3)  # Coluna C (Disponibilidade)
                        disp_cell.number_format = '0.00%'  # Formato percentual
                        gps_cell = worksheet.cell(row=row, column=4)  # Coluna D (GPS)
                        gps_cell.number_format = '0.00%'  # Formato percentual
            
            elif sheet_name == '2_Eficiência Energética':
                for row in range(2, worksheet.max_row + 1):
                    # Verificar se "Frente" é a primeira coluna (reorganizada)
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        operador_cell = worksheet.cell(row=row, column=2)  # Coluna B (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        efic_cell = worksheet.cell(row=row, column=3)  # Coluna C (Eficiência)
                        efic_cell.number_format = '0.00%'  # Formato percentual
                    else:
                        operador_cell = worksheet.cell(row=row, column=1)  # Coluna A (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        efic_cell = worksheet.cell(row=row, column=3)  # Coluna C (Eficiência)
                        efic_cell.number_format = '0.00%'  # Formato percentual
            
            elif sheet_name == '3_Motor Ocioso':
                for row in range(2, worksheet.max_row + 1):
                    # Verificar se "Frente" é a primeira coluna (reorganizada)
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        operador_cell = worksheet.cell(row=row, column=2)  # Coluna B (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        perc_cell = worksheet.cell(row=row, column=3)  # Coluna C (Porcentagem)
                        perc_cell.number_format = '0.00%'  # Formato percentual
                        tempo_ligado_cell = worksheet.cell(row=row, column=4)  # Coluna D (Tempo Ligado)
                        tempo_ligado_cell.number_format = '0.00'  # Formato decimal
                        tempo_ocioso_cell = worksheet.cell(row=row, column=5)  # Coluna E (Tempo Ocioso)
                        tempo_ocioso_cell.number_format = '0.00'  # Formato decimal
                    else:
                        operador_cell = worksheet.cell(row=row, column=1)  # Coluna A (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        perc_cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                        perc_cell.number_format = '0.00%'  # Formato percentual
                        tempo_ligado_cell = worksheet.cell(row=row, column=3)  # Coluna C (Tempo Ligado)
                        tempo_ligado_cell.number_format = '0.00'  # Formato decimal
                        tempo_ocioso_cell = worksheet.cell(row=row, column=4)  # Coluna D (Tempo Ocioso)
                        tempo_ocioso_cell.number_format = '0.00'  # Formato decimal
            
            elif sheet_name == '4_Falta Apontamento':
                for row in range(2, worksheet.max_row + 1):
                    # Verificar se "Frente" é a primeira coluna (reorganizada)
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        operador_cell = worksheet.cell(row=row, column=2)  # Coluna B (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        perc_cell = worksheet.cell(row=row, column=3)  # Coluna C (Porcentagem)
                        perc_cell.number_format = '0.00%'  # Formato percentual
                    else:
                        operador_cell = worksheet.cell(row=row, column=1)  # Coluna A (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        perc_cell = worksheet.cell(row=row, column=3)  # Coluna C (Porcentagem)
                        perc_cell.number_format = '0.00%'  # Formato percentual
            
            elif sheet_name == '5_Uso GPS':
                for row in range(2, worksheet.max_row + 1):
                    # Verificar se "Frente" é a primeira coluna (reorganizada)
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        operador_cell = worksheet.cell(row=row, column=2)  # Coluna B (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        perc_cell = worksheet.cell(row=row, column=3)  # Coluna C (Porcentagem)
                        perc_cell.number_format = '0.00%'  # Formato percentual
                    else:
                        operador_cell = worksheet.cell(row=row, column=1)  # Coluna A (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        perc_cell = worksheet.cell(row=row, column=3)  # Coluna C (Porcentagem)
                        perc_cell.number_format = '0.00%'  # Formato percentual
            
            elif sheet_name == 'Média Velocidade':
                for row in range(2, worksheet.max_row + 1):
                    # Verificar se "Frente" é a primeira coluna (reorganizada)
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        operador_cell = worksheet.cell(row=row, column=2)  # Coluna B (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        # Demais colunas como decimal (velocidades)
                        for col in range(3, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.number_format = '0.00'  # Formato decimal
                    else:
                        operador_cell = worksheet.cell(row=row, column=1)  # Coluna A (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        # Demais colunas como decimal (velocidades)
                        for col in range(2, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.number_format = '0.00'  # Formato decimal
            
            elif sheet_name == 'Horas por Frota':
                for row in range(2, worksheet.max_row + 1):
                    # Formatar coluna Frota como texto
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        frota_cell = worksheet.cell(row=row, column=2)  # Coluna B (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        # Demais colunas como decimal (colunas de tempo)
                        for col in range(3, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.number_format = '0.00'  # Formato decimal
                    else:
                        frota_cell = worksheet.cell(row=row, column=1)  # Coluna A (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        # Demais colunas como decimal (colunas de tempo)
                        for col in range(2, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.number_format = '0.00'  # Formato decimal
            
            elif sheet_name == 'Ofensores':
                for row in range(2, worksheet.max_row + 1):
                    # Verificar se "Frente" é a primeira coluna (reorganizada)
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        operador_cell = worksheet.cell(row=row, column=2)  # Coluna B (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        # Coluna C (Tempo)
                        cell = worksheet.cell(row=row, column=4)  # Coluna D (Tempo)
                        cell.number_format = '0.00'  # Formato decimal
                        # Coluna D (Porcentagem)
                        cell = worksheet.cell(row=row, column=5)  # Coluna E (Porcentagem)
                        cell.number_format = '0.00%'  # Formato percentual
                    else:
                        operador_cell = worksheet.cell(row=row, column=1)  # Coluna A (Operador)
                        operador_cell.number_format = '@'  # Formato texto
                        # Coluna C (Tempo)
                        cell = worksheet.cell(row=row, column=3)  # Coluna C (Tempo)
                        cell.number_format = '0.00'  # Formato decimal
                        # Coluna D (Porcentagem)
                        cell = worksheet.cell(row=row, column=4)  # Coluna D (Porcentagem)
                        cell.number_format = '0.00%'  # Formato percentual
                        
            elif sheet_name == 'TDH':
                for row in range(2, worksheet.max_row + 1):
                    # Frota (coluna A se "Frente" for primeira, senão coluna B) - formato texto
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        frota_cell = worksheet.cell(row=row, column=2)  # Coluna B (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        tdh_cell = worksheet.cell(row=row, column=3)  # Coluna C (TDH)
                        tdh_cell.number_format = '0.0000'  # 4 casas decimais
                    else:
                        frota_cell = worksheet.cell(row=row, column=1)  # Coluna A (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        tdh_cell = worksheet.cell(row=row, column=3)  # Coluna C (TDH)
                        tdh_cell.number_format = '0.0000'  # 4 casas decimais
            
            elif sheet_name == 'Diesel':
                for row in range(2, worksheet.max_row + 1):
                    # Frota (coluna A se "Frente" for primeira, senão coluna B) - formato texto
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        frota_cell = worksheet.cell(row=row, column=2)  # Coluna B (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        diesel_cell = worksheet.cell(row=row, column=3)  # Coluna C (Diesel)
                        diesel_cell.number_format = '0.0000'  # 4 casas decimais
                    else:
                        frota_cell = worksheet.cell(row=row, column=1)  # Coluna A (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        diesel_cell = worksheet.cell(row=row, column=3)  # Coluna C (Diesel)
                        diesel_cell.number_format = '0.0000'  # 4 casas decimais
            
            elif sheet_name == 'Impureza Vegetal':
                for row in range(2, worksheet.max_row + 1):
                    # Frota (coluna A se "Frente" for primeira, senão coluna B) - formato texto
                    if worksheet.cell(row=1, column=1).value == 'Frente':
                        frota_cell = worksheet.cell(row=row, column=2)  # Coluna B (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        impureza_cell = worksheet.cell(row=row, column=3)  # Coluna C (Impureza)
                        impureza_cell.number_format = '0.00'  # 2 casas decimais
                    else:
                        frota_cell = worksheet.cell(row=row, column=1)  # Coluna A (Frota)
                        frota_cell.number_format = '@'  # Formato texto
                        impureza_cell = worksheet.cell(row=row, column=3)  # Coluna C (Impureza)
                        impureza_cell.number_format = '0.00'  # 2 casas decimais
            
            elif sheet_name == 'Base Calculo':
                colunas_porcentagem = ['% Parado com motor ligado', '% Utilização GPS', '% Falta de Apontamento']
                colunas_tempo = ['Horas totais', 'Motor Ligado', 'Parado com motor ligado', 'GPS', 'Horas Produtivas', 'Falta de Apontamento']
                
                for row in range(2, worksheet.max_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        header = worksheet.cell(row=1, column=col).value
                        cell = worksheet.cell(row=row, column=col)
                        
                        # Verificar se é a coluna Frota ou Equipamento e formatá-la como texto
                        if header in ['Frota', 'Equipamento']:
                            cell.number_format = '@'  # Formato texto
                        elif header in colunas_porcentagem:
                            cell.number_format = '0.00%'
                        elif header in colunas_tempo:
                            cell.number_format = '0.00'

def extrair_arquivo_zip(caminho_zip, pasta_destino=None):
    """
    Extrai o conteúdo de um arquivo ZIP para uma pasta temporária ou destino especificado.
    Renomeia os arquivos extraídos para terem o mesmo nome do arquivo ZIP original.
    
    Args:
        caminho_zip (str): Caminho para o arquivo ZIP
        pasta_destino (str, optional): Pasta onde os arquivos serão extraídos.
                                       Se None, usa uma pasta temporária.
    
    Returns:
        list: Lista de caminhos dos arquivos extraídos e renomeados (apenas TXT e CSV)
        str: Caminho da pasta temporária (se criada) ou None
    """
    # Se pasta_destino não foi especificada, criar uma pasta temporária
    pasta_temp = None
    if pasta_destino is None:
        pasta_temp = tempfile.mkdtemp()
        pasta_destino = pasta_temp
    
    arquivos_extraidos = []
    nome_zip_sem_extensao = os.path.splitext(os.path.basename(caminho_zip))[0]
    
    try:
        with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
            # Extrair todos os arquivos do ZIP
            zip_ref.extractall(pasta_destino)
            
            # Processar cada arquivo extraído (apenas TXT e CSV)
            for arquivo in zip_ref.namelist():
                caminho_completo = os.path.join(pasta_destino, arquivo)
                # Verificar se é um arquivo e não uma pasta
                if os.path.isfile(caminho_completo):
                    # Verificar extensão
                    extensao = os.path.splitext(arquivo)[1].lower()
                    if extensao in ['.txt', '.csv']:
                        # Criar novo nome: nome do ZIP + extensão original
                        novo_nome = f"{nome_zip_sem_extensao}{extensao}"
                        novo_caminho = os.path.join(pasta_destino, novo_nome)
                        
                        # Renomear o arquivo extraído
                        try:
                            # Se já existe um arquivo com esse nome, remover primeiro
                            if os.path.exists(novo_caminho):
                                os.remove(novo_caminho)
                            # Renomear o arquivo
                            os.rename(caminho_completo, novo_caminho)
                            arquivos_extraidos.append(novo_caminho)
                            print(f"Arquivo extraído renomeado: {novo_nome}")
                        except Exception as e:
                            print(f"Erro ao renomear arquivo {arquivo} para {novo_nome}: {str(e)}")
                            arquivos_extraidos.append(caminho_completo)  # Adicionar o caminho original em caso de erro
        
        return arquivos_extraidos, pasta_temp
    
    except Exception as e:
        print(f"Erro ao extrair o arquivo ZIP {caminho_zip}: {str(e)}")
        # Se houve erro e criamos uma pasta temporária, tentar limpá-la
        if pasta_temp:
            try:
                shutil.rmtree(pasta_temp)
            except:
                pass
        return [], None

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT, CSV ou ZIP de transbordos nas pastas dados e dados/transbordos.
    Busca arquivos que começam com "RV Transbordo", "frente" e "transbordos" com extensão .csv, .txt ou .zip.
    Ignora arquivos que contenham "colhedora" no nome.
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_transbordos = os.path.join(diretorio_raiz, "dados", "transbordos")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    
    # Verificar se os diretórios existem, caso contrário criar
    if not os.path.exists(diretorio_dados):
        os.makedirs(diretorio_dados)
    if not os.path.exists(diretorio_transbordos):
        os.makedirs(diretorio_transbordos)
    if not os.path.exists(diretorio_saida):
        os.makedirs(diretorio_saida)
    
    # Lista de diretórios para buscar arquivos
    diretorios_busca = [diretorio_dados, diretorio_transbordos]
    
    # Encontrar todos os arquivos TXT/CSV/ZIP de transbordos em ambos os diretórios
    arquivos = []
    arquivos_zip = []
    
    for diretorio in diretorios_busca:
        # Adicionar arquivos TXT sempre
        arquivos += glob.glob(os.path.join(diretorio, "RV Transbordo*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "*transbordo*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "frente*transbordos*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "transbordo*.txt"))
        
        # Adicionar arquivos CSV apenas se processCsv for True
        if processCsv:
            arquivos += glob.glob(os.path.join(diretorio, "RV Transbordo*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "*transbordo*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "frente*transbordos*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "transbordo*.csv"))
        
        # Adicionar arquivos ZIP
        arquivos_zip += glob.glob(os.path.join(diretorio, "RV Transbordo*.zip"))
        arquivos_zip += glob.glob(os.path.join(diretorio, "*transbordo*.zip"))
        arquivos_zip += glob.glob(os.path.join(diretorio, "frente*transbordos*.zip"))
        arquivos_zip += glob.glob(os.path.join(diretorio, "transbordo*.zip"))
    
    # Filtrar arquivos que contenham "colhedora" no nome (case insensitive)
    arquivos = [arquivo for arquivo in arquivos if "colhedora" not in os.path.basename(arquivo).lower()]
    arquivos_zip = [arquivo for arquivo in arquivos_zip if "colhedora" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos = list(set(arquivos))
    arquivos_zip = list(set(arquivos_zip))
    
    if not arquivos and not arquivos_zip:
        print("Nenhum arquivo de transbordos encontrado nas pastas dados ou dados/transbordos!")
        return
    
    print(f"Encontrados {len(arquivos)} arquivos de transbordos (TXT/CSV) para processar.")
    print(f"Encontrados {len(arquivos_zip)} arquivos ZIP de transbordos para processar.")
    
    # Processar cada arquivo TXT/CSV
    for arquivo in arquivos:
        processar_arquivo(arquivo, diretorio_saida)
    
    # Processar cada arquivo ZIP
    for arquivo_zip in arquivos_zip:
        print(f"\nProcessando arquivo ZIP: {os.path.basename(arquivo_zip)}")
        
        # Extrair arquivo ZIP para pasta temporária
        arquivos_extraidos, pasta_temp = extrair_arquivo_zip(arquivo_zip)
        
        if not arquivos_extraidos:
            print(f"Nenhum arquivo TXT ou CSV encontrado no ZIP {os.path.basename(arquivo_zip)}")
            continue
        
        print(f"Extraídos {len(arquivos_extraidos)} arquivos do ZIP.")
        
        # Processar cada arquivo extraído
        for arquivo_extraido in arquivos_extraidos:
            # Filtrar arquivos que contenham "colhedora" no nome
            if "colhedora" not in os.path.basename(arquivo_extraido).lower():
                processar_arquivo(arquivo_extraido, diretorio_saida)
        
        # Limpar pasta temporária se foi criada
        if pasta_temp:
            try:
                shutil.rmtree(pasta_temp)
                print(f"Pasta temporária removida: {pasta_temp}")
            except Exception as e:
                print(f"Erro ao remover pasta temporária {pasta_temp}: {str(e)}")

def processar_arquivo(caminho_arquivo, diretorio_saida):
    """
    Processa um arquivo de transbordos e gera um arquivo Excel com as métricas calculadas.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo a ser processado
        diretorio_saida (str): Diretório onde o arquivo de saída será salvo
    """
    # Obter apenas o nome do arquivo (sem caminho e sem extensão)
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    
    # Nome de saída igual ao original, mas com sufixo "_min" e extensão .xlsx na pasta output
    arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}-unificado.xlsx")
    
    print(f"\nProcessando arquivo: {os.path.basename(caminho_arquivo)}")
    print(f"Arquivo de saída: {os.path.basename(arquivo_saida)}")
    
    # Processar o arquivo base
    df_base = processar_arquivo_base(caminho_arquivo)
    if df_base is None:
        print(f"Erro ao processar {os.path.basename(caminho_arquivo)}. Pulando para o próximo arquivo.")
        return
    
    # Identificar operadores com IDs que começam com 133 e têm 7 dígitos
    mapeamento_duplicados, df_duplicados = identificar_operadores_duplicados(df_base)
    
    # Carregar substituições de operadores
    substituicoes = carregar_substituicoes_operadores()
    
    # Combinar as substituições manuais com as automáticas
    substituicoes_combinadas = {**substituicoes, **mapeamento_duplicados}
    
    # Aplicar as substituições usando a nova função
    print("\nAplicando substituições de operadores...")
    df_base, df_substituicoes = aplicar_substituicao_operadores(df_base, substituicoes_combinadas)
    
    # Se o DataFrame estiver vazio, gerar apenas a planilha BASE
    if len(df_base) == 0:
        writer = pd.ExcelWriter(arquivo_saida, engine='openpyxl')
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        if not df_duplicados.empty:
            df_duplicados.to_excel(writer, sheet_name='IDs Encontradas', index=False)
        writer.close()
        print(f"Arquivo {arquivo_saida} gerado com apenas a planilha BASE (sem dados).")
        return
    
    # Primeiro, aplicar o novo cálculo de motor ocioso no DataFrame
    # Esta função modifica o df_base adicionando a coluna 'Motor Ocioso' e também retorna um DataFrame formatado para a planilha
    df_base_com_motor_ocioso = calcular_motor_ocioso_para_base(df_base)
    
    # Agora calculamos a Base Calculo
    base_calculo = calcular_base_calculo(df_base_com_motor_ocioso)
    
    # Calcular as métricas auxiliares
    disp_mecanica = calcular_disponibilidade_mecanica(df_base_com_motor_ocioso)
    eficiencia_energetica = calcular_eficiencia_energetica(base_calculo)
    motor_ocioso = calcular_motor_ocioso(base_calculo, df_base_com_motor_ocioso)  # Agora passa a Base Calculo diretamente
    falta_apontamento = calcular_falta_apontamento(base_calculo)
    uso_gps = calcular_uso_gps(base_calculo)
    horas_por_frota = calcular_horas_por_frota(df_base_com_motor_ocioso)
    
    # Calcular média de velocidade por operador
    media_velocidade = calcular_media_velocidade(df_base_com_motor_ocioso)
    
    # Criar o arquivo Excel com todas as planilhas
    criar_excel_com_planilhas(
        df_base_com_motor_ocioso, base_calculo, disp_mecanica, eficiencia_energetica,
        motor_ocioso, falta_apontamento, uso_gps, horas_por_frota, arquivo_saida,
        df_duplicados,  # Adicionar a tabela de IDs duplicadas
        media_velocidade,  # Adicionar a tabela de média de velocidade
        df_substituicoes  # Adicionar a tabela de IDs substituídas
    )
    
    print(f"Arquivo {arquivo_saida} gerado com sucesso!")

def salvar_planilha_base(df, caminho_saida):
    """
    Salva o DataFrame em um arquivo Excel, aplicando formatação adequada.
    
    Args:
        df (DataFrame): DataFrame a ser salvo
        caminho_saida (str): Caminho do arquivo Excel de saída
    """
    try:
        # Criar uma cópia do DataFrame para não modificar o original
        df_copy = df.copy()
        
        # Identificar colunas de tempo
        colunas_tempo = ['Diferença_Hora', 'Horas Produtivas']
        
        # Criar arquivo Excel
        writer = pd.ExcelWriter(caminho_saida, engine='openpyxl')
        
        # Salvar DataFrame
        df_copy.to_excel(writer, index=False)
        
        # Ajustar largura das colunas e aplicar formatação
        worksheet = writer.book.active
        
        # Ajustar largura das colunas
        for idx, col in enumerate(df_copy.columns):
            max_length = max(
                df_copy[col].astype(str).apply(len).max(),
                len(str(col))
            )
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width
            
            # Aplicar formatação de tempo
            if col in colunas_tempo:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=idx + 1)
                    cell.number_format = '0.00'  # Formato decimal
        
        # Salvar arquivo
        writer.close()
        print(f"Arquivo {caminho_saida} gerado com sucesso!")
        
    except Exception as e:
        print(f"Erro ao gerar arquivo {caminho_saida}: {str(e)}")
        print(f"Arquivo {caminho_saida} não foi gerado.")

def calcular_motor_ocioso_para_base(df):
    """
    Calcula o tempo de motor ocioso de acordo com as novas regras e modifica o DataFrame original.
    Esta função MODIFICA o DataFrame passado, adicionando a coluna 'Motor Ocioso'.
    Considera as configurações de exclusão de operações e grupos de operação do arquivo de configuração.
    
    Args:
        df (DataFrame): DataFrame com os dados de operação
        
    Returns:
        DataFrame: DataFrame modificado com a coluna 'Motor Ocioso' atualizada
    """
    # Carregar configurações de cálculos
    config = carregar_config_calculos()
    tipo_equipamento = "TT"  # Para transbordos
    
    # Obter operações e grupos excluídos da configuração
    operacoes_excluidas = []
    grupos_operacao_excluidos = []
    
    if tipo_equipamento in config and "motor_ocioso" in config[tipo_equipamento]:
        operacoes_excluidas = config[tipo_equipamento]["motor_ocioso"].get("operacoes_excluidas", [])
        grupos_operacao_excluidos = config[tipo_equipamento]["motor_ocioso"].get("grupos_operacao_excluidos", [])
    
    print(f"Operações excluídas do cálculo de motor ocioso: {operacoes_excluidas}")
    print(f"Grupos de operação excluídos do cálculo de motor ocioso: {grupos_operacao_excluidos}")
    
    # Criar uma cópia para modificar
    df_resultado = df.copy()
    
    # Converter a coluna de diferença para minutos
    df_resultado['Diferença_Minutos'] = df_resultado['Diferença_Hora'] * 60
    
    # Inicializar colunas com tipos corretos
    df_resultado['Motor Ocioso'] = 0.0  # float
    df_resultado['Em_Intervalo'] = False  # bool
    df_resultado['Soma_Intervalo'] = 0.0  # float
    
    # Filtrar operações e grupos de operação excluídos
    df_filtrado_config = df_resultado.copy()
    if operacoes_excluidas:
        df_filtrado_config = df_filtrado_config[~df_filtrado_config['Operacao'].isin(operacoes_excluidas)]
    if grupos_operacao_excluidos:
        df_filtrado_config = df_filtrado_config[~df_filtrado_config['Grupo Operacao'].isin(grupos_operacao_excluidos)]
    
    print(f"Total de registros antes da filtragem: {len(df_resultado)}")
    print(f"Total de registros após filtragem por operações e grupos excluídos: {len(df_filtrado_config)}")
    
    # Variáveis para controle do intervalo atual
    em_intervalo = False
    soma_intervalo = 0
    inicio_intervalo = None
    
    # Iterar sobre as linhas do DataFrame filtrado
    for i, row in df_filtrado_config.iterrows():
        parada_motor = row['Parado com motor ligado']
        diferenca = row['Diferença_Minutos']
        
        # Se não estamos em um intervalo
        if not em_intervalo:
            # Se encontrar Parado com Motor Ligado = 1, inicia novo intervalo
            if parada_motor == 1:
                em_intervalo = True
                soma_intervalo = diferenca
                inicio_intervalo = i
                df_resultado.at[i, 'Em_Intervalo'] = True
                df_resultado.at[i, 'Soma_Intervalo'] = float(soma_intervalo)  # Converter para float
        
        # Se estamos em um intervalo
        else:
            # Se encontrar Parado com Motor Ligado = 0
            if parada_motor == 0:
                # Se a duração for > 1 minuto, fecha o intervalo
                if diferenca > 1:
                    # Se o total acumulado > 1 minuto, subtrai 1 minuto
                    if soma_intervalo > 1:
                        tempo_ocioso = soma_intervalo - 1
                        # Atribui o tempo ocioso à primeira linha do intervalo
                        # IMPORTANTE: Converter de minutos para horas antes de atribuir
                        df_resultado.at[inicio_intervalo, 'Motor Ocioso'] = float(tempo_ocioso / 60.0)  # Dividir por 60 para converter minutos em horas
                    
                    # Reseta o intervalo
                    em_intervalo = False
                    soma_intervalo = 0
                    inicio_intervalo = None
                else:
                    # Se <= 1 minuto, soma ao intervalo atual
                    soma_intervalo += diferenca
                    df_resultado.at[i, 'Em_Intervalo'] = True
                    df_resultado.at[i, 'Soma_Intervalo'] = float(soma_intervalo)  # Converter para float
            
            # Se encontrar Parado com Motor Ligado = 1
            else:
                soma_intervalo += diferenca
                df_resultado.at[i, 'Em_Intervalo'] = True
                df_resultado.at[i, 'Soma_Intervalo'] = float(soma_intervalo)  # Converter para float
    
    # Tratar último intervalo aberto, se houver
    if em_intervalo and soma_intervalo > 1:
        tempo_ocioso = soma_intervalo - 1
        # Converter de minutos para horas antes de atribuir
        df_resultado.at[inicio_intervalo, 'Motor Ocioso'] = float(tempo_ocioso / 60.0)  # Dividir por 60 para converter minutos em horas
    
    # Garantir que o tempo ocioso nunca seja maior que o tempo ligado para cada registro
    for i in range(len(df_resultado)):
        if df_resultado.iloc[i]['Motor Ocioso'] > 0:
            # Para transbordos, Motor Ligado é 'LIGADO' ou 'DESLIGADO', não 1 ou 0
            motor_ligado = df_resultado.iloc[i]['Motor Ligado'] == 'LIGADO'
            # Se o motor estiver ligado, limitar o tempo ocioso ao tempo ligado
            if motor_ligado:
                tempo_hora = df_resultado.iloc[i]['Diferença_Hora']
                if df_resultado.iloc[i]['Motor Ocioso'] > tempo_hora:
                    df_resultado.at[i, 'Motor Ocioso'] = tempo_hora
            else:
                # Se o motor não estiver ligado, o tempo ocioso deve ser zero
                df_resultado.at[i, 'Motor Ocioso'] = 0.0
    
    # Remover colunas auxiliares
    df_resultado = df_resultado.drop(['Diferença_Minutos', 'Em_Intervalo', 'Soma_Intervalo'], axis=1)
    
    return df_resultado

def calcular_base_calculo(df):
    """
    Calcula a Base Calculo necessária para as métricas.
    
    Args:
        df (DataFrame): DataFrame com os dados processados
        
    Returns:
        DataFrame: Base Calculo com as colunas
            ['Operador', 'Equipamento', 'Grupo Equipamento/Frente', 'Horas totais', 'Motor Ligado', 'Parado com motor ligado',
             '% Parado com motor ligado', 'GPS', '% Utilização GPS', 'Horas Produtivas', 
             'Falta de Apontamento', '% Falta de Apontamento']
    """
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Filtrar operadores excluídos
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)].copy()
    
    # Agrupar por Operador, Equipamento e Grupo Equipamento/Frente
    operadores = df_filtrado['Operador'].unique()
    equipamentos = df_filtrado['Equipamento'].unique()
    
    resultados_base_calculo = []
    
    for operador in operadores:
        for equipamento in equipamentos:
            dados = df_filtrado[(df_filtrado['Operador'] == operador) & 
                               (df_filtrado['Equipamento'] == equipamento)]
            
            # Se não houver dados para este operador e equipamento, pular
            if len(dados) == 0:
                continue
            
            # Obter o valor de Grupo Equipamento/Frente (assumindo que é consistente para cada combinação operador+equipamento)
            grupo_frente = dados['Grupo Equipamento/Frente'].iloc[0] if 'Grupo Equipamento/Frente' in dados.columns else 'Padrão'
            
            # Calcular métricas
            horas_totais = dados['Diferença_Hora'].sum()
            
            # Para transbordos, Motor Ligado é 'LIGADO' ou 'DESLIGADO', não 1 ou 0
            motor_ligado = dados[dados['Motor Ligado'] == 'LIGADO']['Diferença_Hora'].sum()
            parado_motor_ligado = dados['Motor Ocioso'].sum()  # Usando a coluna calculada
            
            # Calcular a porcentagem de parado com motor ligado
            porcentagem_parado = calcular_porcentagem(parado_motor_ligado, motor_ligado) if motor_ligado > 0 else 0
            
            # Horas produtivas
            horas_produtivas = dados['Horas Produtivas'].sum()
            
            # GPS
            horas_gps = dados['GPS'].sum()
            porcentagem_gps = calcular_porcentagem(horas_gps, horas_produtivas) if horas_produtivas > 0 else 0
            
            # Falta de apontamento - CORRIGIDO
            # Agora somamos apenas as horas onde operação é "8340 - FALTA DE APONTAMENTO"
            falta_apontamento = dados[dados['Operacao'] == '8340 - FALTA DE APONTAMENTO']['Diferença_Hora'].sum()
            
            # Calcular a porcentagem de falta de apontamento em relação ao tempo total
            porcentagem_falta = calcular_porcentagem(falta_apontamento, horas_totais) if horas_totais > 0 else 0
            
            # Adicionar aos resultados - INCLUINDO a coluna Grupo Equipamento/Frente
            resultados_base_calculo.append({
                'Operador': operador,
                'Equipamento': equipamento,
                'Grupo Equipamento/Frente': grupo_frente,
                'Horas totais': horas_totais,
                'Motor Ligado': motor_ligado,
                'Parado com motor ligado': parado_motor_ligado,
                '% Parado com motor ligado': porcentagem_parado,
                'GPS': horas_gps,
                '% Utilização GPS': porcentagem_gps,
                'Horas Produtivas': horas_produtivas,
                'Falta de Apontamento': falta_apontamento,
                '% Falta de Apontamento': porcentagem_falta
            })
    
    # Criar DataFrame com os resultados
    return pd.DataFrame(resultados_base_calculo)

def calcular_motor_ocioso(base_calculo, df_base=None):
    """
    Extrai o percentual de motor ocioso por operador e frente da Base Calculo, sem realizar novos cálculos.
    Agrega os dados por operador e frente, calculando a média quando um operador aparece em múltiplas situações.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
        df_base (DataFrame): DataFrame base (não usado mais, mantido para compatibilidade)
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador e frente (agregado)
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in base_calculo.columns:
        base_calculo = base_calculo.copy()
        base_calculo['Frente'] = base_calculo['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Selecionar apenas as colunas relevantes
    df_temp = base_calculo[['Operador', 'Frente', 'Motor Ligado', 'Parado com motor ligado', '% Parado com motor ligado']].copy()
    
    # Agrupar por operador e frente
    agrupado = df_temp.groupby(['Operador', 'Frente']).agg({
        'Motor Ligado': 'sum',
        'Parado com motor ligado': 'sum',
        '% Parado com motor ligado': 'mean'  # Média ponderada do percentual
    }).reset_index()
    
    # Renomear as colunas para o formato esperado no relatório
    agrupado.rename(columns={
        '% Parado com motor ligado': 'Porcentagem',
        'Parado com motor ligado': 'Tempo Ocioso',
        'Motor Ligado': 'Tempo Ligado'
    }, inplace=True)
    
    # Colunas de saída no formato esperado
    resultado = agrupado[['Operador', 'Frente', 'Porcentagem', 'Tempo Ligado', 'Tempo Ocioso']]
    
    print("\n=== DETALHAMENTO DO MOTOR OCIOSO POR FRENTE (EXTRAÍDO DA BASE CALCULO) ===")
    for _, row in resultado.iterrows():
        print(f"\nOperador: {row['Operador']} - Frente: {row['Frente']}")
        print(f"Tempo Ocioso = {row['Tempo Ocioso']:.6f} horas")
        print(f"Tempo Ligado = {row['Tempo Ligado']:.6f} horas")
        print(f"Porcentagem = {row['Porcentagem']:.6f} ({row['Porcentagem']*100:.2f}%)")
        print("-" * 60)
    
    # Ordenar primeiro por frente, depois por porcentagem (decrescente)
    return resultado.sort_values(['Frente', 'Porcentagem'], ascending=[True, False])

def reordenar_colunas_frente_primeiro(df):
    """
    Reordena as colunas do DataFrame colocando 'Frente' como primeira coluna.
    Se não houver coluna 'Frente', retorna o DataFrame inalterado.
    
    Args:
        df (DataFrame): DataFrame a ser reordenado
    
    Returns:
        DataFrame: DataFrame com coluna 'Frente' como primeira
    """
    if 'Frente' not in df.columns:
        return df
    
    # Criar lista de colunas com Frente primeiro
    colunas = ['Frente'] + [col for col in df.columns if col != 'Frente']
    return df[colunas]

def adicionar_frotas_ao_operador(df_base, resultado_df, nome_coluna_operador='Operador'):
    """
    Adiciona as frotas (equipamentos) utilizadas ao nome do operador.
    
    Args:
        df_base (DataFrame): DataFrame base com dados de Operador e Equipamento
        resultado_df (DataFrame): DataFrame com resultados calculados
        nome_coluna_operador (str): Nome da coluna do operador no resultado_df
    
    Returns:
        DataFrame: DataFrame com operadores incluindo frotas utilizadas
    """
    if df_base is None or df_base.empty or resultado_df.empty:
        return resultado_df
    
    # Criar mapeamento operador -> frotas
    mapeamento_frotas = {}
    
    if 'Operador' in df_base.columns and 'Equipamento' in df_base.columns:
        for operador in df_base['Operador'].unique():
            dados_operador = df_base[df_base['Operador'] == operador]
            # Coletar todas as frotas (equipamentos) que este operador utilizou e converter para string
            frotas = sorted([str(eq) for eq in dados_operador['Equipamento'].unique()])
            
            # Montar o nome do operador com as frotas
            if len(frotas) > 0:
                frotas_str = ', '.join(map(str, frotas))
                operador_com_frotas = f"{operador} ({frotas_str})"
            else:
                operador_com_frotas = operador
            
            mapeamento_frotas[operador] = operador_com_frotas
    
    # Aplicar mapeamento ao resultado
    resultado_modificado = resultado_df.copy()
    if nome_coluna_operador in resultado_modificado.columns:
        resultado_modificado[nome_coluna_operador] = resultado_modificado[nome_coluna_operador].map(
            lambda x: mapeamento_frotas.get(x, x)
        )
    
    return resultado_modificado

if __name__ == "__main__":
    print("="*80)
    print("Iniciando processamento de arquivos de transbordos...")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Este script processa arquivos de transbordos e gera planilhas Excel com métricas")
    print("Suporta arquivos TXT, CSV e ZIP")
    print("Ignorando arquivos que contenham 'colhedora' no nome")
    print("="*80)
    processar_todos_arquivos()
    print("\nProcessamento concluído!") 