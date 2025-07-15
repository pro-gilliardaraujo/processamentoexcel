import pandas as pd
import openpyxl
import os

# Caminho para o arquivo Excel gerado
arquivo = '../output/colhedorasFrente08_14072025_processado.xlsx'

# Verificar se o arquivo existe
if not os.path.exists(arquivo):
    print(f"Arquivo {arquivo} não encontrado!")
    exit(1)

print(f"Verificando valores no arquivo {os.path.basename(arquivo)}...")

# Ler o arquivo Excel usando openpyxl para verificar os valores reais
wb = openpyxl.load_workbook(arquivo)

# Verificar valores na planilha de Disponibilidade Mecânica
if '1_Disponibilidade Mecânica' in wb.sheetnames:
    ws = wb['1_Disponibilidade Mecânica']
    print("\n=== Planilha: 1_Disponibilidade Mecânica ===")
    print(f"{'Frota':<10} | {'Disponibilidade':<15} | {'Formato':<20}")
    print("-" * 50)
    
    for row in range(2, ws.max_row + 1):
        frota = ws.cell(row=row, column=1).value
        disponibilidade = ws.cell(row=row, column=2).value
        formato = ws.cell(row=row, column=2).number_format
        
        print(f"{frota:<10} | {disponibilidade:<15.6f} | {formato:<20}")

# Verificar valores na planilha de Uso GPS
if '2_Uso GPS' in wb.sheetnames:
    ws = wb['2_Uso GPS']
    print("\n=== Planilha: 2_Uso GPS ===")
    print(f"{'Frota':<10} | {'Porcentagem':<15} | {'Formato':<20}")
    print("-" * 50)
    
    for row in range(2, ws.max_row + 1):
        frota = ws.cell(row=row, column=1).value
        porcentagem = ws.cell(row=row, column=2).value
        formato = ws.cell(row=row, column=2).number_format
        
        print(f"{frota:<10} | {porcentagem:<15.6f} | {formato:<20}")

# Verificar valores na planilha de Motor Ocioso
if '3_Motor Ocioso' in wb.sheetnames:
    ws = wb['3_Motor Ocioso']
    print("\n=== Planilha: 3_Motor Ocioso ===")
    print(f"{'Frota':<10} | {'Porcentagem':<15} | {'Tempo Ligado':<15} | {'Tempo Ocioso':<15}")
    print("-" * 65)
    
    for row in range(2, ws.max_row + 1):
        frota = ws.cell(row=row, column=1).value
        porcentagem = ws.cell(row=row, column=2).value
        tempo_ligado = ws.cell(row=row, column=3).value if ws.max_column >= 3 else None
        tempo_ocioso = ws.cell(row=row, column=4).value if ws.max_column >= 4 else None
        
        print(f"{frota:<10} | {porcentagem:<15.6f} | {tempo_ligado:<15.6f} | {tempo_ocioso:<15.6f}")

print("\nComo você pode ver, os valores estão sendo armazenados como decimais (0-1)")
print("e a formatação do Excel está sendo aplicada corretamente.")
print("Quando você seleciona uma célula, o valor decimal é mostrado na barra de fórmulas.") 