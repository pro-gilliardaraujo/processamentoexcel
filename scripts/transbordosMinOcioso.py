#!/usr/bin/env python
# -*- coding: utf-8 -*-

# =============================================================================
# CONFIGURAÇÕES GERAIS
# =============================================================================
# Ativar ou desativar processamento de arquivos CSV
processCsv = True

# Configurações de manobras
tempoMinimoManobras = 15  # Tempo mínimo para considerar uma manobra válida (em segundos)
velocidadeMinimaManobras = 1.0  # Velocidade mínima para considerar uma manobra válida (em km/h)

import os
import sys
import glob
import zipfile
import tempfile
import shutil
import pandas as pd
import numpy as np
import json
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

"""
Script para processamento completo de dados de monitoramento de transbordos.
Lê arquivos TXT ou CSV na pasta raiz, processa-os e gera arquivos Excel com planilhas auxiliares prontas.
Também processa arquivos ZIP contendo TXT ou CSV.
"""

# Constantes
COLUNAS_REMOVER = [
    'Latitude',
    'Longitude',
    'Regional',
    'Unidade',
    'Centro de Custo',
    'Fazenda', 
    'Zona', 
    'Talhao'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento', 'Descricao Equipamento', 'Estado', 'Estado Operacional',
    'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro', 'Motor Ligado', 'Operacao', 'Operador',
    'RPM Motor', 'Tipo de Equipamento', 'Velocidade', 'Parado com motor ligado',
    'Diferença_Hora', 'Horas Produtivas', 'GPS', 'Motor Ocioso'
]

# Valores a serem filtrados
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO", "1 - SEM OPERADOR"]

# Adicionar função para extrair frente antes da função processar_arquivo_base()
def extrair_frente(grupo_equipamento_frente):
    """
    Extrai a frente da coluna 'Grupo Equipamento/Frente'.
    
    Args:
        grupo_equipamento_frente (str): Valor da coluna 'Grupo Equipamento/Frente'
    
    Returns:
        str: Nome da frente extraído ou 'Não informado' se não conseguir extrair
    """
    if pd.isna(grupo_equipamento_frente) or grupo_equipamento_frente == '':
        return 'Não informado'
    
    # Tentar extrair a frente (assumindo formato "GRUPO/FRENTE" ou similar)
    try:
        # Se contém "/", pega a parte após a barra
        if '/' in str(grupo_equipamento_frente):
            return str(grupo_equipamento_frente).split('/')[-1].strip()
        # Se contém "-", pega a parte após o traço
        elif '-' in str(grupo_equipamento_frente):
            return str(grupo_equipamento_frente).split('-')[-1].strip()
        # Caso contrário, usa o valor completo
        else:
            return str(grupo_equipamento_frente).strip()
    except:
        return 'Não informado'

def processar_arquivo_base(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV de transbordos e retorna o DataFrame com as transformações necessárias.
    Usando exatamente o mesmo método do Codigo_Base_TT.py para cálculo da Diferença_Hora.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo (TXT ou CSV são tratados da mesma forma se usarem separador ';')
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio (apenas cabeçalhos sem dados)
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                # Retornar o DataFrame vazio mas com as colunas, em vez de None
                # Garantir que todas as colunas desejadas existam
                for col in COLUNAS_DESEJADAS:
                    if col not in df.columns:
                        df[col] = np.nan
                # Reorganizar as colunas na ordem desejada
                colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Padronizar valores da coluna Grupo Operacao
            if 'Grupo Operacao' in df.columns:
                df['Grupo Operacao'] = df['Grupo Operacao'].str.strip()
                # Mapear valores para garantir consistência
                mapa_grupo_operacao = {
                    'SEM APONTAMENTO': 'Sem Apontamento',
                    'PRODUTIVA': 'Produtiva',
                    'MANUTENCAO': 'Manutenção',
                    'MANUTENÇÃO': 'Manutenção'
                }
                df['Grupo Operacao'] = df['Grupo Operacao'].replace(mapa_grupo_operacao)
            
            # Verificar se 'Data/Hora' existe, caso ainda não tenha sido separado
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Remover colunas conforme solicitado no Codigo_Base_TT.py
            colunas_remover = ['Unidade', 'Centro de Custo', 'Fazenda', 'Zona', 'Talhao']
            df = df.drop(columns=colunas_remover, errors='ignore')
            
            # MÉTODO EXATO DO Codigo_Base_TT.py para cálculo da Diferença_Hora para garantir mesmos resultados
            # Conversão de Hora para datetime (apenas se ainda não for)
            if df['Hora'].dtype != 'datetime64[ns]':
                df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular Diferença_Hora sem arredondamentos usando o método EXATO do Codigo_Base_TT.py
            # NOTA: Removida a regra que zerava valores > 0.5, pois não existe no Codigo_Base_TT.py
            # e estava causando perda de aproximadamente 16 horas no total
            df['Diferença_Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S').diff()
            df['Diferença_Hora'] = pd.to_timedelta(df['Diferença_Hora'], errors='coerce')
            df['Diferença_Hora'] = df['Diferença_Hora'].dt.total_seconds() / 3600  # Conversor para horas
            df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: x if x >= 0 else 0)
            
            # Soma total para verificação de precisão (mesma lógica do Codigo_Base_TT.py)
            print(f"Diferença_Hora calculada usando método exato do Codigo_Base_TT.py. Soma total: {df['Diferença_Hora'].sum():.8f} horas")
            
            # Conversão de Motor Ligado para formato conforme Codigo_Base_TT.py
            for col in ['Motor Ligado']:
                if col in df.columns:
                    df[col] = df[col].replace({1: 'LIGADO', 0: 'DESLIGADO'})
            
            # Criar a coluna "Parado com motor ligado" exatamente como no Codigo_Base_TT.py
            df['Parado com motor ligado'] = ((df['Velocidade'] == 0) & (df['Motor Ligado'] == 'LIGADO')).astype(int)
            
            # Verifica se Horas Produtivas já existe, senão calcula usando método do Codigo_Base_TT.py
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                # Calcular horas produtivas sem arredondamento, mantendo a precisão completa
                df['Horas Produtivas'] = df.apply(
                    lambda row: row['Diferença_Hora'] if row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
                # Soma total de horas produtivas para verificação
                print(f"Total de horas produtivas: {df['Horas Produtivas'].sum():.8f}")
            else:
                # Limpa e converte para número
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # IMPORTANTE: Zerar horas produtivas dos operadores excluídos para garantir que não sejam contabilizadas
            df.loc[df['Operador'].isin(OPERADORES_EXCLUIR), 'Horas Produtivas'] = 0
            print(f"Total de horas produtivas após exclusão de operadores: {df['Horas Produtivas'].sum():.8f}")
            
            # Coluna de GPS - Para transbordos, vamos considerar GPS quando houver "RTK (Piloto Automatico)" 
            # e Velocidade > 0 (se a coluna existir)
            if 'RTK (Piloto Automatico)' in df.columns:
                # Primeiro, garantir que RTK seja convertido para valores numéricos (1/0)
                df['RTK (Piloto Automatico)'] = df['RTK (Piloto Automatico)'].apply(lambda x: 
                    1 if (isinstance(x, bool) and x) or 
                         (isinstance(x, (int, float)) and x == 1) or
                         (isinstance(x, str) and str(x).upper().strip() in ['1', 'SIM', 'S', 'VERDADEIRO', 'TRUE', 'LIGADO'])
                    else 0
                )
                
                # Verificar se a coluna GPS já existe e tem valores
                if 'GPS' not in df.columns or df['GPS'].isna().all() or (df['GPS'] == 0).all():
                    print("Calculando coluna GPS com base em RTK (Piloto Automático)...")
                    # Agora aplicar a lógica de GPS usando o RTK normalizado
                    df['GPS'] = df.apply(
                        lambda row: row['Diferença_Hora'] if row['RTK (Piloto Automatico)'] == 1 
                        and row['Velocidade'] > 0 and row['Grupo Operacao'] == 'Produtiva' else 0,
                        axis=1
                    )
                else:
                    print("Coluna GPS já existe com valores. Mantendo os valores originais.")
                
                # Criar coluna alias 'RTK' (0/1) para compatibilidade com cálculos posteriores
                if 'RTK' not in df.columns:
                    df['RTK'] = df['RTK (Piloto Automatico)']
                else:
                    # Se já existe, garantir que seja numérica 0/1
                    df['RTK'] = pd.to_numeric(df['RTK'], errors='coerce').fillna(0).astype(int)
                
                # IMPORTANTE: Zerar GPS dos operadores excluídos para garantir que não sejam contabilizados  
                df.loc[df['Operador'].isin(OPERADORES_EXCLUIR), 'GPS'] = 0
                    
                # Verificar o total de horas com GPS ativo
                total_gps = df['GPS'].sum()
                print(f"Total de horas com GPS ativo após exclusão de operadores: {total_gps:.4f}")
            else:
                # Se não tiver a coluna RTK, criar uma coluna GPS zerada
                df['GPS'] = 0
                print("Coluna RTK (Piloto Automatico) não encontrada. GPS definido como zero.")
            
            # Conversão de colunas binárias para valores numéricos (garantindo que sejam números)
            for col in ['Esteira Ligada', 'Field Cruiser', 'Implemento Ligado']:  # RTK já foi tratado acima
                if col in df.columns and col != 'Motor Ligado':  # Motor Ligado já foi tratado acima
                    # Se a coluna for texto (LIGADO/DESLIGADO), converter para 1/0
                    if df[col].dtype == 'object':
                        df[col] = df[col].replace({'LIGADO': 1, 'DESLIGADO': 0})
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    # Se chegou aqui, todas as tentativas de codificação falharam
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_motor_ocioso_novo(df):
    """
    Extrai os dados de motor ocioso da Base Calculo.
    Esta função não faz cálculos adicionais, apenas formata os dados
    já calculados na Base Calculo para exibição na planilha auxiliar.
    Os dados são agregados por operador calculando a média quando o mesmo 
    operador aparece em diferentes equipamentos/grupos.
    
    Args:
        df (DataFrame): Base Calculo com os dados já processados
        
    Returns:
        DataFrame: DataFrame com as colunas 'Operador', 'Porcentagem', 'Tempo Ligado', 'Tempo Ocioso'
    """
    # Agrupar por operador e calcular as médias
    agrupado = df.groupby('Operador').agg({
        'Motor Ligado': 'mean',  # Média do tempo ligado
        'Parado com motor ligado': 'mean',  # Média do tempo ocioso
        'Equipamento': 'count'  # Conta quantas vezes o operador aparece
    })
    
    resultado_motor_ocioso = []
    
    print("\n=== DETALHAMENTO DO MOTOR OCIOSO (DADOS DA BASE CALCULO) ===")
    
    # Para cada operador na Base Calculo
    for operador, row in agrupado.iterrows():
        tempo_ligado = row['Motor Ligado']
        tempo_ocioso = row['Parado com motor ligado']
        ocorrencias = row['Equipamento']  # Número de vezes que o operador aparece
        
        # Calcular a porcentagem usando os valores médios
        porcentagem = tempo_ocioso / tempo_ligado if tempo_ligado > 0 else 0
        
        print(f"\nOperador: {operador} (média de {ocorrencias} registros)")
        print(f"Tempo Ligado (média): {tempo_ligado:.6f}")
        print(f"Tempo Ocioso (média): {tempo_ocioso:.6f}")
        print(f"Porcentagem: {porcentagem:.6f}")
        print("-" * 50)
        
        resultado_motor_ocioso.append({
            'Operador': operador,
            'Porcentagem': porcentagem,
            'Tempo Ligado': tempo_ligado,
            'Tempo Ocioso': tempo_ocioso
        })
    
    # Retornar DataFrame formatado para a planilha
    return pd.DataFrame(resultado_motor_ocioso)

def carregar_config_calculos():
    """
    Retorna as configurações de cálculos embutidas diretamente no código.
    As configurações são as mesmas definidas no arquivo calculos_config.json.
    """
    # Configurações embutidas diretamente no código
    config = {
        "CD": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "8490 - LAVAGEM",
                    "MANUTENCAO",
                    "LAVAGEM",
                    "INST CONFIG TECNOL EMBARCADAS",
                    "1055 - MANOBRA"
                ],
                "grupos_operacao_excluidos": [
                    "Manutenção", 
                    "Inaptidão"
                ],
                "operadores_excluidos": []
            },
            "equipamentos_excluidos": []
        },
        "TT": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "9016 - ENCH SISTEMA FREIO",
                    "6340 - BASCULANDO  TRANSBORDAGEM",
                    "9024 - DESATOLAMENTO",
                    "MANUTENCAO",
                    "MANUTENÇÃO",
                    "INST CONFIG TECNOL EMBARCADAS",
                    "1055 - MANOBRA"
                ],
                "grupos_operacao_excluidos": [
                    "Manutenção", 
                    "Inaptidão"
                ],
                "operadores_excluidos": []
            },
            "equipamentos_excluidos": []
        }
    }
    
    print("Usando configurações embutidas no código, ignorando o arquivo calculos_config.json")
    return config

def carregar_substituicoes_operadores():
    """
    Carrega o arquivo substituiroperadores.json que contém os mapeamentos 
    de substituição de operadores.
    
    Returns:
        dict: Dicionário com mapeamento {operador_origem: operador_destino}
        ou dicionário vazio se o arquivo não existir ou for inválido
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Caminho para o arquivo de substituição
    arquivo_substituicao = os.path.join(diretorio_raiz, "config", "substituiroperadores.json")
    
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo_substituicao):
        print(f"Arquivo de substituição de operadores não encontrado: {arquivo_substituicao}")
        return {}
    
    try:
        # Carregar o arquivo JSON
        with open(arquivo_substituicao, 'r', encoding='utf-8') as f:
            substituicoes = json.load(f)
        
        # Criar dicionário de substituições
        mapeamento = {item['operador_origem']: item['operador_destino'] for item in substituicoes}
        
        print(f"Carregadas {len(mapeamento)} substituições de operadores.")
        return mapeamento
        
    except Exception as e:
        print(f"Erro ao carregar arquivo de substituição de operadores: {str(e)}")
        return {}

def carregar_substituicoes_operadores_horario():
    """
    Carrega o arquivo substituiroperadores_horario.json que contém os mapeamentos 
    de substituição de operadores com intervalos de horário.
    
    Returns:
        list: Lista de dicionários com mapeamentos 
              {operador_origem, operador_destino, hora_inicio, hora_fim, frota_origem}
        ou lista vazia se o arquivo não existir ou for inválido.
        O campo frota_origem é opcional:
        - Se presente e não vazio, a substituição é aplicada apenas a registros daquela frota específica
        - Se ausente ou vazio, a substituição é aplicada a todos os registros do operador
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Caminho para o arquivo de substituição
    arquivo_substituicao = os.path.join(diretorio_raiz, "config", "substituiroperadores_horario.json")
    
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo_substituicao):
        print(f"Arquivo de substituição de operadores com horário não encontrado: {arquivo_substituicao}")
        return []
    
    try:
        # Carregar o arquivo JSON
        with open(arquivo_substituicao, 'r', encoding='utf-8') as f:
            substituicoes = json.load(f)
        
        # Converter strings de hora para objetos datetime.time
        for item in substituicoes:
            if 'hora_inicio' in item and isinstance(item['hora_inicio'], str):
                hora_str = item['hora_inicio']
                # Adicionar segundos se não estiverem presentes
                if len(hora_str.split(':')) == 2:
                    hora_str += ':00'
                item['hora_inicio_obj'] = datetime.strptime(hora_str, '%H:%M:%S').time()
            
            if 'hora_fim' in item and isinstance(item['hora_fim'], str):
                hora_str = item['hora_fim']
                # Adicionar segundos se não estiverem presentes
                if len(hora_str.split(':')) == 2:
                    hora_str += ':00'
                item['hora_fim_obj'] = datetime.strptime(hora_str, '%H:%M:%S').time()
        
        print(f"Carregadas {len(substituicoes)} substituições de operadores com intervalos de horário.")
        return substituicoes
        
    except Exception as e:
        print(f"Erro ao carregar arquivo de substituição de operadores com horário: {str(e)}")
        return []

def aplicar_substituicao_operadores(df, mapeamento_substituicoes, mapeamento_horario=None):
    """
    Aplica as substituições de operadores no DataFrame.
    
    Args:
        df (DataFrame): DataFrame a ser processado
        mapeamento_substituicoes (dict): Dicionário com mapeamento {operador_origem: operador_destino}
        mapeamento_horario (list, optional): Lista de dicionários com mapeamentos
            {operador_origem, operador_destino, hora_inicio, hora_fim, frota_origem}.
            O campo frota_origem é opcional:
            - Se presente e não vazio, a substituição é aplicada apenas a registros daquela frota específica
            - Se ausente ou vazio, a substituição é aplicada a todos os registros do operador
    
    Returns:
        tuple: (DataFrame com substituições aplicadas, DataFrame com registro das substituições)
    """
    if (not mapeamento_substituicoes and not mapeamento_horario) or 'Operador' not in df.columns:
        return df, pd.DataFrame(columns=['ID Original', 'Nome Original', 'ID Nova', 'Nome Novo', 'Registros Afetados'])
    
    # Criar uma cópia para não alterar o DataFrame original
    df_modificado = df.copy()
    
    # Lista para armazenar as substituições realizadas
    substituicoes_realizadas = []
    total_registros_substituidos = 0
    
    # Verificar operadores antes da substituição para relatório
    operadores_antes = df_modificado['Operador'].unique()
    print(f"\nOperadores antes da substituição: {len(operadores_antes)}")
    
    # Aplicar as substituições por horário se disponíveis e se o DataFrame tiver coluna de data/hora
    if mapeamento_horario and 'Data' in df_modificado.columns and 'Hora' in df_modificado.columns:
        # Criar uma cópia backup dos operadores originais
        df_modificado['Operador_Original'] = df_modificado['Operador'].copy()
        
        # Para cada linha no DataFrame
        for idx, row in df_modificado.iterrows():
            # Tenta extrair a hora do registro
            try:
                if isinstance(row['Hora'], str):
                    hora_registro = datetime.strptime(row['Hora'], '%H:%M:%S').time()
                elif isinstance(row['Hora'], datetime.time):
                    hora_registro = row['Hora']
                else:
                    continue  # Pula se não conseguir converter
                
                operador = row['Operador']
                
                # Verificar todas as regras de substituição por horário
                for regra in mapeamento_horario:
                    # Verificar se a condição básica é atendida: operador de origem correto e horário dentro do intervalo
                    condicao_basica = (operador == regra['operador_origem'] and 
                                      regra['hora_inicio_obj'] <= hora_registro <= regra['hora_fim_obj'])
                    
                    # Verificar condição de frota se ela existir na regra
                    condicao_frota = True
                    if 'frota_origem' in regra and regra['frota_origem'] and 'Equipamento' in df_modificado.columns:
                        condicao_frota = (row['Equipamento'] == regra['frota_origem'])
                    # Caso contrário, aplica a todos os registros do operador
                    
                    # Aplicar substituição apenas se ambas condições forem atendidas
                    if condicao_basica and condicao_frota:
                        # Aplicar a substituição
                        df_modificado.at[idx, 'Operador'] = regra['operador_destino']
                        break
            except Exception as e:
                print(f"Erro ao processar substituição com horário para linha {idx}: {str(e)}")
        
        # Contar substituições realizadas por horário
        substituicoes_contagem = {}
        for idx, row in df_modificado[df_modificado['Operador'] != df_modificado['Operador_Original']].iterrows():
            origem = row['Operador_Original']
            destino = row['Operador']
            chave = (origem, destino)
            if chave in substituicoes_contagem:
                substituicoes_contagem[chave] += 1
            else:
                substituicoes_contagem[chave] = 1
        
        # Adicionar as substituições com horário à lista de substituições realizadas
        for (origem, destino), count in substituicoes_contagem.items():
            total_registros_substituidos += count
            id_original = origem.split(' - ')[0] if ' - ' in origem else origem
            nome_original = origem.split(' - ')[1] if ' - ' in origem else ''
            id_nova = destino.split(' - ')[0] if ' - ' in destino else destino
            nome_novo = destino.split(' - ')[1] if ' - ' in destino else ''
            
            substituicoes_realizadas.append({
                'ID Original': id_original,
                'Nome Original': nome_original,
                'ID Nova': id_nova,
                'Nome Novo': nome_novo,
                'Registros Afetados': count,
                'Por Horário': True
            })
            print(f"Operador '{origem}' substituído por '{destino}' em {count} registros (por horário)")
        
        # Remover a coluna temporária
        df_modificado.drop('Operador_Original', axis=1, inplace=True)
    
    # Contar operadores e registros antes da substituição padrão
    contagem_antes = df_modificado['Operador'].value_counts().to_dict()
    
    # Aplicar as substituições padrão (sem horário)
    for origem, destino in mapeamento_substituicoes.items():
        # Verificar se o operador de origem existe no DataFrame
        registros_afetados = df_modificado[df_modificado['Operador'] == origem].shape[0]
        
        if registros_afetados > 0:
            # Substituir o operador
            df_modificado.loc[df_modificado['Operador'] == origem, 'Operador'] = destino
            
            total_registros_substituidos += registros_afetados
            
            # Extrair IDs e nomes
            id_original = origem.split(' - ')[0] if ' - ' in origem else origem
            nome_original = origem.split(' - ')[1] if ' - ' in origem else ''
            id_nova = destino.split(' - ')[0] if ' - ' in destino else destino
            nome_novo = destino.split(' - ')[1] if ' - ' in destino else ''
            
            substituicoes_realizadas.append({
                'ID Original': id_original,
                'Nome Original': nome_original,
                'ID Nova': id_nova, 
                'Nome Novo': nome_novo,
                'Registros Afetados': registros_afetados,
                'Por Horário': False
            })
            print(f"Operador '{origem}' substituído por '{destino}' em {registros_afetados} registros")
    
    # Verificar operadores após substituição
    operadores_depois = df_modificado['Operador'].unique()
    print(f"Operadores após substituição: {len(operadores_depois)}")
    print(f"Total de registros substituídos: {total_registros_substituidos}")
    
    # Criar DataFrame com as substituições realizadas
    df_substituicoes = pd.DataFrame(substituicoes_realizadas)
    
    return df_modificado, df_substituicoes

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento.
    Fórmula: (Total Geral - Manutenção) / Total Geral
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento
    """
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        
        # CORREÇÃO: Usar soma total direta, não média diária
        # Calcular Total Geral (soma de todas as diferenças de hora)
        total_geral = dados_equip['Diferença_Hora'].sum()
        
        # Calcular horas de manutenção (Grupo Operacao = 'Manutenção')
        horas_manutencao = dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum()
        
        # CORREÇÃO: Fórmula exata como no Excel: (Total Geral - Manutenção) / Total Geral
        # A disponibilidade mecânica é: (Total - Manutenção) / Total
        if total_geral > 0:
            disp_mecanica = (total_geral - horas_manutencao) / total_geral
        else:
            disp_mecanica = 0.0
        
        # Debug: mostrar valores para verificação
        print(f"Equipamento: {equipamento}")
        print(f"  Total Geral: {total_geral:.6f}")
        print(f"  Manutenção: {horas_manutencao:.6f}")
        print(f"  Disponibilidade: {disp_mecanica:.6f} ({disp_mecanica*100:.2f}%)")
        print(f"  Fórmula: ({total_geral:.6f} - {horas_manutencao:.6f}) / {total_geral:.6f} = {disp_mecanica:.6f}")
        
        resultados.append({
            'Frota': equipamento,
            'Disponibilidade': disp_mecanica
        })
    
    return pd.DataFrame(resultados)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Calcula médias diárias considerando os dias efetivos de cada frota.
    Esta função NÃO aplica qualquer filtro de operador.
    Também identifica as faltas de horário por dia específico.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota com detalhamento por dia
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    # Obter todos os dias únicos no dataset
    dias_unicos = sorted(df['Data'].unique()) if 'Data' in df.columns else []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        
        # Determinar número de dias efetivos para este equipamento
        dias_equip = dados_equip['Data'].nunique() if 'Data' in dados_equip.columns else 1
        
        total_horas = dados_equip['Diferença_Hora'].sum()
        
        # Se houver múltiplos dias, usar média diária
        if dias_equip > 1:
            total_horas = total_horas / dias_equip
        
        # Calcular a diferença para 24 horas
        diferenca_24h = max(24 - total_horas, 0)
        
        # Criar o resultado básico (colunas originais mantidas)
        resultado = {
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        }
        
        # Adicionar detalhamento por dia (novas colunas)
        if len(dias_unicos) > 0:
            for dia in dias_unicos:
                dados_dia = dados_equip[dados_equip['Data'] == dia]
                
                # Se não houver dados para este dia e equipamento, a diferença é 24h
                if len(dados_dia) == 0:
                    resultado[f'Falta {dia}'] = 24.0
                    continue
                
                # Calcular horas registradas neste dia
                horas_dia = dados_dia['Diferença_Hora'].sum()
                
                # Calcular a diferença para 24 horas neste dia
                diferenca_dia = max(24 - horas_dia, 0)
                
                # Adicionar ao resultado apenas se houver falta (diferença > 0)
                if diferenca_dia > 0:
                    resultado[f'Falta {dia}'] = diferenca_dia
                else:
                    resultado[f'Falta {dia}'] = 0.0
        
        resultados.append(resultado)
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica(base_calculo):
    """
    Extrai e agrega a eficiência energética por operador da tabela Base Calculo.
    Não realiza novos cálculos, apenas agrupa os valores já calculados por operador.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador (agregado)
    """
    # Verificar se as colunas necessárias existem
    colunas_necessarias = ['Operador', 'Horas Produtivas', 'Horas totais']
    if not all(coluna in base_calculo.columns for coluna in colunas_necessarias):
        print(f"AVISO: Algumas colunas necessárias para o cálculo de eficiência energética não foram encontradas.")
        print(f"Colunas disponíveis: {base_calculo.columns.tolist()}")
        print(f"Colunas necessárias: {colunas_necessarias}")
        # Retornar DataFrame vazio com as colunas esperadas
        return pd.DataFrame(columns=['Operador', 'Eficiência'])
    
    # Selecionar apenas as colunas relevantes
    df_temp = base_calculo[colunas_necessarias].copy()
    
    # Determinar nome da coluna de frota na base (Equipamento ou Frota)
    col_frota = 'Equipamento' if 'Equipamento' in base_calculo.columns else 'Frota'
    
    # Agrupar por operador e calcular a soma
    agrupado = df_temp.groupby('Operador').sum().reset_index()
    
    # Calcular eficiência a partir dos valores agrupados
    agrupado['Eficiência'] = agrupado.apply(
        lambda row: row['Horas Produtivas'] / row['Horas totais'] if row['Horas totais'] > 0 else 0,
        axis=1
    )
    
    resultados = []
    for _, row in agrupado.iterrows():
        operador = row['Operador']
        frotas = sorted(base_calculo[base_calculo['Operador'] == operador][col_frota].astype(str).unique())
        operador_nome = f"{operador} ({', '.join(frotas)})" if frotas else operador
        resultados.append({'Operador': operador_nome, 'Eficiência': row['Eficiência']})
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso(base_calculo, df_base=None):
    """
    Extrai o percentual de motor ocioso por operador da Base Calculo, sem realizar novos cálculos.
    Agrega os dados por operador, calculando a média quando um operador aparece em múltiplas frotas.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
        df_base (DataFrame): DataFrame base (não usado mais, mantido para compatibilidade)
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador (agregado)
    """
    # Verificar se as colunas necessárias existem
    colunas_necessarias = ['Operador', 'Motor Ligado', 'Parado com motor ligado', '% Parado com motor ligado']
    if not all(coluna in base_calculo.columns for coluna in colunas_necessarias):
        print(f"AVISO: Algumas colunas necessárias para o cálculo de motor ocioso não foram encontradas.")
        print(f"Colunas disponíveis: {base_calculo.columns.tolist()}")
        print(f"Colunas necessárias: {colunas_necessarias}")
        # Retornar DataFrame vazio com as colunas esperadas
        return pd.DataFrame(columns=['Operador', 'Porcentagem', 'Tempo Ligado', 'Tempo Ocioso'])
    
    # Selecionar apenas as colunas relevantes
    df_temp = base_calculo[colunas_necessarias].copy()
    
    col_frota = 'Equipamento' if 'Equipamento' in base_calculo.columns else 'Frota'
    
    # Agrupar por operador
    agrupado = df_temp.groupby('Operador').agg({
        'Motor Ligado': 'sum',
        'Parado com motor ligado': 'sum',
        '% Parado com motor ligado': 'mean'  # Média ponderada do percentual
    }).reset_index()
    
    # Renomear as colunas para o formato esperado no relatório
    agrupado.rename(columns={
        '% Parado com motor ligado': 'Porcentagem',
        'Parado com motor ligado': 'Tempo Ocioso',
        'Motor Ligado': 'Tempo Ligado'
    }, inplace=True)
    
    resultado = []
    for _, row in agrupado.iterrows():
        operador = row['Operador']
        frotas = sorted(base_calculo[base_calculo['Operador'] == operador][col_frota].astype(str).unique())
        operador_nome = f"{operador} ({', '.join(frotas)})" if frotas else operador
        resultado.append({
            'Operador': operador_nome,
            'Porcentagem': row['Porcentagem'],
            'Tempo Ligado': row['Tempo Ligado'],
            'Tempo Ocioso': row['Tempo Ocioso']
        })
    
    resultado = pd.DataFrame(resultado)
    
    print("\n=== DETALHAMENTO DO MOTOR OCIOSO (EXTRAÍDO DA BASE CALCULO) ===")
    for _, row in resultado.iterrows():
        print(f"\nOperador: {row['Operador']}")
        print(f"Tempo Ocioso = {row['Tempo Ocioso']:.6f} horas")
        print(f"Tempo Ligado = {row['Tempo Ligado']:.6f} horas")
        print(f"Porcentagem = {row['Porcentagem']:.6f} ({row['Porcentagem']*100:.2f}%)")
        print("-" * 60)
    
    return resultado

def calcular_falta_apontamento(base_calculo):
    """
    Extrai o percentual de falta de apontamento por operador da Base Calculo, sem realizar novos cálculos.
    Agrega os dados por operador, calculando a média quando um operador aparece em múltiplas frotas.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de falta de apontamento por operador (agregado)
    """
    # Verificar se as colunas necessárias existem
    colunas_necessarias = ['Operador', '% Falta de Apontamento']
    if not all(coluna in base_calculo.columns for coluna in colunas_necessarias):
        print(f"AVISO: Algumas colunas necessárias para o cálculo de falta de apontamento não foram encontradas.")
        print(f"Colunas disponíveis: {base_calculo.columns.tolist()}")
        print(f"Colunas necessárias: {colunas_necessarias}")
        # Retornar DataFrame vazio com as colunas esperadas
        return pd.DataFrame(columns=['Operador', 'Porcentagem'])
    
    # Selecionar apenas as colunas relevantes
    df_temp = base_calculo[colunas_necessarias].copy()
    
    col_frota = 'Equipamento' if 'Equipamento' in base_calculo.columns else 'Frota'
    
    # Agrupar por operador e calcular a média
    agrupado = df_temp.groupby('Operador')['% Falta de Apontamento'].mean().reset_index()
    
    # Garantir coluna 'Porcentagem'
    agrupado.rename(columns={'% Falta de Apontamento': 'Porcentagem'}, inplace=True)
    
    # Adicionar sufixo de frotas
    resultados = []
    for _, row in agrupado.iterrows():
        op = row['Operador']
        frotas = sorted(base_calculo[base_calculo['Operador'] == op][col_frota].astype(str).unique())
        op_nome = f"{op} ({', '.join(frotas)})" if frotas else op
        resultados.append({'Operador': op_nome, 'Porcentagem': row['Porcentagem']})

    resultado_df = pd.DataFrame(resultados)
    print("\n=== DETALHAMENTO DE FALTA DE APONTAMENTO (EXTRAÍDO DA BASE CALCULO) ===")
    for _, row in resultado_df.iterrows():
        print(f"Operador: {row['Operador']}, Porcentagem: {row['Porcentagem']:.6f}")
    print("-" * 60)
    return resultado_df

def calcular_uso_gps(base_calculo):
    """
    Extrai o percentual de uso de GPS por operador da Base Calculo, sem realizar novos cálculos.
    Agrega os dados por operador, calculando a média quando um operador aparece em múltiplas frotas.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador (agregado)
    """
    # Verificar se as colunas necessárias existem
    colunas_necessarias = ['Operador', '% Utilização GPS']
    if not all(coluna in base_calculo.columns for coluna in colunas_necessarias):
        print(f"AVISO: Algumas colunas necessárias para o cálculo de uso de GPS não foram encontradas.")
        print(f"Colunas disponíveis: {base_calculo.columns.tolist()}")
        print(f"Colunas necessárias: {colunas_necessarias}")
        # Retornar DataFrame vazio com as colunas esperadas
        return pd.DataFrame(columns=['Operador', 'Porcentagem'])
    
    # Selecionar apenas as colunas relevantes
    df_temp = base_calculo[colunas_necessarias].copy()
    
    col_frota = 'Equipamento' if 'Equipamento' in base_calculo.columns else 'Frota'
    
    # Agrupar por operador e calcular a média ponderada
    agrupado = df_temp.groupby('Operador')['% Utilização GPS'].mean().reset_index()
    
    # Garantir coluna 'Porcentagem'
    agrupado.rename(columns={'% Utilização GPS': 'Porcentagem'}, inplace=True)
    
    # Adicionar sufixo de frotas
    resultados = []
    for _, row in agrupado.iterrows():
        op = row['Operador']
        frotas = sorted(base_calculo[base_calculo['Operador'] == op][col_frota].astype(str).unique())
        op_nome = f"{op} ({', '.join(frotas)})" if frotas else op
        resultados.append({'Operador': op_nome, 'Porcentagem': row['Porcentagem']})

    resultado_df = pd.DataFrame(resultados)
    print("\n=== DETALHAMENTO DE UTILIZAÇÃO DE GPS (EXTRAÍDO DA BASE CALCULO) ===")
    for _, row in resultado_df.iterrows():
        print(f"Operador: {row['Operador']}, Porcentagem: {row['Porcentagem']:.6f}")
    print("-" * 60)
    return resultado_df

def calcular_media_velocidade(df):
    """
    Calcula a média de velocidade para cada operador, separando por tipo de deslocamento:
    - Deslocamento Carregado (Estado Operacional = "DESLOCAMENTO CARREGADO")
    - Deslocamento Vazio (Estado Operacional = "DESLOCAMENTO VAZIO")
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame com a média de velocidade por operador e tipo de deslocamento
    """
    # Filtrar operadores excluídos
    df = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # DIAGNÓSTICO: Verificar dados antes da filtragem
    print(f"Total de registros antes da filtragem: {len(df)}")
    
    # Verificar se as colunas necessárias existem
    colunas_necessarias = ['Estado Operacional', 'Velocidade']
    for coluna in colunas_necessarias:
        if coluna not in df.columns:
            print(f"ERRO: Coluna '{coluna}' não encontrada no DataFrame!")
            return pd.DataFrame(columns=['Operador', 'Velocidade Geral', 'Velocidade Carregado', 'Velocidade Vazio'])
    
    # Identificar registros com velocidade > 0 (sem filtros adicionais inicialmente)
    registros_velocidade = df['Velocidade'] > 0
    
    # DIAGNÓSTICO: Mostrar estatísticas de velocidade
    print(f"Registros com velocidade > 0: {registros_velocidade.sum()}")
    print(f"Média de velocidade geral: {df[registros_velocidade]['Velocidade'].mean()}")
    
    # DIAGNÓSTICO: Verificar estados operacionais únicos disponíveis
    estados_unicos = df['Estado Operacional'].unique()
    print(f"Estados operacionais únicos disponíveis: {estados_unicos}")
    
    # Estados operacionais específicos - tratando possíveis variações
    estado_carregado = "DESLOCAMENTO CARREGADO"
    estado_vazio = "DESLOCAMENTO VAZIO"
    
    # DIAGNÓSTICO: Verificar quantos registros existem para cada estado operacional
    # antes de aplicar outros filtros
    for estado in estados_unicos:
        registros = len(df[df['Estado Operacional'] == estado])
        velocidade_media = df[df['Estado Operacional'] == estado]['Velocidade'].mean()
        print(f"Estado '{estado}': {registros} registros, velocidade média: {velocidade_media}")
    
    # DIAGNÓSTICO: Verificar especificamente para os estados que nos interessam
    print(f"Registros para '{estado_carregado}' com velocidade > 0: {len(df[(df['Estado Operacional'] == estado_carregado) & (df['Velocidade'] > 0)])}")
    print(f"Registros para '{estado_vazio}' com velocidade > 0: {len(df[(df['Estado Operacional'] == estado_vazio) & (df['Velocidade'] > 0)])}")
    
    # Aplicar filtro de velocidade > 0 (mínimo necessário)
    df_validos = df[df['Velocidade'] > 0].copy()
    
    # Estatísticas gerais para toda a população
    print("\nESTATÍSTICAS GERAIS DE VELOCIDADE:")
    print(f"Mínimo: {df_validos['Velocidade'].min()}")
    print(f"Máximo: {df_validos['Velocidade'].max()}")
    print(f"Média: {df_validos['Velocidade'].mean()}")
    print(f"Mediana: {df_validos['Velocidade'].median()}")
    print(f"Desvio Padrão: {df_validos['Velocidade'].std()}")
    print(f"Total de registros: {len(df_validos)}")
    
    # Estatísticas para cada estado operacional
    for estado in [estado_carregado, estado_vazio]:
        df_estado = df_validos[df_validos['Estado Operacional'] == estado]
        if not df_estado.empty:
            print(f"\nEstatísticas gerais para {estado}:")
            print(f"Mínimo: {df_estado['Velocidade'].min()}")
            print(f"Máximo: {df_estado['Velocidade'].max()}")
            print(f"Média: {df_estado['Velocidade'].mean()}")
            print(f"Mediana: {df_estado['Velocidade'].median()}")
            print(f"Desvio Padrão: {df_estado['Velocidade'].std()}")
            print(f"Total de registros: {len(df_estado)}")
        else:
            print(f"\nNenhum registro para {estado}")
    
    # Inicializar DataFrame para resultado
    todos_operadores = df['Operador'].unique()
    resultado = pd.DataFrame({'Operador': todos_operadores})
    resultado['Velocidade Geral'] = 0
    resultado['Velocidade Carregado'] = 0
    resultado['Velocidade Vazio'] = 0
    resultado['Tipo Deslocamento'] = ''  # Nova coluna para indicar o tipo de deslocamento
    
    # Calcular média geral de velocidade por operador
    if not df_validos.empty:
        # Média geral (apenas filtro de velocidade > 0)
        media_geral = df_validos.groupby('Operador')['Velocidade'].mean()
        
        # Calcular média de velocidade para deslocamento carregado
        df_carregado = df_validos[df_validos['Estado Operacional'] == estado_carregado]
        if not df_carregado.empty:
            # Estatísticas detalhadas para carregado
            stats_carregado = df_carregado.groupby('Operador')['Velocidade'].agg(['min', 'max', 'mean', 'count'])
            print("\nEstatísticas detalhadas para DESLOCAMENTO CARREGADO:")
            print(stats_carregado)
            
            media_carregado = stats_carregado['mean']
            print(f"Média de velocidade carregado calculada para {len(media_carregado)} operadores.")
            print(f"Exemplo de registros carregado: {df_carregado.head(3)[['Operador', 'Estado Operacional', 'Velocidade']]}")
        else:
            media_carregado = pd.Series(dtype='float64')
            print("Nenhum registro para cálculo de média carregado.")
        
        # Calcular média de velocidade para deslocamento vazio
        df_vazio = df_validos[df_validos['Estado Operacional'] == estado_vazio]
        if not df_vazio.empty:
            # Estatísticas detalhadas para vazio
            stats_vazio = df_vazio.groupby('Operador')['Velocidade'].agg(['min', 'max', 'mean', 'count'])
            print("\nEstatísticas detalhadas para DESLOCAMENTO VAZIO:")
            print(stats_vazio)
            
            media_vazio = stats_vazio['mean']
            print(f"Média de velocidade vazio calculada para {len(media_vazio)} operadores.")
            print(f"Exemplo de registros vazio: {df_vazio.head(3)[['Operador', 'Estado Operacional', 'Velocidade']]}")
        else:
            media_vazio = pd.Series(dtype='float64')
            print("Nenhum registro para cálculo de média vazio.")
        
        # Preencher resultados para cada operador
        for operador in todos_operadores:
            # Média geral
            if operador in media_geral:
                resultado.loc[resultado['Operador'] == operador, 'Velocidade Geral'] = media_geral[operador]
            
            # Média carregado
            tem_carregado = False
            if operador in media_carregado:
                resultado.loc[resultado['Operador'] == operador, 'Velocidade Carregado'] = media_carregado[operador]
                tem_carregado = media_carregado[operador] > 0
            
            # Média vazio
            tem_vazio = False
            if operador in media_vazio:
                resultado.loc[resultado['Operador'] == operador, 'Velocidade Vazio'] = media_vazio[operador]
                tem_vazio = media_vazio[operador] > 0
            
            # Definir tipo de deslocamento
            if tem_carregado and tem_vazio:
                tipo = "Ambos"
            elif tem_carregado:
                tipo = "Apenas Carregado"
            elif tem_vazio:
                tipo = "Apenas Vazio"
            else:
                tipo = "Nenhum"
                
            resultado.loc[resultado['Operador'] == operador, 'Tipo Deslocamento'] = tipo
    
    # Ordenar por operador
    resultado = resultado.sort_values('Operador')
    
    # Adicionar sufixo de frotas ao nome do operador
    def _op_frotas_mv(op):
        frotas = sorted(df[df['Operador'] == op]['Equipamento'].astype(str).unique())
        return f"{op} ({', '.join(frotas)})" if frotas else op
    resultado['Operador'] = resultado['Operador'].apply(_op_frotas_mv)
    
    # DIAGNÓSTICO: Verificar resultado final
    print(f"\nVelocidades calculadas para {len(resultado)} operadores.")
    print("Operadores com velocidade vazio > 0:", len(resultado[resultado['Velocidade Vazio'] > 0]))
    print("Operadores com velocidade carregado > 0:", len(resultado[resultado['Velocidade Carregado'] > 0]))
    
    # Identificar operadores com apenas um tipo de deslocamento
    apenas_carregado = resultado[(resultado['Velocidade Carregado'] > 0) & (resultado['Velocidade Vazio'] == 0)]
    apenas_vazio = resultado[(resultado['Velocidade Carregado'] == 0) & (resultado['Velocidade Vazio'] > 0)]
    ambos = resultado[(resultado['Velocidade Carregado'] > 0) & (resultado['Velocidade Vazio'] > 0)]
    nenhum = resultado[(resultado['Velocidade Carregado'] == 0) & (resultado['Velocidade Vazio'] == 0)]
    
    print(f"\nDistribuição dos operadores:")
    print(f"Apenas deslocamento carregado: {len(apenas_carregado)} operadores")
    print(f"Apenas deslocamento vazio: {len(apenas_vazio)} operadores")
    print(f"Ambos os deslocamentos: {len(ambos)} operadores")
    print(f"Nenhum deslocamento: {len(nenhum)} operadores")
    
    if len(apenas_carregado) > 0:
        print("\nExemplos de operadores apenas com deslocamento carregado:")
        print(apenas_carregado.head(3)[['Operador', 'Velocidade Carregado', 'Velocidade Vazio']])
        
    if len(apenas_vazio) > 0:
        print("\nExemplos de operadores apenas com deslocamento vazio:")
        print(apenas_vazio.head(3)[['Operador', 'Velocidade Carregado', 'Velocidade Vazio']])
    
    return resultado

def identificar_operadores_duplicados(df, substituicoes=None):
    """
    Identifica operadores que começam com '133' e têm 7 dígitos.
    Verifica se já existe uma substituição no arquivo JSON, caso contrário, registra como ID encontrada.
    
    Args:
        df (DataFrame): DataFrame com os dados dos operadores
        substituicoes (dict): Dicionário com as substituições do arquivo JSON
    
    Returns:
        dict: Dicionário com mapeamento {id_incorreta: id_correta}
        DataFrame: DataFrame com as IDs encontradas para relatório
    """
    if 'Operador' not in df.columns or len(df) == 0:
        return {}, pd.DataFrame(columns=['ID Encontrada', 'Nome', 'Status', 'ID Substituição'])
    
    # Extrair operadores únicos
    operadores = df['Operador'].unique()
    
    # Lista para armazenar as IDs encontradas
    ids_encontradas = []
    mapeamento = {}
    
    for op in operadores:
        if ' - ' in op:
            try:
                id_parte, nome_parte = op.split(' - ', 1)
                # Verificar se a ID começa com 133 e tem 7 dígitos
                if id_parte.startswith('133') and len(id_parte) == 7:
                    # Verificar se existe uma substituição no arquivo JSON
                    if substituicoes and op in substituicoes:
                        status = "Substituição encontrada"
                        id_substituicao = substituicoes[op].split(' - ')[0] if ' - ' in substituicoes[op] else substituicoes[op]
                        mapeamento[op] = substituicoes[op]
                    else:
                        status = "Sem substituição definida"
                        id_substituicao = ""
                    
                    # Adicionar à lista de IDs encontradas, mesmo se for "NAO CADASTRADO"
                    ids_encontradas.append({
                        'ID Encontrada': id_parte,
                        'Nome': nome_parte,
                        'Status': status,
                        'ID Substituição': id_substituicao
                    })
                    
                    print(f"ID encontrada: {id_parte} - {nome_parte}")
            except Exception as e:
                print(f"Erro ao processar operador {op}: {str(e)}")
                continue
    
    print(f"Encontradas {len(ids_encontradas)} IDs começando com 133 e 7 dígitos.")
    for id_enc in ids_encontradas:
        print(f"  - {id_enc['ID Encontrada']} - {id_enc['Nome']} ({id_enc['Status']})")
    
    # Criar o DataFrame e ordenar por ID Encontrada
    df_encontradas = pd.DataFrame(ids_encontradas)
    if not df_encontradas.empty:
        df_encontradas = df_encontradas.sort_values('ID Encontrada')
    
    return mapeamento, df_encontradas

def criar_planilha_tdh(df):
    """
    Cria uma planilha vazia para TDH, contendo apenas as frotas encontradas.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame vazio com as colunas 'Frota' e 'TDH'
    """
    # Obter todas as frotas únicas
    frotas = df['Equipamento'].unique()
    
    # Criar DataFrame vazio com as frotas
    df_tdh = pd.DataFrame({'Frota': frotas, 'TDH': ''})
    
    return df_tdh

def criar_planilha_diesel(df):
    """
    Cria uma planilha vazia para Diesel, contendo apenas as frotas encontradas.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame vazio com as colunas 'Frota' e 'Diesel'
    """
    # Obter todas as frotas únicas
    frotas = df['Equipamento'].unique()
    
    # Criar DataFrame vazio com as frotas
    df_diesel = pd.DataFrame({'Frota': frotas, 'Diesel': ''})
    
    return df_diesel

def criar_planilha_impureza(df):
    """
    Cria uma planilha vazia para Impureza Vegetal, contendo apenas as frotas encontradas.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame vazio com as colunas 'Frota' e 'Impureza'
    """
    # Obter todas as frotas únicas
    frotas = df['Equipamento'].unique()
    
    # Criar DataFrame vazio com as frotas
    df_impureza = pd.DataFrame({'Frota': frotas, 'Impureza': ''})
    
    return df_impureza

def calcular_ofensores(df):
    """
    Calcula os top 5 ofensores gerais.
    Agrupa apenas por Operacao onde Estado Operacional é 'PARADA',
    soma a Diferença_Hora, classifica do maior para o menor e seleciona os top 5.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame com os top 5 ofensores gerais
    """
    # Filtrar apenas os registros com Estado Operacional PARADA
    df_paradas = df[df['Estado Operacional'] == 'PARADA'].copy()
    
    # Se não houver dados de parada, retornar DataFrame vazio
    if len(df_paradas) == 0:
        return pd.DataFrame(columns=['Operação', 'Tempo', 'Porcentagem'])
    
    # Agrupar apenas por Operacao, somar o tempo
    paradas_agrupadas = df_paradas.groupby('Operacao')['Diferença_Hora'].sum().reset_index()
    
    # Calcular o tempo total de todas as paradas
    tempo_total = paradas_agrupadas['Diferença_Hora'].sum()
    
    # Adicionar coluna de porcentagem
    paradas_agrupadas['Porcentagem'] = paradas_agrupadas.apply(
        lambda row: row['Diferença_Hora'] / tempo_total if tempo_total > 0 else 0,
        axis=1
    )
    
    # Ordenar por tempo (decrescente)
    paradas_agrupadas = paradas_agrupadas.sort_values(by='Diferença_Hora', ascending=False)
    
    # Selecionar os top 5 gerais
    resultado = paradas_agrupadas.head(5)
    
    # Renomear colunas para melhor visualização
    resultado = resultado.rename(columns={
        'Operacao': 'Operação',
        'Diferença_Hora': 'Tempo'
    })
    
    return resultado

def criar_excel_com_planilhas(df_base, base_calculo, disp_mecanica, eficiencia_energetica,
                            motor_ocioso, falta_apontamento, uso_gps, horas_por_frota, caminho_saida,
                            df_duplicados=None, media_velocidade=None, df_substituicoes=None):
    """
    Cria um arquivo Excel com todas as planilhas necessárias.
    """
    # Definir função de ajuste de largura de colunas
    def ajustar_largura_colunas(worksheet):
        """Ajusta a largura das colunas da planilha"""
        for col in worksheet.columns:
            max_length = 10
            column = col[0].column_letter
            header_text = str(col[0].value)
            if header_text:
                max_length = max(max_length, len(header_text) + 2)
            for cell in col[1:min(20, len(col))]:
                if cell.value:
                    cell_text = str(cell.value)
                    max_length = max(max_length, len(cell_text) + 2)
            max_length = min(max_length, 40)
            worksheet.column_dimensions[column].width = max_length
    
    # Criar planilhas adicionais
    df_tdh = criar_planilha_tdh(df_base)
    df_diesel = criar_planilha_diesel(df_base)
    df_impureza = criar_planilha_impureza(df_base)
    
    # Calcular ofensores
    df_ofensores = calcular_ofensores(df_base)
    
    # ===== CÁLCULO DE MANOBRAS (média simples) =====
    if {'Estado', 'Diferença_Hora'}.issubset(df_base.columns):
        # Filtrar manobras pelo tempo mínimo configurado (converter de segundos para horas)
        tempo_minimo_horas = tempoMinimoManobras / 3600
        df_manobras = df_base[(df_base['Estado'] == 'MANOBRA') & 
                             (df_base['Diferença_Hora'] >= tempo_minimo_horas) & 
                             (df_base['Velocidade'] >= velocidadeMinimaManobras)]
        
        # Mostrar informações sobre o filtro aplicado
        total_manobras_original = len(df_base[df_base['Estado'] == 'MANOBRA'])
        total_manobras_filtradas = len(df_manobras)
        print(f"\nFiltros de manobras aplicados:")
        print(f"- Tempo mínimo: {tempoMinimoManobras} segundos ({tempo_minimo_horas:.6f} horas)")
        print(f"- Velocidade mínima: {velocidadeMinimaManobras} km/h")
        print(f"Total de manobras antes dos filtros: {total_manobras_original}")
        print(f"Total de manobras após os filtros: {total_manobras_filtradas}")
        print(f"Manobras removidas: {total_manobras_original - total_manobras_filtradas} ({((total_manobras_original - total_manobras_filtradas) / total_manobras_original * 100) if total_manobras_original > 0 else 0:.2f}%)\n")
    else:
        df_manobras = pd.DataFrame()

    if not df_manobras.empty:
        # Operador
        if 'Operador' in df_manobras.columns:
            # Calcular métricas por operador
            df_manobras_operador = df_manobras.groupby('Operador').agg({
                'Diferença_Hora': ['count', 'mean', 'sum'],
                'RPM Motor': 'mean',
                'Velocidade': 'mean'
            }).reset_index()
            
            # Renomear colunas
            df_manobras_operador.columns = [
                'Operador', 
                'Quantidade Manobras', 
                'Tempo Médio Manobras', 
                'Tempo Total Manobras',
                'RPM Médio', 
                'Velocidade Média'
            ]
            
            # Ordenar por tempo médio
            df_manobras_operador = df_manobras_operador.sort_values('Tempo Médio Manobras', ascending=False)
            
            # Adicionar sufixo de frotas ao nome do operador
            def _op_frotas(op):
                frotas = sorted(df_manobras[df_manobras['Operador'] == op]['Equipamento'].astype(str).unique())
                return f"{op} ({', '.join(frotas)})" if frotas else op
            df_manobras_operador['Operador'] = df_manobras_operador['Operador'].apply(_op_frotas)
        else:
            df_manobras_operador = pd.DataFrame(columns=[
                'Operador', 
                'Quantidade Manobras', 
                'Tempo Médio Manobras', 
                'Tempo Total Manobras',
                'RPM Médio', 
                'Velocidade Média'
            ])
            
        # Frota
        if 'Equipamento' in df_manobras.columns:
            # Calcular métricas por frota
            df_manobras_frota = df_manobras.groupby('Equipamento').agg({
                'Diferença_Hora': ['count', 'mean', 'sum'],
                'RPM Motor': 'mean',
                'Velocidade': 'mean'
            }).reset_index()
            
            # Renomear colunas
            df_manobras_frota.columns = [
                'Frota', 
                'Quantidade Manobras', 
                'Tempo Médio Manobras', 
                'Tempo Total Manobras',
                'RPM Médio', 
                'Velocidade Média'
            ]
            
            # Ordenar por tempo médio
            df_manobras_frota = df_manobras_frota.sort_values('Tempo Médio Manobras', ascending=False)
        else:
            df_manobras_frota = pd.DataFrame(columns=[
                'Frota', 
                'Quantidade Manobras', 
                'Tempo Médio Manobras', 
                'Tempo Total Manobras',
                'RPM Médio', 
                'Velocidade Média'
            ])
    else:
        df_manobras_operador = pd.DataFrame(columns=[
            'Operador', 
            'Quantidade Manobras', 
            'Tempo Médio Manobras', 
            'Tempo Total Manobras',
            'RPM Médio', 
            'Velocidade Média'
        ])
        df_manobras_frota = pd.DataFrame(columns=[
            'Frota', 
            'Quantidade Manobras', 
            'Tempo Médio Manobras', 
            'Tempo Total Manobras',
            'RPM Médio', 
            'Velocidade Média'
        ])
    # ===== FIM CÁLCULO DE MANOBRAS =====
    
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        # Salvar cada DataFrame em uma planilha separada
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        base_calculo.to_excel(writer, sheet_name='Base Calculo', index=False)
        disp_mecanica.to_excel(writer, sheet_name='1_Disponibilidade Mecânica', index=False)
        eficiencia_energetica.to_excel(writer, sheet_name='2_Eficiência Energética', index=False)
        motor_ocioso.to_excel(writer, sheet_name='3_Motor Ocioso', index=False)
        falta_apontamento.to_excel(writer, sheet_name='4_Falta Apontamento', index=False)
        uso_gps.to_excel(writer, sheet_name='5_Uso GPS', index=False)
        horas_por_frota.to_excel(writer, sheet_name='Horas por Frota', index=False)
        
        # Adicionar nova planilha de ofensores
        df_ofensores.to_excel(writer, sheet_name='Ofensores', index=False)
        
        # Adicionar novas planilhas
        df_tdh.to_excel(writer, sheet_name='TDH', index=False)
        df_diesel.to_excel(writer, sheet_name='Diesel', index=False)
        df_impureza.to_excel(writer, sheet_name='Impureza Vegetal', index=False)
        
        # Adicionar planilhas de Manobras
        df_manobras_operador.to_excel(writer, sheet_name='Manobras Operador', index=False)
        df_manobras_frota.to_excel(writer, sheet_name='Manobras Frotas', index=False)
        
        if media_velocidade is None:
            media_velocidade = pd.DataFrame(columns=['Operador', 'Velocidade Geral', 'Velocidade Carregado', 'Velocidade Vazio'])
        media_velocidade.to_excel(writer, sheet_name='Média Velocidade', index=False)
        
        # IDs duplicadas e substituídas
        if df_duplicados is not None and not df_duplicados.empty:
            df_duplicados.to_excel(writer, sheet_name='IDs Encontradas', index=False)
        if df_substituicoes is not None and not df_substituicoes.empty:
            df_substituicoes.to_excel(writer, sheet_name='IDs Substituídas', index=False)
        
        # Formatar cada planilha
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            ajustar_largura_colunas(worksheet)
            
            if sheet_name == '1_Disponibilidade Mecânica':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Disponibilidade)
                    cell.number_format = '0.00%'
            
            elif sheet_name == '2_Eficiência Energética':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Eficiência)
                    cell.number_format = '0.00%'
            
            elif sheet_name == '3_Motor Ocioso':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
                    cell = worksheet.cell(row=row, column=3)  # Coluna C (Tempo Ligado)
                    cell.number_format = '0.00'  # Formato decimal
                    cell = worksheet.cell(row=row, column=4)  # Coluna D (Tempo Ocioso)
                    cell.number_format = '0.00'  # Formato decimal
            
            elif sheet_name == '4_Falta Apontamento':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
            
            elif sheet_name == '5_Uso GPS':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
            
            elif sheet_name == 'Média Velocidade':
                for row in range(2, worksheet.max_row + 1):
                    # Coluna B (Velocidade Geral)
                    cell = worksheet.cell(row=row, column=2)
                    cell.number_format = '0.00'
                    
                    # Coluna C (Velocidade Carregado)
                    cell = worksheet.cell(row=row, column=3)
                    cell.number_format = '0.00'
                    
                    # Coluna D (Velocidade Vazio)
                    cell = worksheet.cell(row=row, column=4)
                    cell.number_format = '0.00'
            
            elif sheet_name == 'Horas por Frota':
                for row in range(2, worksheet.max_row + 1):
                    for col in range(2, worksheet.max_column + 1):  # Todas as colunas de tempo
                        cell = worksheet.cell(row=row, column=col)
                        cell.number_format = '0.00'  # Formato decimal
            
            elif sheet_name == 'Ofensores':
                for row in range(2, worksheet.max_row + 1):
                    # Coluna B (Tempo)
                    cell = worksheet.cell(row=row, column=2)
                    cell.number_format = '0.00'  # Formato decimal
                    
                    # Coluna C (Porcentagem)
                    cell = worksheet.cell(row=row, column=3)
                    cell.number_format = '0.00%'  # Formato percentual
                        
            elif sheet_name == 'TDH':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (TDH)
                    cell.number_format = '0.0000'  # 4 casas decimais
            
            elif sheet_name == 'Diesel':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Diesel)
                    cell.number_format = '0.0000'  # 4 casas decimais
            
            elif sheet_name == 'Impureza Vegetal':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Impureza)
                    cell.number_format = '0.00'  # 2 casas decimais
            
            elif sheet_name in ['Manobras Operador', 'Manobras Frotas']:
                # Cabeçalho da nova coluna para formato de hora
                worksheet.cell(row=1, column=worksheet.max_column + 1).value = 'Tempo Médio (hh:mm)'
                worksheet.cell(row=1, column=worksheet.max_column).value = 'Tempo Total (hh:mm)'
                
                # Formatar colunas numéricas
                for row in range(2, worksheet.max_row + 1):
                    # Quantidade de Manobras (coluna 2)
                    cell = worksheet.cell(row=row, column=2)
                    cell.number_format = '0'
                    
                    # Tempo Médio (coluna 3)
                    dec_cell = worksheet.cell(row=row, column=3)
                    dec_cell.number_format = '0.0000'
                    
                    # Tempo Total (coluna 4)
                    total_cell = worksheet.cell(row=row, column=4)
                    total_cell.number_format = '0.0000'
                    
                    # RPM Médio (coluna 5)
                    rpm_cell = worksheet.cell(row=row, column=5)
                    rpm_cell.number_format = '0'
                    
                    # Velocidade Média (coluna 6) - Sem formatação específica
                    vel_cell = worksheet.cell(row=row, column=6)
                    vel_cell.number_format = 'General'
                    
                    # Adicionar colunas com tempo formatado como horas
                    # Tempo Médio (hh:mm)
                    worksheet.cell(row=row, column=worksheet.max_column).value = dec_cell.value / 24 if dec_cell.value else 0
                    worksheet.cell(row=row, column=worksheet.max_column).number_format = 'h:mm:ss'
                    
                    # Tempo Total (hh:mm)
                    worksheet.cell(row=row, column=worksheet.max_column - 1).value = total_cell.value / 24 if total_cell.value else 0
                    worksheet.cell(row=row, column=worksheet.max_column - 1).number_format = 'h:mm:ss'
                
                # Reajustar largura das colunas
                ajustar_largura_colunas(worksheet)
            
            elif sheet_name == 'Base Calculo':
                colunas_porcentagem = ['% Parado com motor ligado', '% Utilização GPS', '% Falta de Apontamento']
                colunas_tempo = ['Horas totais', 'Motor Ligado', 'Parado com motor ligado', 'GPS', 'Horas Produtivas', 'Falta de Apontamento']
                
                for row in range(2, worksheet.max_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        header = worksheet.cell(row=1, column=col).value
                        cell = worksheet.cell(row=row, column=col)
                        
                        if header in colunas_porcentagem:
                            cell.number_format = '0.00%'
                        elif header in colunas_tempo:
                            cell.number_format = '0.00'

def extrair_arquivo_zip(caminho_zip, pasta_destino=None):
    """
    Extrai o conteúdo de um arquivo ZIP para uma pasta temporária ou destino especificado.
    Renomeia os arquivos extraídos para terem o mesmo nome do arquivo ZIP original.
    
    Args:
        caminho_zip (str): Caminho para o arquivo ZIP
        pasta_destino (str, optional): Pasta onde os arquivos serão extraídos.
                                       Se None, usa uma pasta temporária.
    
    Returns:
        list: Lista de caminhos dos arquivos extraídos e renomeados (apenas TXT e CSV)
        str: Caminho da pasta temporária (se criada) ou None
    """
    # Se pasta_destino não foi especificada, criar uma pasta temporária
    pasta_temp = None
    if pasta_destino is None:
        pasta_temp = tempfile.mkdtemp()
        pasta_destino = pasta_temp
    
    arquivos_extraidos = []
    nome_zip_sem_extensao = os.path.splitext(os.path.basename(caminho_zip))[0]
    
    try:
        with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
            # Extrair todos os arquivos do ZIP
            zip_ref.extractall(pasta_destino)
            
            # Processar cada arquivo extraído (apenas TXT e CSV)
            for arquivo in zip_ref.namelist():
                caminho_completo = os.path.join(pasta_destino, arquivo)
                # Verificar se é um arquivo e não uma pasta
                if os.path.isfile(caminho_completo):
                    # Verificar extensão
                    extensao = os.path.splitext(arquivo)[1].lower()
                    if extensao in ['.txt', '.csv']:
                        # Criar novo nome: nome do ZIP + extensão original
                        novo_nome = f"{nome_zip_sem_extensao}{extensao}"
                        novo_caminho = os.path.join(pasta_destino, novo_nome)
                        
                        # Renomear o arquivo extraído
                        try:
                            # Se já existe um arquivo com esse nome, remover primeiro
                            if os.path.exists(novo_caminho):
                                os.remove(novo_caminho)
                            # Renomear o arquivo
                            os.rename(caminho_completo, novo_caminho)
                            arquivos_extraidos.append(novo_caminho)
                            print(f"Arquivo extraído renomeado: {novo_nome}")
                        except Exception as e:
                            print(f"Erro ao renomear arquivo {arquivo} para {novo_nome}: {str(e)}")
                            arquivos_extraidos.append(caminho_completo)  # Adicionar o caminho original em caso de erro
        
        return arquivos_extraidos, pasta_temp
    
    except Exception as e:
        print(f"Erro ao extrair o arquivo ZIP {caminho_zip}: {str(e)}")
        # Se houve erro e criamos uma pasta temporária, tentar limpá-la
        if pasta_temp:
            try:
                shutil.rmtree(pasta_temp)
            except:
                pass
        return [], None

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT, CSV ou ZIP de transbordos nas pastas dados e dados/transbordos.
    Busca arquivos que começam com "RV Transbordo", "frente" e "transbordos" com extensão .csv, .txt ou .zip.
    Ignora arquivos que contenham "colhedora" no nome.
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_transbordos = os.path.join(diretorio_raiz, "dados", "transbordos")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    
    # Verificar se os diretórios existem, caso contrário criar
    if not os.path.exists(diretorio_dados):
        os.makedirs(diretorio_dados)
    if not os.path.exists(diretorio_transbordos):
        os.makedirs(diretorio_transbordos)
    if not os.path.exists(diretorio_saida):
        os.makedirs(diretorio_saida)
    
    # Lista de diretórios para buscar arquivos
    diretorios_busca = [diretorio_dados, diretorio_transbordos]
    
    # Encontrar todos os arquivos TXT/CSV/ZIP de transbordos em ambos os diretórios
    arquivos = []
    arquivos_zip = []
    
    for diretorio in diretorios_busca:
        # Adicionar arquivos TXT sempre
        arquivos += glob.glob(os.path.join(diretorio, "RV Transbordo*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "*transbordo*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "frente*transbordos*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "transbordo*.txt"))
        
        # Adicionar arquivos CSV apenas se processCsv for True
        if processCsv:
            arquivos += glob.glob(os.path.join(diretorio, "RV Transbordo*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "*transbordo*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "frente*transbordos*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "transbordo*.csv"))
        
        # Adicionar arquivos ZIP
        arquivos_zip += glob.glob(os.path.join(diretorio, "RV Transbordo*.zip"))
        arquivos_zip += glob.glob(os.path.join(diretorio, "*transbordo*.zip"))
        arquivos_zip += glob.glob(os.path.join(diretorio, "frente*transbordos*.zip"))
        arquivos_zip += glob.glob(os.path.join(diretorio, "transbordo*.zip"))
    
    # Filtrar arquivos que contenham "colhedora" no nome (case insensitive)
    arquivos = [arquivo for arquivo in arquivos if "colhedora" not in os.path.basename(arquivo).lower()]
    arquivos_zip = [arquivo for arquivo in arquivos_zip if "colhedora" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos = list(set(arquivos))
    arquivos_zip = list(set(arquivos_zip))
    
    if not arquivos and not arquivos_zip:
        print("Nenhum arquivo de transbordos encontrado nas pastas dados ou dados/transbordos!")
        return
    
    print(f"Encontrados {len(arquivos)} arquivos de transbordos (TXT/CSV) para processar.")
    print(f"Encontrados {len(arquivos_zip)} arquivos ZIP de transbordos para processar.")
    
    # Processar cada arquivo TXT/CSV
    for arquivo in arquivos:
        processar_arquivo(arquivo, diretorio_saida)
    
    # Processar cada arquivo ZIP
    for arquivo_zip in arquivos_zip:
        print(f"\nProcessando arquivo ZIP: {os.path.basename(arquivo_zip)}")
        
        # Extrair arquivo ZIP para pasta temporária
        arquivos_extraidos, pasta_temp = extrair_arquivo_zip(arquivo_zip)
        
        if not arquivos_extraidos:
            print(f"Nenhum arquivo TXT ou CSV encontrado no ZIP {os.path.basename(arquivo_zip)}")
            continue
        
        print(f"Extraídos {len(arquivos_extraidos)} arquivos do ZIP.")
        
        # Processar cada arquivo extraído
        for arquivo_extraido in arquivos_extraidos:
            # Filtrar arquivos que contenham "colhedora" no nome
            if "colhedora" not in os.path.basename(arquivo_extraido).lower():
                processar_arquivo(arquivo_extraido, diretorio_saida)
        
        # Limpar pasta temporária se foi criada
        if pasta_temp:
            try:
                shutil.rmtree(pasta_temp)
                print(f"Pasta temporária removida: {pasta_temp}")
            except Exception as e:
                print(f"Erro ao remover pasta temporária {pasta_temp}: {str(e)}")

def processar_arquivo(caminho_arquivo, diretorio_saida):
    """
    Processa um arquivo de transbordos e gera um arquivo Excel com as métricas calculadas.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo a ser processado
        diretorio_saida (str): Diretório onde o arquivo de saída será salvo
    """
    # Obter apenas o nome do arquivo (sem caminho e sem extensão)
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    
    # Nome de saída igual ao original, mas com sufixo "_processado" e extensão .xlsx na pasta output
    arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}_processado.xlsx")
    
    print(f"\nProcessando arquivo: {os.path.basename(caminho_arquivo)}")
    print(f"Arquivo de saída: {os.path.basename(arquivo_saida)}")
    
    # Processar o arquivo base
    df_base = processar_arquivo_base(caminho_arquivo)
    if df_base is None:
        print(f"Erro ao processar {os.path.basename(caminho_arquivo)}. Pulando para o próximo arquivo.")
        return
    
    # Identificar operadores com IDs que começam com 133 e têm 7 dígitos
    mapeamento_duplicados, df_duplicados = identificar_operadores_duplicados(df_base)
    
    # Carregar substituições de operadores
    substituicoes = carregar_substituicoes_operadores()
    substituicoes_horario = carregar_substituicoes_operadores_horario()
    
    # Combinar as substituições manuais com as automáticas
    substituicoes_combinadas = {**substituicoes, **mapeamento_duplicados}
    
    # Aplicar as substituições usando a nova função
    print("\nAplicando substituições de operadores...")
    df_base, df_substituicoes = aplicar_substituicao_operadores(df_base, substituicoes_combinadas, substituicoes_horario)
    
    # Se o DataFrame estiver vazio, gerar apenas a planilha BASE
    if len(df_base) == 0:
        writer = pd.ExcelWriter(arquivo_saida, engine='openpyxl')
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        if not df_duplicados.empty:
            df_duplicados.to_excel(writer, sheet_name='IDs Encontradas', index=False)
        writer.close()
        print(f"Arquivo {arquivo_saida} gerado com apenas a planilha BASE (sem dados).")
        return
    
    # Primeiro, aplicar o novo cálculo de motor ocioso no DataFrame
    # Esta função modifica o df_base adicionando a coluna 'Motor Ocioso' e também retorna um DataFrame formatado para a planilha
    df_base_com_motor_ocioso = calcular_motor_ocioso_para_base(df_base)
    
    # Agora calculamos a Base Calculo
    base_calculo = calcular_base_calculo(df_base_com_motor_ocioso)
    
    # Calcular as métricas auxiliares
    disp_mecanica = calcular_disponibilidade_mecanica(df_base_com_motor_ocioso)
    eficiencia_energetica = calcular_eficiencia_energetica(base_calculo)
    motor_ocioso = calcular_motor_ocioso(base_calculo, df_base_com_motor_ocioso)  # Agora passa a Base Calculo diretamente
    falta_apontamento = calcular_falta_apontamento(base_calculo)
    uso_gps = calcular_uso_gps(base_calculo)
    horas_por_frota = calcular_horas_por_frota(df_base_com_motor_ocioso)
    
    # Calcular média de velocidade por operador
    media_velocidade = calcular_media_velocidade(df_base_com_motor_ocioso)
    
    # Criar o arquivo Excel com todas as planilhas
    criar_excel_com_planilhas(
        df_base_com_motor_ocioso, base_calculo, disp_mecanica, eficiencia_energetica,
        motor_ocioso, falta_apontamento, uso_gps, horas_por_frota, arquivo_saida,
        df_duplicados,  # Adicionar a tabela de IDs duplicadas
        media_velocidade,  # Adicionar a tabela de média de velocidade
        df_substituicoes  # Adicionar a tabela de IDs substituídas
    )
    
    print(f"Arquivo {arquivo_saida} gerado com sucesso!")

def salvar_planilha_base(df, caminho_saida):
    """
    Salva o DataFrame em um arquivo Excel, aplicando formatação adequada.
    
    Args:
        df (DataFrame): DataFrame a ser salvo
        caminho_saida (str): Caminho do arquivo Excel de saída
    """
    try:
        # Criar uma cópia do DataFrame para não modificar o original
        df_copy = df.copy()
        
        # Identificar colunas de tempo
        colunas_tempo = ['Diferença_Hora', 'Horas Produtivas']
        
        # Criar arquivo Excel
        writer = pd.ExcelWriter(caminho_saida, engine='openpyxl')
        
        # Salvar DataFrame
        df_copy.to_excel(writer, index=False)
        
        # Ajustar largura das colunas e aplicar formatação
        worksheet = writer.book.active
        
        # Ajustar largura das colunas
        for idx, col in enumerate(df_copy.columns):
            max_length = max(
                df_copy[col].astype(str).apply(len).max(),
                len(str(col))
            )
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width
            
            # Aplicar formatação de tempo
            if col in colunas_tempo:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=idx + 1)
                    cell.number_format = '0.00'  # Formato decimal
        
        # Salvar arquivo
        writer.close()
        print(f"Arquivo {caminho_saida} gerado com sucesso!")
        
    except Exception as e:
        print(f"Erro ao gerar arquivo {caminho_saida}: {str(e)}")
        print(f"Arquivo {caminho_saida} não foi gerado.")

def calcular_motor_ocioso_para_base(df):
    """
    Calcula o tempo de motor ocioso de acordo com as novas regras e modifica o DataFrame original.
    Esta função MODIFICA o DataFrame passado, adicionando a coluna 'Motor Ocioso'.
    Considera as configurações de exclusão de operações e grupos de operação do arquivo de configuração.
    
    Args:
        df (DataFrame): DataFrame com os dados de operação
        
    Returns:
        DataFrame: DataFrame modificado com a coluna 'Motor Ocioso' atualizada
    """
    # Carregar configurações de cálculos
    config = carregar_config_calculos()
    tipo_equipamento = "TT"  # Para transbordos
    
    # Obter operações e grupos excluídos da configuração
    operacoes_excluidas = []
    grupos_operacao_excluidos = []
    
    if tipo_equipamento in config and "motor_ocioso" in config[tipo_equipamento]:
        operacoes_excluidas = config[tipo_equipamento]["motor_ocioso"].get("operacoes_excluidas", [])
        grupos_operacao_excluidos = config[tipo_equipamento]["motor_ocioso"].get("grupos_operacao_excluidos", [])
    
    print(f"Operações excluídas do cálculo de motor ocioso: {operacoes_excluidas}")
    print(f"Grupos de operação excluídos do cálculo de motor ocioso: {grupos_operacao_excluidos}")
    
    # Criar uma cópia para modificar
    df_resultado = df.copy()
    
    # Converter a coluna de diferença para minutos
    df_resultado['Diferença_Minutos'] = df_resultado['Diferença_Hora'] * 60
    
    # Inicializar colunas
    df_resultado['Motor Ocioso'] = 0
    df_resultado['Em_Intervalo'] = False
    df_resultado['Soma_Intervalo'] = 0
    
    # Filtrar operações e grupos de operação excluídos
    df_filtrado_config = df_resultado.copy()
    if operacoes_excluidas:
        df_filtrado_config = df_filtrado_config[~df_filtrado_config['Operacao'].isin(operacoes_excluidas)]
    if grupos_operacao_excluidos:
        df_filtrado_config = df_filtrado_config[~df_filtrado_config['Grupo Operacao'].isin(grupos_operacao_excluidos)]
    
    print(f"Total de registros antes da filtragem: {len(df_resultado)}")
    print(f"Total de registros após filtragem por operações e grupos excluídos: {len(df_filtrado_config)}")
    
    # Variáveis para controle do intervalo atual
    em_intervalo = False
    soma_intervalo = 0
    inicio_intervalo = None
    
    # Iterar sobre as linhas do DataFrame filtrado
    for i, row in df_filtrado_config.iterrows():
        parada_motor = row['Parado com motor ligado']
        diferenca = row['Diferença_Minutos']
        
        # Se não estamos em um intervalo
        if not em_intervalo:
            # Se encontrar Parado com Motor Ligado = 1, inicia novo intervalo
            if parada_motor == 1:
                em_intervalo = True
                soma_intervalo = diferenca
                inicio_intervalo = i
                df_resultado.at[i, 'Em_Intervalo'] = True
                df_resultado.at[i, 'Soma_Intervalo'] = soma_intervalo
        
        # Se estamos em um intervalo
        else:
            # Se encontrar Parado com Motor Ligado = 0
            if parada_motor == 0:
                # Se a duração for > 1 minuto, fecha o intervalo
                if diferenca > 1:
                    # Se o total acumulado > 1 minuto, subtrai 1 minuto
                    if soma_intervalo > 1:
                        tempo_ocioso = soma_intervalo - 1
                        # Atribui o tempo ocioso à primeira linha do intervalo
                        # IMPORTANTE: Converter de minutos para horas antes de atribuir
                        df_resultado.at[inicio_intervalo, 'Motor Ocioso'] = tempo_ocioso / 60.0  # Dividir por 60 para converter minutos em horas
                    
                    # Reseta o intervalo
                    em_intervalo = False
                    soma_intervalo = 0
                    inicio_intervalo = None
                else:
                    # Se <= 1 minuto, soma ao intervalo atual
                    soma_intervalo += diferenca
                    df_resultado.at[i, 'Em_Intervalo'] = True
                    df_resultado.at[i, 'Soma_Intervalo'] = soma_intervalo
            
            # Se encontrar Parado com Motor Ligado = 1
            else:
                soma_intervalo += diferenca
                df_resultado.at[i, 'Em_Intervalo'] = True
                df_resultado.at[i, 'Soma_Intervalo'] = soma_intervalo
    
    # Tratar último intervalo aberto, se houver
    if em_intervalo and soma_intervalo > 1:
        tempo_ocioso = soma_intervalo - 1
        # Converter de minutos para horas antes de atribuir
        df_resultado.at[inicio_intervalo, 'Motor Ocioso'] = tempo_ocioso / 60.0  # Dividir por 60 para converter minutos em horas
    
    # Garantir que o tempo ocioso nunca seja maior que o tempo ligado para cada registro
    for i in range(len(df_resultado)):
        if df_resultado.iloc[i]['Motor Ocioso'] > 0:
            # Para transbordos, Motor Ligado é 'LIGADO' ou 'DESLIGADO', não 1 ou 0
            motor_ligado = df_resultado.iloc[i]['Motor Ligado'] == 'LIGADO'
            # Se o motor estiver ligado, limitar o tempo ocioso ao tempo ligado
            if motor_ligado:
                tempo_hora = df_resultado.iloc[i]['Diferença_Hora']
                if df_resultado.iloc[i]['Motor Ocioso'] > tempo_hora:
                    df_resultado.at[i, 'Motor Ocioso'] = tempo_hora
            else:
                # Se o motor não estiver ligado, o tempo ocioso deve ser zero
                df_resultado.at[i, 'Motor Ocioso'] = 0
    
    # Remover colunas auxiliares
    df_resultado = df_resultado.drop(['Diferença_Minutos', 'Em_Intervalo', 'Soma_Intervalo'], axis=1)
    
    return df_resultado

def calcular_base_calculo(df):
    """
    Calcula as métricas base para cada operador/frota.
    Retorna um DataFrame com as métricas calculadas.
    """
    # Função utilitária local para evitar divisão por zero
    def calcular_porcentagem(numerador, denominador, precisao: int = 4):
        """Retorna numerador/denominador arredondado, ou 0 se denominador==0"""
        if denominador and denominador > 0:
            return round(numerador / denominador, precisao)
        return 0.0
    
    resultados_base_calculo = []
    
    # Verificar se existem as colunas necessárias
    colunas_necessarias = ['Operador', 'Equipamento', 'Diferença_Hora', 'Estado', 'Estado Operacional', 'RTK']
    if not all(coluna in df.columns for coluna in colunas_necessarias):
        print("AVISO: Colunas necessárias para cálculo da base não encontradas.")
        return pd.DataFrame()
    
    # Identificar operadores únicos
    operadores = df['Operador'].unique()
    
    # Calcular métricas para cada operador
    for operador in operadores:
        # Filtrar dados do operador
        df_operador = df[df['Operador'] == operador]
        
        # Identificar frotas do operador
        frotas = df_operador['Equipamento'].unique()
        
        for frota in frotas:
            # Filtrar dados da frota
            df_frota = df_operador[df_operador['Equipamento'] == frota]
            
            # Calcular horas totais
            horas_totais = df_frota['Diferença_Hora'].sum()
            
            # Calcular horas com motor ligado (todos os estados exceto DESLIGADO)
            df_motor_ligado = df_frota[df_frota['Estado'] != 'DESLIGADO']
            motor_ligado = df_motor_ligado['Diferença_Hora'].sum()
            
            # Calcular horas parado com motor ligado
            df_parado = df_frota[(df_frota['Estado'] == 'PARADO') & (df_frota['Estado Operacional'] == 'PARADA')]
            parado_motor_ligado = df_parado['Diferença_Hora'].sum()
            
            # Calcular porcentagem de tempo parado com motor ligado
            pct_parado_motor_ligado = calcular_porcentagem(parado_motor_ligado, motor_ligado)
            
            # Calcular horas de GPS ativo (RTK = 1)
            df_gps = df_frota[df_frota['RTK'] == 1]
            horas_gps = df_gps['Diferença_Hora'].sum()
            
            # Calcular porcentagem de utilização de GPS
            pct_utilizacao_gps = calcular_porcentagem(horas_gps, horas_totais)
            
            # Calcular horas produtivas (todos os estados exceto PARADO)
            df_produtivo = df_frota[df_frota['Estado Operacional'] != 'PARADA']
            horas_produtivas = df_produtivo['Diferença_Hora'].sum()
            
            # Calcular falta de apontamento
            df_falta_apontamento = df_frota[df_frota['Estado Operacional'] == 'PARADA']
            falta_apontamento = df_falta_apontamento['Diferença_Hora'].sum()
            
            # Calcular porcentagem de falta de apontamento
            pct_falta_apontamento = calcular_porcentagem(falta_apontamento, horas_totais)
            
            # Adicionar resultado à lista
            resultados_base_calculo.append({
                'Operador': operador,
                'Frota': frota,
                'Horas totais': horas_totais,
                'Motor Ligado': motor_ligado,
                'Parado com motor ligado': parado_motor_ligado,
                '% Parado com motor ligado': pct_parado_motor_ligado,
                'GPS': horas_gps,
                '% Utilização GPS': pct_utilizacao_gps,
                'Horas Produtivas': horas_produtivas,
                'Falta de Apontamento': falta_apontamento,
                '% Falta de Apontamento': pct_falta_apontamento
            })
    
    return pd.DataFrame(resultados_base_calculo)

if __name__ == "__main__":
    print("="*80)
    print("Iniciando processamento de arquivos de transbordos...")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Este script processa arquivos de transbordos e gera planilhas Excel com métricas")
    print("Suporta arquivos TXT, CSV e ZIP")
    print("Ignorando arquivos que contenham 'colhedora' no nome")
    print("="*80)
    processar_todos_arquivos()
    print("\nProcessamento concluído!") 