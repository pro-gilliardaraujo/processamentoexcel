"""
Script para processamento completo de dados de monitoramento de colhedoras e transbordos.
Lê arquivos TXT ou CSV na pasta raiz, processa-os e gera arquivos Excel com planilhas auxiliares prontas.
Também processa arquivos ZIP contendo TXT ou CSV.
"""

import pandas as pd
import numpy as np
import os
import glob
import zipfile
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import json
import traceback

# Configurações
processCsv = False  # Altere para True quando quiser processar arquivos CSV

# Constantes
COLUNAS_REMOVER = [
    'Justificativa Corte Base Desligado',
    'Latitude',
    'Longitude',
    'Regional',
    'Tipo de Equipamento',
    'Unidade',
    'Centro de Custo'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento', 'Apertura do Rolo', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo', 'Corte Base Automatico/Manual',
    'Descricao Equipamento', 'Estado', 'Estado Operacional', 'Esteira Ligada',
    'Field Cruiser', 'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Implemento Ligado', 'Motor Ligado', 'Operacao', 'Operador', 'Pressao de Corte',
    'RPM Extrator', 'RPM Motor', 'RTK (Piloto Automatico)', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parada com Motor Ligado',
    'Horas Produtivas'
]

# Valores a serem filtrados
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO"]

def carregar_config_calculos():
    """
    Carrega as configurações de cálculos do arquivo JSON.
    Se o arquivo não existir, retorna configurações padrão.
    """
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config", "calculos_config.json")
    
    # Configuração padrão
    config_padrao = {
        "CD": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "8490 - LAVAGEM",
                    "MANUTENCAO",
                    "LAVAGEM",
                    "INST CONFIG TECNOL EMBARCADAS"
                ],
                "grupos_operacao_excluidos": ["Manutenção"]
            },
            "operadores_excluidos": ["9999 - TROCA DE TURNO"],
            "equipamentos_excluidos": []
        },
        "TT": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "9016 - ENCH SISTEMA FREIO",
                    "6340 - BASCULANDO  TRANSBORDAGEM",
                    "9024 - DESATOLAMENTO",
                    "MANUTENÇÃO",
                    "INST CONFIG TECNOL EMBARCADAS",
                    "DESATOLAMENTO"
                ],
                "grupos_operacao_excluidos": ["Manutenção"]
            },
            "operadores_excluidos": ["9999 - TROCA DE TURNO"],
            "equipamentos_excluidos": []
        }
    }
    
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                
                # Garantir que os equipamentos excluídos sejam tratados como texto
                if "CD" in config and "equipamentos_excluidos" in config["CD"]:
                    config["CD"]["equipamentos_excluidos"] = [str(eq).replace('.0', '') for eq in config["CD"]["equipamentos_excluidos"]]
                
                if "TT" in config and "equipamentos_excluidos" in config["TT"]:
                    config["TT"]["equipamentos_excluidos"] = [str(eq).replace('.0', '') for eq in config["TT"]["equipamentos_excluidos"]]
                
                return config
        else:
            print(f"Arquivo de configuração não encontrado em {config_path}. Usando configuração padrão.")
            return config_padrao
    except Exception as e:
        print(f"Erro ao carregar configurações: {str(e)}. Usando configuração padrão.")
        return config_padrao

def identificar_tipo_equipamento(arquivo, df=None):
    """
    Identifica se o arquivo é de colhedora ou transbordo.
    
    Args:
        arquivo (str): Nome do arquivo
        df (DataFrame, optional): DataFrame com os dados se já carregado
        
    Returns:
        str: 'colhedora' ou 'transbordo'
    """
    # Verificar pelo nome do arquivo primeiro
    nome_arquivo = os.path.basename(arquivo).lower()
    
    if 'colhedora' in nome_arquivo and 'transbordo' not in nome_arquivo:
        return 'colhedora'
    
    if 'transbordo' in nome_arquivo and 'colhedora' not in nome_arquivo:
        return 'transbordo'
    
    # Se não conseguiu determinar pelo nome, tentar pelo DataFrame
    if df is not None and len(df) > 0:
        # Verificar pela descrição do equipamento se existir
        if 'Descricao Equipamento' in df.columns:
            descricoes = df['Descricao Equipamento'].astype(str).str.lower()
            if descricoes.str.contains('colhedora').any() and not descricoes.str.contains('transbordo').any():
                return 'colhedora'
            if descricoes.str.contains('transbordo').any() and not descricoes.str.contains('colhedora').any():
                return 'transbordo'
        
        # Verificar pelo grupo de equipamento/frente
        if 'Grupo Equipamento/Frente' in df.columns:
            grupos = df['Grupo Equipamento/Frente'].astype(str).str.lower()
            if grupos.str.contains('colhedora').any() and not grupos.str.contains('transbordo').any():
                return 'colhedora'
            if grupos.str.contains('transbordo').any() and not grupos.str.contains('colhedora').any():
                return 'transbordo'
    
    # Se ainda não identificou, assumir como colhedora por padrão (pode ser ajustado)
    # Também é possível retornar None e tratar isso como um caso especial
    print(f"AVISO: Não foi possível identificar claramente o tipo de equipamento para {nome_arquivo}. Assumindo como COLHEDORA.")
    return 'colhedora'

def processar_arquivo_base(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
        str: Tipo de equipamento identificado ('colhedora' ou 'transbordo')
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo (TXT ou CSV são tratados da mesma forma se usarem separador ';')
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio (apenas cabeçalhos sem dados)
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                # Retornar o DataFrame vazio mas com as colunas, em vez de None
                # Garantir que todas as colunas desejadas existam
                for col in COLUNAS_DESEJADAS:
                    if col not in df.columns:
                        df[col] = np.nan
                # Reorganizar as colunas na ordem desejada
                colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
                
                # Identificar tipo de equipamento pelo nome do arquivo
                tipo_equipamento = identificar_tipo_equipamento(caminho_arquivo)
                return df[colunas_existentes + colunas_extras], tipo_equipamento
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Verificar se 'Data/Hora' existe e processá-la
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Conversão e cálculo de diferenças de hora
            if isinstance(df['Hora'].iloc[0], str):  # Se ainda for string, converter para datetime
                df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular a diferença de hora se ainda não existir
            if 'Diferença_Hora' not in df.columns or df['Diferença_Hora'].isna().any():
                df['Diferença_Hora'] = df['Hora'].diff().dt.total_seconds() / 3600
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if pd.isna(x) or x < 0 else x)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            else:
                # Certifica-se de que a coluna 'Diferença_Hora' esteja limpa e como número
                df['Diferença_Hora'] = pd.to_numeric(df['Diferença_Hora'].astype(str).str.strip(), errors='coerce')
                df['Diferença_Hora'] = df['Diferença_Hora'].fillna(0)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            
            # Conversão de colunas binárias para valores numéricos
            for col in ['Esteira Ligada', 'Motor Ligado', 'Field Cruiser', 'RTK (Piloto Automatico)', 'Implemento Ligado', 'Corte Base Automatico/Manual']:
                if col in df.columns:
                    if df[col].dtype == 'object':
                        df[col] = df[col].replace({'LIGADO': 1, 'DESLIGADO': 0, 'AUTOMATICO': 1, 'MANUAL': 0})
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
            # Tratar coluna Equipamento como texto e remover sufixo ".0"
            if 'Equipamento' in df.columns:
                df['Equipamento'] = df['Equipamento'].astype(str)
                df['Equipamento'] = df['Equipamento'].str.replace('.0$', '', regex=True)
            
            # Verificar e calcular "Parada com Motor Ligado" se necessário
            if 'Parada com Motor Ligado' not in df.columns:
                RPM_MINIMO = 300  # RPM mínimo considerado como motor ligado
                df['Parada com Motor Ligado'] = ((df['Velocidade'] == 0) & 
                                                (df['RPM Motor'] >= RPM_MINIMO) &
                                                (df['Motor Ligado'] == 1)).astype(int)
            
            # Verificar se Horas Produtivas já existe
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                # Para colhedoras e transbordos, horas produtivas são registros onde Estado = 'TRABALHANDO' ou 'COLHEITA'
                condicao_produtiva = (df['Estado'] == 'TRABALHANDO') | (df['Estado'] == 'COLHEITA')
                df['Horas Produtivas'] = df['Diferença_Hora'] * condicao_produtiva.astype(int)
            else:
                # Limpa e converte para número
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
            df = df[colunas_existentes + colunas_extras]
            
            # Identificar tipo de equipamento
            tipo_equipamento = identificar_tipo_equipamento(caminho_arquivo, df)
            
            return df, tipo_equipamento
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None, None

def calcular_base_calculo(df):
    """
    Calcula a tabela Base Calculo a partir do DataFrame base processado.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com cálculos por frota
    """
    # Criar cópia do DataFrame para não alterar o original
    df_temp = df.copy()
    
    # Agrupar por frota
    resultado = []
    
    # Agrupar por Frota/Grupo Equipamento
    agrupado = df_temp.groupby('Grupo Equipamento/Frente')
    
    for grupo, dados in agrupado:
        # Cálculos básicos
        horas_totais = dados['Diferença_Hora'].sum()
        
        # Filtrar para cálculo de horas de elevador (apenas para colhedoras)
        if 'Elevador Ativado' in dados.columns:
            filtro_elevador = dados['Elevador Ativado'] == 1
            horas_elevador = dados.loc[filtro_elevador, 'Diferença_Hora'].sum() if filtro_elevador.any() else 0
            perc_elevador = (horas_elevador / horas_totais) if horas_totais > 0 else 0
        else:
            horas_elevador = 0
            perc_elevador = 0
        
        # RTK ativado
        if 'RTK Ativado' in dados.columns:
            filtro_rtk = dados['RTK Ativado'] == 1
            horas_rtk = dados.loc[filtro_rtk, 'Diferença_Hora'].sum() if filtro_rtk.any() else 0
        else:
            horas_rtk = 0
        
        # Horas produtivas (quando elevador está ativado para colhedoras)
        horas_produtivas = horas_elevador if horas_elevador > 0 else 0
        
        # Perc utilização RTK
        perc_rtk = (horas_rtk / horas_totais) if horas_totais > 0 else 0
        
        # Motor Ligado
        filtro_motor = dados['Motor'] == 1
        motor_ligado = dados.loc[filtro_motor, 'Diferença_Hora'].sum() if filtro_motor.any() else 0
        
        # Eficiência do Elevador
        eficiencia_elevador = (horas_elevador / motor_ligado) if motor_ligado > 0 else 0
        
        # Parado com motor ligado
        filtro_parado = dados['Parada com Motor Ligado'] == 1
        parado_motor_ligado = dados.loc[filtro_parado, 'Diferença_Hora'].sum() if filtro_parado.any() else 0
        
        # Percentual parado com motor ligado
        perc_parado_motor = (parado_motor_ligado / motor_ligado) if motor_ligado > 0 else 0
        
        # Adicionar os resultados
        resultado.append({
            'Grupo Equipamento/Frente': grupo,
            'Frota': grupo,  # Duplicando para manter o padrão
            'Descrição': dados['Descrição Equipamento'].iloc[0] if 'Descrição Equipamento' in dados.columns else grupo,
            'Horas totais': horas_totais,
            'Horas elevador': horas_elevador,
            '%': perc_elevador,
            'RTK': horas_rtk,
            'Horas Produtivas': horas_produtivas,
            '% Utilização RTK': perc_rtk,
            'Motor Ligado': motor_ligado,
            '% Eficiência Elevador': eficiencia_elevador,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': perc_parado_motor
        })
    
    # Criar DataFrame final
    df_resultado = pd.DataFrame(resultado)
    
    # Ordenar pelo nome da frota
    df_resultado = df_resultado.sort_values('Frota')
    
    return df_resultado

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por grupo de equipamento
    """
    # Criar cópia do DataFrame para não alterar o original
    df_temp = df.copy()
    
    # Inicializar lista para armazenar resultados
    resultados = []
    
    # Agrupar por frota
    agrupado = df_temp.groupby('Grupo Equipamento/Frente')
    
    for grupo, dados in agrupado:
        # Verificar se temos as colunas necessárias
        if 'Tempo Aguard.Manutenção' in dados.columns and 'Tempo Manut. Prevent.' in dados.columns and 'Tempo Manut. Corret.' in dados.columns:
            # Calcular a soma total das horas
            total_horas = dados['Diferença_Hora'].sum()
            
            # Calcular a soma das horas de manutenção
            horas_aguardando = dados['Tempo Aguard.Manutenção'].sum()
            horas_preventiva = dados['Tempo Manut. Prevent.'].sum()
            horas_corretiva = dados['Tempo Manut. Corret.'].sum()
            
            # Total de horas de manutenção
            total_manutencao = horas_aguardando + horas_preventiva + horas_corretiva
            
            # Calcular disponibilidade
            if total_horas > 0:
                disponibilidade = 1 - (total_manutencao / total_horas)
            else:
                disponibilidade = 0
        else:
            # Se não temos as colunas necessárias, assumir disponibilidade de 0
            disponibilidade = 0
        
        # Adicionar à lista de resultados
        resultados.append({
            'Grupo Equipamento/Frente': grupo,
            'Disponibilidade': disponibilidade
        })
    
    # Criar DataFrame com os resultados
    df_disponibilidade = pd.DataFrame(resultados)
    
    # Ordenar pela disponibilidade (do maior para o menor)
    df_disponibilidade = df_disponibilidade.sort_values('Disponibilidade', ascending=False)
    
    return df_disponibilidade

def calcular_eficiencia_energetica(base_calculo):
    """
    Calcula a eficiência energética.
    
    Args:
        base_calculo (DataFrame): DataFrame da Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por grupo de equipamento
    """
    # Criar cópia do DataFrame para não alterar o original
    df_temp = base_calculo.copy()
    
    # Calcular a eficiência energética (usando % Eficiência Elevador da Base Calculo)
    eficiencia = df_temp[['Grupo Equipamento/Frente', '% Eficiência Elevador']].copy()
    eficiencia = eficiencia.rename(columns={'% Eficiência Elevador': 'Eficiência'})
    
    # Ordenar pela eficiência (do maior para o menor)
    eficiencia = eficiencia.sort_values('Eficiência', ascending=False)
    
    return eficiencia

def calcular_hora_elevador(df, base_calculo):
    """
    Calcula as horas de elevador (específico para colhedoras).
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): DataFrame da Base Calculo
    
    Returns:
        DataFrame: Horas de elevador por frota
    """
    # Verificar se temos a coluna necessária
    if 'Elevador Ativado' not in df.columns:
        return None
    
    # Pegar as horas de elevador da Base Calculo
    hora_elevador = base_calculo[['Grupo Equipamento/Frente', 'Horas elevador']].copy()
    hora_elevador = hora_elevador.rename(columns={'Horas elevador': 'Horas'})
    
    # Ordenar pelo número de horas (do maior para o menor)
    hora_elevador = hora_elevador.sort_values('Horas', ascending=False)
    
    return hora_elevador

def calcular_motor_ocioso(base_calculo):
    """
    Calcula o percentual de motor ocioso.
    
    Args:
        base_calculo (DataFrame): DataFrame da Base Calculo
    
    Returns:
        DataFrame: Percentual de motor ocioso por grupo de equipamento
    """
    # Criar cópia do DataFrame para não alterar o original
    df_temp = base_calculo.copy()
    
    # Usar o percentual de parado com motor ligado da Base Calculo
    motor_ocioso = df_temp[['Grupo Equipamento/Frente', '% Parado com motor ligado']].copy()
    motor_ocioso = motor_ocioso.rename(columns={'% Parado com motor ligado': 'Porcentagem'})
    
    # Ordenar pelo percentual (do maior para o menor)
    motor_ocioso = motor_ocioso.sort_values('Porcentagem', ascending=False)
    
    return motor_ocioso

def calcular_uso_gps(df, base_calculo):
    """
    Calcula o percentual de uso de GPS.
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): DataFrame da Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por grupo de equipamento
    """
    # Criar cópia do DataFrame para não alterar o original
    df_temp = base_calculo.copy()
    
    # Usar o percentual de utilização RTK da Base Calculo
    uso_gps = df_temp[['Grupo Equipamento/Frente', '% Utilização RTK']].copy()
    uso_gps = uso_gps.rename(columns={'% Utilização RTK': 'Porcentagem'})
    
    # Ordenar pelo percentual (do maior para o menor)
    uso_gps = uso_gps.sort_values('Porcentagem', ascending=False)
    
    return uso_gps

def calcular_horas_por_frota(df):
    """
    Calcula as horas registradas por frota e a diferença para 24h.
    
    Args:
        df (DataFrame): DataFrame base processado
    
    Returns:
        DataFrame: Horas registradas por frota e diferença para 24h
    """
    # Criar cópia do DataFrame para não alterar o original
    df_temp = df.copy()
    
    # Inicializar lista para armazenar resultados
    resultados = []
    
    # Agrupar por frota e data
    agrupado = df_temp.groupby(['Grupo Equipamento/Frente', 'Data'])
    
    for (frota, data), dados in agrupado:
        # Calcular a soma das horas
        total_horas = dados['Diferença_Hora'].sum()
        
        # Calcular a diferença para 24h
        diferenca_24h = 24 - total_horas
        
        # Adicionar à lista de resultados
        resultados.append({
            'Frota': frota,
            'Data': data,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        })
    
    # Criar DataFrame com os resultados
    df_horas = pd.DataFrame(resultados)
    
    # Ordenar pelo nome da frota e data
    df_horas = df_horas.sort_values(['Frota', 'Data'])
    
    return df_horas

def criar_excel_com_planilhas(df_base, base_calculo, disp_mecanica, eficiencia_energetica, 
                             hora_elevador, motor_ocioso, uso_gps, horas_por_frota, 
                             motor_ocioso_por_operacao, caminho_saida, tipo_equipamento):
    """
    Cria um arquivo Excel com todas as planilhas auxiliares.
    Gera planilhas diferentes dependendo do tipo de equipamento.
    
    Args:
        df_base (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
        disp_mecanica (DataFrame): Disponibilidade mecânica
        eficiencia_energetica (DataFrame): Eficiência energética
        hora_elevador (DataFrame): Horas de elevador (apenas para colhedoras, None para transbordos)
        motor_ocioso (DataFrame): Motor ocioso
        uso_gps (DataFrame): Uso GPS
        horas_por_frota (DataFrame): Horas totais registradas por frota
        motor_ocioso_por_operacao (DataFrame): Motor ocioso por operação
        caminho_saida (str): Caminho do arquivo Excel de saída
        tipo_equipamento (str): 'colhedora' ou 'transbordo'
    """
    writer = pd.ExcelWriter(caminho_saida, engine='openpyxl')
    
    # Arredondamento fixo para 2 casas decimais em todas as colunas numéricas antes de exportar
    # Base Calculo - garantir que todas as colunas numéricas tenham 2 casas decimais
    colunas_numericas = ['Horas totais', 'Horas elevador', '%', 'RTK', 'Horas Produtivas', 
                         '% Utilização RTK', 'Motor Ligado', '% Eficiência Elevador', 
                         'Parado Com Motor Ligado', '% Parado com motor ligado']
    
    for col in colunas_numericas:
        if col in base_calculo.columns:
            base_calculo[col] = base_calculo[col].apply(lambda x: round(x, 2))
    
    # Arredondar valores nas outras planilhas
    disp_mecanica['Disponibilidade'] = disp_mecanica['Disponibilidade'].apply(lambda x: round(x, 4))
    eficiencia_energetica['Eficiência'] = eficiencia_energetica['Eficiência'].apply(lambda x: round(x, 4))
    if hora_elevador is not None:
        hora_elevador['Horas'] = hora_elevador['Horas'].apply(lambda x: round(x, 2))
    motor_ocioso['Porcentagem'] = motor_ocioso['Porcentagem'].apply(lambda x: round(x, 4))
    uso_gps['Porcentagem'] = uso_gps['Porcentagem'].apply(lambda x: round(x, 4))
    horas_por_frota['Horas Registradas'] = horas_por_frota['Horas Registradas'].apply(lambda x: round(x, 2))
    horas_por_frota['Diferença para 24h'] = horas_por_frota['Diferença para 24h'].apply(lambda x: round(x, 2))
    
    # Arredondar valor na planilha de Motor Ocioso por Operação
    if 'Horas Ociosas' in motor_ocioso_por_operacao.columns:
        motor_ocioso_por_operacao['Horas Ociosas'] = motor_ocioso_por_operacao['Horas Ociosas'].apply(lambda x: round(x, 2))
    
    # Salvar DataFrame base e Base Calculo em planilhas separadas
    df_base.to_excel(writer, sheet_name='BASE', index=False)
    base_calculo.to_excel(writer, sheet_name='Base Calculo', index=False)
    
    # Planilhas comuns para ambos os tipos
    disp_mecanica.to_excel(writer, sheet_name='1_Disponibilidade Mecânica', index=False)
    eficiencia_energetica.to_excel(writer, sheet_name='2_Eficiência Energética', index=False)
    
    # Planilhas específicas para cada tipo de equipamento
    if tipo_equipamento == 'colhedora':
        # Colhedoras têm planilha de Hora Elevador
        if hora_elevador is not None:
            hora_elevador.to_excel(writer, sheet_name='3_Hora Elevador', index=False)
            motor_ocioso.to_excel(writer, sheet_name='4_Motor Ocioso', index=False)
            uso_gps.to_excel(writer, sheet_name='5_Uso GPS', index=False)
            motor_ocioso_por_operacao.to_excel(writer, sheet_name='6_Motor Ocioso por Operação', index=False)
    else:  # transbordo
        # Transbordos não têm Hora Elevador, então ajustamos os índices
        motor_ocioso.to_excel(writer, sheet_name='3_Motor Ocioso', index=False)
        uso_gps.to_excel(writer, sheet_name='4_Uso GPS', index=False)
        motor_ocioso_por_operacao.to_excel(writer, sheet_name='5_Motor Ocioso por Operação', index=False)
    
    # Planilha comum
    horas_por_frota.to_excel(writer, sheet_name='Horas por Frota', index=False)
    
    # Aplicar formatação nas planilhas
    workbook = writer.book
    
    # Formatar planilha de Disponibilidade Mecânica
    worksheet = workbook['1_Disponibilidade Mecânica']
    for row in range(2, worksheet.max_row + 1):  # Começando da linha 2 (ignorando cabeçalho)
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Disponibilidade)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Eficiência Energética
    worksheet = workbook['2_Eficiência Energética']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Eficiência)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatação específica para cada tipo de equipamento
    if tipo_equipamento == 'colhedora':
        # Formatar planilha de Hora Elevador
        if '3_Hora Elevador' in workbook.sheetnames:
            worksheet = workbook['3_Hora Elevador']
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=2)  # Coluna B (Horas)
                cell.number_format = '0.00'  # Formato decimal com 2 casas
        
        # Formatar planilha de Motor Ocioso
        worksheet = workbook['4_Motor Ocioso']
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
            cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
        
        # Formatar planilha de Uso GPS
        worksheet = workbook['5_Uso GPS']
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
            cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
        
        # Formatar planilha de Motor Ocioso por Operação
        worksheet = workbook['6_Motor Ocioso por Operação']
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=2)  # Coluna B (Horas Ociosas)
            cell.number_format = '0.00'  # Formato decimal com 2 casas
    else:  # transbordo
        # Formatar planilha de Motor Ocioso
        worksheet = workbook['3_Motor Ocioso']
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
            cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
        
        # Formatar planilha de Uso GPS
        worksheet = workbook['4_Uso GPS']
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
            cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
        
        # Formatar planilha de Motor Ocioso por Operação
        worksheet = workbook['5_Motor Ocioso por Operação']
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=2)  # Coluna B (Horas Ociosas)
            cell.number_format = '0.00'  # Formato decimal com 2 casas
    
    # Formatar planilha de Base Calculo
    worksheet = workbook['Base Calculo']
    for row in range(2, worksheet.max_row + 1):
        # Formatar colunas decimais
        columns_decimal = [4, 5, 7, 8, 10, 12]  # Colunas D, E, G, H, J, L (Horas totais, Horas elevador, etc.)
        for col in columns_decimal:
            if col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                cell.number_format = '0.00'  # Formato decimal com 2 casas
        
        # Formatar colunas de porcentagem
        columns_percent = [6, 9, 11, 13]  # Colunas F, I, K, M (%, % Utilização RTK, etc.)
        for col in columns_percent:
            if col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Horas por Frota
    worksheet = workbook['Horas por Frota']
    for row in range(2, worksheet.max_row + 1):
        # Coluna B (Horas Registradas)
        cell_b = worksheet.cell(row=row, column=2)
        cell_b.number_format = '0.00'  # Formato decimal com 2 casas
        
        # Coluna C (Diferença para 24h)
        cell_c = worksheet.cell(row=row, column=3)
        cell_c.number_format = '0.00'  # Formato decimal com 2 casas
    
    writer.close()
    print(f"Arquivo Excel salvo com sucesso em {caminho_saida}")

def calcular_motor_ocioso_colhedora(df, config):
    """
    Calcula o motor ocioso especificamente para colhedoras.
    
    Args:
        df (DataFrame): DataFrame processado
        config (dict): Configurações carregadas do arquivo JSON
    
    Returns:
        DataFrame: Tempo de motor ocioso por operação para colhedoras
    """
    # Verificar se temos as colunas necessárias
    if 'Parada com Motor Ligado' not in df.columns:
        return pd.DataFrame(columns=['Operacao', 'Diferença_Hora', 'Tipo'])
    
    # Configurações específicas para colhedoras
    operacoes_excluidas = config.get('CD', {}).get('motor_ocioso', {}).get('operacoes_excluidas', [])
    grupos_operacao_excluidos = config.get('CD', {}).get('motor_ocioso', {}).get('grupos_operacao_excluidos', ['Manutenção'])
    
    # Filtrar registros com motor ocioso
    filtro = (df['Parada com Motor Ligado'] == 1)
    
    # Excluir operações específicas se configurado
    if operacoes_excluidas and 'Operacao' in df.columns:
        for operacao in operacoes_excluidas:
            if ' - ' in operacao:
                codigo, nome = operacao.split(' - ', 1)
                # Filtrar por código ou nome
                if 'Codigo da Operacao' in df.columns:
                    filtro = filtro & (~df['Codigo da Operacao'].astype(str).eq(codigo))
                # Filtrar por nome da operação (contém o texto)
                filtro = filtro & (~df['Operacao'].str.contains(nome, case=False, na=False))
            else:
                # Se não tiver o formato código - nome, considerar como nome
                if 'Operacao' in df.columns:
                    filtro = filtro & (~df['Operacao'].str.contains(operacao, case=False, na=False))
    
    # Excluir grupos de operação se configurado
    if grupos_operacao_excluidos and 'Grupo Operacao' in df.columns:
        filtro = filtro & (~df['Grupo Operacao'].isin(grupos_operacao_excluidos))
    
    # Aplicar filtro
    df_filtrado = df[filtro]
    
    # Verificar se temos dados após filtragem
    if df_filtrado.empty:
        return pd.DataFrame(columns=['Operacao', 'Diferença_Hora', 'Tipo'])
    
    # Agrupar por operação e somar o tempo ocioso
    if 'Operacao' in df_filtrado.columns:
        tempo_por_operacao = df_filtrado.groupby('Operacao')['Diferença_Hora'].sum().reset_index()
    else:
        # Se não temos a coluna Operacao, agrupar por Estado ou outro campo disponível
        campo_agrupamento = 'Estado' if 'Estado' in df_filtrado.columns else 'Grupo Operacao' if 'Grupo Operacao' in df_filtrado.columns else None
        
        if campo_agrupamento:
            tempo_por_operacao = df_filtrado.groupby(campo_agrupamento)['Diferença_Hora'].sum().reset_index()
            tempo_por_operacao = tempo_por_operacao.rename(columns={campo_agrupamento: 'Operacao'})
        else:
            # Se não temos nenhum campo para agrupar, retornar total
            tempo_total = df_filtrado['Diferença_Hora'].sum()
            tempo_por_operacao = pd.DataFrame([{'Operacao': 'Total', 'Diferença_Hora': tempo_total}])
    
    tempo_por_operacao['Tipo'] = 'Colhedora'
    
    return tempo_por_operacao

def calcular_motor_ocioso_transbordo(df, config):
    """
    Calcula o motor ocioso especificamente para transbordos.
    
    Args:
        df (DataFrame): DataFrame processado
        config (dict): Configurações carregadas do arquivo JSON
    
    Returns:
        DataFrame: Tempo de motor ocioso por operação para transbordos
    """
    # Verificar se temos as colunas necessárias
    if 'Parada com Motor Ligado' not in df.columns:
        return pd.DataFrame(columns=['Operacao', 'Diferença_Hora', 'Tipo'])
    
    # Configurações específicas para transbordos
    operacoes_excluidas = config.get('TT', {}).get('motor_ocioso', {}).get('operacoes_excluidas', [])
    grupos_operacao_excluidos = config.get('TT', {}).get('motor_ocioso', {}).get('grupos_operacao_excluidos', ['Manutenção'])
    
    # Filtrar registros com motor ocioso
    filtro = (df['Parada com Motor Ligado'] == 1)
    
    # Excluir operações específicas se configurado
    if operacoes_excluidas and 'Operacao' in df.columns:
        for operacao in operacoes_excluidas:
            if ' - ' in operacao:
                codigo, nome = operacao.split(' - ', 1)
                # Filtrar por código ou nome
                if 'Codigo da Operacao' in df.columns:
                    filtro = filtro & (~df['Codigo da Operacao'].astype(str).eq(codigo))
                # Filtrar por nome da operação (contém o texto)
                filtro = filtro & (~df['Operacao'].str.contains(nome, case=False, na=False))
            else:
                # Se não tiver o formato código - nome, considerar como nome
                if 'Operacao' in df.columns:
                    filtro = filtro & (~df['Operacao'].str.contains(operacao, case=False, na=False))
    
    # Excluir grupos de operação se configurado
    if grupos_operacao_excluidos and 'Grupo Operacao' in df.columns:
        filtro = filtro & (~df['Grupo Operacao'].isin(grupos_operacao_excluidos))
    
    # Aplicar filtro
    df_filtrado = df[filtro]
    
    # Verificar se temos dados após filtragem
    if df_filtrado.empty:
        return pd.DataFrame(columns=['Operacao', 'Diferença_Hora', 'Tipo'])
    
    # Agrupar por operação e somar o tempo ocioso
    if 'Operacao' in df_filtrado.columns:
        tempo_por_operacao = df_filtrado.groupby('Operacao')['Diferença_Hora'].sum().reset_index()
    else:
        # Se não temos a coluna Operacao, agrupar por Estado ou outro campo disponível
        campo_agrupamento = 'Estado' if 'Estado' in df_filtrado.columns else 'Grupo Operacao' if 'Grupo Operacao' in df_filtrado.columns else None
        
        if campo_agrupamento:
            tempo_por_operacao = df_filtrado.groupby(campo_agrupamento)['Diferença_Hora'].sum().reset_index()
            tempo_por_operacao = tempo_por_operacao.rename(columns={campo_agrupamento: 'Operacao'})
        else:
            # Se não temos nenhum campo para agrupar, retornar total
            tempo_total = df_filtrado['Diferença_Hora'].sum()
            tempo_por_operacao = pd.DataFrame([{'Operacao': 'Total', 'Diferença_Hora': tempo_total}])
    
    tempo_por_operacao['Tipo'] = 'Transbordo'
    
    return tempo_por_operacao

def extrair_txt_de_zip(arquivo_zip, diretorio_destino):
    """
    Extrai o primeiro arquivo TXT ou CSV de um arquivo ZIP e o renomeia com o nome do ZIP.
    
    Args:
        arquivo_zip (str): Caminho para o arquivo ZIP
        diretorio_destino (str): Diretório onde salvar o arquivo extraído
    
    Returns:
        str: Caminho do arquivo extraído, ou None se nenhum arquivo TXT/CSV for encontrado
    """
    try:
        print(f"Extraindo arquivos de {arquivo_zip}...")
        
        # Garantir que o diretório de destino existe
        if not os.path.exists(diretorio_destino):
            os.makedirs(diretorio_destino)
            print(f"Diretório de destino criado: {diretorio_destino}")
        
        # Nome base do arquivo ZIP (sem extensão)
        nome_base = os.path.splitext(os.path.basename(arquivo_zip))[0]
        
        # Abrir o arquivo ZIP
        with zipfile.ZipFile(arquivo_zip, 'r') as zip_ref:
            # Listar arquivos no ZIP
            arquivos = zip_ref.namelist()
            print(f"Encontrados {len(arquivos)} arquivos dentro do ZIP.")
            
            # Procurar por arquivos TXT primeiro
            arquivos_txt = [f for f in arquivos if f.lower().endswith('.txt')]
            
            # Se não encontrar TXT, procurar por CSV
            if not arquivos_txt:
                arquivos_txt = [f for f in arquivos if f.lower().endswith('.csv')]
            
            # Se ainda não encontrar nenhum arquivo adequado
            if not arquivos_txt:
                print(f"Nenhum arquivo TXT ou CSV encontrado em {arquivo_zip}")
                return None
            
            # Pegar o primeiro arquivo encontrado
            arquivo_para_extrair = arquivos_txt[0]
            print(f"Extraindo arquivo: {arquivo_para_extrair}")
            
            # Determinar a extensão do arquivo encontrado
            extensao = os.path.splitext(arquivo_para_extrair)[1]
            
            # Caminho de destino com o nome do ZIP
            arquivo_destino = os.path.join(diretorio_destino, f"{nome_base}{extensao}")
            
            # Extrair e renomear
            try:
                with zip_ref.open(arquivo_para_extrair) as arquivo_zip_aberto, open(arquivo_destino, 'wb') as arquivo_saida:
                    shutil.copyfileobj(arquivo_zip_aberto, arquivo_saida)
                
                # Verificar se o arquivo foi extraído corretamente
                if os.path.exists(arquivo_destino) and os.path.getsize(arquivo_destino) > 0:
                    print(f"Arquivo extraído com sucesso: {arquivo_destino}")
                    return arquivo_destino
                else:
                    print(f"Erro: Arquivo extraído {arquivo_destino} está vazio ou não existe.")
                    return None
            except Exception as e:
                print(f"Erro ao extrair o arquivo {arquivo_para_extrair}: {str(e)}")
                return None
    
    except zipfile.BadZipFile:
        print(f"Erro: {arquivo_zip} não é um arquivo ZIP válido.")
        return None
    except Exception as e:
        print(f"Erro ao extrair arquivo de {arquivo_zip}: {str(e)}")
        traceback.print_exc()
        return None

def detectar_tipo_equipamento(df):
    """
    Detecta automaticamente se o dataframe é de colhedora ou transbordo.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        str: 'colhedora' ou 'transbordo'
    """
    # Verificar colunas que são específicas para cada equipamento
    colunas_colhedora = ['Elevador Ativado', 'Elevador Invertido']
    colunas_transbordo = ['Peso da Carga', 'Velocidade de Descarga']
    
    # Contar quantas colunas específicas de cada tipo estão presentes
    cont_colhedora = sum(1 for col in colunas_colhedora if col in df.columns)
    cont_transbordo = sum(1 for col in colunas_transbordo if col in df.columns)
    
    # Se temos mais colunas de colhedora que de transbordo, é uma colhedora
    if cont_colhedora > cont_transbordo:
        return 'colhedora'
    # Se temos mais colunas de transbordo que de colhedora, é um transbordo
    elif cont_transbordo > cont_colhedora:
        return 'transbordo'
    
    # Se empatou ou não temos nenhuma coluna específica, verificar os valores
    # Verificar se há menção de 'Colhedora' ou 'Transbordo' em campos de equipamento ou frota
    if 'Equipamento' in df.columns:
        # Verificar se algum valor contém 'colhedora' (case insensitive)
        if df['Equipamento'].str.contains('colhedora', case=False, na=False).any():
            return 'colhedora'
        # Verificar se algum valor contém 'transbordo' (case insensitive)
        if df['Equipamento'].str.contains('transbordo', case=False, na=False).any():
            return 'transbordo'
    
    if 'Frota' in df.columns:
        # Verificar se algum valor contém 'colhedora' (case insensitive)
        if df['Frota'].str.contains('colhedora', case=False, na=False).any():
            return 'colhedora'
        # Verificar se algum valor contém 'transbordo' (case insensitive)
        if df['Frota'].str.contains('transbordo', case=False, na=False).any():
            return 'transbordo'
    
    # Se ainda não conseguimos determinar, verificar a frequência de 'Elevador Ativado' se existir
    if 'Elevador Ativado' in df.columns:
        # Se mais de 1% dos registros têm o elevador ativado, é provavelmente uma colhedora
        if df['Elevador Ativado'].mean() > 0.01:
            return 'colhedora'
    
    # Default para colhedora se não conseguirmos determinar
    return 'colhedora'

def processar_colhedora(df_original, config, caminho_arquivo_saida=None):
    """
    Processa dados específicos de colhedoras e gera relatórios.
    
    Args:
        df_original (DataFrame): DataFrame processado
        config (dict): Configurações carregadas do arquivo JSON
        caminho_arquivo_saida (str, opcional): Caminho para salvar o arquivo Excel
        
    Returns:
        tuple: (base_calculo, dict com DataFrames de métricas calculadas)
    """
    # Criar uma cópia para não modificar o original
    df = df_original.copy()
    
    # Verificar colunas necessárias
    colunas_necessarias = ['Data', 'Frota', 'Equipamento', 'Diferença_Hora']
    for coluna in colunas_necessarias:
        if coluna not in df.columns:
            print(f"Aviso: Coluna {coluna} não encontrada para cálculos de colhedora")
            # Criar coluna vazia se não existir
            if coluna in ['Data', 'Frota', 'Equipamento']:
                df[coluna] = 'Não Informado'
            else:
                df[coluna] = 0
    
    # Calcular base de cálculo
    base_calculo = calcular_base_calculo(df)
    
    # Calcular métricas específicas para colhedoras
    metricas = {}
    
    # Disponibilidade mecânica
    metricas['disponibilidade_mecanica'] = calcular_disponibilidade_mecanica(df)
    
    # Eficiência energética
    metricas['eficiencia_energetica'] = calcular_eficiencia_energetica(base_calculo)
    
    # Hora de elevador
    metricas['hora_elevador'] = calcular_hora_elevador(df, base_calculo)
    
    # Motor ocioso
    metricas['motor_ocioso'] = calcular_motor_ocioso_colhedora(df, config)
    
    # Uso de GPS
    metricas['uso_gps'] = calcular_uso_gps(df, base_calculo)
    
    # Horas por frota
    metricas['horas_por_frota'] = calcular_horas_por_frota(df)
    
    # Salvar os resultados em Excel se um caminho foi fornecido
    if caminho_arquivo_saida:
        criar_excel_com_planilhas(df_original, base_calculo, metricas['disponibilidade_mecanica'], metricas['eficiencia_energetica'],
                                 metricas['hora_elevador'], metricas['motor_ocioso'], metricas['uso_gps'], metricas['horas_por_frota'],
                                 metricas['motor_ocioso_por_operacao'], caminho_arquivo_saida, detectar_tipo_equipamento(df_original))
    
    return base_calculo, metricas

def processar_transbordo(df_original, config, caminho_arquivo_saida=None):
    """
    Processa dados específicos de transbordos e gera relatórios.
    
    Args:
        df_original (DataFrame): DataFrame processado
        config (dict): Configurações carregadas do arquivo JSON
        caminho_arquivo_saida (str, opcional): Caminho para salvar o arquivo Excel
        
    Returns:
        tuple: (base_calculo, dict com DataFrames de métricas calculadas)
    """
    # Criar uma cópia para não modificar o original
    df = df_original.copy()
    
    # Verificar colunas necessárias
    colunas_necessarias = ['Data', 'Frota', 'Equipamento', 'Diferença_Hora']
    for coluna in colunas_necessarias:
        if coluna not in df.columns:
            print(f"Aviso: Coluna {coluna} não encontrada para cálculos de transbordo")
            # Criar coluna vazia se não existir
            if coluna in ['Data', 'Frota', 'Equipamento']:
                df[coluna] = 'Não Informado'
            else:
                df[coluna] = 0
    
    # Calcular base de cálculo
    base_calculo = calcular_base_calculo(df)
    
    # Calcular métricas específicas para transbordos
    metricas = {}
    
    # Disponibilidade mecânica
    metricas['disponibilidade_mecanica'] = calcular_disponibilidade_mecanica(df)
    
    # Eficiência energética
    metricas['eficiencia_energetica'] = calcular_eficiencia_energetica(base_calculo)
    
    # Motor ocioso para transbordos
    metricas['motor_ocioso'] = calcular_motor_ocioso_transbordo(df, config)
    
    # Uso de GPS
    metricas['uso_gps'] = calcular_uso_gps(df, base_calculo)
    
    # Horas por frota
    metricas['horas_por_frota'] = calcular_horas_por_frota(df)
    
    # Métricas específicas para transbordos podem ser adicionadas aqui
    
    # Salvar os resultados em Excel se um caminho foi fornecido
    if caminho_arquivo_saida:
        criar_excel_com_planilhas(df_original, base_calculo, metricas['disponibilidade_mecanica'], metricas['eficiencia_energetica'],
                                 None, metricas['motor_ocioso'], metricas['uso_gps'], metricas['horas_por_frota'],
                                 metricas['motor_ocioso_por_operacao'], caminho_arquivo_saida, detectar_tipo_equipamento(df_original))
    
    return base_calculo, metricas

def processar_arquivo(caminho_arquivo, config, diretorio_saida=None, tipo_forcado=None):
    """
    Processa um arquivo TXT, CSV ou ZIP e gera um relatório Excel.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo a ser processado
        config (dict): Configurações carregadas do arquivo JSON
        diretorio_saida (str, opcional): Diretório onde salvar o arquivo de saída
        tipo_forcado (str, opcional): Forçar o tipo de equipamento ('colhedora' ou 'transbordo')
        
    Returns:
        str: Caminho do arquivo Excel gerado ou None em caso de erro
    """
    try:
        print(f"Processando arquivo: {caminho_arquivo}")
        
        # Criar diretório de saída se não existir
        if diretorio_saida and not os.path.exists(diretorio_saida):
            os.makedirs(diretorio_saida)
        
        arquivo_para_processar = caminho_arquivo
        
        # Se for um arquivo ZIP, extrair primeiro
        if caminho_arquivo.lower().endswith('.zip'):
            temp_dir = tempfile.mkdtemp() if not diretorio_saida else diretorio_saida
            arquivo_extraido = extrair_txt_de_zip(caminho_arquivo, temp_dir)
            
            if not arquivo_extraido:
                print(f"Nenhum arquivo válido encontrado no ZIP: {caminho_arquivo}")
                return None
            
            arquivo_para_processar = arquivo_extraido
        
        # Carregar o arquivo como DataFrame
        print(f"Carregando arquivo: {arquivo_para_processar}")
        df, tipo_equipamento_detectado = processar_arquivo_base(arquivo_para_processar)
        
        if df is None or len(df) == 0:
            print(f"Erro ao processar o arquivo ou arquivo vazio: {arquivo_para_processar}")
            return None
        
        # Usar o tipo forçado se fornecido, caso contrário usar o tipo detectado
        tipo_equipamento = tipo_forcado if tipo_forcado else tipo_equipamento_detectado
        print(f"Tipo de equipamento: {tipo_equipamento}")
        
        # Nome base para o arquivo de saída (sem extensão)
        nome_base = os.path.splitext(os.path.basename(arquivo_para_processar))[0]
        
        # Caminho completo do arquivo de saída
        caminho_saida = None
        if diretorio_saida:
            caminho_saida = os.path.join(diretorio_saida, f"{nome_base}.xlsx")
        
        # Calcular métricas básicas
        base_calculo = calcular_base_calculo(df)
        disp_mecanica = calcular_disponibilidade_mecanica(df)
        eficiencia_energetica = calcular_eficiencia_energetica(base_calculo)
        uso_gps = calcular_uso_gps(df, base_calculo)
        horas_por_frota = calcular_horas_por_frota(df)
        
        # Processar conforme o tipo de equipamento
        if tipo_equipamento == 'colhedora':
            hora_elevador = calcular_hora_elevador(df, base_calculo)
            motor_ocioso = calcular_motor_ocioso_colhedora(df, config)
            motor_ocioso_por_operacao = calcular_motor_ocioso_colhedora(df, config)
            
            # Criar o arquivo Excel
            if caminho_saida:
                try:
                    criar_excel_com_planilhas(
                        df, base_calculo, disp_mecanica, eficiencia_energetica,
                        hora_elevador, motor_ocioso, uso_gps, horas_por_frota, 
                        motor_ocioso_por_operacao, caminho_saida, tipo_equipamento
                    )
                    
                    # Verificar se o arquivo foi realmente criado
                    if os.path.exists(caminho_saida):
                        return caminho_saida
                    else:
                        print(f"Erro: Arquivo Excel não foi criado em {caminho_saida}")
                        return None
                except Exception as e:
                    print(f"Erro ao criar arquivo Excel: {str(e)}")
                    return None
        else:  # transbordo
            motor_ocioso = calcular_motor_ocioso_transbordo(df, config)
            motor_ocioso_por_operacao = calcular_motor_ocioso_transbordo(df, config)
            
            # Criar o arquivo Excel (sem hora elevador para transbordos)
            if caminho_saida:
                try:
                    criar_excel_com_planilhas(
                        df, base_calculo, disp_mecanica, eficiencia_energetica,
                        None, motor_ocioso, uso_gps, horas_por_frota, 
                        motor_ocioso_por_operacao, caminho_saida, tipo_equipamento
                    )
                    
                    # Verificar se o arquivo foi realmente criado
                    if os.path.exists(caminho_saida):
                        return caminho_saida
                    else:
                        print(f"Erro: Arquivo Excel não foi criado em {caminho_saida}")
                        return None
                except Exception as e:
                    print(f"Erro ao criar arquivo Excel: {str(e)}")
                    return None
        
        return caminho_saida
    
    except Exception as e:
        print(f"Erro ao processar arquivo {caminho_arquivo}: {str(e)}")
        traceback.print_exc()
        return None

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT ou CSV na pasta de dados e gera arquivos processados na pasta de saída.
    Também processa arquivos ZIP, extraindo os TXT ou CSV contidos dentro deles.
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Diretórios para dados de entrada e saída
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    diretorio_temp = os.path.join(diretorio_raiz, "temp")
    diretorio_config = os.path.join(diretorio_raiz, "config")
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    
    # Verificar se os diretórios existem, caso contrário criar
    for diretorio in [diretorio_saida, diretorio_temp, diretorio_config, diretorio_dados]:
        if not os.path.exists(diretorio):
            os.makedirs(diretorio)
            print(f"Diretório criado: {diretorio}")
    
    # Criar arquivo de configuração padrão se não existir
    arquivo_config = os.path.join(diretorio_config, "calculos_config.json")
    if not os.path.exists(arquivo_config):
        config_padrao = {
            "CD": {
                "motor_ocioso": {
                    "tipo_calculo": "Remover do cálculo",
                    "operacoes_excluidas": [
                        "8490 - LAVAGEM",
                        "MANUTENCAO",
                        "LAVAGEM",
                        "INST CONFIG TECNOL EMBARCADAS"
                    ],
                    "grupos_operacao_excluidos": ["Manutenção"]
                },
                "operadores_excluidos": ["9999 - TROCA DE TURNO"],
                "equipamentos_excluidos": []
            },
            "TT": {
                "motor_ocioso": {
                    "tipo_calculo": "Remover do cálculo",
                    "operacoes_excluidas": [
                        "9016 - ENCH SISTEMA FREIO",
                        "6340 - BASCULANDO  TRANSBORDAGEM",
                        "9024 - DESATOLAMENTO",
                        "MANUTENÇÃO",
                        "INST CONFIG TECNOL EMBARCADAS",
                        "DESATOLAMENTO"
                    ],
                    "grupos_operacao_excluidos": ["Manutenção"]
                },
                "operadores_excluidos": ["9999 - TROCA DE TURNO"],
                "equipamentos_excluidos": []
            }
        }
        
        try:
            with open(arquivo_config, 'w', encoding='utf-8') as f:
                json.dump(config_padrao, f, indent=2, ensure_ascii=False)
            print(f"Arquivo de configuração padrão criado: {arquivo_config}")
        except Exception as e:
            print(f"Erro ao criar arquivo de configuração padrão: {str(e)}")
    
    # Carregar configurações
    config = carregar_config_calculos()
    
    # Verificar arquivos em todas as pastas de dados
    diretorios_dados = [
        os.path.join(diretorio_raiz, "dados"),
        os.path.join(diretorio_raiz, "dados", "colhedoras"),
        os.path.join(diretorio_raiz, "dados", "transbordos")
    ]
    
    # Lista para armazenar todos os arquivos a serem processados
    arquivos_processar = []
    
    # Lista para armazenar arquivos ZIP e seus respectivos arquivos extraídos
    arquivos_zip_info = []  # Lista de tuplas (arquivo_zip, arquivo_extraido, arquivo_saida)
    
    # Primeiro, procurar e extrair arquivos ZIP
    for diretorio in diretorios_dados:
        if os.path.exists(diretorio):
            # Encontrar todos os arquivos ZIP
            arquivos_zip = glob.glob(os.path.join(diretorio, "*.zip"))
            
            for arquivo_zip in arquivos_zip:
                # Extrair TXT ou CSV do ZIP para a pasta temp
                arquivo_extraido = extrair_txt_de_zip(arquivo_zip, diretorio_temp)
                
                if arquivo_extraido:
                    # Copiar o arquivo extraído para a pasta dados
                    nome_arquivo = os.path.basename(arquivo_extraido)
                    arquivo_dados = os.path.join(diretorio_dados, nome_arquivo)
                    try:
                        shutil.copy2(arquivo_extraido, arquivo_dados)
                        print(f"Arquivo copiado para pasta dados: {arquivo_dados}")
                        # Adicionar o arquivo copiado à lista de processamento
                        arquivos_processar.append(arquivo_dados)
                        # Calcular nome do arquivo de saída
                        nome_base = os.path.splitext(nome_arquivo)[0]
                        arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}.xlsx")
                        # Guardar informações para verificação posterior
                        arquivos_zip_info.append((arquivo_zip, arquivo_extraido, arquivo_saida))
                    except Exception as e:
                        print(f"Erro ao copiar arquivo para pasta dados: {str(e)}")
                        # Se falhou a cópia, usa o arquivo na pasta temp
                        arquivos_processar.append(arquivo_extraido)
                        nome_base = os.path.splitext(os.path.basename(arquivo_extraido))[0]
                        arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}.xlsx")
                        arquivos_zip_info.append((arquivo_zip, arquivo_extraido, arquivo_saida))
    
    # Agora, procurar arquivos TXT e CSV regulares
    for diretorio in diretorios_dados:
        if os.path.exists(diretorio):
            # Adicionar todos os TXT e CSV encontrados à lista de processamento
            arquivos_txt = glob.glob(os.path.join(diretorio, "*.txt"))
            for txt in arquivos_txt:
                if txt not in arquivos_processar:  # Evitar duplicatas
                    arquivos_processar.append(txt)
            
            if processCsv:
                arquivos_csv = glob.glob(os.path.join(diretorio, "*.csv"))
                for csv in arquivos_csv:
                    if csv not in arquivos_processar:  # Evitar duplicatas
                        arquivos_processar.append(csv)
    
    if not arquivos_processar:
        print("Nenhum arquivo TXT ou CSV encontrado nas pastas de dados ou dentro de ZIPs!")
        return
    
    print(f"Encontrados {len(arquivos_processar)} arquivos para processar.")
    
    # Processamentos bem-sucedidos (arquivos Excel gerados)
    arquivos_gerados = []
    
    # Processar cada arquivo
    for arquivo in arquivos_processar:
        nome_base = os.path.splitext(os.path.basename(arquivo))[0]
        # Nome de saída igual ao original, mas com extensão .xlsx na pasta output
        arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}.xlsx")
        
        extensao = os.path.splitext(arquivo)[1].lower()
        print(f"\nProcessando arquivo: {nome_base}{extensao}")
        print(f"Arquivo de saída esperado: {os.path.basename(arquivo_saida)}")
        
        # Determinar o tipo de equipamento pelo caminho
        tipo_forcado = None
        if "colhedora" in arquivo.lower():
            tipo_forcado = "colhedora"
        elif "transbordo" in arquivo.lower():
            tipo_forcado = "transbordo"
        
        # Processar o arquivo
        resultado = processar_arquivo(arquivo, config, diretorio_saida, tipo_forcado)
        
        # Verificar se o arquivo Excel foi gerado com sucesso
        if resultado and os.path.exists(arquivo_saida):
            print(f"Arquivo {arquivo_saida} gerado com sucesso!")
            arquivos_gerados.append(arquivo_saida)
        else:
            print(f"ATENÇÃO: Arquivo {arquivo_saida} NÃO foi gerado corretamente!")
    
    # Excluir os arquivos ZIP originais apenas se seus respectivos arquivos Excel foram gerados
    for arquivo_zip, arquivo_extraido, arquivo_saida in arquivos_zip_info:
        if arquivo_saida in arquivos_gerados:
            try:
                os.remove(arquivo_zip)
                print(f"Excluído ZIP original: {os.path.basename(arquivo_zip)}")
                # Também remover o arquivo extraído se estiver na pasta temp
                if os.path.exists(arquivo_extraido) and diretorio_temp in arquivo_extraido:
                    os.remove(arquivo_extraido)
                    print(f"Excluído arquivo temporário: {os.path.basename(arquivo_extraido)}")
            except Exception as e:
                print(f"Erro ao excluir {os.path.basename(arquivo_zip)}: {str(e)}")
        else:
            print(f"ATENÇÃO: Mantendo ZIP original {os.path.basename(arquivo_zip)} pois o Excel não foi gerado corretamente!")
    
    print("\nProcessamento concluído!")

if __name__ == "__main__":
    print("Iniciando processamento de arquivos...")
    processar_todos_arquivos()
    print("\nProcessamento concluído!") 