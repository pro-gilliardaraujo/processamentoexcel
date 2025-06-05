"""
Script para processamento completo de dados de monitoramento de colhedoras.
Lê arquivos TXT ou CSV na pasta raiz, processa-os e gera arquivos Excel com planilhas auxiliares prontas.
Também processa arquivos ZIP contendo TXT ou CSV.
"""

import pandas as pd
import numpy as np
import os
import glob
import zipfile
import tempfile
import shutil
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurações
processCsv = False  # Altere para True quando quiser processar arquivos CSV

# Constantes
COLUNAS_REMOVER = [
    'Justificativa Corte Base Desligado',
    'Regional',
    'Tipo de Equipamento',
    'Unidade',
    'Centro de Custo',
    'Trabalhando em File',
    'Trabalhando Frente Dividida',
    'Trabalhando em Fila'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento', 'Apertura do Rolo', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo', 'Corte Base Automatico/Manual',
    'Descricao Equipamento', 'Estado', 'Estado Operacional', 'Esteira Ligada',
    'Field Cruiser', 'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Implemento Ligado', 'Motor Ligado', 'Operacao', 'Operador', 'Pressao de Corte',
    'RPM Extrator', 'RPM Motor', 'RTK (Piloto Automatico)', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parada com Motor Ligado',
    'Horas Produtivas',
]

# Valores a serem filtrados
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO", "1 - SEM OPERADOR"]

# Mapeamento de valores booleanos para 1/0
MAPEAMENTO_BOOLEANO = {
    'VERDADEIRO': 1, 'FALSO': 0,
    'TRUE': 1, 'FALSE': 0,
    'LIGADO': 1, 'DESLIGADO': 0,
    True: 1, False: 0,
    1: 1, 0: 0
}

# Adicionar função para extrair frente antes da função carregar_config_calculos()
def extrair_frente(grupo_equipamento_frente):
    """
    Extrai a frente da coluna 'Grupo Equipamento/Frente'.
    
    Args:
        grupo_equipamento_frente (str): Valor da coluna 'Grupo Equipamento/Frente'
    
    Returns:
        str: Nome da frente extraído ou 'Não informado' se não conseguir extrair
    """
    if pd.isna(grupo_equipamento_frente) or grupo_equipamento_frente == '':
        return 'Não informado'
    
    # Tentar extrair a frente (assumindo formato "GRUPO/FRENTE" ou similar)
    try:
        # Se contém "/", pega a parte após a barra
        if '/' in str(grupo_equipamento_frente):
            return str(grupo_equipamento_frente).split('/')[-1].strip()
        # Se contém "-", pega a parte após o traço
        elif '-' in str(grupo_equipamento_frente):
            return str(grupo_equipamento_frente).split('-')[-1].strip()
        # Caso contrário, usa o valor completo
        else:
            return str(grupo_equipamento_frente).strip()
    except:
        return 'Não informado'

def carregar_config_calculos():
    """
    Retorna as configurações de cálculos embutidas diretamente no código.
    As configurações são as mesmas definidas no arquivo calculos_config.json.
    """
    # Configurações embutidas diretamente no código
    config = {
        "CD": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "8490 - LAVAGEM",
                    "MANUTENCAO",
                    "LAVAGEM",
                    "INST CONFIG TECNOL EMBARCADAS",
                    "1055 - MANOBRA"
                ],
                "grupos_operacao_excluidos": [
                    "Manutenção", 
                    "Inaptidão"
                ]
            },
            "equipamentos_excluidos": []
        },
        "TT": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [
                    "9016 - ENCH SISTEMA FREIO",
                    "6340 - BASCULANDO  TRANSBORDAGEM",
                    "9024 - DESATOLAMENTO",
                    "MANUTENCAO",
                    "MANUTENÇÃO",
                    "INST CONFIG TECNOL EMBARCADAS",
                    "1055 - MANOBRA"
                ],
                "grupos_operacao_excluidos": [
                    "Manutenção", 
                    "Inaptidão"
                ]
            },
            "equipamentos_excluidos": []
        }
    }
    
    print("Usando configurações embutidas no código, ignorando o arquivo calculos_config.json")
    return config

def calcular_motor_ocioso_novo(df):
    """
    Calcula o tempo de motor ocioso de acordo com as novas regras:
    1. Intervalo é fechado quando encontra 'Parada com Motor Ligado = 0' com duração > 1 minuto
    2. Soma sequências de 'Parada com Motor Ligado = 1' com 'Parada com Motor Ligado = 0' < 1 minuto
    3. Se o total > 1 minuto, subtrai 1 minuto; se menor, descarta o intervalo
    
    Args:
        df (DataFrame): DataFrame com os dados de operação
        
    Returns:
        DataFrame: DataFrame com a coluna 'Motor Ocioso' atualizada
    """
    # Converter a coluna de diferença para minutos
    df['Diferença_Minutos'] = df['Diferença_Hora'] * 60
    
    # Inicializar colunas com tipos corretos
    df['Motor Ocioso'] = 0.0  # float
    df['Em_Intervalo'] = False  # bool
    df['Soma_Intervalo'] = 0.0  # float
    
    # Variáveis para controle do intervalo atual
    em_intervalo = False
    soma_intervalo = 0
    inicio_intervalo = None
    
    # Iterar sobre as linhas do DataFrame
    for i in range(len(df)):
        parada_motor = df.iloc[i]['Parada com Motor Ligado']
        diferenca = df.iloc[i]['Diferença_Minutos']
        
        # Se não estamos em um intervalo
        if not em_intervalo:
            # Se encontrar Parada com Motor Ligado = 1, inicia novo intervalo
            if parada_motor == 1:
                em_intervalo = True
                soma_intervalo = diferenca
                inicio_intervalo = i
                df.at[i, 'Em_Intervalo'] = True
                df.at[i, 'Soma_Intervalo'] = float(soma_intervalo)  # Converter para float
        
        # Se estamos em um intervalo
        else:
            # Se encontrar Parada com Motor Ligado = 0
            if parada_motor == 0:
                # Se a duração for > 1 minuto, fecha o intervalo
                if diferenca > 1:
                    # Se o total acumulado > 1 minuto, subtrai 1 minuto
                    if soma_intervalo > 1:
                        tempo_ocioso = soma_intervalo - 1
                        # IMPORTANTE: Converter de minutos para horas antes de atribuir
                        df.at[inicio_intervalo, 'Motor Ocioso'] = float(tempo_ocioso / 60.0)  # Dividir por 60 para converter minutos em horas
                    
                    # Reseta o intervalo
                    em_intervalo = False
                    soma_intervalo = 0
                    inicio_intervalo = None
                else:
                    # Se <= 1 minuto, soma ao intervalo atual
                    soma_intervalo += diferenca
                    df.at[i, 'Em_Intervalo'] = True
                    df.at[i, 'Soma_Intervalo'] = float(soma_intervalo)  # Converter para float
            
            # Se encontrar Parada com Motor Ligado = 1
            else:
                soma_intervalo += diferenca
                df.at[i, 'Em_Intervalo'] = True
                df.at[i, 'Soma_Intervalo'] = float(soma_intervalo)  # Converter para float
    
    # Tratar último intervalo aberto, se houver
    if em_intervalo and soma_intervalo > 1:
        tempo_ocioso = soma_intervalo - 1
        # CORREÇÃO: Converter de minutos para horas antes de atribuir
        df.at[inicio_intervalo, 'Motor Ocioso'] = float(tempo_ocioso / 60.0)  # Dividir por 60 para converter minutos em horas
    
    # Remover colunas auxiliares
    df = df.drop(['Diferença_Minutos', 'Em_Intervalo', 'Soma_Intervalo'], axis=1)
    
    # Garantir que o tempo ocioso nunca seja maior que o tempo ligado para cada registro
    for i in range(len(df)):
        if df.iloc[i]['Motor Ocioso'] > 0:
            # Se o motor estiver ligado, limitar o tempo ocioso ao tempo ligado
            if df.iloc[i]['Motor Ligado'] == 1:
                tempo_hora = df.iloc[i]['Diferença_Hora']
                if df.iloc[i]['Motor Ocioso'] > tempo_hora:
                    df.at[i, 'Motor Ocioso'] = tempo_hora
            else:
                # Se o motor não estiver ligado, o tempo ocioso deve ser zero
                df.at[i, 'Motor Ocioso'] = 0.0
    
    return df

def processar_arquivo_base(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo (TXT ou CSV são tratados da mesma forma se usarem separador ';')
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio (apenas cabeçalhos sem dados)
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                # Retornar o DataFrame vazio mas com as colunas, em vez de None
                # Garantir que todas as colunas desejadas existam
                for col in COLUNAS_DESEJADAS:
                    if col not in df.columns:
                        df[col] = np.nan
                # Reorganizar as colunas na ordem desejada
                colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Verificar se 'Data/Hora' existe e processá-la
            if 'Data/Hora' in df.columns:
                # Processamento de data e hora
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
                
                # Conversão e cálculo de diferenças de hora
                df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S')
                df['Diferença_Hora'] = df['Hora'].diff().dt.total_seconds() / 3600
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: max(x, 0))
                
                # Nova regra: se Diferença_Hora > 0.50, então 0
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else x)
            
            # Cálculos adicionais
            RPM_MINIMO = 300  # Definindo constante para RPM mínimo

            # Carregar configurações
            config = carregar_config_calculos()
            operacoes_excluidas = config['CD']['motor_ocioso']['operacoes_excluidas']
            grupos_operacao_excluidos = config['CD']['motor_ocioso']['grupos_operacao_excluidos']

            # Filtrar dados para cálculo de motor ocioso
            df_motor_ocioso = df[
                ~df['Operacao'].isin(operacoes_excluidas) & 
                ~df['Grupo Operacao'].isin(grupos_operacao_excluidos)
            ]

            # Calcular Parada com Motor Ligado usando dados filtrados
            df['Parada com Motor Ligado'] = 0  # Inicializa com 0
            df.loc[df_motor_ocioso.index, 'Parada com Motor Ligado'] = (
                (df_motor_ocioso['Velocidade'] == 0) & 
                (df_motor_ocioso['RPM Motor'] >= RPM_MINIMO)
            ).astype(int)
            
            # Aplicar o novo cálculo de motor ocioso com intervalos de 1 minuto
            df = calcular_motor_ocioso_novo(df)
            
            # Verifica se Horas Produtivas já existe
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                # Calcular horas produtivas sem arredondamento, mantendo a precisão completa
                df['Horas Produtivas'] = df.apply(
                    lambda row: row['Diferença_Hora'] if row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
            else:
                # Limpa e converte para número
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # IMPORTANTE: Zerar horas produtivas dos operadores excluídos para garantir que não sejam contabilizadas
            df.loc[df['Operador'].isin(OPERADORES_EXCLUIR), 'Horas Produtivas'] = 0
            print(f"Total de horas produtivas após exclusão de operadores: {df['Horas Produtivas'].sum():.2f}")
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    # Se chegou aqui, todas as tentativas de codificação falharam
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado.
    Calcula médias diárias considerando os dias efetivos de trabalho de cada operador.
    
    Cálculos principais:
    - Horas totais: soma de Diferença_Hora
    - Horas elevador: soma de Diferença_Hora onde Esteira Ligada = 1 E Pressão de Corte > 400
    - Motor Ligado: soma de Diferença_Hora onde Motor Ligado = 1
    - Parado Com Motor Ligado: MÉTODO AVANÇADO - soma da coluna Motor Ocioso, que usa o cálculo com intervalos
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Detectar número de dias totais nos dados (apenas para informação)
    dias_unicos_total = df['Data'].nunique() if 'Data' in df.columns else 1
    print(f"Detectados {dias_unicos_total} dias distintos na base de dados.")
    
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Determinar o número de dias efetivos para este operador
        dias_operador = dados_filtrados['Data'].nunique() if 'Data' in dados_filtrados.columns else 1
        
        # Horas totais - soma de Diferença_Hora (IGUAL AO ORIGINAL)
        horas_totais = dados_filtrados['Diferença_Hora'].sum()
        if dias_operador > 1:
            horas_totais = horas_totais / dias_operador
        
        # Motor Ligado - soma de Diferença_Hora onde Motor Ligado = 1 (IGUAL AO ORIGINAL)
        motor_ligado = dados_filtrados[
            dados_filtrados['Motor Ligado'] == 1
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            motor_ligado = motor_ligado / dias_operador
        
        # Horas elevador - soma de Diferença_Hora onde Esteira Ligada = 1 E Pressão de Corte > 400 (IGUAL AO ORIGINAL)
        horas_elevador = dados_filtrados[
            (dados_filtrados['Esteira Ligada'] == 1) & 
            (dados_filtrados['Pressao de Corte'] > 400)
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            horas_elevador = horas_elevador / dias_operador
        
        # Percentual horas elevador (em decimal 0-1)
        percent_elevador = calcular_porcentagem(horas_elevador, horas_totais)
        
        # RTK - soma de Diferença_Hora onde todas as condições são atendidas (IGUAL AO ORIGINAL)
        rtk = dados_filtrados[
            (dados_filtrados['Grupo Operacao'] == 'Produtiva') &
            (dados_filtrados['Pressao de Corte'] > 300) &
            (dados_filtrados['RTK (Piloto Automatico)'] == 1) &
            (dados_filtrados['Esteira Ligada'] == 1)
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            rtk = rtk / dias_operador
        
        # Horas Produtivas (IGUAL AO ORIGINAL)
        horas_produtivas = dados_filtrados[
            dados_filtrados['Grupo Operacao'] == 'Produtiva'
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            horas_produtivas = horas_produtivas / dias_operador
        
        # % Utilização RTK (em decimal 0-1)
        utilizacao_rtk = calcular_porcentagem(rtk, horas_produtivas)
        
        # % Eficiência Elevador (em decimal 0-1)
        eficiencia_elevador = calcular_porcentagem(horas_elevador, motor_ligado)
        
        # NOVO MÉTODO: Parado com Motor Ligado - usando o valor calculado pela função calcular_motor_ocioso_novo
        # A coluna 'Motor Ocioso' contém o tempo ocioso após aplicar a lógica de intervalos e tolerância
        parado_motor_ligado = dados_filtrados['Motor Ocioso'].sum()
        if dias_operador > 1:
            parado_motor_ligado = parado_motor_ligado / dias_operador
        
        # % Parado com motor ligado (em decimal 0-1)
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        # Debug para verificar os valores
        print(f"\nOperador: {operador} em {equipamento}")
        print(f"Motor Ligado: {motor_ligado:.6f}")
        print(f"Parado com Motor Ligado (método avançado): {parado_motor_ligado:.6f}")
        print(f"% Parado com motor ligado: {percent_parado_motor:.6f}")
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas elevador': horas_elevador,
            '%': percent_elevador,
            'RTK': rtk,
            'Horas Produtivas': horas_produtivas,
            '% Utilização RTK': utilizacao_rtk,
            'Motor Ligado': motor_ligado,
            '% Eficiência Elevador': eficiencia_elevador,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor
        })
    
    return pd.DataFrame(resultados)

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento e frente.
    Calcula médias diárias considerando os dias efetivos de cada equipamento.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento e frente
    """
    # Filtramos os dados excluindo os operadores da lista
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Extrair frente da coluna 'Grupo Equipamento/Frente'
    df_filtrado = df_filtrado.copy()
    df_filtrado['Frente'] = df_filtrado['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por Equipamento e Frente
    grupos = df_filtrado.groupby(['Equipamento', 'Frente'])
    resultados = []
    
    for (equipamento, frente), dados_grupo in grupos:
        # Determinar número de dias efetivos para este equipamento e frente
        dias_grupo = dados_grupo['Data'].nunique() if 'Data' in dados_grupo.columns else 1
        
        total_horas = round(dados_grupo['Diferença_Hora'].sum(), 4)
        
        # Calcular horas de manutenção
        manutencao = round(dados_grupo[dados_grupo['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum(), 4)
        
        # Se houver múltiplos dias, usar médias diárias
        if dias_grupo > 1:
            total_horas = round(total_horas / dias_grupo, 4)
            manutencao = round(manutencao / dias_grupo, 4)
            print(f"Equipamento: {equipamento}, Frente: {frente}, Dias efetivos: {dias_grupo}, Média diária: {total_horas:.2f} horas")
        
        # A disponibilidade mecânica é o percentual de tempo fora de manutenção
        disp_mecanica = calcular_porcentagem(total_horas - manutencao, total_horas)
        
        resultados.append({
            'Frota': equipamento,
            'Frente': frente,
            'Disponibilidade': disp_mecanica
        })
    
    # Ordenar primeiro por frente, depois por disponibilidade (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values(['Frente', 'Disponibilidade'], ascending=[True, False])

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e frente, e a diferença para 24 horas.
    Calcula médias diárias considerando os dias efetivos de cada frota e frente.
    Esta função NÃO aplica qualquer filtro de operador.
    Também identifica as faltas de horário por dia específico.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota e frente com detalhamento por dia
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente'
    df = df.copy()
    df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Agrupar por Equipamento e Frente
    grupos = df.groupby(['Equipamento', 'Frente'])
    resultados = []
    
    # Obter todos os dias únicos no dataset
    dias_unicos = sorted(df['Data'].unique()) if 'Data' in df.columns else []
    
    for (equipamento, frente), dados_grupo in grupos:
        # Determinar número de dias efetivos para este equipamento e frente
        dias_grupo = dados_grupo['Data'].nunique() if 'Data' in dados_grupo.columns else 1
        
        total_horas = round(dados_grupo['Diferença_Hora'].sum(), 2)
        
        # Se houver múltiplos dias, usar média diária
        if dias_grupo > 1:
            total_horas = round(total_horas / dias_grupo, 2)
        
        # Calcular a diferença para 24 horas
        diferenca_24h = round(max(24 - total_horas, 0), 2)
        
        # Criar o resultado básico (colunas originais mantidas)
        resultado = {
            'Frota': equipamento,
            'Frente': frente,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        }
        
        # Adicionar detalhamento por dia (novas colunas)
        if len(dias_unicos) > 0:
            for dia in dias_unicos:
                dados_dia = dados_grupo[dados_grupo['Data'] == dia]  # CORREÇÃO: usar dados_grupo em vez de dados_equip
                
                # Se não houver dados para este dia e equipamento, a diferença é 24h
                if len(dados_dia) == 0:
                    resultado[f'Falta {dia}'] = 24.0
                    continue
                
                # Calcular horas registradas neste dia
                horas_dia = round(dados_dia['Diferença_Hora'].sum(), 2)
                
                # Calcular a diferença para 24 horas neste dia
                diferenca_dia = round(max(24 - horas_dia, 0), 2)
                
                # Adicionar ao resultado apenas se houver falta (diferença > 0)
                if diferenca_dia > 0:
                    resultado[f'Falta {dia}'] = diferenca_dia
                else:
                    resultado[f'Falta {dia}'] = 0.0
        
        resultados.append(resultado)
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso(base_calculo, df_base):
    """
    Calcula o percentual de motor ocioso por operador e frente usando os dados da Base Calculo.
    Agrega os dados por operador e frente, calculando a média quando um operador aparece em múltiplas situações.
    
    Regras de cálculo (Método Avançado):
    1. Existe uma tolerância de 1 minuto para operações de "Parada com motor ligado" (valor 1)
    2. Se uma operação tem "Parada com motor ligado = 1" e dura menos de 1 minuto, ela é desconsiderada se:
       - A próxima operação tem "Parada com motor ligado = 0" e dura mais de 1 minuto
    3. Se existem duas ou mais operações com "Parada com motor ligado = 1" em sequência (ou com outras operações entre elas que duram menos de 1 minuto):
       - Somamos o tempo total dessas operações
       - Subtraímos 1 minuto do total
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
        df_base (DataFrame): DataFrame base para aplicar filtros de operações
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador e frente (agregado)
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in base_calculo.columns:
        base_calculo = base_calculo.copy()
        base_calculo['Frente'] = base_calculo['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Agrupar por operador e frente
    agrupado = base_calculo.groupby(['Operador', 'Frente']).agg({
        'Motor Ligado': 'sum',
        'Parado Com Motor Ligado': 'sum'
    }).reset_index()
    
    resultados = []
    print("\n=== DETALHAMENTO DO CÁLCULO DE MOTOR OCIOSO POR FRENTE (MÉTODO AVANÇADO) ===")
    print("Utilizando valores da coluna Motor Ocioso calculados com a lógica de intervalos e tolerância")
    print("=" * 80)

    for _, row in agrupado.iterrows():
        operador = row['Operador']
        frente = row['Frente']
        tempo_ligado = row['Motor Ligado']
        tempo_ocioso = row['Parado Com Motor Ligado']
        
        # Calcular porcentagem de tempo ocioso
        porcentagem = tempo_ocioso / tempo_ligado if tempo_ligado > 0 else 0
        
        print(f"\nOperador: {operador} - Frente: {frente}")
        print(f"Tempo Ocioso (método avançado) = {tempo_ocioso:.6f} horas")
        print(f"Tempo Ligado = {tempo_ligado:.6f} horas")
        print(f"Porcentagem = {porcentagem:.6f} ({porcentagem*100:.2f}%)")
        print("-" * 80)
        
        resultados.append({
            'Operador': operador,
            'Frente': frente,
            'Porcentagem': porcentagem,
            'Tempo Ligado': tempo_ligado,
            'Tempo Ocioso': tempo_ocioso
        })
    
    # Ordenar primeiro por frente, depois por porcentagem (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values(['Frente', 'Porcentagem'], ascending=[True, False])

def calcular_eficiencia_energetica(base_calculo):
    """
    Calcula a eficiência energética por operador e frente usando os dados da Base Calculo.
    Agrega os dados por operador e frente, calculando a média ponderada quando um operador aparece em múltiplas situações.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador e frente (agregado)
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in base_calculo.columns:
        base_calculo = base_calculo.copy()
        base_calculo['Frente'] = base_calculo['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Agrupar por operador e frente
    agrupado = base_calculo.groupby(['Operador', 'Frente']).agg({
        'Horas elevador': 'sum',
        'Motor Ligado': 'sum'
    }).reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        eficiencia = row['Horas elevador'] / row['Motor Ligado'] if row['Motor Ligado'] > 0 else 0
        resultados.append({
            'Operador': row['Operador'],
            'Frente': row['Frente'],
            'Eficiência': eficiencia
        })
    
    # Ordenar primeiro por frente, depois por eficiência (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values(['Frente', 'Eficiência'], ascending=[True, False])

def calcular_hora_elevador(df_base, base_calculo):
    """
    Extrai as horas de elevador da Base Calculo.
    Agrega os dados por operador e frente, somando quando um operador aparece em múltiplas situações.
    
    Args:
        df_base: Não usado mais, mantido para compatibilidade
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Horas de elevador por operador e frente (agregado)
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in base_calculo.columns:
        base_calculo = base_calculo.copy()
        base_calculo['Frente'] = base_calculo['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Agrupar por operador e frente
    agrupado = base_calculo.groupby(['Operador', 'Frente'])['Horas elevador'].sum().reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        resultados.append({
            'Operador': row['Operador'],
            'Frente': row['Frente'],
            'Horas': row['Horas elevador']
        })
    
    # Ordenar primeiro por frente, depois por horas (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values(['Frente', 'Horas'], ascending=[True, False])

def calcular_uso_gps(df_base, base_calculo):
    """
    Extrai o uso de GPS da Base Calculo.
    Agrega os dados por operador e frente, calculando a média ponderada quando um operador aparece em múltiplas situações.
    
    Args:
        df_base: Não usado mais, mantido para compatibilidade
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador e frente (agregado)
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente' se não existir
    if 'Frente' not in base_calculo.columns:
        base_calculo = base_calculo.copy()
        base_calculo['Frente'] = base_calculo['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Agrupar por operador e frente
    agrupado = base_calculo.groupby(['Operador', 'Frente']).agg({
        'RTK': 'sum',
        'Horas Produtivas': 'sum'
    }).reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        porcentagem = row['RTK'] / row['Horas Produtivas'] if row['Horas Produtivas'] > 0 else 0
        resultados.append({
            'Operador': row['Operador'],
            'Frente': row['Frente'],
            'Porcentagem': porcentagem
        })
    
    # Ordenar primeiro por frente, depois por porcentagem (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values(['Frente', 'Porcentagem'], ascending=[True, False])

def calcular_media_velocidade(df):
    """
    Calcula a média de velocidade para cada operador e frente considerando apenas registros produtivos e deslocamentos.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: DataFrame com a média de velocidade por operador e frente
    """
    # Filtramos os dados excluindo os operadores da lista de exclusão
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Extrair frente da coluna 'Grupo Equipamento/Frente'
    df_filtrado = df_filtrado.copy()
    df_filtrado['Frente'] = df_filtrado['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Identificar registros produtivos e deslocamentos
    # Produtivos: onde Grupo Operacao é 'Produtiva'
    # Deslocamentos: onde Estado é 'DESLOCAMENTO' ou velocidade > 0 em operações não produtivas
    registros_validos = (
        (df_filtrado['Grupo Operacao'] == 'Produtiva') | 
        (df_filtrado['Estado'] == 'DESLOCAMENTO') |
        ((df_filtrado['Velocidade'] > 0) & (df_filtrado['Estado'] != 'PARADO'))
    )
    
    df_velocidade = df_filtrado[registros_validos]
    
    # Obter combinações únicas de operador e frente
    combinacoes = df_filtrado[['Operador', 'Frente']].drop_duplicates()
    resultados = []
    
    for _, comb in combinacoes.iterrows():
        operador = comb['Operador']
        frente = comb['Frente']
        
        # Filtrar dados para este operador e frente
        dados_op = df_velocidade[(df_velocidade['Operador'] == operador) & (df_velocidade['Frente'] == frente)]
        
        # Se não houver dados válidos para esta combinação, adicionar com velocidade zero
        if len(dados_op) == 0:
            resultados.append({
                'Operador': operador,
                'Frente': frente,
                'Velocidade': 0
            })
            continue
        
        # Calcular a média de velocidade (ponderada pelo tempo, se houver a coluna Diferença_Hora)
        if 'Diferença_Hora' in dados_op.columns:
            # Média ponderada pelo tempo 
            velocidade_media = round(
                (dados_op['Velocidade'] * dados_op['Diferença_Hora']).sum() / dados_op['Diferença_Hora'].sum()
                if dados_op['Diferença_Hora'].sum() > 0 else 0,
                2
            )
        else:
            # Média simples
            velocidade_media = round(dados_op['Velocidade'].mean(), 2)
        
        resultados.append({
            'Operador': operador,
            'Frente': frente,
            'Velocidade': velocidade_media
        })
    
    # Ordenar primeiro por frente, depois por velocidade (decrescente)
    df_resultado = pd.DataFrame(resultados)
    return df_resultado.sort_values(['Frente', 'Velocidade'], ascending=[True, False])

def identificar_operadores_duplicados(df):
    """
    Identifica operadores que aparecem com IDs diferentes no mesmo conjunto de dados.
    Detecta principalmente IDs que começam com '133' e têm 7 dígitos, verificando se
    existe outra ID com o mesmo nome mas com menos dígitos.
    
    Args:
        df (DataFrame): DataFrame com os dados dos operadores
    
    Returns:
        dict: Dicionário com mapeamento {id_incorreta: id_correta}
        DataFrame: DataFrame com as duplicidades encontradas para relatório
    """
    if 'Operador' not in df.columns or len(df) == 0:
        return {}, pd.DataFrame(columns=['ID Incorreta', 'ID Correta', 'Nome'])
    
    # Extrair operadores únicos
    operadores = df['Operador'].unique()
    
    # Mapear nomes para IDs
    nomes_para_ids = {}
    for op in operadores:
        if ' - ' in op:
            try:
                id_parte, nome_parte = op.split(' - ', 1)
                # Normalizar nome para comparação (maiúsculo e sem espaços extras)
                nome_normalizado = nome_parte.upper().strip()
                if nome_normalizado not in nomes_para_ids:
                    nomes_para_ids[nome_normalizado] = []
                nomes_para_ids[nome_normalizado].append(op)
            except:
                continue
    
    # Encontrar nomes com múltiplas IDs
    duplicidades = []
    mapeamento = {}
    
    for nome, ids in nomes_para_ids.items():
        if len(ids) > 1:
            print(f"Encontrado operador duplicado: {nome} com {len(ids)} IDs diferentes")
            
            # Separar IDs que começam com '133' e têm 7 dígitos
            ids_suspeitas = [id_op for id_op in ids if ' - ' in id_op and id_op.split(' - ')[0].startswith('133') and len(id_op.split(' - ')[0]) == 7]
            ids_normais = [id_op for id_op in ids if id_op not in ids_suspeitas]
            
            # Se temos IDs suspeitas e normais, considerar a suspeita como incorreta
            if ids_suspeitas and ids_normais:
                for id_suspeita in ids_suspeitas:
                    # Usar a ID normal mais curta como destino (geralmente a correta)
                    id_correta = min(ids_normais, key=lambda x: len(x.split(' - ')[0]) if ' - ' in x else float('inf'))
                    
                    # Adicionar ao mapeamento
                    mapeamento[id_suspeita] = id_correta
                    
                    # Extrair as partes para o relatório
                    id_incorreta_parte = id_suspeita.split(' - ')[0]
                    id_correta_parte = id_correta.split(' - ')[0]
                    
                    print(f"  - Mapeando: {id_suspeita} -> {id_correta}")
                    
                    # Adicionar à lista de duplicidades para o relatório
                    duplicidades.append({
                        'ID Incorreta': id_suspeita,
                        'ID Correta': id_correta,
                        'Nome': nome
                    })
            
            # Caso especial: todas as IDs são suspeitas ou todas são normais
            # Neste caso, apresentamos no relatório mas não fazemos substituição automática
            else:
                print(f"  - Múltiplas IDs do mesmo tipo encontradas para {nome}, sem ação automática")
                # Ainda assim, adicionar ao relatório para conhecimento
                for i, id1 in enumerate(ids):
                    for id2 in ids[i+1:]:
                        duplicidades.append({
                            'ID Incorreta': id1, 
                            'ID Correta': id2,
                            'Nome': nome,
                            'Observação': "Duplicidade detectada, verificar manualmente"
                        })
    
    print(f"Encontrados {len(duplicidades)} operadores com IDs duplicadas.")
    return mapeamento, pd.DataFrame(duplicidades)

def calcular_horas_motor_ligado_total(df):
    """
    Calcula o total de horas com motor ligado por frota e frente durante todo o período,
    e também a média diária.
    Usa a coluna 'Diferença_Hora' para calcular o tempo real com motor ligado.
    
    Args:
        df (DataFrame): DataFrame com os dados de operação
        
    Returns:
        DataFrame: DataFrame com o total de horas com motor ligado por frota e a média diária
    """
    try:
        # Extrair frente da coluna 'Grupo Equipamento/Frente'
        df = df.copy()
        df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
        
        # Verificar se as colunas necessárias existem
        if 'Data' not in df.columns:
            print("Aviso: Coluna 'Data' não encontrada. Usando o número total de registros como divisor.")
            # Criar uma coluna temporária para calcular as horas com motor ligado
            df_temp = df.copy()
            df_temp['Horas_Motor_Ligado'] = df_temp.apply(
                lambda row: row['Diferença_Hora'] if row['Motor Ligado'] == 1 else 0, 
                axis=1
            )
            
            # Agrupa por frota e frente, e soma as horas com motor ligado
            df_horas_motor = df_temp.groupby(['Equipamento', 'Frente']).agg({
                'Horas_Motor_Ligado': 'sum'
            }).reset_index()
            
            # Renomeia as colunas para maior clareza
            df_horas_motor.columns = ['Frota', 'Frente', 'Total Horas Motor Ligado']
            
            # Ordena pelo total de horas (decrescente)
            df_horas_motor = df_horas_motor.sort_values('Total Horas Motor Ligado', ascending=False)
            
            return df_horas_motor
        
        # Criar uma coluna temporária para calcular as horas com motor ligado
        df_temp = df.copy()
        df_temp['Horas_Motor_Ligado'] = df_temp.apply(
            lambda row: row['Diferença_Hora'] if row['Motor Ligado'] == 1 else 0, 
            axis=1
        )
        
        # Agrupa por frota e frente, calcula o total de horas com motor ligado e o número de dias
        df_horas_motor = df_temp.groupby(['Equipamento', 'Frente']).agg({
            'Horas_Motor_Ligado': 'sum',
            'Data': 'nunique'  # Conta o número de dias únicos para cada frota e frente
        }).reset_index()
        
        # Renomeia as colunas para maior clareza
        df_horas_motor.columns = ['Frota', 'Frente', 'Total Horas Motor Ligado', 'Dias Registrados']
        
        # Calcula a média diária de horas com motor ligado
        df_horas_motor['Horas Motor Ligado por Dia'] = df_horas_motor['Total Horas Motor Ligado'] / df_horas_motor['Dias Registrados']
        
        # Ordena pelo total de horas (decrescente)
        df_horas_motor = df_horas_motor.sort_values('Total Horas Motor Ligado', ascending=False)
        
        # Mantém todas as colunas relevantes
        df_horas_motor = df_horas_motor[['Frota', 'Frente', 'Total Horas Motor Ligado', 'Horas Motor Ligado por Dia', 'Dias Registrados']]
        
        return df_horas_motor
        
    except Exception as e:
        print(f"Erro ao calcular horas totais de motor ligado: {str(e)}")
        return pd.DataFrame(columns=['Frota', 'Frente', 'Total Horas Motor Ligado', 'Horas Motor Ligado por Dia', 'Dias Registrados'])

def criar_planilha_tdh(df):
    """
    Cria uma planilha vazia para TDH, contendo as frotas e frentes encontradas.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame vazio com as colunas 'Frota', 'Frente' e 'TDH'
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente'
    df = df.copy()
    df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Obter todas as combinações únicas de frota e frente
    combinacoes = df[['Equipamento', 'Frente']].drop_duplicates().sort_values(['Equipamento', 'Frente'])
    
    # Criar DataFrame vazio com as frotas e frentes
    df_tdh = pd.DataFrame({
        'Frota': combinacoes['Equipamento'].values,
        'Frente': combinacoes['Frente'].values,
        'TDH': ''
    })
    
    # Ordenar primeiro por frente, depois por frota
    return df_tdh.sort_values(['Frente', 'Frota'], ascending=[True, True])

def criar_planilha_diesel(df):
    """
    Cria uma planilha vazia para Diesel, contendo as frotas e frentes encontradas.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame vazio com as colunas 'Frota', 'Frente' e 'Diesel'
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente'
    df = df.copy()
    df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Obter todas as combinações únicas de frota e frente
    combinacoes = df[['Equipamento', 'Frente']].drop_duplicates().sort_values(['Equipamento', 'Frente'])
    
    # Criar DataFrame vazio com as frotas e frentes
    df_diesel = pd.DataFrame({
        'Frota': combinacoes['Equipamento'].values,
        'Frente': combinacoes['Frente'].values,
        'Diesel': ''
    })
    
    # Ordenar primeiro por frente, depois por frota
    return df_diesel.sort_values(['Frente', 'Frota'], ascending=[True, True])

def criar_planilha_impureza(df):
    """
    Cria uma planilha vazia para Impureza Vegetal, contendo as frotas e frentes encontradas.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame vazio com as colunas 'Frota', 'Frente' e 'Impureza'
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente'
    df = df.copy()
    df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Obter todas as combinações únicas de frota e frente
    combinacoes = df[['Equipamento', 'Frente']].drop_duplicates().sort_values(['Equipamento', 'Frente'])
    
    # Criar DataFrame vazio com as frotas e frentes
    df_impureza = pd.DataFrame({
        'Frota': combinacoes['Equipamento'].values,
        'Frente': combinacoes['Frente'].values,
        'Impureza': ''
    })
    
    # Ordenar primeiro por frente, depois por frota
    return df_impureza.sort_values(['Frente', 'Frota'], ascending=[True, True])

def calcular_ofensores(df):
    """
    Calcula os top 5 ofensores por frente.
    Agrupa por Frente e Operacao onde Estado Operacional é 'PARADA',
    soma a Diferença_Hora, classifica do maior para o menor.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame com os ofensores por frente
    """
    # Extrair frente da coluna 'Grupo Equipamento/Frente'
    df = df.copy()
    df['Frente'] = df['Grupo Equipamento/Frente'].apply(extrair_frente)
    
    # Filtrar apenas os registros com Estado Operacional PARADA
    df_paradas = df[df['Estado Operacional'] == 'PARADA'].copy()
    
    # Se não houver dados de parada, retornar DataFrame vazio
    if len(df_paradas) == 0:
        return pd.DataFrame(columns=['Frente', 'Operação', 'Tempo', 'Porcentagem'])
    
    # Agrupar por Frente e Operacao, somar o tempo
    paradas_agrupadas = df_paradas.groupby(['Frente', 'Operacao'])['Diferença_Hora'].sum().reset_index()
    
    # Calcular o tempo total de todas as paradas por frente
    tempo_total_por_frente = df_paradas.groupby('Frente')['Diferença_Hora'].sum().to_dict()
    
    # Adicionar coluna de porcentagem (porcentagem dentro da frente)
    paradas_agrupadas['Porcentagem'] = paradas_agrupadas.apply(
        lambda row: row['Diferença_Hora'] / tempo_total_por_frente.get(row['Frente'], 1) if tempo_total_por_frente.get(row['Frente'], 0) > 0 else 0,
        axis=1
    )
    
    # Ordenar primeiro por frente, depois por tempo (decrescente)
    paradas_agrupadas = paradas_agrupadas.sort_values(['Frente', 'Diferença_Hora'], ascending=[True, False])
    
    # Renomear colunas para melhor visualização
    resultado = paradas_agrupadas.rename(columns={
        'Operacao': 'Operação',
        'Diferença_Hora': 'Tempo'
    })
    
    return resultado

def criar_planilha_coordenadas(df_base):
    """
    Cria uma planilha com coordenadas das frotas, ordenada por hora e por frota.
    
    Args:
        df_base (DataFrame): DataFrame com os dados base
        
    Returns:
        DataFrame: DataFrame com as colunas Frota, Hora, Latitude e Longitude
    """
    # Verificar se as colunas necessárias existem
    colunas_necessarias = ['Equipamento', 'Hora', 'Latitude', 'Longitude']
    for coluna in colunas_necessarias:
        if coluna not in df_base.columns:
            print(f"Aviso: Coluna '{coluna}' não encontrada para criar planilha de coordenadas.")
            # Criar um DataFrame vazio com as colunas necessárias
            return pd.DataFrame(columns=['Frota', 'Hora', 'Latitude', 'Longitude'])
    
    # Criar um novo DataFrame apenas com as colunas necessárias
    df_coordenadas = df_base[colunas_necessarias].copy()
    
    # Renomear a coluna Equipamento para Frota
    df_coordenadas.rename(columns={'Equipamento': 'Frota'}, inplace=True)
    
    # Garantir que a coluna Hora esteja no formato correto (hh:mm:ss)
    if df_coordenadas['Hora'].dtype == 'datetime64[ns]':
        df_coordenadas['Hora'] = df_coordenadas['Hora'].dt.strftime('%H:%M:%S')
    
    # Ordenar por Frota (como texto) e por Hora
    df_coordenadas['Frota'] = df_coordenadas['Frota'].astype(str)
    
    # Certificar que podemos ordenar por hora (convertendo temporariamente)
    df_coordenadas['Hora_temp'] = pd.to_datetime(df_coordenadas['Hora'], format='%H:%M:%S', errors='coerce')
    df_coordenadas = df_coordenadas.sort_values(['Frota', 'Hora_temp'])
    df_coordenadas.drop('Hora_temp', axis=1, inplace=True)
    
    # Garantir que as coordenadas sejam numéricas
    df_coordenadas['Latitude'] = pd.to_numeric(df_coordenadas['Latitude'], errors='coerce')
    df_coordenadas['Longitude'] = pd.to_numeric(df_coordenadas['Longitude'], errors='coerce')
    
    # Formatar as coordenadas como strings com ponto decimal
    df_coordenadas['Latitude'] = df_coordenadas['Latitude'].apply(lambda x: f"{x:.9f}" if pd.notnull(x) else '')
    df_coordenadas['Longitude'] = df_coordenadas['Longitude'].apply(lambda x: f"{x:.9f}" if pd.notnull(x) else '')
    
    # Remover duplicatas completas para reduzir tamanho da planilha
    df_coordenadas = df_coordenadas.drop_duplicates()
    
    return df_coordenadas

def criar_excel_com_planilhas(df_base, base_calculo, disp_mecanica, eficiencia_energetica,
                            hora_elevador, motor_ocioso, uso_gps, horas_por_frota, caminho_saida,
                            df_duplicados=None, media_velocidade=None, df_substituicoes=None):
    """
    Cria um arquivo Excel com todas as planilhas necessárias.
    """
    # Definir função de ajuste de largura de colunas
    def ajustar_largura_colunas(worksheet):
        """Ajusta a largura das colunas da planilha"""
        for col in worksheet.columns:
            max_length = 10
            column = col[0].column_letter
            header_text = str(col[0].value)
            if header_text:
                max_length = max(max_length, len(header_text) + 2)
            for cell in col[1:min(20, len(col))]:
                if cell.value:
                    cell_text = str(cell.value)
                    max_length = max(max_length, len(cell_text) + 2)
            max_length = min(max_length, 40)
            worksheet.column_dimensions[column].width = max_length
    
    # Calcular horas totais de motor ligado por frota
    # df_horas_motor_total = calcular_horas_motor_ligado_total(df_base)  # Removido
    
    # Criar planilhas adicionais
    df_tdh = criar_planilha_tdh(df_base)
    df_diesel = criar_planilha_diesel(df_base)
    df_impureza = criar_planilha_impureza(df_base)
    
    # Calcular ofensores
    df_ofensores = calcular_ofensores(df_base)
    
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        # Salvar cada DataFrame em uma planilha separada
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        reordenar_colunas_frente_primeiro(base_calculo).to_excel(writer, sheet_name='Base Calculo', index=False)
        reordenar_colunas_frente_primeiro(disp_mecanica).to_excel(writer, sheet_name='1_Disponibilidade Mecânica', index=False)
        reordenar_colunas_frente_primeiro(eficiencia_energetica).to_excel(writer, sheet_name='2_Eficiência Energética', index=False)
        reordenar_colunas_frente_primeiro(hora_elevador).to_excel(writer, sheet_name='3_Hora Elevador', index=False)
        
        # Garantir que os valores numéricos do motor_ocioso sejam mantidos como números
        motor_ocioso['Tempo Ligado'] = pd.to_numeric(motor_ocioso['Tempo Ligado'], errors='coerce')
        motor_ocioso['Tempo Ocioso'] = pd.to_numeric(motor_ocioso['Tempo Ocioso'], errors='coerce')
        motor_ocioso['Porcentagem'] = pd.to_numeric(motor_ocioso['Porcentagem'], errors='coerce')
        reordenar_colunas_frente_primeiro(motor_ocioso).to_excel(writer, sheet_name='4_Motor Ocioso', index=False)
        
        reordenar_colunas_frente_primeiro(uso_gps).to_excel(writer, sheet_name='5_Uso GPS', index=False)
        reordenar_colunas_frente_primeiro(horas_por_frota).to_excel(writer, sheet_name='Horas por Frota', index=False)
        
        # Adicionar a nova planilha com horas totais de motor ligado
        # df_horas_motor_total.to_excel(writer, sheet_name='Horas Motor', index=False)  # Removido
        
        # Adicionar nova planilha de ofensores
        reordenar_colunas_frente_primeiro(df_ofensores).to_excel(writer, sheet_name='Ofensores', index=False)
        
        # Adicionar novas planilhas
        reordenar_colunas_frente_primeiro(df_tdh).to_excel(writer, sheet_name='TDH', index=False)
        reordenar_colunas_frente_primeiro(df_diesel).to_excel(writer, sheet_name='Diesel', index=False)
        reordenar_colunas_frente_primeiro(df_impureza).to_excel(writer, sheet_name='Impureza Vegetal', index=False)
        
        if media_velocidade is None:
            media_velocidade = pd.DataFrame(columns=['Operador', 'Frente', 'Velocidade'])
        reordenar_colunas_frente_primeiro(media_velocidade).to_excel(writer, sheet_name='Média Velocidade', index=False)
        
        # IDs duplicadas e substituídas
        if df_duplicados is not None and not df_duplicados.empty:
            df_duplicados.to_excel(writer, sheet_name='IDs Encontradas', index=False)
        if df_substituicoes is not None and not df_substituicoes.empty:
            df_substituicoes.to_excel(writer, sheet_name='IDs Substituídas', index=False)
        
        # Formatar cada planilha
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            ajustar_largura_colunas(worksheet)
            
            if sheet_name == '1_Disponibilidade Mecânica':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Disponibilidade)
                    cell.number_format = '0.00%'
            
            elif sheet_name == '2_Eficiência Energética':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Eficiência)
                    cell.number_format = '0.00%'
            
            elif sheet_name == '3_Hora Elevador':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Horas)
                    cell.number_format = '0.00'
            
            elif sheet_name == '4_Motor Ocioso':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
                    cell = worksheet.cell(row=row, column=3)  # Coluna C (Tempo Ligado)
                    cell.number_format = '0.00'
                    cell = worksheet.cell(row=row, column=4)  # Coluna D (Tempo Ocioso)
                    cell.number_format = '0.00'
            
            elif sheet_name == '5_Uso GPS':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
            
            elif sheet_name == 'Média Velocidade':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Velocidade)
                    cell.number_format = '0.00'
            
            elif sheet_name == 'Horas por Frota':
                for row in range(2, worksheet.max_row + 1):
                    for col in range(2, worksheet.max_column + 1):  # Todas as colunas de tempo
                        cell = worksheet.cell(row=row, column=col)
                        cell.number_format = '0.00'
            
            elif sheet_name == 'Ofensores':
                for row in range(2, worksheet.max_row + 1):
                    # Coluna C (Tempo)
                    cell = worksheet.cell(row=row, column=3)
                    cell.number_format = '0.00'  # Formato decimal
                    
                    # Coluna D (Porcentagem)
                    cell = worksheet.cell(row=row, column=4)
                    cell.number_format = '0.00%'  # Formato percentual
            
            elif sheet_name == 'TDH':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (TDH)
                    cell.number_format = '0.0000'  # 4 casas decimais
            
            elif sheet_name == 'Diesel':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Diesel)
                    cell.number_format = '0.0000'  # 4 casas decimais
            
            elif sheet_name == 'Impureza Vegetal':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Impureza)
                    cell.number_format = '0.00'  # 2 casas decimais
            
            elif sheet_name == 'Base Calculo':
                colunas_porcentagem = ['%', '% Utilização RTK', '% Eficiência Elevador', '% Parado com motor ligado']
                colunas_tempo = ['Horas totais', 'Horas elevador', 'RTK', 'Horas Produtivas', 'Motor Ligado', 'Parado Com Motor Ligado']
                
                # Adicionar uma nota explicativa sobre o método avançado de cálculo
                max_row = worksheet.max_row
                nota_row = max_row + 3  # Deixar algumas linhas em branco
                
                # Título
                cell = worksheet.cell(row=nota_row, column=1)
                cell.value = "MÉTODO AVANÇADO DE CÁLCULO DE MOTOR OCIOSO"
                cell.font = Font(bold=True, size=12)
                
                # Descrição do método
                worksheet.cell(row=nota_row + 2, column=1).value = "Regras aplicadas:"
                worksheet.cell(row=nota_row + 3, column=1).value = "1. Tolerância de 1 minuto aplicada a cada sequência de paradas"
                worksheet.cell(row=nota_row + 4, column=1).value = "2. Sequências de paradas com menos de 1 minuto são ignoradas"
                worksheet.cell(row=nota_row + 5, column=1).value = "3. Paradas de motor com velocidade zero são agrupadas em intervalos"
                worksheet.cell(row=nota_row + 7, column=1).value = "Obs: Os valores na coluna 'Parado Com Motor Ligado' já refletem este método avançado"
                
                # Formatar a nota
                for r in range(nota_row, nota_row + 8):
                    cell = worksheet.cell(row=r, column=1)
                    cell.font = Font(size=10)
                    if r == nota_row + 7:  # A observação em itálico
                        cell.font = Font(size=10, italic=True)
                
                # Formatar as células de dados
                for row in range(2, max_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        header = worksheet.cell(row=1, column=col).value
                        cell = worksheet.cell(row=row, column=col)
                        
                        if header in colunas_porcentagem:
                            cell.number_format = '0.00%'
                        elif header in colunas_tempo:
                            cell.number_format = '0.00'

def processar_arquivo(caminho_arquivo, diretorio_saida):
    """
    Processa um único arquivo e gera o Excel de saída.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo a ser processado
        diretorio_saida (str): Diretório onde o arquivo de saída será salvo
    """
    # Obter apenas o nome do arquivo (sem caminho e sem extensão)
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    
    # Nome de saída igual ao original, mas com "-rastros" no final e extensão .xlsx na pasta output
    arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}-unificado.xlsx")
    
    print(f"\nProcessando arquivo: {os.path.basename(caminho_arquivo)}")
    print(f"Arquivo de saída: {os.path.basename(arquivo_saida)}")
    
    # Processar o arquivo base
    df_base = processar_arquivo_base(caminho_arquivo)
    if df_base is None:
        print(f"Erro ao processar {os.path.basename(caminho_arquivo)}. Pulando para o próximo arquivo.")
        return
    
    # Identificar operadores duplicados antes de aplicar as substituições manuais
    mapeamento_duplicados, df_duplicados = identificar_operadores_duplicados(df_base)
    
    # Carregar e aplicar substituições de operadores (manuais + automáticas)
    substituicoes = carregar_substituicoes_operadores()
    
    # Combinar as substituições manuais com as automáticas (automáticas têm precedência)
    substituicoes_combinadas = {**substituicoes, **mapeamento_duplicados}
    
    print("\n=== APLICANDO SUBSTITUIÇÕES DE OPERADORES ===")
    if substituicoes_combinadas:
        df_base, df_substituicoes = aplicar_substituicao_operadores(df_base, substituicoes_combinadas)
    else:
        df_substituicoes = pd.DataFrame(columns=['ID Original', 'Nome Original', 'ID Nova', 'Nome Novo', 'Registros Afetados'])
    print("=== FIM DAS SUBSTITUIÇÕES ===\n")
    
    # Se o DataFrame estiver vazio, gerar apenas a planilha BASE
    if len(df_base) == 0:
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
            df_base.to_excel(writer, sheet_name='BASE', index=False)
            if not df_duplicados.empty:
                df_duplicados.to_excel(writer, sheet_name='IDs Duplicadas', index=False)
        print(f"Arquivo {arquivo_saida} gerado com apenas a planilha BASE (sem dados).")
        return
    
    # ORDENAR o DataFrame base por Equipamento e depois por Hora
    # Isso garante que os registros sejam agrupados por frota na visualização
    if 'Data' in df_base.columns:
        # Ordenar por Equipamento, Data e depois Hora
        df_base = df_base.sort_values(by=['Equipamento', 'Data', 'Hora'])
    else:
        # Se não tiver Data, ordenar apenas por Equipamento e Hora
        df_base = df_base.sort_values(by=['Equipamento', 'Hora'])
    
    # Calcular a Base Calculo com os operadores já substituídos
    print("Calculando Base Calculo com operadores consolidados...")
    base_calculo = calcular_base_calculo(df_base)
    
    # Calcular as métricas auxiliares
    print("Calculando métricas auxiliares com operadores consolidados...")
    disp_mecanica = calcular_disponibilidade_mecanica(df_base)
    eficiencia_energetica = calcular_eficiencia_energetica(base_calculo)
    hora_elevador = calcular_hora_elevador(df_base, base_calculo)
    motor_ocioso = calcular_motor_ocioso(base_calculo, df_base)
    uso_gps = calcular_uso_gps(df_base, base_calculo)
    horas_por_frota = calcular_horas_por_frota(df_base)
    
    # Calcular média de velocidade por operador
    media_velocidade = calcular_media_velocidade(df_base)
    
    # Criar o arquivo Excel com todas as planilhas
    criar_excel_com_planilhas(
        df_base, base_calculo, disp_mecanica, eficiencia_energetica,
        hora_elevador, motor_ocioso, uso_gps, horas_por_frota, arquivo_saida,
        df_duplicados,  # Adicionar a tabela de IDs duplicadas
        media_velocidade,  # Adicionar a tabela de média de velocidade
        df_substituicoes  # Adicionar a tabela de IDs substituídas
    )
    
    print(f"Arquivo {arquivo_saida} gerado com sucesso!")

def extrair_arquivo_zip(caminho_zip, pasta_destino=None):
    """
    Extrai o conteúdo de um arquivo ZIP para uma pasta temporária ou destino especificado.
    Renomeia os arquivos extraídos para terem o mesmo nome do arquivo ZIP original.
    
    Args:
        caminho_zip (str): Caminho para o arquivo ZIP
        pasta_destino (str, optional): Pasta onde os arquivos serão extraídos.
                                       Se None, usa uma pasta temporária.
    
    Returns:
        list: Lista de caminhos dos arquivos extraídos e renomeados (apenas TXT e CSV)
        str: Caminho da pasta temporária (se criada) ou None
    """
    # Se pasta_destino não foi especificada, criar uma pasta temporária
    pasta_temp = None
    if pasta_destino is None:
        pasta_temp = tempfile.mkdtemp()
        pasta_destino = pasta_temp
    
    arquivos_extraidos = []
    nome_zip_sem_extensao = os.path.splitext(os.path.basename(caminho_zip))[0]
    
    try:
        with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
            # Extrair todos os arquivos do ZIP
            zip_ref.extractall(pasta_destino)
            
            # Processar cada arquivo extraído (apenas TXT e CSV)
            for arquivo in zip_ref.namelist():
                caminho_completo = os.path.join(pasta_destino, arquivo)
                # Verificar se é um arquivo e não uma pasta
                if os.path.isfile(caminho_completo):
                    # Verificar extensão
                    extensao = os.path.splitext(arquivo)[1].lower()
                    if extensao in ['.txt', '.csv']:
                        # Criar novo nome: nome do ZIP + extensão original
                        novo_nome = f"{nome_zip_sem_extensao}{extensao}"
                        novo_caminho = os.path.join(pasta_destino, novo_nome)
                        
                        # Renomear o arquivo extraído
                        try:
                            # Se já existe um arquivo com esse nome, remover primeiro
                            if os.path.exists(novo_caminho):
                                os.remove(novo_caminho)
                            # Renomear o arquivo
                            os.rename(caminho_completo, novo_caminho)
                            arquivos_extraidos.append(novo_caminho)
                            print(f"Arquivo extraído renomeado: {novo_nome}")
                        except Exception as e:
                            print(f"Erro ao renomear arquivo {arquivo} para {novo_nome}: {str(e)}")
                            arquivos_extraidos.append(caminho_completo)  # Adicionar o caminho original em caso de erro
        
        return arquivos_extraidos, pasta_temp
    
    except Exception as e:
        print(f"Erro ao extrair o arquivo ZIP {caminho_zip}: {str(e)}")
        # Se houve erro e criamos uma pasta temporária, tentar limpá-la
        if pasta_temp:
            try:
                print(f"Removendo pasta temporária: {pasta_temp}")
                shutil.rmtree(pasta_temp)
                print(f"Pasta temporária removida com sucesso")
            except Exception as e:
                print(f"Erro ao remover pasta temporária {pasta_temp}: {str(e)}")
        return [], None

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT, CSV ou ZIP de colhedoras na pasta dados.
    Ignora arquivos que contenham "transbordo" no nome.
    """
    print("\nIniciando processamento de arquivos...")
    
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    print(f"Diretório do script: {diretorio_script}")
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    print(f"Diretório raiz: {diretorio_raiz}")
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    print(f"Diretório de dados: {diretorio_dados}")
    print(f"Diretório de saída: {diretorio_saida}")
    
    # Verificar se os diretórios existem, caso contrário criar
    if not os.path.exists(diretorio_dados):
        print(f"Criando diretório de dados: {diretorio_dados}")
        os.makedirs(diretorio_dados)
    if not os.path.exists(diretorio_saida):
        print(f"Criando diretório de saída: {diretorio_saida}")
        os.makedirs(diretorio_saida)
    
    # Encontrar todos os arquivos TXT/CSV/ZIP de colhedoras
    arquivos = []
    arquivos_zip = []
    
    # Adicionar arquivos TXT sempre
    arquivos += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.txt"))
    arquivos += glob.glob(os.path.join(diretorio_dados, "*colhedora*.txt"))
    arquivos += glob.glob(os.path.join(diretorio_dados, "colhedora*.txt"))
    
    # Adicionar arquivos CSV apenas se processCsv for True
    if processCsv:
        arquivos += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.csv"))
        arquivos += glob.glob(os.path.join(diretorio_dados, "*colhedora*.csv"))
        arquivos += glob.glob(os.path.join(diretorio_dados, "colhedora*.csv"))
    
    # Adicionar arquivos ZIP
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "RV Colhedora*.zip"))
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "*colhedora*.zip"))
    arquivos_zip += glob.glob(os.path.join(diretorio_dados, "colhedora*.zip"))
    
    print("\nArquivos encontrados antes da filtragem:")
    print(f"TXT/CSV: {[os.path.basename(a) for a in arquivos]}")
    print(f"ZIP: {[os.path.basename(a) for a in arquivos_zip]}")
    
    # Filtrar arquivos que contenham "transbordo" no nome (case insensitive)
    arquivos = [arquivo for arquivo in arquivos if "transbordo" not in os.path.basename(arquivo).lower()]
    arquivos_zip = [arquivo for arquivo in arquivos_zip if "transbordo" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos = list(set(arquivos))
    arquivos_zip = list(set(arquivos_zip))
    
    print("\nArquivos encontrados após a filtragem:")
    print(f"TXT/CSV: {[os.path.basename(a) for a in arquivos]}")
    print(f"ZIP: {[os.path.basename(a) for a in arquivos_zip]}")
    
    if not arquivos and not arquivos_zip:
        print("Nenhum arquivo de colhedoras encontrado na pasta dados!")
        return
    
    print(f"\nEncontrados {len(arquivos)} arquivos de colhedoras (TXT/CSV) para processar.")
    print(f"Encontrados {len(arquivos_zip)} arquivos ZIP de colhedoras para processar.")
    
    # Processar cada arquivo TXT/CSV
    for arquivo in arquivos:
        print(f"\nProcessando arquivo TXT/CSV: {os.path.basename(arquivo)}")
        processar_arquivo(arquivo, diretorio_saida)
    
    # Processar cada arquivo ZIP
    for arquivo_zip in arquivos_zip:
        print(f"\nProcessando arquivo ZIP: {os.path.basename(arquivo_zip)}")
        
        # Extrair arquivo ZIP para pasta temporária
        print(f"Extraindo arquivo ZIP: {arquivo_zip}")
        arquivos_extraidos, pasta_temp = extrair_arquivo_zip(arquivo_zip)
        
        if not arquivos_extraidos:
            print(f"Nenhum arquivo TXT ou CSV encontrado no ZIP {os.path.basename(arquivo_zip)}")
            continue
        
        print(f"Extraídos {len(arquivos_extraidos)} arquivos do ZIP:")
        for arquivo in arquivos_extraidos:
            print(f"  - {os.path.basename(arquivo)}")
        
        # Processar cada arquivo extraído
        for arquivo_extraido in arquivos_extraidos:
            # Filtrar arquivos que contenham "transbordo" no nome
            if "transbordo" not in os.path.basename(arquivo_extraido).lower():
                print(f"\nProcessando arquivo extraído: {os.path.basename(arquivo_extraido)}")
                processar_arquivo(arquivo_extraido, diretorio_saida)
        
        # Limpar pasta temporária se foi criada
        if pasta_temp:
            try:
                print(f"Removendo pasta temporária: {pasta_temp}")
                shutil.rmtree(pasta_temp)
                print(f"Pasta temporária removida com sucesso")
            except Exception as e:
                print(f"Erro ao remover pasta temporária {pasta_temp}: {str(e)}")
    
    print("\nProcessamento de todos os arquivos concluído!")

def carregar_substituicoes_operadores():
    """
    Carrega o arquivo substituiroperadores.json que contém os mapeamentos 
    de substituição de operadores.
    
    Returns:
        dict: Dicionário com mapeamento {operador_origem: operador_destino}
        ou dicionário vazio se o arquivo não existir ou for inválido
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Caminho para o arquivo de substituição
    arquivo_substituicao = os.path.join(diretorio_raiz, "config", "substituiroperadores.json")
    
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo_substituicao):
        print(f"Arquivo de substituição de operadores não encontrado: {arquivo_substituicao}")
        return {}
    
    try:
        # Carregar o arquivo JSON
        with open(arquivo_substituicao, 'r', encoding='utf-8') as f:
            substituicoes = json.load(f)
        
        # Criar dicionário de substituições
        mapeamento = {item['operador_origem']: item['operador_destino'] for item in substituicoes}
        
        print(f"Carregadas {len(mapeamento)} substituições de operadores.")
        return mapeamento
        
    except Exception as e:
        print(f"Erro ao carregar arquivo de substituição de operadores: {str(e)}")
        return {}

def aplicar_substituicao_operadores(df, mapeamento_substituicoes):
    """
    Aplica as substituições de operadores no DataFrame.
    
    Args:
        df (DataFrame): DataFrame a ser processado
        mapeamento_substituicoes (dict): Dicionário com mapeamento {operador_origem: operador_destino}
    
    Returns:
        tuple: (DataFrame com substituições aplicadas, DataFrame com registro das substituições)
    """
    if not mapeamento_substituicoes or 'Operador' not in df.columns:
        return df, pd.DataFrame(columns=['ID Original', 'Nome Original', 'ID Nova', 'Nome Novo', 'Registros Afetados'])
    
    # Criar uma cópia para não alterar o DataFrame original
    df_modificado = df.copy()
    
    # Lista para armazenar as substituições realizadas
    substituicoes_realizadas = []
    total_registros_substituidos = 0
    
    # Verificar operadores antes da substituição para relatório
    operadores_antes = df_modificado['Operador'].unique()
    print(f"\nOperadores antes da substituição: {len(operadores_antes)}")
    
    # Contar operadores e registros antes da substituição
    contagem_antes = df_modificado['Operador'].value_counts().to_dict()
    
    # Aplicar as substituições
    for origem, destino in mapeamento_substituicoes.items():
        # Verificar se o operador de origem existe no DataFrame
        registros_afetados = df_modificado[df_modificado['Operador'] == origem].shape[0]
        
        if registros_afetados > 0:
            # Substituir o operador
            df_modificado.loc[df_modificado['Operador'] == origem, 'Operador'] = destino
            
            total_registros_substituidos += registros_afetados
            
            # Extrair IDs e nomes
            id_original = origem.split(' - ')[0] if ' - ' in origem else origem
            nome_original = origem.split(' - ')[1] if ' - ' in origem else ''
            id_nova = destino.split(' - ')[0] if ' - ' in destino else destino
            nome_novo = destino.split(' - ')[1] if ' - ' in destino else ''
            
            substituicoes_realizadas.append({
                'ID Original': id_original,
                'Nome Original': nome_original,
                'ID Nova': id_nova, 
                'Nome Novo': nome_novo,
                'Registros Afetados': registros_afetados
            })
            print(f"Operador '{origem}' substituído por '{destino}' em {registros_afetados} registros")
    
    # Verificar operadores após substituição
    operadores_depois = df_modificado['Operador'].unique()
    print(f"Operadores após substituição: {len(operadores_depois)}")
    print(f"Total de registros substituídos: {total_registros_substituidos}")
    
    # Criar DataFrame com as substituições realizadas
    df_substituicoes = pd.DataFrame(substituicoes_realizadas)
    
    return df_modificado, df_substituicoes

def reordenar_colunas_frente_primeiro(df):
    """
    Reordena as colunas do DataFrame colocando 'Frente' como primeira coluna.
    Se não houver coluna 'Frente', retorna o DataFrame inalterado.
    
    Args:
        df (DataFrame): DataFrame a ser reordenado
    
    Returns:
        DataFrame: DataFrame com coluna 'Frente' como primeira
    """
    if 'Frente' not in df.columns:
        return df
    
    # Criar lista de colunas com Frente primeiro
    colunas = ['Frente'] + [col for col in df.columns if col != 'Frente']
    return df[colunas]

if __name__ == "__main__":
    print("="*80)
    print("Iniciando processamento de arquivos de colhedoras...")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Este script processa arquivos de colhedoras e gera planilhas Excel com métricas")
    print("Suporta arquivos TXT, CSV e ZIP")
    print("Ignorando arquivos que contenham 'transbordo' no nome")
    print("="*50)
    print("USANDO MÉTODO AVANÇADO DE CÁLCULO DE MOTOR OCIOSO")
    print("- Tolerância de 1 minuto é aplicada a cada sequência de paradas")
    print("- Sequências de paradas com menos de 1 minuto são ignoradas")
    print("- Paradas de motor com velocidade zero são agrupadas em intervalos")
    print("="*80)
    
    try:
        processar_todos_arquivos()
        print("\nProcessamento concluído com sucesso!")
    except Exception as e:
        print(f"\nErro durante o processamento: {str(e)}")
        print("Detalhes do erro:")
        import traceback
        traceback.print_exc() 