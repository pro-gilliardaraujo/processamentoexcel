"""
Script para processamento unificado de dados de monitoramento de colhedoras e transbordos.
Lê arquivos TXT ou CSV nas pastas especificadas, processa-os e gera um único arquivo Excel 
com todas as planilhas, adicionando prefixos "CD_" para colhedoras e "TT_" para transbordos.
"""

import pandas as pd
import numpy as np
import os
import glob
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configurações
processCsv = True  # Altere para True quando quiser processar arquivos CSV
arquivo_saida_unificado = "relatorio_unificado.xlsx"  # Nome do arquivo de saída unificado

# Carregar configurações
def carregar_config_calculos():
    """
    Carrega as configurações de cálculos do arquivo JSON.
    Se o arquivo não existir, retorna configurações padrão.
    """
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config", "calculos_config.json")
    
    # Configuração padrão
    config_padrao = {
        "CD": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [],
                "grupos_operacao_excluidos": ["Manutenção"]
            },
            "operadores_excluidos": ["9999 - TROCA DE TURNO"],
            "equipamentos_excluidos": []
        },
        "TT": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [],
                "grupos_operacao_excluidos": ["Manutenção"]
            },
            "operadores_excluidos": ["9999 - TROCA DE TURNO"],
            "equipamentos_excluidos": []
        }
    }
    
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            print(f"Arquivo de configuração não encontrado em {config_path}. Usando configuração padrão.")
            return config_padrao
    except Exception as e:
        print(f"Erro ao carregar configurações: {str(e)}. Usando configuração padrão.")
        return config_padrao

# Carregar configurações
CONFIG_CALCULOS = carregar_config_calculos()

# Constantes comuns
# OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO"]
# Removida a constante hardcoded já que estamos usando a configuração dinâmica

# Constantes para colhedoras
COLUNAS_REMOVER_COLHEDORAS = [
    'Justificativa Corte Base Desligado',
    'Latitude',
    'Longitude',
    'Regional',
    'Tipo de Equipamento',
    'Unidade',
    'Centro de Custo'
]

COLUNAS_DESEJADAS_COLHEDORAS = [
    'Data', 'Hora', 'Equipamento', 'Apertura do Rolo', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo', 'Corte Base Automatico/Manual',
    'Descricao Equipamento', 'Estado', 'Estado Operacional', 'Esteira Ligada',
    'Field Cruiser', 'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Implemento Ligado', 'Motor Ligado', 'Operacao', 'Operador', 'Pressao de Corte',
    'RPM Extrator', 'RPM Motor', 'RTK (Piloto Automatico)', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parada com Motor Ligado',
    'Horas Produtivas'
]

# Constantes para transbordos
COLUNAS_REMOVER_TRANSBORDOS = [
    'Latitude',
    'Longitude',
    'Regional',
    'Unidade',
    'Centro de Custo'
]

COLUNAS_DESEJADAS_TRANSBORDOS = [
    'Data', 'Hora', 'Equipamento', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo',
    'Descricao Equipamento', 'Estado', 'Estado Operacional',
    'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Motor Ligado', 'Operacao', 'Operador',
    'RPM Motor', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parado Com Motor Ligado',
    'Horas Produtivas', 'GPS'
]

# Funções comuns
def calcular_porcentagem(numerador, denominador, precisao=4):
    """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
    if denominador > 0:
        return round((numerador / denominador), precisao)
    return 0.0

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento
    """
    # Determinar o tipo de equipamento (CD ou TT) com base nas colunas
    tipo_equipamento = "CD" if "Parada com Motor Ligado" in df.columns else "TT"
    
    # Obter operadores e equipamentos excluídos da configuração
    operadores_excluidos = CONFIG_CALCULOS.get(tipo_equipamento, {}).get("operadores_excluidos", ["9999 - TROCA DE TURNO"])
    equipamentos_excluidos = CONFIG_CALCULOS.get(tipo_equipamento, {}).get("equipamentos_excluidos", [])
    
    # Filtramos os dados excluindo os operadores e equipamentos configurados
    df_filtrado = df[~df['Operador'].isin(operadores_excluidos)]
    df_filtrado = df_filtrado[~df_filtrado['Equipamento'].isin(equipamentos_excluidos)]
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df_filtrado['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 4)
        
        # Calcular horas de manutenção
        manutencao = round(dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum(), 4)
        
        # A disponibilidade mecânica é o percentual de tempo fora de manutenção
        disp_mecanica = calcular_porcentagem(total_horas - manutencao, total_horas)
        
        resultados.append({
            'Frota': equipamento,
            'Disponibilidade': disp_mecanica
        })
    
    return pd.DataFrame(resultados)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Esta função NÃO aplica qualquer filtro de operador.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 2)
        
        # Calcular a diferença para 24 horas
        diferenca_24h = round(max(24 - total_horas, 0), 2)
        
        resultados.append({
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        })
    
    return pd.DataFrame(resultados)

# Funções para colhedoras
def processar_arquivo_colhedora(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias para colhedoras.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                for col in COLUNAS_DESEJADAS_COLHEDORAS:
                    if col not in df.columns:
                        df[col] = np.nan
                colunas_existentes = [col for col in COLUNAS_DESEJADAS_COLHEDORAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS_COLHEDORAS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Verificar se 'Data/Hora' existe e processá-la
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Conversão e cálculo de diferenças de hora
            df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular a diferença de hora em segundos para maior precisão e depois converter para horas
            if 'Diferença_Hora' not in df.columns or df['Diferença_Hora'].isna().any():
                df['Diferença_Hora'] = df['Hora'].diff().dt.total_seconds() / 3600
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if pd.isna(x) or x < 0 else x)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            else:
                df['Diferença_Hora'] = pd.to_numeric(df['Diferença_Hora'].astype(str).str.strip(), errors='coerce')
                df['Diferença_Hora'] = df['Diferença_Hora'].fillna(0)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            
            # Cálculos adicionais
            RPM_MINIMO = 300
            if 'Parada com Motor Ligado' not in df.columns:
                df['Parada com Motor Ligado'] = ((df['Velocidade'] == 0) & 
                                              (df['RPM Motor'] >= RPM_MINIMO)).astype(int)
            
            # Verificar se Horas Produtivas já existe
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                df['Horas Produtivas'] = df.apply(
                    lambda row: round(row['Diferença_Hora'], 4) if row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
            else:
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # Conversão de colunas binárias para valores numéricos
            for col in ['Esteira Ligada', 'Motor Ligado', 'Field Cruiser', 'RTK (Piloto Automatico)', 'Implemento Ligado']:
                if col in df.columns:
                    if df[col].dtype == 'object':
                        df[col] = df[col].replace({'LIGADO': 1, 'DESLIGADO': 0})
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER_COLHEDORAS, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS_COLHEDORAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS_COLHEDORAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS_COLHEDORAS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo_colhedora(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado para colhedoras.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Obter operadores e equipamentos excluídos da configuração
    operadores_excluidos = CONFIG_CALCULOS.get("CD", {}).get("operadores_excluidos", ["9999 - TROCA DE TURNO"])
    equipamentos_excluidos = CONFIG_CALCULOS.get("CD", {}).get("equipamentos_excluidos", [])
    
    # Filtrar equipamentos excluídos
    df = df[~df['Equipamento'].isin(equipamentos_excluidos)]
    
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(operadores_excluidos)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Verificar se há dados suficientes para esta combinação
        if len(dados_filtrados) == 0:
            continue
        
        # Horas totais - manter mais casas decimais para cálculos intermediários
        horas_totais = round(dados_filtrados['Diferença_Hora'].sum(), 4)
        
        # Horas elevador (Esteira Ligada = 1 E Pressão de Corte > 400)
        horas_elevador = round(dados_filtrados[
            (dados_filtrados['Esteira Ligada'] == 1) & 
            (dados_filtrados['Pressao de Corte'] > 400)
        ]['Diferença_Hora'].sum(), 4)
        
        # Percentual horas elevador (em decimal 0-1)
        percent_elevador = calcular_porcentagem(horas_elevador, horas_totais)
        
        # RTK (Piloto Automático = 1 e Field Cruiser = 1)
        rtk = round(dados_filtrados[(dados_filtrados['RTK (Piloto Automatico)'] == 1) & 
                             (dados_filtrados['Field Cruiser'] == 1)]['Diferença_Hora'].sum(), 4)
        
        # Horas Produtivas
        horas_produtivas = round(dados_filtrados['Horas Produtivas'].sum(), 4)
        
        # % Utilização RTK (em decimal 0-1)
        utilizacao_rtk = calcular_porcentagem(rtk, horas_produtivas)
        
        # Motor Ligado
        motor_ligado = round(dados_filtrados[dados_filtrados['Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # % Eficiência Elevador (em decimal 0-1)
        eficiencia_elevador = calcular_porcentagem(horas_elevador, motor_ligado)
        
        # Parado com Motor Ligado
        parado_motor_ligado = round(dados_filtrados[dados_filtrados['Parada com Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # % Parado com motor ligado (em decimal 0-1)
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas elevador': horas_elevador,
            '%': percent_elevador,
            'RTK': rtk,
            'Horas Produtivas': horas_produtivas,
            '% Utilização RTK': utilizacao_rtk,
            'Motor Ligado': motor_ligado,
            '% Eficiência Elevador': eficiencia_elevador,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor
        })
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica_colhedora(base_calculo):
    """
    Calcula a eficiência energética por operador para colhedoras.
    Eficiência energética = Horas elevador / Horas motor ligado
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Eficiência Energética = horas elevador / motor ligado
        horas_elevador_sum = round(dados_op['Horas elevador'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        # Calcular eficiência - já está em decimal, não precisa multiplicar por 100
        eficiencia = calcular_porcentagem(horas_elevador_sum, motor_ligado_sum)
        
        # Garantir que não ultrapasse 100%
        eficiencia = min(eficiencia, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Eficiência': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_hora_elevador(df, base_calculo):
    """
    Calcula as horas de elevador por operador.
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Horas de elevador por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Somar horas de elevador da base de cálculo
        horas_elevador_sum = round(dados_op['Horas elevador'].sum(), 2)
        
        resultados.append({
            'Operador': operador,
            'Horas': horas_elevador_sum
        })
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso_colhedora(base_calculo, df_base):
    """
    Calcula o percentual de motor ocioso por operador para colhedoras.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
        df_base (DataFrame): DataFrame base processado
    
    Returns:
        DataFrame: Dados de motor ocioso por operador incluindo tempo operação, tempo ocioso e porcentagem
    """
    # Carregar configuração para saber quais operações e grupos devem ser excluídos
    config = carregar_config_calculos()
    operacoes_excluidas = config.get("CD", {}).get("motor_ocioso", {}).get("operacoes_excluidas", [])
    grupos_operacao_excluidos = config.get("CD", {}).get("motor_ocioso", {}).get("grupos_operacao_excluidos", ["Manutenção"])
    
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro_base = (df_base['Operador'] == operador) & (df_base['Grupo Equipamento/Frente'] == grupo)
        df_operador = df_base[filtro_base].copy()
        
        # Filtrar dados - excluir operações específicas
        if operacoes_excluidas:
            df_operador = df_operador[~df_operador['Operacao'].isin(operacoes_excluidas)]
        
        # Excluir grupos de operação específicos
        if grupos_operacao_excluidos:
            df_operador = df_operador[~df_operador['Grupo Operacao'].isin(grupos_operacao_excluidos)]
        
        # Calcular motor ocioso
        tempo_motor_ligado = round(df_operador[df_operador['Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # Parado com Motor Ligado (já calculado previamente na preparação dos dados)
        tempo_parado_motor_ligado = round(df_operador[df_operador['Parada com Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # Calcular percentual motor ocioso
        if tempo_motor_ligado > 0:
            percentual = round(tempo_parado_motor_ligado / tempo_motor_ligado, 4)
        else:
            percentual = 0.0
        
        resultados.append({
            'Operador': operador,
            'Tempo Operação': tempo_motor_ligado,
            'Tempo Ocioso': tempo_parado_motor_ligado,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def calcular_uso_gps_colhedora(df, base_calculo):
    """
    Calcula o percentual de uso de GPS por operador para colhedoras.
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados base para este operador e grupo
        filtro_base = (df['Operador'] == operador) & (df['Grupo Equipamento/Frente'] == grupo)
        dados_op_base = df[filtro_base]
        
        # Calcular tempo total trabalhando
        tempo_trabalhando = round(dados_op_base[
            (dados_op_base['Estado'].isin(['TRABALHANDO', 'COLHEITA']))
        ]['Diferença_Hora'].sum(), 4)
        
        # Calcular tempo com GPS ativo
        tempo_gps_ativo = round(dados_op_base[
            (dados_op_base['Estado'].isin(['TRABALHANDO', 'COLHEITA'])) &
            (dados_op_base['RTK (Piloto Automatico)'] == 1) &
            (dados_op_base['Velocidade'] > 0)
        ]['Diferença_Hora'].sum(), 4)
        
        # Calcular percentual em formato decimal (0-1)
        percentual = calcular_porcentagem(tempo_gps_ativo, tempo_trabalhando)
        
        # Garantir que não ultrapasse 100% (1.0)
        percentual = min(percentual, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

# Funções para transbordos
def processar_arquivo_transbordo(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias para transbordos.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                for col in COLUNAS_DESEJADAS_TRANSBORDOS:
                    if col not in df.columns:
                        df[col] = np.nan
                colunas_existentes = [col for col in COLUNAS_DESEJADAS_TRANSBORDOS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS_TRANSBORDOS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Verificar se 'Data/Hora' existe e processá-la
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Conversão e cálculo de diferenças de hora
            if isinstance(df['Hora'].iloc[0], str):
                df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular a diferença de hora se ainda não existir
            if 'Diferença_Hora' not in df.columns or df['Diferença_Hora'].isna().any():
                df['Diferença_Hora'] = df['Hora'].diff().dt.total_seconds() / 3600
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if pd.isna(x) or x < 0 else x)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            else:
                df['Diferença_Hora'] = pd.to_numeric(df['Diferença_Hora'].astype(str).str.strip(), errors='coerce')
                df['Diferença_Hora'] = df['Diferença_Hora'].fillna(0)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            
            # Conversão de Motor Ligado para formato numérico
            if 'Motor Ligado' in df.columns:
                if df['Motor Ligado'].dtype == 'object':
                    df['Motor Ligado'] = df['Motor Ligado'].replace({'LIGADO': 1, 'DESLIGADO': 0})
                df['Motor Ligado'] = pd.to_numeric(df['Motor Ligado'], errors='coerce').fillna(0).astype(int)
            
            # Cálculos específicos para transbordos
            RPM_MINIMO = 300
            
            # Verificar e calcular "Parado Com Motor Ligado" se necessário
            if 'Parado Com Motor Ligado' not in df.columns:
                df['Parado Com Motor Ligado'] = ((df['Velocidade'] == 0) & 
                                               (df['RPM Motor'] >= RPM_MINIMO)).astype(int)
            
            # Verificar se Horas Produtivas já existe
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                df['Horas Produtivas'] = df.apply(
                    lambda row: round(row['Diferença_Hora'], 4) if row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
            else:
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # Coluna de GPS - Para transbordos
            if 'RTK (Piloto Automatico)' in df.columns:
                df['GPS'] = df.apply(
                    lambda row: row['Diferença_Hora'] if row.get('RTK (Piloto Automatico)', 0) == 1 
                    and row['Velocidade'] > 0 and row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
            else:
                df['GPS'] = 0
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER_TRANSBORDOS, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS_TRANSBORDOS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS_TRANSBORDOS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS_TRANSBORDOS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo_transbordo(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado para transbordos.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Obter operadores e equipamentos excluídos da configuração
    operadores_excluidos = CONFIG_CALCULOS.get("TT", {}).get("operadores_excluidos", ["9999 - TROCA DE TURNO"])
    equipamentos_excluidos = CONFIG_CALCULOS.get("TT", {}).get("equipamentos_excluidos", [])
    
    # Filtrar equipamentos excluídos
    df = df[~df['Equipamento'].isin(equipamentos_excluidos)]
    
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(operadores_excluidos)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Verificar se há dados suficientes para esta combinação
        if len(dados_filtrados) == 0:
            continue
        
        # Horas totais - manter mais casas decimais para cálculos intermediários
        horas_totais = round(dados_filtrados['Diferença_Hora'].sum(), 4)
        
        # Horas Produtivas
        horas_produtivas = round(dados_filtrados['Horas Produtivas'].sum(), 4)
        
        # GPS para transbordos
        gps = round(dados_filtrados['GPS'].sum(), 4)
        
        # % Utilização GPS (em decimal 0-1)
        utilizacao_gps = calcular_porcentagem(gps, horas_produtivas)
        
        # Motor Ligado
        motor_ligado = round(dados_filtrados[dados_filtrados['Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # Parado com Motor Ligado
        parado_motor_ligado = round(dados_filtrados[dados_filtrados['Parado Com Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # % Parado com motor ligado (em decimal 0-1)
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        # Falta de Apontamento - Contabilizar apenas registros explicitamente marcados
        falta_apontamento = round(dados_filtrados[
            (dados_filtrados['Motor Ligado'] == 1) & 
            (
                (dados_filtrados['Codigo da Operacao'] == 8340) |
                (dados_filtrados['Codigo da Operacao'].astype(str).str.startswith('8340')) |
                (dados_filtrados['Operacao'].astype(str).str.contains('FALTA DE APONTAMENTO', case=False))
            )
        ]['Diferença_Hora'].sum(), 4)
        
        # % Falta de apontamento (em decimal 0-1)
        percent_falta_apontamento = calcular_porcentagem(falta_apontamento, motor_ligado)
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas Produtivas': horas_produtivas,
            'GPS': gps,
            '% Utilização GPS': utilizacao_gps,
            'Motor Ligado': motor_ligado,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor,
            'Falta de Apontamento': falta_apontamento,
            '% Falta de Apontamento': percent_falta_apontamento
        })
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica_transbordo(base_calculo):
    """
    Calcula a eficiência energética por operador para transbordos.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Eficiência Energética para transbordos = Horas Produtivas / Horas Totais
        horas_produtivas_sum = round(dados_op['Horas Produtivas'].sum(), 4)
        horas_totais_sum = round(dados_op['Horas totais'].sum(), 4)
        
        # Calcular eficiência
        eficiencia = calcular_porcentagem(horas_produtivas_sum, horas_totais_sum)
        
        # Garantir que não ultrapasse 100%
        eficiencia = min(eficiencia, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Eficiência': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso_transbordo(base_calculo, df_base):
    """
    Calcula o percentual de motor ocioso por operador para transbordos.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
        df_base (DataFrame): DataFrame base processado
    
    Returns:
        DataFrame: Dados de motor ocioso por operador incluindo tempo operação, tempo ocioso e porcentagem
    """
    # Carregar configuração para saber quais operações devem ser excluídas do cálculo de motor ocioso
    config = carregar_config_calculos()
    operacoes_excluidas = config.get("TT", {}).get("motor_ocioso", {}).get("operacoes_excluidas", [])
    grupos_operacao_excluidos = config.get("TT", {}).get("motor_ocioso", {}).get("grupos_operacao_excluidos", ["Manutenção"])
    
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados base para este operador e grupo
        filtro_base = (df_base['Operador'] == operador) & (df_base['Grupo Equipamento/Frente'] == grupo)
        df_operador = df_base[filtro_base].copy()
        
        # Filtrar dados - excluir operações específicas ou grupos de operação
        if operacoes_excluidas:
            df_operador = df_operador[~df_operador['Operacao'].isin(operacoes_excluidas)]
        
        if grupos_operacao_excluidos:
            df_operador = df_operador[~df_operador['Grupo Operacao'].isin(grupos_operacao_excluidos)]
        
        # Calcular motor ocioso
        # Condição de parado com motor ligado: Velocidade = 0, RPM Motor >= 300, Motor Ligado = 1
        tempo_motor_ligado = round(df_operador[df_operador['Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # Parado com Motor Ligado
        tempo_parado_motor_ligado = round(df_operador[
            (df_operador['Velocidade'] == 0) & 
            (df_operador['RPM Motor'] >= 300) & 
            (df_operador['Motor Ligado'] == 1)
        ]['Diferença_Hora'].sum(), 4)
        
        # Calcular percentual motor ocioso
        if tempo_motor_ligado > 0:
            percentual = round(tempo_parado_motor_ligado / tempo_motor_ligado, 4)
        else:
            percentual = 0.0
        
        resultados.append({
            'Operador': operador,
            'Tempo Operação': tempo_motor_ligado,
            'Tempo Ocioso': tempo_parado_motor_ligado,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def calcular_falta_apontamento(base_calculo):
    """
    Calcula o percentual de falta de apontamento por operador.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de falta de apontamento por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Usar os valores já calculados em base_calculo
        falta_apontamento_sum = round(dados_op['Falta de Apontamento'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        # Calcular percentual
        percentual = calcular_porcentagem(falta_apontamento_sum, motor_ligado_sum)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def calcular_uso_gps_transbordo(base_calculo):
    """
    Calcula o percentual de uso de GPS por operador para transbordos.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Uso GPS = GPS / Horas Produtivas
        gps_sum = round(dados_op['GPS'].sum(), 4)
        horas_produtivas_sum = round(dados_op['Horas Produtivas'].sum(), 4)
        
        percentual = calcular_porcentagem(gps_sum, horas_produtivas_sum)
        
        # Garantir que não ultrapasse 100% (1.0)
        percentual = min(percentual, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def adicionar_planilhas_ao_excel(writer, planilhas_dados, tipo):
    """
    Adiciona planilhas ao arquivo Excel e formata células.
    
    Args:
        writer: ExcelWriter para o arquivo
        planilhas_dados: Dicionário com os dados a serem adicionados
        tipo: Tipo de equipamento (CD ou TT)
    """
    # Adicionar cada planilha ao Excel
    for nome_planilha, dados in planilhas_dados.items():
        # Pular planilhas vazias
        if dados is None or len(dados) == 0:
            continue
        
        # Arredondar valores numéricos para melhor visualização
        if nome_planilha == "Base Calculo":
            for col in dados.columns:
                if col in ['Horas totais', 'Horas Produtivas', 'GPS', 'Motor Ligado', 'Parado Com Motor Ligado', 'Horas elevador']:
                    dados[col] = dados[col].apply(lambda x: round(float(x), 2) if pd.notnull(x) else x)
                elif col in ['% Utilização GPS', '% Parado com motor ligado', '% Hora Elevador', '% Eficiência Energética']:
                    dados[col] = dados[col].apply(lambda x: round(float(x), 4) if pd.notnull(x) else x)
        
        # Arredondar valores para outras planilhas específicas
        if nome_planilha.startswith("Disponibilidade"):
            if 'Disponibilidade' in dados.columns:
                dados['Disponibilidade'] = dados['Disponibilidade'].apply(lambda x: round(float(x), 4) if pd.notnull(x) else x)
        
        if nome_planilha.startswith("Eficiência"):
            if 'Eficiência' in dados.columns:
                dados['Eficiência'] = dados['Eficiência'].apply(lambda x: round(float(x), 4) if pd.notnull(x) else x)
        
        if nome_planilha.startswith("Hora Elevador"):
            if 'Horas' in dados.columns:
                dados['Horas'] = dados['Horas'].apply(lambda x: round(float(x), 2) if pd.notnull(x) else x)
        
        if nome_planilha.startswith("Motor Ocioso"):
            if 'Porcentagem' in dados.columns:
                dados['Porcentagem'] = dados['Porcentagem'].apply(lambda x: round(float(x), 4) if pd.notnull(x) else x)
            if 'Tempo Operação' in dados.columns:
                dados['Tempo Operação'] = dados['Tempo Operação'].apply(lambda x: round(float(x), 2) if pd.notnull(x) else x)
            if 'Tempo Ocioso' in dados.columns:
                dados['Tempo Ocioso'] = dados['Tempo Ocioso'].apply(lambda x: round(float(x), 2) if pd.notnull(x) else x)
        
        if nome_planilha.startswith("Falta Apontamento") or nome_planilha.startswith("Uso GPS"):
            if 'Porcentagem' in dados.columns:
                dados['Porcentagem'] = dados['Porcentagem'].apply(lambda x: round(float(x), 4) if pd.notnull(x) else x)
        
        if nome_planilha.startswith("Horas por Frota"):
            if 'Horas Registradas' in dados.columns:
                dados['Horas Registradas'] = dados['Horas Registradas'].apply(lambda x: round(float(x), 2) if pd.notnull(x) else x)
            if 'Diferença para 24h' in dados.columns:
                dados['Diferença para 24h'] = dados['Diferença para 24h'].apply(lambda x: round(float(x), 2) if pd.notnull(x) else x)
        
        # Adicionar a planilha ao Excel
        dados.to_excel(writer, sheet_name=nome_planilha, index=False)
        
        # Aplicar formatação às células
        worksheet = writer.sheets[nome_planilha]
        
        # Aplicar formatação de acordo com o tipo de planilha
        if nome_planilha.startswith("Disponibilidade"):
            # Formatar coluna de porcentagem
            for row in range(2, len(dados) + 2):  # +2 porque Excel começa em 1 e tem cabeçalho
                cell = worksheet.cell(row=row, column=2)  # Coluna B
                cell.number_format = '0.00%'  # Formato de porcentagem
        
        elif nome_planilha.startswith("Eficiência"):
            # Formatar coluna de porcentagem
            for row in range(2, len(dados) + 2):
                cell = worksheet.cell(row=row, column=2)  # Coluna B
                cell.number_format = '0.00%'  # Formato de porcentagem
        
        elif nome_planilha.startswith("Hora Elevador"):
            # Formatar coluna de horas
            for row in range(2, len(dados) + 2):
                cell = worksheet.cell(row=row, column=2)  # Coluna B
                cell.number_format = '0.00'  # Formato de número com 2 casas decimais
        
        elif nome_planilha.startswith("Motor Ocioso"):
            # Formatar colunas de tempo e porcentagem
            for row in range(2, len(dados) + 2):
                # Tempo Operação (coluna B)
                cell_tempo_op = worksheet.cell(row=row, column=2)
                cell_tempo_op.number_format = '0.00'  # Formato decimal com 2 casas
                
                # Tempo Ocioso (coluna C)
                cell_tempo_oc = worksheet.cell(row=row, column=3)
                cell_tempo_oc.number_format = '0.00'  # Formato decimal com 2 casas
                
                # Porcentagem (coluna D)
                cell_porc = worksheet.cell(row=row, column=4)
                cell_porc.number_format = '0.00%'  # Formato de porcentagem com 2 casas
        
        elif nome_planilha.startswith("Falta Apontamento") or nome_planilha.startswith("Uso GPS"):
            # Formatar coluna de porcentagem
            for row in range(2, len(dados) + 2):
                cell = worksheet.cell(row=row, column=2)  # Coluna B
                cell.number_format = '0.00%'  # Formato de porcentagem
        
        elif nome_planilha.startswith("Horas por Frota"):
            # Formatar colunas de horas
            for row in range(2, len(dados) + 2):
                cell_b = worksheet.cell(row=row, column=2)  # Coluna B (Horas Registradas)
                cell_b.number_format = '0.00'  # Formato de número com 2 casas decimais
                
                cell_c = worksheet.cell(row=row, column=3)  # Coluna C (Diferença para 24h)
                cell_c.number_format = '0.00'  # Formato de número com 2 casas decimais

def calcular_eficiencia_operacional(df):
    """
    Calcula a eficiência operacional para cada equipamento.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Eficiência operacional por equipamento
    """
    # Determinar o tipo de equipamento (CD ou TT) com base nas colunas
    tipo_equipamento = "CD" if "Parada com Motor Ligado" in df.columns else "TT"
    
    # Obter operadores e equipamentos excluídos da configuração
    operadores_excluidos = CONFIG_CALCULOS.get(tipo_equipamento, {}).get("operadores_excluidos", ["9999 - TROCA DE TURNO"])
    equipamentos_excluidos = CONFIG_CALCULOS.get(tipo_equipamento, {}).get("equipamentos_excluidos", [])
    
    # Filtramos os dados excluindo os operadores e equipamentos da configuração
    df_filtrado = df[~df['Operador'].isin(operadores_excluidos)]
    df_filtrado = df_filtrado[~df_filtrado['Equipamento'].isin(equipamentos_excluidos)]
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df_filtrado['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 4)
        
        # Calcular horas produtivas (Colheita) ou Efetivas (Transbordamento)
        if tipo_equipamento == "CD":
            tempo_produtivo = round(dados_equip[dados_equip['Grupo Operacao'] == 'Colheita']['Diferença_Hora'].sum(), 4)
        else:  # TT
            tempo_produtivo = round(dados_equip[dados_equip['Grupo Operacao'] == 'Transbordamento']['Diferença_Hora'].sum(), 4)
        
        # A eficiência operacional é o percentual do tempo produtivo
        eficiencia = calcular_porcentagem(tempo_produtivo, total_horas)
        
        resultados.append({
            'Frota': equipamento,
            'Eficiencia': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_metricas_por_equipamento(df):
    """
    Calcula várias métricas por equipamento.
    
    Args:
        df (DataFrame): DataFrame com dados processados
        
    Returns:
        DataFrame: Métricas por equipamento
    """
    # Determinar o tipo de equipamento (CD ou TT) com base nas colunas
    tipo_equipamento = "CD" if "Parada com Motor Ligado" in df.columns else "TT"
    
    # Obter operadores e equipamentos excluídos da configuração
    operadores_excluidos = CONFIG_CALCULOS.get(tipo_equipamento, {}).get("operadores_excluidos", ["9999 - TROCA DE TURNO"])
    equipamentos_excluidos = CONFIG_CALCULOS.get(tipo_equipamento, {}).get("equipamentos_excluidos", [])
    
    # Filtramos os dados excluindo os operadores e equipamentos da configuração
    df_filtrado = df[~df['Operador'].isin(operadores_excluidos)]
    df_filtrado = df_filtrado[~df_filtrado['Equipamento'].isin(equipamentos_excluidos)]
    
    # Agrupar por Equipamento e calcular métricas
    resultados = []
    equipamentos = df_filtrado['Equipamento'].unique()
    
    for equipamento in equipamentos:
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 4)
        
        # Calcular métricas específicas por tipo de equipamento
        if tipo_equipamento == "CD":
            # Para Colhedoras
            colheita = round(dados_equip[dados_equip['Grupo Operacao'] == 'Colheita']['Diferença_Hora'].sum(), 4)
            manutencao = round(dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum(), 4)
            abastecimento = round(dados_equip[dados_equip['Grupo Operacao'] == 'Abastecimento']['Diferença_Hora'].sum(), 4)
            manobra = round(dados_equip[dados_equip['Grupo Operacao'] == 'Manobra']['Diferença_Hora'].sum(), 4)
            transito = round(dados_equip[dados_equip['Grupo Operacao'] == 'Transito']['Diferença_Hora'].sum(), 4)
            outros = total_horas - (colheita + manutencao + abastecimento + manobra + transito)
            
            eficiencia = calcular_porcentagem(colheita, total_horas)
            disponibilidade = calcular_porcentagem(total_horas - manutencao, total_horas)
            utilizacao = calcular_porcentagem(colheita, total_horas - manutencao)
            
            # Adicionar aos resultados
            resultados.append({
                'Equipamento': equipamento,
                'Total Horas': total_horas,
                'Colheita': colheita,
                'Manutenção': manutencao,
                'Abastecimento': abastecimento,
                'Manobra': manobra,
                'Transito': transito,
                'Outros': outros,
                'Eficiência': eficiencia,
                'Disponibilidade': disponibilidade,
                'Utilização': utilizacao
            })
        else:
            # Para Transbordos
            transbordamento = round(dados_equip[dados_equip['Grupo Operacao'] == 'Transbordamento']['Diferença_Hora'].sum(), 4)
            manutencao = round(dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum(), 4)
            abastecimento = round(dados_equip[dados_equip['Grupo Operacao'] == 'Abastecimento']['Diferença_Hora'].sum(), 4)
            paradas = round(dados_equip[dados_equip['Grupo Operacao'] == 'Paradas']['Diferença_Hora'].sum(), 4)
            transito = round(dados_equip[dados_equip['Grupo Operacao'] == 'Transito']['Diferença_Hora'].sum(), 4)
            outros = total_horas - (transbordamento + manutencao + abastecimento + paradas + transito)
            
            eficiencia = calcular_porcentagem(transbordamento, total_horas)
            disponibilidade = calcular_porcentagem(total_horas - manutencao, total_horas)
            utilizacao = calcular_porcentagem(transbordamento, total_horas - manutencao)
            
            # Adicionar aos resultados
            resultados.append({
                'Equipamento': equipamento,
                'Total Horas': total_horas,
                'Transbordamento': transbordamento,
                'Manutenção': manutencao,
                'Abastecimento': abastecimento,
                'Paradas': paradas,
                'Transito': transito,
                'Outros': outros,
                'Eficiência': eficiencia,
                'Disponibilidade': disponibilidade,
                'Utilização': utilizacao
            })
    
    return pd.DataFrame(resultados)

def processar_arquivos_unificados():
    """
    Processa todos os arquivos de colhedoras e transbordos, combinando-os em um único arquivo Excel
    com prefixos 'CD_' para colhedoras e 'TT_' para transbordos.
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_colhedoras = os.path.join(diretorio_raiz, "dados", "colhedoras")
    diretorio_transbordos = os.path.join(diretorio_raiz, "dados", "transbordos")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    
    # Verificar se os diretórios existem, caso contrário criar
    for diretorio in [diretorio_dados, diretorio_colhedoras, diretorio_transbordos, diretorio_saida]:
        if not os.path.exists(diretorio):
            os.makedirs(diretorio)
    
    # Caminho completo para o arquivo Excel unificado
    caminho_saida_unificado = os.path.join(diretorio_saida, arquivo_saida_unificado)
    
    # Criar o objeto ExcelWriter para o arquivo unificado
    writer = pd.ExcelWriter(caminho_saida_unificado, engine='openpyxl')
    workbook = writer.book
    
    # Lista para controlar quais equipamentos já foram processados
    equipamentos_processados = []
    
    # Processar arquivos de colhedoras
    print("="*80)
    print("Processando arquivos de COLHEDORAS...")
    
    # Lista de diretórios para buscar arquivos de colhedoras
    diretorios_busca_colhedoras = [diretorio_dados, diretorio_colhedoras]
    arquivos_colhedoras = []
    
    for diretorio in diretorios_busca_colhedoras:
        # Adicionar arquivos TXT sempre
        arquivos_colhedoras += glob.glob(os.path.join(diretorio, "RV Colhedora*.txt"))
        arquivos_colhedoras += glob.glob(os.path.join(diretorio, "*colhedora*.txt"))
        arquivos_colhedoras += glob.glob(os.path.join(diretorio, "colhedora*.txt"))
        
        # Adicionar arquivos CSV apenas se processCsv for True
        if processCsv:
            arquivos_colhedoras += glob.glob(os.path.join(diretorio, "RV Colhedora*.csv"))
            arquivos_colhedoras += glob.glob(os.path.join(diretorio, "*colhedora*.csv"))
            arquivos_colhedoras += glob.glob(os.path.join(diretorio, "colhedora*.csv"))
    
    # Filtrar arquivos que contenham "transbordo" no nome (case insensitive)
    arquivos_colhedoras = [arquivo for arquivo in arquivos_colhedoras if "transbordo" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos_colhedoras = list(set(arquivos_colhedoras))
    
    print(f"Encontrados {len(arquivos_colhedoras)} arquivos de colhedoras para processar.")
    
    # Processar cada arquivo de colhedora
    for arquivo in arquivos_colhedoras:
        print(f"\nProcessando arquivo de colhedora: {os.path.basename(arquivo)}")
        
        # Processar o arquivo base
        df_base = processar_arquivo_colhedora(arquivo)
        if df_base is None or len(df_base) == 0:
            print(f"Arquivo {os.path.basename(arquivo)} sem dados válidos. Pulando.")
            continue
        
        # Extrair os equipamentos deste arquivo
        equipamentos = df_base['Equipamento'].unique()
        
        # Verificar se algum equipamento já foi processado e remover dos dados
        equipamentos_a_processar = [eq for eq in equipamentos if eq not in equipamentos_processados]
        
        if not equipamentos_a_processar:
            print(f"Todos os equipamentos de {os.path.basename(arquivo)} já foram processados. Pulando.")
            continue
        
        # Filtrar dados apenas para equipamentos ainda não processados
        df_base_filtrado = df_base[df_base['Equipamento'].isin(equipamentos_a_processar)]
        
        # Calcular a Base Calculo
        base_calculo = calcular_base_calculo_colhedora(df_base_filtrado)
        
        # Calcular as métricas auxiliares
        disp_mecanica = calcular_disponibilidade_mecanica(df_base_filtrado)
        eficiencia_energetica = calcular_eficiencia_energetica_colhedora(base_calculo)
        hora_elevador = calcular_hora_elevador(df_base_filtrado, base_calculo)
        motor_ocioso = calcular_motor_ocioso_colhedora(base_calculo, df_base_filtrado)
        uso_gps = calcular_uso_gps_colhedora(df_base_filtrado, base_calculo)
        horas_por_frota = calcular_horas_por_frota(df_base_filtrado)
        
        # Preparar dicionário de planilhas para adicionar
        planilhas_colhedoras = {
            'BASE': df_base_filtrado,
            'Base Calculo': base_calculo,
            '1_Disponibilidade Mecânica': disp_mecanica,
            '2_Eficiência Energética': eficiencia_energetica,
            '3_Hora Elevador': hora_elevador,
            '4_Motor Ocioso': motor_ocioso,
            '5_Uso GPS': uso_gps,
            'Horas por Frota': horas_por_frota
        }
        
        # Adicionar planilhas ao Excel
        adicionar_planilhas_ao_excel(writer, planilhas_colhedoras, 'CD')
        
        # Marcar estes equipamentos como processados
        equipamentos_processados.extend(equipamentos_a_processar)
    
    # Processar arquivos de transbordos
    print("\n" + "="*80)
    print("Processando arquivos de TRANSBORDOS...")
    
    # Lista de diretórios para buscar arquivos de transbordos
    diretorios_busca_transbordos = [diretorio_dados, diretorio_transbordos]
    arquivos_transbordos = []
    
    for diretorio in diretorios_busca_transbordos:
        # Adicionar arquivos TXT sempre
        arquivos_transbordos += glob.glob(os.path.join(diretorio, "RV Transbordo*.txt"))
        arquivos_transbordos += glob.glob(os.path.join(diretorio, "*transbordo*.txt"))
        arquivos_transbordos += glob.glob(os.path.join(diretorio, "frente*transbordos*.txt"))
        arquivos_transbordos += glob.glob(os.path.join(diretorio, "transbordo*.txt"))
        
        # Adicionar arquivos CSV apenas se processCsv for True
        if processCsv:
            arquivos_transbordos += glob.glob(os.path.join(diretorio, "RV Transbordo*.csv"))
            arquivos_transbordos += glob.glob(os.path.join(diretorio, "*transbordo*.csv"))
            arquivos_transbordos += glob.glob(os.path.join(diretorio, "frente*transbordos*.csv"))
            arquivos_transbordos += glob.glob(os.path.join(diretorio, "transbordo*.csv"))
    
    # Filtrar arquivos que contenham "colhedora" no nome (case insensitive)
    arquivos_transbordos = [arquivo for arquivo in arquivos_transbordos if "colhedora" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos_transbordos = list(set(arquivos_transbordos))
    
    print(f"Encontrados {len(arquivos_transbordos)} arquivos de transbordos para processar.")
    
    # Limpar a lista de equipamentos processados para transbordos
    equipamentos_processados = []
    
    # Processar cada arquivo de transbordo
    for arquivo in arquivos_transbordos:
        print(f"\nProcessando arquivo de transbordo: {os.path.basename(arquivo)}")
        
        # Processar o arquivo base
        df_base = processar_arquivo_transbordo(arquivo)
        if df_base is None or len(df_base) == 0:
            print(f"Arquivo {os.path.basename(arquivo)} sem dados válidos. Pulando.")
            continue
        
        # Extrair os equipamentos deste arquivo
        equipamentos = df_base['Equipamento'].unique()
        
        # Verificar se algum equipamento já foi processado e remover dos dados
        equipamentos_a_processar = [eq for eq in equipamentos if eq not in equipamentos_processados]
        
        if not equipamentos_a_processar:
            print(f"Todos os equipamentos de {os.path.basename(arquivo)} já foram processados. Pulando.")
            continue
        
        # Filtrar dados apenas para equipamentos ainda não processados
        df_base_filtrado = df_base[df_base['Equipamento'].isin(equipamentos_a_processar)]
        
        # Calcular a Base Calculo
        base_calculo = calcular_base_calculo_transbordo(df_base_filtrado)
        
        # Calcular as métricas auxiliares
        disp_mecanica = calcular_disponibilidade_mecanica(df_base_filtrado)
        eficiencia_energetica = calcular_eficiencia_energetica_transbordo(base_calculo)
        motor_ocioso = calcular_motor_ocioso_transbordo(base_calculo, df_base_filtrado)
        falta_apontamento = calcular_falta_apontamento(base_calculo)
        uso_gps = calcular_uso_gps_transbordo(base_calculo)
        horas_por_frota = calcular_horas_por_frota(df_base_filtrado)
        
        # Preparar dicionário de planilhas para adicionar
        planilhas_transbordos = {
            'BASE': df_base_filtrado,
            'Base Calculo': base_calculo,
            '1_Disponibilidade Mecânica': disp_mecanica,
            '2_Eficiência Energética': eficiencia_energetica,
            '3_Motor Ocioso': motor_ocioso,
            '4_Falta de Apontamento': falta_apontamento,
            '5_Uso GPS': uso_gps,
            'Horas por Frota': horas_por_frota
        }
        
        # Adicionar planilhas ao Excel
        adicionar_planilhas_ao_excel(writer, planilhas_transbordos, 'TT')
        
        # Marcar estes equipamentos como processados
        equipamentos_processados.extend(equipamentos_a_processar)
    
    # Salvar o arquivo Excel unificado
    writer.close()
    print("\n" + "="*80)
    print(f"Arquivo unificado gerado com sucesso em: {caminho_saida_unificado}")
    print("="*80)

if __name__ == "__main__":
    print("="*80)
    print("PROCESSAMENTO UNIFICADO DE COLHEDORAS E TRANSBORDOS")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Gerando um único arquivo Excel contendo todas as planilhas com prefixos:")
    print("  - 'CD_' para dados de colhedoras")
    print("  - 'TT_' para dados de transbordos")
    print("="*80)
    processar_arquivos_unificados()
    print("\nProcessamento concluído!") 