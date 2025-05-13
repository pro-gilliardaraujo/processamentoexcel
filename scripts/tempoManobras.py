#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import json
import pandas as pd
import numpy as np
import zipfile
import tempfile
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def encontrar_arquivos_zip_manobras(pasta_dados):
    """
    Encontra arquivos ZIP que contenham 'manobras' no nome.
    """
    arquivos_zip = []
    for arquivo in os.listdir(pasta_dados):
        if 'manobras' in arquivo.lower() and arquivo.endswith('.zip'):
            arquivos_zip.append(os.path.join(pasta_dados, arquivo))
    return arquivos_zip

def processar_csv_do_zip(arquivo_zip):
    """
    Extrai e processa os arquivos CSV de dentro do ZIP.
    """
    print(f"\nProcessando arquivo ZIP: {arquivo_zip}")
    dfs = []
    
    with tempfile.TemporaryDirectory() as temp_dir:
        with zipfile.ZipFile(arquivo_zip, 'r') as zip_ref:
            # Listar todos os arquivos no ZIP
            arquivos_csv = [f for f in zip_ref.namelist() if f.lower().endswith('.csv')]
            print(f"Arquivos CSV encontrados no ZIP: {len(arquivos_csv)}")
            
            for csv_file in arquivos_csv:
                print(f"\nExtraindo e processando: {csv_file}")
                zip_ref.extract(csv_file, temp_dir)
                
                # Ler o CSV
                caminho_csv = os.path.join(temp_dir, csv_file)
                try:
                    df = pd.read_csv(caminho_csv, sep=';', encoding='latin1')
                    print(f"Arquivo lido com sucesso: {len(df)} linhas")
                    
                    # Converter Data/Hora para datetime e separar em duas colunas
                    if 'Data/Hora' in df.columns:
                        try:
                            # Converter para datetime
                            df['Data/Hora'] = pd.to_datetime(df['Data/Hora'], format='%d/%m/%Y %H:%M:%S')
                        except ValueError:
                            try:
                                df['Data/Hora'] = pd.to_datetime(df['Data/Hora'], format='%d/%m/%Y %H:%M')
                            except ValueError as e:
                                print(f"Erro ao converter datas: {str(e)}")
                                print("Tentando inferir formato...")
                                df['Data/Hora'] = pd.to_datetime(df['Data/Hora'])
                        
                        # Separar em Data e Hora
                        df['Data'] = df['Data/Hora'].dt.strftime('%d/%m/%Y')
                        df['Hora'] = df['Data/Hora'].dt.strftime('%H:%M:%S')
                        print("Coluna Data/Hora separada em Data e Hora")
                    
                    # Ordenar por Equipamento e Data/Hora
                    if 'Equipamento' in df.columns and 'Data/Hora' in df.columns:
                        df = df.sort_values(['Equipamento', 'Data/Hora']).reset_index(drop=True)
                        print("Dados ordenados por Equipamento e Data/Hora")
                        
                        # Calcular diferença de tempo entre registros consecutivos
                        df['Diferença_Hora'] = df.groupby('Equipamento')['Data/Hora'].diff().dt.total_seconds()
                        # Converter para horas sem arredondamento
                        df['Diferença_Hora'] = df['Diferença_Hora'].astype('float64') / 3600.0
                        print("Diferenças de tempo calculadas")
                    
                    # Remover coluna Data/Hora original após separação
                    if 'Data/Hora' in df.columns:
                        df = df.drop('Data/Hora', axis=1)
                    
                    dfs.append(df)
                    print(f"Arquivo processado e adicionado ao conjunto de dados")
                except Exception as e:
                    print(f"Erro ao processar {csv_file}: {str(e)}")
                    continue
    
    if dfs:
        # Concatenar todos os DataFrames
        df_final = pd.concat(dfs, ignore_index=True)
        print(f"\nTotal de registros após concatenação: {len(df_final)}")
        
        # Ordenar novamente após a concatenação
        if 'Equipamento' in df_final.columns:
            df_final = df_final.sort_values(['Equipamento', 'Data', 'Hora']).reset_index(drop=True)
            print("Dados finais ordenados por Equipamento, Data e Hora")
        
        return df_final
    return None

def calcular_intervalos_manobra(df):
    """
    Calcula os intervalos de manobra, considerando:
    - Um intervalo começa com estado MANOBRA
    - Enquanto o intervalo está aberto, soma TODOS os tempos (MANOBRA e não-MANOBRA)
    - Se tempo em não-MANOBRA > 1 minuto, fecha o intervalo
    - Ao fechar, soma apenas até a última entrada válida (não inclui o tempo que fechou)
    """
    print("\nCalculando intervalos de manobra...")
    
    # Lista para armazenar os intervalos
    intervalos = []
    
    # Processar cada equipamento separadamente
    for equipamento in df['Equipamento'].unique():
        print(f"\nProcessando equipamento: {equipamento}")
        
        # Filtrar dados do equipamento
        df_equip = df[df['Equipamento'] == equipamento].copy()
        
        # Variáveis para controle do intervalo
        em_intervalo = False
        inicio_intervalo = None
        soma_intervalo = 0
        tempo_nao_manobra = 0
        ultima_data = None
        ultima_hora = None
        ultima_diferenca = 0  # Armazena a última diferença de tempo
        
        # Processar cada registro
        for idx, row in df_equip.iterrows():
            eh_manobra = row['Estado'] == 'MANOBRA'
            diferenca = row['Diferença_Hora'] * 3600  # Converter para segundos
            
            # Se não estamos em um intervalo
            if not em_intervalo:
                if eh_manobra:
                    # Iniciar novo intervalo
                    em_intervalo = True
                    inicio_intervalo = idx
                    soma_intervalo = diferenca if not pd.isna(diferenca) else 0
                    ultima_data = row['Data']
                    ultima_hora = row['Hora']
                    ultima_diferenca = diferenca
            
            # Se estamos em um intervalo
            else:
                if pd.isna(diferenca) or diferenca <= 0:
                    continue
                
                if eh_manobra:
                    # Resetar contador de tempo não-manobra
                    tempo_nao_manobra = 0
                    # Somar o tempo ao intervalo
                    soma_intervalo += diferenca
                else:
                    # Acumular tempo não-manobra
                    tempo_nao_manobra += diferenca
                    
                    # Se tempo não-manobra > 1 minuto, fechar intervalo
                    if tempo_nao_manobra > 60:
                        # Subtrair o último tempo que fechou o intervalo
                        soma_intervalo += (tempo_nao_manobra - diferenca)
                        
                        # Calcular todos os tempos em diferentes formatos
                        tempo_segundos = soma_intervalo
                        tempo_horas = soma_intervalo / 3600
                        tempo_formato_hora = tempo_horas / 24
                        
                        # Registrar intervalo
                        intervalos.append({
                            'Equipamento': equipamento,
                            'Data Início': df_equip.loc[inicio_intervalo, 'Data'],
                            'Hora Início': df_equip.loc[inicio_intervalo, 'Hora'],
                            'Data Fim': ultima_data,
                            'Hora Fim': ultima_hora,
                            'Tempo Total (s)': tempo_segundos,
                            'Tempo Total (s) Formatado': tempo_segundos / (24 * 3600),
                            'Tempo Total (h)': tempo_horas,
                            'Tempo Total (h) Formatado': tempo_formato_hora
                        })
                        
                        # Resetar controles
                        em_intervalo = False
                        inicio_intervalo = None
                        soma_intervalo = 0
                        tempo_nao_manobra = 0
                    else:
                        # Se não fechou o intervalo, soma normalmente
                        soma_intervalo += diferenca
                
                # Atualizar última entrada válida
                ultima_data = row['Data']
                ultima_hora = row['Hora']
                ultima_diferenca = diferenca
        
        # Fechar último intervalo se estiver aberto
        if em_intervalo and soma_intervalo > 0:
            # Calcular todos os tempos em diferentes formatos
            tempo_segundos = soma_intervalo
            tempo_horas = soma_intervalo / 3600
            tempo_formato_hora = tempo_horas / 24
            
            intervalos.append({
                'Equipamento': equipamento,
                'Data Início': df_equip.loc[inicio_intervalo, 'Data'],
                'Hora Início': df_equip.loc[inicio_intervalo, 'Hora'],
                'Data Fim': ultima_data,
                'Hora Fim': ultima_hora,
                'Tempo Total (s)': tempo_segundos,
                'Tempo Total (s) Formatado': tempo_segundos / (24 * 3600),
                'Tempo Total (h)': tempo_horas,
                'Tempo Total (h) Formatado': tempo_formato_hora
            })
    
    # Criar DataFrame com os intervalos
    if intervalos:
        df_intervalos = pd.DataFrame(intervalos)
        print(f"\nTotal de intervalos encontrados: {len(df_intervalos)}")
        
        # Calcular estatísticas
        stats = df_intervalos.groupby('Equipamento').agg({
            'Tempo Total (h)': ['count', 'sum', 'mean'],
            'Tempo Total (s)': ['min', 'max'],
            'Tempo Total (h) Formatado': ['sum', 'mean'],
            'Tempo Total (s) Formatado': ['sum', 'mean']
        }).round(6)
        
        stats.columns = [
            'Quantidade Intervalos',
            'Tempo Total (h)',
            'Tempo Médio (h)',
            'Menor Intervalo (s)',
            'Maior Intervalo (s)',
            'Tempo Total (h) Formatado',
            'Tempo Médio (h) Formatado',
            'Tempo Total (s) Formatado',
            'Tempo Médio (s) Formatado'
        ]
        
        return df_intervalos, stats.reset_index()
    
    return pd.DataFrame(), pd.DataFrame()

def calcular_estatisticas_agrupadas(df_intervalos, df_base):
    """
    Calcula estatísticas agrupadas por diferentes dimensões, incluindo médias de RPM e Velocidade
    """
    # Função auxiliar para calcular estatísticas
    def calc_stats(grupo):
        return pd.Series({
            'Quantidade Intervalos': len(grupo),
            'Tempo Total (h)': grupo['Tempo Total (h)'].sum(),
            'Tempo Médio (h)': grupo['Tempo Total (h)'].mean(),
            'Tempo Total (h) Formatado': grupo['Tempo Total (h) Formatado'].sum(),
            'Tempo Médio (h) Formatado': grupo['Tempo Total (h) Formatado'].mean(),
            'RPM Motor Médio': grupo['RPM Motor'].mean(),
            'Velocidade Média': grupo['Velocidade'].mean()
        })
    
    # Preparar DataFrame base com RPM e Velocidade
    df_completo = df_intervalos.copy()
    
    # Converter colunas numéricas
    def converter_para_numerico(valor):
        if isinstance(valor, str):
            try:
                # Remover referências múltiplas de vírgulas (problema nos dados)
                valor_limpo = valor.split(',')[0]
                return float(valor_limpo.replace(',', '.'))
            except:
                return 0.0
        elif pd.isna(valor):
            return 0.0
        return float(valor)
    
    # Converter RPM e Velocidade para numérico
    try:
        df_base['RPM Motor'] = df_base['RPM Motor'].apply(converter_para_numerico)
        df_base['Velocidade'] = df_base['Velocidade'].apply(converter_para_numerico)
    except Exception as e:
        print(f"AVISO: Erro ao converter colunas numéricas: {str(e)}")
    
    # Adicionar médias de RPM e Velocidade para cada intervalo
    for idx, row in df_intervalos.iterrows():
        mask = (
            (df_base['Equipamento'] == row['Equipamento']) &
            (df_base['Data'] >= row['Data Início']) &
            (df_base['Data'] <= row['Data Fim']) &
            (df_base['Hora'] >= row['Hora Início']) &
            (df_base['Hora'] <= row['Hora Fim'])
        )
        df_completo.loc[idx, 'RPM Motor'] = df_base.loc[mask, 'RPM Motor'].mean()
        df_completo.loc[idx, 'Velocidade'] = df_base.loc[mask, 'Velocidade'].mean()
    
    # Verificar se as colunas de agrupamento existem
    colunas_disponiveis = df_completo.columns.tolist()
    stats_equip = pd.DataFrame()
    stats_tipo = pd.DataFrame()
    stats_frente = pd.DataFrame()
    stats_operador = pd.DataFrame()
    
    # Calcular estatísticas por diferentes agrupamentos
    try:
        stats_equip = df_completo.groupby('Equipamento').apply(calc_stats).round(2)
    except Exception as e:
        print(f"AVISO: Não foi possível calcular estatísticas por Equipamento - {str(e)}")
    
    if 'Tipo de Equipamento' in colunas_disponiveis:
        try:
            stats_tipo = df_completo.groupby('Tipo de Equipamento').apply(calc_stats).round(2)
        except Exception as e:
            print(f"AVISO: Não foi possível calcular estatísticas por Tipo de Equipamento - {str(e)}")
    
    if 'Frente' in colunas_disponiveis:
        try:
            stats_frente = df_completo.groupby('Frente').apply(calc_stats).round(2)
        except Exception as e:
            print(f"AVISO: Não foi possível calcular estatísticas por Frente - {str(e)}")
    
    if 'Operador' in colunas_disponiveis:
        try:
            stats_operador = df_completo.groupby('Operador').apply(calc_stats).round(2)
        except Exception as e:
            print(f"AVISO: Não foi possível calcular estatísticas por Operador - {str(e)}")
    
    return stats_equip, stats_tipo, stats_frente, stats_operador

def preparar_dados_para_analise(df_intervalos, df_base):
    """
    Prepara os dados para análise, calculando médias de RPM e Velocidade para cada intervalo
    """
    print("\nPreparando dados para análise...")
    
    # Preparar DataFrame com intervales e dados operacionais
    df_completo = df_intervalos.copy()
    
    # Converter colunas numéricas
    def converter_para_numerico(valor):
        if isinstance(valor, str):
            try:
                # Remover referências múltiplas de vírgulas (problema nos dados)
                valor_limpo = valor.split(',')[0]
                return float(valor_limpo.replace(',', '.'))
            except:
                return 0.0
        elif pd.isna(valor):
            return 0.0
        return float(valor)
    
    # Converter RPM e Velocidade para numérico
    try:
        df_base['RPM Motor'] = df_base['RPM Motor'].apply(converter_para_numerico)
        df_base['Velocidade'] = df_base['Velocidade'].apply(converter_para_numerico)
    except Exception as e:
        print(f"AVISO: Erro ao converter colunas numéricas: {str(e)}")
    
    # Garantir que temos as colunas necessárias
    if 'Operador' not in df_base.columns:
        print("Adicionando coluna Operador (vazia)")
        df_base['Operador'] = 'Não Informado'
    
    # Usar coluna "Grupo" como informação de "Frente"
    if 'Grupo' not in df_base.columns:
        print("AVISO: Coluna 'Grupo' não encontrada, adicionando valor padrão")
        df_base['Grupo'] = 'Não Informado'
    else:
        print("Usando coluna 'Grupo' para informação de frentes")
    
    # Inicializar colunas no df_completo
    df_completo['Operador'] = 'Não Informado'
    df_completo['Frente'] = 'Não Informado'
    df_completo['RPM Motor'] = 0.0
    df_completo['Velocidade'] = 0.0
    
    # Adicionar colunas de Operador e Frente ao DataFrame de intervalos
    for idx, row in df_intervalos.iterrows():
        # Obter o operador e frente mais frequentes no intervalo
        mask = (
            (df_base['Equipamento'] == row['Equipamento']) &
            (df_base['Data'] >= row['Data Início']) &
            (df_base['Data'] <= row['Data Fim']) &
            (df_base['Hora'] >= row['Hora Início']) &
            (df_base['Hora'] <= row['Hora Fim'])
        )
        
        # Calcular médias de RPM e Velocidade para o intervalo
        df_completo.loc[idx, 'RPM Motor'] = df_base.loc[mask, 'RPM Motor'].mean()
        df_completo.loc[idx, 'Velocidade'] = df_base.loc[mask, 'Velocidade'].mean()
        
        # Obter operador e frente (grupo) mais frequentes
        if mask.any():
            operadores = df_base.loc[mask, 'Operador'].value_counts()
            frentes = df_base.loc[mask, 'Grupo'].value_counts()  # Usar Grupo como Frente
            
            df_completo.loc[idx, 'Operador'] = operadores.index[0] if not operadores.empty else 'Não Informado'
            df_completo.loc[idx, 'Frente'] = frentes.index[0] if not frentes.empty else 'Não Informado'
    
    return df_completo

def criar_planilha_por_frota(df_completo):
    """
    Cria análise por frota (equipamento)
    """
    try:
        # Agrupar por Equipamento
        stats = df_completo.groupby('Equipamento').agg({
            'Tempo Total (min)': ['count', 'mean'],
            'RPM Motor': 'mean',
            'Velocidade': 'mean'
        }).round(2)
        
        # Renomear colunas
        stats.columns = [
            'Quantidade Manobras',
            'Tempo Médio (min)',
            'RPM Motor Médio',
            'Velocidade Média'
        ]
        
        # Resetar índice para transformar Equipamento em coluna
        stats.reset_index(inplace=True)
        
        return stats
    except Exception as e:
        print(f"ERRO ao criar planilha por frota: {str(e)}")
        # Criar DataFrame vazio com colunas padrão
        return pd.DataFrame(columns=[
            'Equipamento', 'Quantidade Manobras', 'Tempo Médio (min)', 
            'RPM Motor Médio', 'Velocidade Média'
        ])

def criar_planilha_por_operador(df_completo):
    """
    Cria análise por operador
    """
    try:
        # Agrupar por Operador e Equipamento
        stats = df_completo.groupby(['Operador', 'Equipamento']).agg({
            'Tempo Total (min)': ['count', 'mean'],
            'RPM Motor': 'mean',
            'Velocidade': 'mean'
        }).round(2)
        
        # Renomear colunas
        stats.columns = [
            'Quantidade Manobras',
            'Tempo Médio (min)',
            'RPM Motor Médio',
            'Velocidade Média'
        ]
        
        # Resetar índice para transformar agrupamentos em colunas
        stats.reset_index(inplace=True)
        
        return stats
    except Exception as e:
        print(f"ERRO ao criar planilha por operador: {str(e)}")
        # Criar DataFrame vazio com colunas padrão
        return pd.DataFrame(columns=[
            'Operador', 'Equipamento', 'Quantidade Manobras', 'Tempo Médio (min)', 
            'RPM Motor Médio', 'Velocidade Média'
        ])

def criar_planilha_por_frente(df_completo):
    """
    Cria análise por frente
    """
    try:
        # Agrupar por Frente e Equipamento
        stats = df_completo.groupby(['Frente', 'Equipamento']).agg({
            'Tempo Total (min)': ['count', 'mean'],
            'RPM Motor': 'mean',
            'Velocidade': 'mean'
        }).round(2)
        
        # Renomear colunas
        stats.columns = [
            'Quantidade Manobras',
            'Tempo Médio (min)',
            'RPM Motor Médio',
            'Velocidade Média'
        ]
        
        # Resetar índice para transformar agrupamentos em colunas
        stats.reset_index(inplace=True)
        
        return stats
    except Exception as e:
        print(f"ERRO ao criar planilha por frente: {str(e)}")
        # Criar DataFrame vazio com colunas padrão
        return pd.DataFrame(columns=[
            'Frente', 'Equipamento', 'Quantidade Manobras', 'Tempo Médio (min)', 
            'RPM Motor Médio', 'Velocidade Média'
        ])

def formatar_excel_tabela_dinamica(worksheet, df):
    """
    Aplica formatação estilo de tabela dinâmica a uma planilha
    """
    # Aplicar formatação básica
    formatar_planilha(worksheet, df)
    
    # Obter as colunas de agrupamento (primeiras colunas antes de 'Quantidade')
    colunas_grupo = []
    for col in df.columns:
        if col == 'Quantidade Manobras':
            break
        colunas_grupo.append(col)
    
    if len(colunas_grupo) >= 2:
        # Temos pelo menos duas colunas de agrupamento, podemos formatar como tabela dinâmica
        col1_idx = df.columns.get_loc(colunas_grupo[0]) + 1
        
        # Definir cores alternadas para grupos
        cores_grupo = ['E2EFDA', 'D9E1F2']  # Verde claro e Azul claro
        cor_atual = 0
        valor_anterior = None
        
        # Aplicar cores de fundo para grupos
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1)):
            valor_atual = worksheet.cell(row=row_idx+2, column=col1_idx).value
            
            # Mudar cor quando mudar o grupo
            if valor_atual != valor_anterior:
                cor_atual = (cor_atual + 1) % 2
                valor_anterior = valor_atual
            
            # Aplicar cor de fundo na linha
            for cell in row:
                cell.fill = PatternFill(start_color=cores_grupo[cor_atual], 
                                       end_color=cores_grupo[cor_atual], 
                                       fill_type='solid')

def salvar_excel(df, caminho_saida):
    """
    Salva o DataFrame processado como Excel com as planilhas.
    """
    print(f"\nSalvando resultado em: {caminho_saida}")
    
    # Definir colunas para manter
    colunas_manter = [
        'Equipamento',
        'Data',
        'Hora',
        'Estado',
        'Grupo',
        'Tipo de Equipamento',
        'RPM Motor',
        'Velocidade',
        'Diferença_Hora',
        'Operador',
        'Frente'
    ]
    
    # Verificar se todas as colunas necessárias existem
    colunas_faltando = [col for col in colunas_manter if col not in df.columns]
    if colunas_faltando:
        print(f"AVISO: Colunas faltando no DataFrame: {colunas_faltando}")
        colunas_manter = [col for col in colunas_manter if col in df.columns]
    
    try:
        # Calcular intervalos de manobra
        df_intervalos, df_stats = calcular_intervalos_manobra(df)
        
        if df_intervalos.empty:
            print("Nenhum intervalo encontrado para salvar!")
            return False
        
        # Converter tempos para minutos para visualização mais fácil
        df_intervalos['Tempo Total (min)'] = df_intervalos['Tempo Total (s)'] / 60
        df_intervalos['Tempo Total (min) Formatado'] = df_intervalos['Tempo Total (min)'] / (24 * 60)
        
        # Arredondar para 2 casas decimais
        df_intervalos = df_intervalos.round({
            'Tempo Total (h)': 4,
            'Tempo Total (min)': 2,
        })
        
        # Preparar dados para análises
        df_completo = preparar_dados_para_analise(df_intervalos, df)
        
        # Criar análises específicas
        df_por_frota = criar_planilha_por_frota(df_completo)
        df_por_operador = criar_planilha_por_operador(df_completo)
        df_por_frente = criar_planilha_por_frente(df_completo)
        
        # Criar Excel writer
        with pd.ExcelWriter(caminho_saida, engine='openpyxl', mode='w') as writer:
            # Salvar planilha BASE (todas as colunas)
            df.to_excel(writer, sheet_name='BASE', index=False)
            
            # Criar e salvar planilha Tratamento (apenas colunas necessárias)
            df_tratamento = df[colunas_manter].copy()
            df_tratamento.to_excel(writer, sheet_name='Tratamento', index=False)
            
            # Salvar planilha de Intervalos
            df_intervalos.to_excel(writer, sheet_name='Intervalos Manobra', index=False)
            df_stats.to_excel(writer, sheet_name='Resumo Intervalos', index=False)
            
            # Salvar novas planilhas de análise
            df_por_frota.to_excel(writer, sheet_name='Por Frota', index=False)
            df_por_operador.to_excel(writer, sheet_name='Por Operador', index=False)
            df_por_frente.to_excel(writer, sheet_name='Por Frente', index=False)
            
            # Formatar planilhas
            workbook = writer.book
            
            # Formatar planilhas básicas
            formatar_planilha(writer.sheets['BASE'], df)
            formatar_planilha(writer.sheets['Tratamento'], df_tratamento)
            formatar_planilha(writer.sheets['Intervalos Manobra'], df_intervalos)
            formatar_planilha(writer.sheets['Resumo Intervalos'], df_stats)
            
            # Formatar planilhas de análise
            formatar_planilha(writer.sheets['Por Frota'], df_por_frota)
            formatar_excel_tabela_dinamica(writer.sheets['Por Operador'], df_por_operador)
            formatar_excel_tabela_dinamica(writer.sheets['Por Frente'], df_por_frente)
            
            # Adicionar filtros nas planilhas de análise
            writer.sheets['Por Frota'].auto_filter.ref = f"A1:{get_column_letter(len(df_por_frota.columns))}{len(df_por_frota)+1}"
            writer.sheets['Por Operador'].auto_filter.ref = f"A1:{get_column_letter(len(df_por_operador.columns))}{len(df_por_operador)+1}"
            writer.sheets['Por Frente'].auto_filter.ref = f"A1:{get_column_letter(len(df_por_frente.columns))}{len(df_por_frente)+1}"
        
        print(f"Arquivo Excel salvo com sucesso: {caminho_saida}")
        return True
    
    except PermissionError:
        print(f"ERRO: Não foi possível salvar o arquivo. Verifique se ele está aberto em outro programa.")
        return False
    except Exception as e:
        print(f"ERRO ao salvar arquivo: {str(e)}")
        return False

def formatar_planilha(worksheet, df):
    """
    Aplica formatação padrão a uma planilha.
    """
    # Ajustar largura das colunas
    for idx, col in enumerate(df.columns):
        max_length = max(
            df[col].astype(str).apply(len).max(),
            len(str(col))
        ) + 2
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)
    
    # Formatar cabeçalho
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Formatar colunas específicas
    for col in df.columns:
        if col in df.columns:
            col_idx = df.columns.get_loc(col) + 1
            
            # Formatar datas
            if col == 'Data' or col.endswith('Data Início') or col.endswith('Data Fim'):
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = 'dd/mm/yyyy'
            
            # Formatar horas
            elif col == 'Hora' or col.endswith('Hora Início') or col.endswith('Hora Fim'):
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = 'hh:mm:ss'
            
            # Formatar diferença de tempo
            elif col == 'Diferença_Hora':
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = '0.000000'
            
            # Formatar tempo em segundos
            elif col.endswith('(s)') and not col.endswith('Formatado'):
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = '0.00'
            
            # Formatar tempo em horas
            elif col.endswith('(h)') and not col.endswith('Formatado'):
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = '0.0000'
            
            # Formatar tempo em minutos
            elif col.endswith('(min)') and not col.endswith('Formatado'):
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = '0.00'
            
            # Formatar campos formatados para exibição de hora
            elif 'Formatado' in col:
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if col.endswith('(min) Formatado'):
                        cell.number_format = '[m]:ss'  # Formato de minutos:segundos
                    else:
                        cell.number_format = '[h]:mm:ss'  # Formato de hora:minutos:segundos

def main():
    print("\nIniciando processamento de arquivos de manobras...")
    
    # Obter diretório de dados
    script_dir = os.path.dirname(os.path.abspath(__file__))
    workspace_dir = os.path.dirname(script_dir)
    dados_dir = os.path.join(workspace_dir, "dados")
    output_dir = os.path.join(workspace_dir, "output")
    
    # Criar diretório de saída se não existir
    os.makedirs(output_dir, exist_ok=True)
    
    # Encontrar arquivos ZIP de manobras
    arquivos_zip = encontrar_arquivos_zip_manobras(dados_dir)
    
    if not arquivos_zip:
        print("Nenhum arquivo ZIP de manobras encontrado!")
        return
    
    print(f"Arquivos ZIP encontrados: {len(arquivos_zip)}")
    
    # Processar cada arquivo ZIP
    for arquivo_zip in arquivos_zip:
        try:
            # Processar o ZIP
            df = processar_csv_do_zip(arquivo_zip)
            
            if df is not None:
                # Gerar nome do arquivo de saída (mesmo nome do ZIP)
                nome_arquivo = os.path.basename(arquivo_zip)
                nome_base = os.path.splitext(nome_arquivo)[0]
                # Adicionar um timestamp ao nome do arquivo para evitar conflitos
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                arquivo_saida = os.path.join(output_dir, f"{nome_base}_{timestamp}.xlsx")
                
                # Salvar resultado
                salvar_excel(df, arquivo_saida)
                print(f"Arquivo processado com sucesso: {arquivo_saida}")
            else:
                print(f"Nenhum dado válido encontrado em: {arquivo_zip}")
        
        except Exception as e:
            print(f"Erro ao processar {arquivo_zip}: {str(e)}")
            continue

if __name__ == "__main__":
    print("=" * 60)
    print("PROCESSADOR DE MANOBRAS")
    print("=" * 60)
    main()
    print("\nProcessamento concluído!") 